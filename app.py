import configparser
from datetime import datetime, timedelta
from enum import Enum
import inspect
import json
from logging.handlers import RotatingFileHandler
import math
import os
import re
import threading
from threading import Event, Timer
import time
import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from tkinter import font
from typing import Optional
from collections.abc import Callable

import requests

import bcrypt
from cryptography.fernet import Fernet
from pal.core.audit import AuditLogger
from pal.core.auth import AuthManager
from pal.core.credentials import SecureCredentialsManager
from pal.core.errors import ErrorCode
from pal.core.license import LicenseChecker, LicenseError
from pal.core.log import set_component_level, set_log_callback
from pal.core.session import SessionManager
from pal.core.updater import UpdateManager
UPDATE_MANAGER_AVAILABLE = True
from pal.infrastructure.database import DatabaseManager
from pal.infrastructure.notification_db_backend import PyodbcNotificationBackend
from pal.services.cache import CacheDescripciones
from pal.services.envios import EnvioProgramado, ProgramadorEnvios
from pal.services.notifications import NotificationManager as CentralNotificationManager, NotificationPriority
from pal.ui.debug_console import DebugConsole
from pal.ui.header import create_header, setup_styles as ui_setup_styles, NotificationBell
from pal.ui.splash import SplashScreen
from pal.ui.clientes_menu import ClientesMenu
from tkcalendar import Calendar, DateEntry
from pal.core.config_manager import ConfigManager


CONFIG_FILE = 'db_config.ini'
LICENSE_CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTdAdOg6pI7tOF-9UdFDzw0P5aSpNRc-jGIYHwOHmXb7qqOtag9QTYAi4JU0U2VoIZLd_TjvK_7cxX9/pub?output=csv"
LICENSE_CLIENT_NAME = "PALPY"
APP_VERSION = "1.6.5" # Versión actual de la aplicación
UPDATE_URL_DEFAULT = "https://raw.githubusercontent.com/MrBearCr/nexus/main/updates"  # URL base por defecto para actualizaciones (formato raw)

def load_update_url():
    """Carga la URL de actualizaciones desde la configuración."""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if 'Updates' in config and 'url' in config['Updates']:
            return config['Updates']['url']
    return UPDATE_URL_DEFAULT

def save_update_url(url: str):
    """Guarda la URL de actualizaciones en la configuración."""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
    if 'Updates' not in config:
        config.add_section('Updates')
    config['Updates']['url'] = url
    with open(CONFIG_FILE, 'w') as f:
        config.write(f)

def load_last_update_check() -> Optional[datetime]:
    """Carga la fecha de la última comprobación de actualizaciones."""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if 'Updates' in config and 'last_check' in config['Updates']:
            try:
                return datetime.fromisoformat(config['Updates']['last_check'])
            except Exception:
                return None
    return None

def save_last_update_check(dt: datetime):
    """Guarda la fecha de la última comprobación de actualizaciones."""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
    if 'Updates' not in config:
        config.add_section('Updates')
    config['Updates']['last_check'] = dt.isoformat()
    with open(CONFIG_FILE, 'w') as f:
        config.write(f)
JERARQUIA_CACHE_FILE = "productos_jerarquia_cache.json"
FAVORITOS_CACHE_FILE = 'favoritos_cache.json'
JERARQUIA_CACHE_TTL = timedelta(hours=15)
LOCATION_GROUPS = {
    'BARINAS': ['0101', '0108'],
    'GUANARE': ['0401', '0402'],
    'CDT': ['0106'],
}

# Mapeo de módulos entre BD (mayúsculas) y flags de la app (minúsculas)
DB_MODULE_TO_FLAG = {
    'STOCK': 'stock',
    'TRA': 'tra',
    'MBRP': 'mbrp',
    'MENSAJES': 'envio_mensajes',
    'ESTADISTICAS': 'estadisticas',
    'CALENDARIO': 'calendario',
    'ADMIN': 'admin',
    'CLIENTES': 'clientes',
    'LOGISTICA': 'logistica',
}
FLAG_TO_DB_MODULE = {v: k for k, v in DB_MODULE_TO_FLAG.items()}


# === Configuración de Depuración por Módulo ===

def load_license_key():
        """Lee y desencripta la License Key desde [License] de db_config.ini.

        Si el valor no puede desencriptarse, se asume que está en texto plano
        y se devuelve tal cual (para compatibilidad hacia atrás).
        """
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)
        if 'License' not in config:
            return None
        enc_key = config['License'].get('key', '').strip()
        if not enc_key:
            return None
        try:
            mgr = SecureCredentialsManager()
            return mgr.decrypt(enc_key)
        except Exception:
            # Compatibilidad: si no se puede desencriptar, devolver el valor crudo
            return enc_key or None

def save_license_key(key: str):
        """Encripta y guarda la License Key en la sección [License] de db_config.ini."""
        key = (key or '').strip()
        if not key:
            return
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)
        if 'License' not in config:
            config.add_section('License')
        try:
            mgr = SecureCredentialsManager()
            enc_key = mgr.encrypt(key)
        except Exception:
            # Si falla la encriptación, guardar en texto plano como último recurso
            enc_key = key
        config['License']['key'] = enc_key
        with open(CONFIG_FILE, 'w') as f:
            config.write(f)

def load_debug_config():
        """Lee la sección [Debug] de db_config.ini y devuelve flags por módulo."""
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)
        if 'Debug' not in config:
            # Valores por defecto: deshabilitado
            config['Debug'] = {
                'tra': 'False',
                'stock': 'False',
                'mbrp': 'False',
                'db': 'False',
            }
            with open(CONFIG_FILE, 'w') as f:
                config.write(f)
        flags = {}
        for key in config['Debug']:
            flags[key] = config.getboolean('Debug', key, fallback=False)
        return flags

def save_debug_config(flags: dict):
        """Guarda banderas de depuración en la sección [Debug]."""
        config = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            config.read(CONFIG_FILE)
        if 'Debug' not in config:
            config.add_section('Debug')
        for key, val in flags.items():
            config['Debug'][key] = 'True' if val else 'False'
        with open(CONFIG_FILE, 'w') as f:
            config.write(f)




    




class DatabaseApp:

    def _ensure_license_with_prompt(self) -> bool:
        """Valida la licencia usando LicenseKey local o pidiéndola al usuario.

        Flujo:
        - Intenta leer LicenseKey desde db_config.ini.
        - Si no existe o la validación falla, pide al usuario que ingrese la licencia.
        - Permite varios intentos; si todos fallan o se cancela, devuelve False.
        """
        from tkinter import simpledialog, messagebox

        # Intentar leer LicenseKey guardada
        current_key = load_license_key()

        # Máximo de intentos interactivos (además del intento con la key guardada)
        max_attempts = 3
        attempt = 0

        while True:
            license_key = (current_key or '').strip()
            if not license_key:
                # Pedir al usuario que ingrese la licencia
                license_key = simpledialog.askstring(
                    "Licencia",
                    "Ingrese la License Key proporcionada:",
                    parent=self.root,
                    show='*',
                )
                if not license_key:
                    # Cancelado o vacío
                    messagebox.showerror("Licencia requerida", "Debe ingresar una License Key válida para continuar.")
                    return False

            # Validar contra el servidor (con 7 días de gracia por cache)
            try:
                checker = LicenseChecker(LICENSE_CSV_URL, LICENSE_CLIENT_NAME)
                checker.ensure_valid(license_key=license_key, allow_cached_days=7)
                # OK: guardar y continuar
                save_license_key(license_key)
                print("[DEBUG] Licencia PALPY validada correctamente", flush=True)
                return True
            except LicenseError as e:
                attempt += 1
                current_key = None  # Forzar nueva entrada en el siguiente ciclo
                messagebox.showerror("Licencia inválida", str(e))
                if attempt >= max_attempts:
                    messagebox.showerror(
                        "Límite de intentos",
                        "Se alcanzó el número máximo de intentos de validación de licencia."
                    )
                    return False
    def __init__(self, root):
        self.root = root
        self.root.withdraw()  # Ocultar ventana principal

        # Mostrar splash screen
        self.splash = SplashScreen(self.root)
        self.splash.start_animation()

        # Validar licencia ANTES de inicializar el resto de la app
        if not self._ensure_license_with_prompt():
            # Si falla la licencia o el usuario cancela, cerramos todo
            try:
                if self.splash:
                    self.splash.destroy()
            except Exception:
                pass
            try:
                self.root.destroy()
            except Exception:
                pass
            return
        
        # Iniciar inicialización en segundo plano
        threading.Thread(target=self._initialize_app, daemon=True).start()

    def _show_initial_settings(self):
        """Muestra diálogo de configuración si no hay BD inicial y cierra el splash.
        Cierra el splash marcando login_success para no bloquear la UI de configuración.
        """
        try:
            if self.root.state() == 'withdrawn':
                self.root.deiconify()
            self.show_settings()
            # Permitir que el splash se cierre aunque no haya login todavía
            try:
                if hasattr(self, 'splash') and self.splash and hasattr(self.splash, 'login_success'):
                    self.splash.login_success.set()
            except Exception:
                pass
        except Exception as e:
            print(f"[ERROR] No se pudo mostrar diálogo de configuración: {e}")

    def shutdown(self):
        """Cierra la aplicación de forma ordenada."""
        self.log("Iniciando cierre ordenado...", "INFO")
        try:
            # Detener hilos y procesos en segundo plano
            if hasattr(self, 'update_manager') and self.update_manager:
                self.update_manager.stop_periodic_check()

            if hasattr(self, 'programador') and hasattr(self.programador, 'detener'):
                self.programador.detener()
            
            # Close database connections from the pool
            if hasattr(self, 'db_manager') and hasattr(self.db_manager, 'close_thread_connections'):
                self.db_manager.close_thread_connections()

            self.log("Hilos de fondo y conexiones de BD detenidos. Cerrando ventana principal...", "INFO")
            
            try:
                self.root.quit()
                self.root.destroy()
            except Exception:
                pass
            
            self.log("Cierre de UI completado.", "INFO")
            # Force exit to ensure all daemon/non-daemon threads are killed
            os._exit(0)
        except Exception as e:
            self.log(f"Error durante el cierre ordenado: {e}", "ERROR")
            # Fallback to a hard exit in case of error during shutdown
            os._exit(1)

    def _initialize_app(self):
        try:
            # Aplicar estilos modernos inmediatamente en el hilo principal
            self.root.after(0, lambda: ui_setup_styles(self))
            
            # Tu lógica de inicialización original
            self.ultimas_notificaciones = set()
            print("[DEBUG] Iniciando carga de la aplicación...", flush=True)
            
            # Reportar progreso: 10%
            self.splash.set_progress(0.10)

            # Inicialización de componentes críticos
            self.cred_manager = SecureCredentialsManager()
            self.enviando = False
            self.session = SessionManager(self.root)
            self.session.start_session()
            
            # Reportar progreso: 20%
            self.splash.set_progress(0.20)
            
            self.modules_enabled = {} # Se cargará desde la BD después del login
            self.audit_log = AuditLogger()
            self.db_manager = DatabaseManager(self.cred_manager)
            self.config_manager = ConfigManager(self.db_manager)
            self.auth = None
            self.permissions = None
            self.current_user = None
            self.session_token = None
            self.settings_window = None
            self.show_pwd_var = None
            self.httpd = None
            
            # Inicializar referencias a widgets de estado para evitar AttributeError
            self.db_status = None
            self.api_status = None
            
            self.favoritos = set()
            self._load_favoritos_cache()
            
            # Inicialización temprana de atributos de paginación
            self.page_size = 250
            self.current_page = 1
            self.current_filter = 'TODAS'
            self.cached_alertas = []
            self.last_refresh = None
         # Carga paralela de stock
            self.stock_full_loading_started = False
            self._stock_loading_in_progress = False
            self.new_stock_break_codes = set()

            self.tra_page_size = 500
            self.tra_current_page = 1
            
            # Configuración de UI y bindings
            self.buttons = {}    
            self.root.protocol("WM_DELETE_WINDOW", self.shutdown)
            
            # Reportar progreso: 30%
            self.splash.set_progress(0.30)
            
            # setup_modern_ui() <- Diferido para después del login (acelera inicio)
            
            # Reportar progreso: 40%
            self.splash.set_progress(0.40)
            # Deshabilitar UI dependiente de BD hasta conectar
            try:
                self._set_ui_connected(False)
            except Exception:
                pass
            self.setup_bindings()
            self.cache = CacheDescripciones()
            
            # Consola de debug flotante (para desarrolladores)
            self.debug_console = DebugConsole(self)
            
            # Flags de carga para evitar duplicados
            self.jerarquias_unificadas_cargadas = False
            
            # Cargar configuración de depuración por módulo
            self.debug_flags = load_debug_config()
            self.tra_debug = self.debug_flags.get('tra', False)
            self.stock_debug = self.debug_flags.get('stock', False)
            self.mbrp_debug = self.debug_flags.get('mbrp', False)
            # Configurar niveles de log por componente para reducir ruido
            try:
                set_component_level("TRA", "DEBUG" if self.tra_debug else "WARNING")
                set_component_level("STOCK", "DEBUG" if self.stock_debug else "WARNING")
                set_component_level("MBRP", "DEBUG" if self.mbrp_debug else "WARNING")
                set_component_level("DB", "DEBUG" if self.debug_flags.get('db', False) else "INFO")
                # Redirigir todos los logs de servicios (pal.core.log) a la UI/consola de la app
                set_log_callback(self.log)
            except Exception:
                pass
            try:
                # Habilitar/Deshabilitar debug en el gestor de BD
                setattr(self.db_manager, 'debug_enabled', self.debug_flags.get('db', False))
            except Exception:
                pass

                # Inicialización diferida post-login para monitoreo
                
                # Inicialización de variables para carga paralela TRA
                self.cached_ventas_tra = []
                self.tra_last_refresh = None
                self.tra_full_loading_started = False
                self.tra_loader_thread = None  # Referencia al hilo de carga TRA
                self.tra_total_neto_scaneado = 0.0
                self.tra_fecha_inicio = None
                self.tra_fecha_fin = None
                self.tra_sede_codigo = None
            
            # MBRP - Movimiento de Baja Rotación de Producto
                # Inicialización diferida post-login para MBRP
            
            # Sistema de Paginacion ya inicializado arriba
            # Sistema de notificaciones (inicializar antes de auto-connect)
            # Usar CentralNotificationManager desde el inicio (sin backend BD aún)
            self.notification_manager = CentralNotificationManager()
            self.help_tooltips = self.HelpTooltips(self.root)
            # Tooltips y update manager se inicializarán después del login para acelerar inicio
            self.update_manager = None
            
            # Reportar progreso: 50%
            self.splash.set_progress(0.50)
            
            # NO habilitar login aquí - esperar a que BD conecte
            # self.root.after(0, lambda: self._enable_early_login())
            
            # Conectar BD (esto habilitará el login cuando tenga éxito)
            # Reportar progreso: 60% al iniciar conexión
            self.splash.set_progress(0.60)
            self.attempt_auto_connect()
            
            # Cargar configuraciones globales (exclusiones)
            try:
                self._load_global_settings()
            except Exception:
                pass

            # Diferir actualizaciones de stock hasta después del login
            # self.programar_actualizaciones_stock()
            
            # Verificar hilos activos en segundo plano
            self.listar_hilos_activos()
            
            print("[DEBUG] Inicialización completada exitosamente", flush=True)
            
        except Exception as e:
            print(f"[ERROR] Fallo durante la inicialización: {e}", flush=True)
            import traceback
            traceback.print_exc()
            
        finally:
            # Marcar inicialización como completada siempre
            self.splash.app_initialized.set()
            # Si no hay conexión BD y no hay login configurado, mostrar settings después del splash
            if not hasattr(self, 'auth') or not self.auth:
                self.root.after(1000, self._show_initial_settings)
            
            # No deiconificar la ventana principal aquí.
            # La ventana principal se mostrará cuando el splash complete login exitoso.

    def _inicializacion_completa(self):
        # Destruir el splash screen
        self.splash.destroy()
    
        # Mostrar ventana principal
        self.root.deiconify()

    def _load_favoritos_cache(self):
        """Carga el archivo JSON de favoritos si existe"""
        try:
            if os.path.exists(FAVORITOS_CACHE_FILE):
                with open(FAVORITOS_CACHE_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self.favoritos = set(data)
            else:
                self.favoritos = set()
        except Exception:
            self.favoritos = set()

    def _save_favoritos_cache(self):
        """Guarda el set de favoritos al archivo JSON"""
        try:
            with open(FAVORITOS_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(list(self.favoritos), f, ensure_ascii=False)
        except Exception:
            pass
    
    def _load_saved_theme(self):
        """Carga y aplica el tema guardado desde archivo local"""
        try:
            from pal.ui.themes import load_saved_theme
            if hasattr(self, 'style') and self.style:
                load_saved_theme(self)
        except Exception as e:
            print(f"Error cargando tema: {e}")

    def _toggle_favorito_local(self, codigo):
        """Alterna un código en el set de favoritos y lo cachea"""
        if codigo in self.favoritos:
            self.favoritos.remove(codigo)
        else:
            self.favoritos.add(codigo)
        self._save_favoritos_cache()
        return True
    
    def _check_product_stock_alert(self, codigo):
        """Verifica si un producto tiene alerta de stock y devuelve el nivel de alerta.
        
        Args:
            codigo (str): Código del producto a verificar
            
        Returns:
            str or None: Nivel de alerta ('Crítica', 'Media', 'Leve') o None si no tiene alerta
        """
        try:
            # Si no hay alertas cacheadas, retornar None
            if not hasattr(self, 'cached_alertas') or not self.cached_alertas:
                return None
            
            # Buscar el producto en las alertas de stock
            codigo_str = str(codigo).strip()
            for alerta_codigo, _, _, nivel in self.cached_alertas:
                if str(alerta_codigo).strip() == codigo_str:
                    return nivel
            
            return None
        except Exception as e:
            self.tra_debug_log(f"Error verificando alerta de stock para {codigo}: {e}")
            return None

    def _get_favoritos_local(self):
        """Devuelve el set de códigos favoritos"""
        return set(self.favoritos)

    def obtener_descripcion_producto(self, codigo: str) -> Optional[str]:
        """Obtiene la descripción de un producto desde la base de datos."""
        try:
            result = self.db_manager.fetch_data(
                "SELECT COALESCE(cu_descripcion_corta, 'SIN DESCRIPCIÓN') FROM MA_PRODUCTOS WHERE C_CODIGO = ?", 
                (codigo,)
            )
            # Devolver cadena directamente, sin formateo adicional
            return str(result[0][0]) if result and result[0][0] else None
        except Exception as e:
            self.log(f"Error obteniendo descripción: {str(e)}", "ERROR")
            return None
        
    def validar_stock_producto(self, codigo: str) -> bool:
        """Valida si un producto tiene stock en el depósito 0301.
        Mantiene la funcionalidad exacta de recup.py
        """
        try:
            result = self.db_manager.fetch_data(
                "SELECT n_cantidad FROM MA_DEPOPROD WHERE c_codarticulo = ? AND c_coddeposito = '0301'",
                (codigo,)
            )
            return result and result[0][0] > 0
        except Exception as e:
            self.log(f"Error validando stock: {str(e)}", "ERROR")
            return False
        
    def obtener_codigo_producto_cliente(self, numero_cliente: str) -> Optional[str]:
        """Obtiene el código de producto asociado a un cliente."""
        try:
            result = self.db_manager.fetch_data(
                "SELECT C_CODIGO FROM pal_clientes WHERE numero_cliente = ?", 
                (numero_cliente,)
            )
            return result[0][0] if result else None
        except Exception as e:
            self.log(f"Error obteniendo código de producto: {str(e)}", "ERROR")
            return None
    
    def actualizar_alertas_stock(self, force_refresh=False):
        """Actualiza las alertas de stock respetando permisos y filtros de rotación"""
        if not self.modules_enabled.get("stock", False):
            if hasattr(self, 'alerts_display_label') and self.alerts_display_label.winfo_exists():
                self.alerts_display_label.pack_forget()
            return

        try:
            if not self.db_manager.ensure_connection():
                return
                
            refresh_needed = force_refresh or not self.last_refresh or (datetime.now() - self.last_refresh).seconds > 1800
        
            if refresh_needed:
                # Obtener parámetros de la UI
                # CONFIGURACIÓN FIJA: 30 DÍAS PARA ANÁLISIS DE QUIEBRES
                dias_context = 30
                try:
                    solo_alta = self.stock_solo_alta_var.get()
                except Exception:
                    solo_alta = True

                # Obtener depósitos tratables
                sedes_config = self.config_manager.get_sedes_config()
                
                quiebres_totales = []
                for sede_name, config in sedes_config.items():
                    depositos = config.get('almacenes_tratables', [])
                    if not depositos: continue
                    
                    # Extraer código de sede
                    cod_sede = config.get('codigo_sede')
                    if not cod_sede:
                        if " - " in sede_name:
                            cod_sede = sede_name.split(" - ")[0]
                        else:
                            cod_sede = depositos[0] if depositos else sede_name
                    
                    # EXCLUSIÓN TOTAL DE ICH: Solo procesar sedes físicas reales
                    if cod_sede in ('00', 'ICH', '%', 'ALL'):
                        continue
                    
                    self.stock_debug_log(f"Consultando quiebres para {sede_name} (Cod: {cod_sede}, Rango: {dias_context}d, SoloAlta: {solo_alta})")
                    
                    # Carga de quiebres con contexto de rotación normalizado
                    quiebres = self.db_manager.obtener_quiebres_directos(
                        depositos, 
                        solo_alta_rotacion=solo_alta,
                        sede_context=cod_sede,
                        dias_context=dias_context,
                        nombre_sede_display=sede_name,
                        exclude_depts=getattr(self, 'excluded_depts', [])
                    )
                    if quiebres:
                        quiebres_totales.extend(quiebres)
                
                # Detectar nuevos quiebres comparando con los códigos anteriores (si existen)
                new_codes = {str(q['codigo']).strip() for q in quiebres_totales}
                if hasattr(self, 'cached_alertas') and self.cached_alertas:
                    old_codes = {str(q[0]).strip() for q in self.cached_alertas}
                    self.new_stock_break_codes = new_codes - old_codes
                else:
                    self.new_stock_break_codes = set()

                # Convertir a tuplas para compatibilidad
                self.cached_alertas = [
                    (q['codigo'], q['descripcion'], q['sede'], q['unidades_perdidas'], q['dias_quiebre'], q['ultima_compra'], q['ultima_venta'])
                    for q in quiebres_totales
                ]
                self.last_refresh = datetime.now()
                self._rebuild_effective_views()
                self.root.after(0, self.aplicar_filtro_stock)
                
                # POPUP AUTOMÁTICO: Informar si hay quiebres críticos detectados
                if self.cached_alertas and len(self.cached_alertas) > 0:
                    # Mostrar alerta en el display de alertas
                    if hasattr(self, 'alerts_display_label') and self.alerts_display_label.winfo_exists():
                        self.alerts_display_label.config(text="🚨 Alerta: Quiebres de stock")
                        self.alerts_display_label.pack(side=tk.LEFT, padx=10)
                        self.toggle_stock_blink()

                    # Agregar notificación al Centro de Notificaciones
                    if hasattr(self, 'notification_manager') and self.notification_manager:
                        try:
                            self.notification_manager.add_notification(
                                title="🚨 Alerta de Quiebres de Stock",
                                message=f"Se detectaron {len(self.cached_alertas)} productos en quiebre de stock. Revisa los detalles para más información.",
                                priority=NotificationPriority.URGENT,
                                module="STOCK",
                                modulo_ruta="stock",
                                accion_etiqueta="Ver Quiebres"
                            )
                            # Actualizar display de alertas
                            self.root.after(100, self._update_alerts_display)
                        except Exception as e:
                            self.log(f"Error agregando notificación de quiebres: {e}", "WARNING")

                    # Lógica de revisión: Solo abrir popup si los datos cambiaron (ID diferente)
                    import hashlib
                    codigos_actuales = "".join(sorted([str(q[0]) for q in self.cached_alertas]))
                    current_id = hashlib.md5(codigos_actuales.encode()).hexdigest()
                    
                    last_id = getattr(self, 'quiebre_revisado_id', None)
                    if current_id != last_id:
                        self.abrir_popup_quiebres()
                    else:
                        self.log("ℹ️ Quiebres detectados pero ya fueron revisados previamente.", "DEBUG")
                else:
                    # Ocultar alerta si no hay quiebres
                    if hasattr(self, 'alerts_display_label') and self.alerts_display_label.winfo_exists():
                        self.alerts_display_label.pack_forget()
                
                self.log(f"Quiebres de stock actualizados automáticamente ({len(self.cached_alertas)} encontrados)", "INFO")

        except Exception as e:
            self.log(f"Error al actualizar alertas: {e}", "ERROR")
    
    def abrir_popup_quiebres(self):
        """Abre el popup de quiebres manualmente desde la alerta o el menú"""
        if not self.modules_enabled.get("stock", False):
            return

        if hasattr(self, 'cached_alertas') and self.cached_alertas:
            from pal.ui.popups import StockBreakPopup
            # Pasar la instancia de la app para poder llamar a marcar_revisado
            popup = StockBreakPopup(self.root, self.cached_alertas)
            # Modificar el botón del popup para que llame a la revisión
            if hasattr(popup, 'top'):
                # Buscamos el botón de cierre para inyectar la lógica de revisión
                for widget in popup.top.winfo_children():
                    if isinstance(widget, tk.Frame): # El footer
                        for btn in widget.winfo_children():
                            if isinstance(btn, tk.Button) and "Entendido" in btn.cget("text"):
                                btn.configure(command=lambda: self.marcar_quiebres_como_revisados(popup))

    def marcar_quiebres_como_revisados(self, popup_instance):
        """Marca los quiebres actuales como revisados para no molestar con el popup"""
        import hashlib
        if self.cached_alertas:
            # Creamos un ID único basado en los códigos de productos en quiebre
            codigos = "".join(sorted([str(q[0]) for q in self.cached_alertas]))
            self.quiebre_revisado_id = hashlib.md5(codigos.encode()).hexdigest()
            self.log("✅ Quiebres marcados como revisados. El popup no aparecerá hasta que cambien los datos.", "INFO")
        
        if popup_instance:
            popup_instance.top.destroy()

    def toggle_stock_blink(self):
        """Efecto de parpadeo para la alerta de stock en la barra de estado"""
        if not hasattr(self, 'alerts_display_label') or not self.alerts_display_label.winfo_exists():
            return

        current_color = self.alerts_display_label.cget("foreground")
        # El color de fondo para "apagar" el texto
        bg_color = self.root.cget("background") 
        
        # Alternar entre rojo y el color de fondo (efecto desaparecer)
        next_color = bg_color if current_color == "#D32F2F" else "#D32F2F"
        self.alerts_display_label.configure(foreground=next_color)
        
        # Solo seguir parpadeando si hay quiebres y la etiqueta es visible
        if hasattr(self, 'cached_alertas') and self.cached_alertas and self.alerts_display_label.winfo_ismapped():
            self.root.after(600, self.toggle_stock_blink)

    def _handle_alert_click(self):
        """Maneja el clic en el display de alertas."""
        # Obtener la primera notificación urgente/warning no leída
        if hasattr(self, 'notification_manager') and self.notification_manager:
            notifications = self.notification_manager.get_notifications()
            for notif in notifications:
                if not notif.read and notif.priority.value in ('urgent', 'warning'):
                    # Navegar al módulo si tiene ruta
                    if notif.modulo_ruta and hasattr(self, 'navigate_to_module'):
                        self.navigate_to_module(notif.modulo_ruta)
                    return
        
        # Si no hay notificación específica, abrir popup de quiebres
        self.abrir_popup_quiebres()

    def _update_alerts_display(self):
        """Actualiza el display de alertas basado en notificaciones."""
        if not hasattr(self, 'alerts_display_label'):
            return
            
        try:
            # Buscar notificaciones urgentes/warnings no leídas
            alert_text = ""
            alert_callback = None
            
            if hasattr(self, 'notification_manager') and self.notification_manager:
                notifications = self.notification_manager.get_notifications()
                for notif in notifications:
                    if not notif.read and notif.priority.value in ('urgent', 'warning'):
                        # Mostrar el título de la primera alerta encontrada
                        alert_text = notif.title
                        alert_callback = notif.modulo_ruta
                        break
            
            if alert_text:
                self.alerts_display_label.config(text=alert_text)
                self.alerts_display_label.pack(side=tk.LEFT, padx=10)
                self._alert_callback = alert_callback
            else:
                self.alerts_display_label.pack_forget()
        except Exception as e:
            print(f"Error actualizando alerts display: {e}")

    def recargar_stock(self):
        """Recarga completamente el módulo de stock"""
        if not self.modules_enabled.get("stock", False): return
        
        from tkinter import messagebox
        if not messagebox.askyesno("Confirmar", "¿Desea recargar los quiebres de stock?"):
            return
            
        self.log("🚀 Recargando módulo de quiebres...", "INFO")
        # Recargar jerarquía y luego alertas
        self.load_stock_filters()
        self.actualizar_alertas_stock(force_refresh=True)
    
    def _recargar_stock_async(self):
        """Ejecuta la recarga completa del stock en segundo plano con actualizaciones progresivas de UI"""
        import time
        start_time = time.perf_counter()
        
        try:
            # FASE 1: Inicialización y limpieza (5%)
            self._update_stock_reload_progress(5, "Inicializando recarga...")
            
            # Resetear estado de carga paralela
            self.stock_full_loading_started = False
            
            # Limpiar caches
            self.cached_alertas = []
            self.last_refresh = None
            if hasattr(self, 'all_jerarquia'):
                self.all_jerarquia = {}
            if hasattr(self, 'producto_jerarquia'):
                self.producto_jerarquia = {}
            
            time.sleep(0.3)  # Dar tiempo para que UI se actualice
            
            # FASE 2: Recargar filtros jerárquicos (25%)
            self._update_stock_reload_progress(10, "Cargando departamentos...")
            self.root.after(0, self._load_stock_filters_async_step1)
            time.sleep(0.5)
            
            self._update_stock_reload_progress(25, "Filtros jerárquicos cargados")
            self.root.after(0, lambda: self.log("✅ Filtros jerárquicos recargados", "SUCCESS"))
            
            # FASE 3: Cargar alertas iniciales (50%)
            self._update_stock_reload_progress(30, "Cargando alertas de stock...")
            
            try:
                # Cargar primer lote de alertas (limitado para rapidez)
                alertas_iniciales = self.db_manager.obtener_alertas_stock(limit=500)
                self.cached_alertas = alertas_iniciales or []
                self.last_refresh = datetime.now()
                self.ultimas_notificaciones.clear()
                
                self._update_stock_reload_progress(50, f"Cargadas {len(self.cached_alertas)} alertas iniciales")
                self.root.after(0, lambda: self.log(f"✅ {len(self.cached_alertas)} alertas de stock cargadas", "SUCCESS"))
                
            except Exception as e:
                self.root.after(0, lambda err=str(e): self.log(f"⚠️ Error cargando alertas: {err}", "WARNING"))
                self.cached_alertas = []
            
            time.sleep(0.3)
            
            # FASE 4: Cargar jerarquía de productos (75%)
            self._update_stock_reload_progress(55, "Cargando jerarquía de productos...")
            
            try:
                from pal.services.stock import load_all_jerarquia, build_producto_jerarquia
                
                # Cargar jerarquía completa
                all_jerarquia = load_all_jerarquia(
                    self.db_manager,
                    JERARQUIA_CACHE_FILE,
                    int(JERARQUIA_CACHE_TTL.total_seconds())
                )
                self.all_jerarquia = all_jerarquia or {}
                
                self._update_stock_reload_progress(65, f"Jerarquía cargada: {len(self.all_jerarquia)} productos")
                
                # Filtrar jerarquía por códigos en alerta
                codigos_en_alerta = {str(r[0]).strip() for r in self.cached_alertas}
                self.producto_jerarquia = build_producto_jerarquia(self.all_jerarquia, codigos_en_alerta)
                
                self._update_stock_reload_progress(75, f"Jerarquía filtrada: {len(self.producto_jerarquia)} productos")
                self.root.after(0, lambda: self.log(f"✅ Jerarquía cargada: {len(self.all_jerarquia)} productos totales", "SUCCESS"))
                
            except Exception as e:
                self.root.after(0, lambda err=str(e): self.log(f"⚠️ Error cargando jerarquía: {err}", "WARNING"))
                self.all_jerarquia = {}
                self.producto_jerarquia = {}
            
            time.sleep(0.3)
            
            # FASE 5: Resetear y aplicar filtros (90%)
            self._update_stock_reload_progress(80, "Aplicando filtros...")
            
            # Resetear página actual
            self.current_page = 1
            
            # Aplicar filtros en el hilo principal
            self.root.after(0, self.aplicar_filtro_stock)
            time.sleep(0.5)
            
            # Reconstruir vistas efectivas (aplicar exclusiones una vez)
            try:
                self._rebuild_effective_views()
            except Exception:
                pass
            self._update_stock_reload_progress(90, "Filtros aplicados")
            
            # FASE 6: Iniciar carga completa en segundo plano (100%)
            self._update_stock_reload_progress(95, "Iniciando carga completa...")
            
          # Iniciar carga completa de alertas en background
            if not getattr(self, 'stock_full_loading_started', False) and not getattr(self, '_stock_loading_in_progress', False):
                self.stock_full_loading_started = True
                threading.Thread(target=self._background_load_alertas_stock, daemon=True).start()
                self.root.after(0, lambda: self.log("✅ Carga paralela de alertas iniciada", "SUCCESS"))
            elif getattr(self, '_stock_loading_in_progress', False):
                self.root.after(0, lambda: self.log("📡 Carga de alertas ya en progreso", "DEBUG"))
            
            # FINALIZACIÓN (100%)
            elapsed_time = time.perf_counter() - start_time
            final_msg = f"Recarga completada en {elapsed_time:.1f}s"
            self._update_stock_reload_progress(100, final_msg)
            
            self.root.after(0, lambda: self.log(f"🎉 {final_msg}", "SUCCESS"))
            
            # Ocultar progreso después de 2 segundos
            time.sleep(2)
            self.root.after(0, self._hide_stock_reload_progress)
            
        except Exception as e:
            error_msg = f"Error en recarga asíncrona de stock: {str(e)}"
            self.root.after(0, lambda: self.log(error_msg, "ERROR"))
            self.root.after(0, lambda: self._update_stock_reload_progress(0, "Error en recarga", error=True))
            time.sleep(3)
            self.root.after(0, self._hide_stock_reload_progress)
    
    def _load_stock_filters_async_step1(self):
        """Carga filtros de stock de forma segura en el hilo principal"""
        try:
            self.load_stock_filters()
        except Exception as e:
            self.log(f"Error cargando filtros: {e}", "ERROR")
    
    def _update_stock_reload_progress(self, percentage, message, error=False):
        """Actualiza el progreso de la recarga de stock en la UI"""
        def update_ui():
            try:
                # Actualizar texto del status
                color = "red" if error else "#004C97" if percentage < 100 else "green"
                status_text = f"Stock: {message} ({percentage}%)"
                
                if hasattr(self, 'api_status'):
                    self.api_status.config(text=status_text, foreground=color)
                
                # Actualizar barra de progreso
                if hasattr(self, 'global_progress'):
                    if percentage == 0 or error:
                        # Ocultar barra en caso de error o reset
                        self.global_progress.pack_forget()
                    else:
                        # Mostrar barra determinada
                        self.global_progress.pack(side=tk.RIGHT, padx=10)
                        self.global_progress.config(mode="determinate", maximum=100, value=percentage)
                        
            except Exception as e:
                print(f"Error actualizando progreso de recarga: {e}")
        
        self.root.after(0, update_ui)
    
    def _hide_stock_reload_progress(self):
        """Oculta los indicadores de progreso de recarga"""
        try:
            if hasattr(self, 'global_progress'):
                self.global_progress.stop()
                self.global_progress.pack_forget()
            
            if hasattr(self, 'api_status'):
                self.api_status.config(text="API: Lista", foreground="green")
                
        except Exception as e:
            print(f"Error ocultando progreso: {e}")
    
    def _background_load_ventas_tra_v3_deprecated(self):
        # Esta función ha sido reemplazada por la versión optimizada al final del archivo
        pass

        
        # Parámetros de cache
        cache_key = f"tra_{self.tra_sede_codigo}_{self.tra_fecha_inicio}_{self.tra_fecha_fin}"
        
        try:
            # Verificar cache primero
            if self._check_tra_cache(cache_key):
                self.log("[TRA] Usando datos desde cache", "INFO")
                self.root.after(0, self.aplicar_filtro_tra)
                return
            
            # Inicializar variables de control
            chunk_count = 0
            total_loaded = 0
            consecutive_failures = 0
            max_consecutive_failures = 3
            
            # Usar dict para evitar duplicados y optimizar búsquedas
            existentes = {r[0]: r for r in (self.cached_ventas_tra or [])}
            initial_count = len(existentes)
            
            self.log(f"🚀 [TRA] Iniciando carga adaptativa - Cache: {cache_key}, Registros existentes: {initial_count}", "INFO")
            
            start_row = 1
            
            while True:
                chunk_start_time = time.perf_counter()
                
                # Obtener chunk usando consulta optimizada con índices
                rows = self._fetch_tra_chunk_optimized(
                    self.tra_fecha_inicio, 
                    self.tra_fecha_fin, 
                    self.tra_sede_codigo, 
                    start_row=start_row, 
                    fetch_size=controller.size
                )
                
                chunk_query_time = time.perf_counter() - chunk_start_time
                
                if not rows:
                    consecutive_failures += 1
                    # Solo loggear si es el primer fallo o el último antes de terminar
                    if consecutive_failures == 1 or consecutive_failures >= max_consecutive_failures:
                        self.tra_debug_log(
                            f"Chunk {chunk_count + 1}: Sin datos (fallo {consecutive_failures}/{max_consecutive_failures})",
                            level="WARNING"
                        )
                    
                    if consecutive_failures >= max_consecutive_failures:
                        self.log(f"[TRA] Carga finalizada - {consecutive_failures} chunks consecutivos sin datos", "INFO")
                        break
                    
                    # Ajustar parámetros para siguiente intento
                    start_row += controller.size
                    time.sleep(1.0)  # Pausa tras error
                    continue
                else:
                    consecutive_failures = 0  # Reset en éxito
                
                chunk_count += 1
                new_records = 0
                updated_records = 0
                
                # Procesar registros del chunk
                for r in rows:
                    codigo_str = str(r[0])
                    if codigo_str in existentes:
                        updated_records += 1
                    else:
                        new_records += 1
                    existentes[codigo_str] = r
                
                # Ajuste adaptativo: delegar en el controlador (EMA + cooldown)
                controller.update(chunk_query_time, len(rows))
                
                # OPTIMIZACIÓN: Guardar en cache sin clasificar (clasificamos solo al final)
                # Esto ahorra muchísimo tiempo en consultas largas (180 días)
                if chunk_count % 10 == 0:  # Cada 10 chunks (menos frecuente)
                    self._save_tra_cache(cache_key, list(existentes.values()))
                
                total_loaded += len(rows)
                
                # Logging optimizado - solo cada 5 chunks o si es importante
                total_time = time.perf_counter() - load_start_time
                avg_latency = total_time / chunk_count
                records_per_sec = total_loaded / total_time if total_time > 0 else 0
                
                # Log cada 5 chunks, al inicio, o si hay problemas de rendimiento
                should_log = (
                    chunk_count <= 2 or  # Primeros 2 chunks siempre
                    chunk_count % 5 == 0 or  # Cada 5 chunks
                    chunk_query_time > target_latency * 1.5 or  # Si es muy lento
                    len(rows) < controller.size  # Último chunk
                )
                
                if should_log:
                    self.tra_debug_log(
                        f"Chunk {chunk_count}: {len(rows)} filas | Nuevos: {new_records} | "
                        f"Total: {len(existentes)} | Latencia: {chunk_query_time:.2f}s | "
                        f"Size: {controller.size} | Velocidad: {records_per_sec:.0f} reg/s",
                        level="INFO",
                        throttle_key="chunk_progress",
                        throttle_seconds=3.0
                    )
                
                # Actualizar UI de forma eficiente (cada 3 chunks o al final para reducir overhead)
                if chunk_count % 3 == 0 or len(rows) < controller.size:
                    try:
                        self.root.after(0, self._update_tra_ui_after_chunk, len(existentes), chunk_count, records_per_sec)
                    except Exception as e:
                        self.tra_debug_log(f"Error actualizando UI TRA: {e}", level="ERROR")
                
                start_row += len(rows)
                
                # Pausa adaptativa según rendimiento (recomendación del controlador)
                time.sleep(controller.recommend_sleep(chunk_query_time))
                
                # Condición de salida optimizada
                if len(rows) < controller.size:
                    self.log(f"[TRA] Último chunk cargado: {len(rows)} registros", "INFO")
                    break
            
            # Clasificación final y guardado en cache
            from pal.services.tra import clasificar_rotacion_tra
            self.cached_ventas_tra = clasificar_rotacion_tra(list(existentes.values()))
            self._save_tra_cache(cache_key, self.cached_ventas_tra)
            
            # Estadísticas finales
            total_time = time.perf_counter() - load_start_time
            final_count = len(existentes)
            net_new = final_count - initial_count
            avg_chunk_time = total_time / chunk_count if chunk_count > 0 else 0
            
            self.log(
                f"✅ [TRA] Carga adaptativa completada: {chunk_count} chunks | "
                f"{final_count} registros totales | {net_new} nuevos | "
                f"Tiempo: {total_time:.2f}s | Promedio/chunk: {avg_chunk_time:.2f}s | "
                f"Velocidad final: {final_count / total_time:.0f} reg/s", 
                "SUCCESS"
            )
            
            # Aplicar filtros con datos completos
            try:
                self.root.after(0, self.aplicar_filtro_tra)
            except Exception as e:
                self.log(f"Error aplicando filtros TRA tras carga: {e}", "ERROR")
            
        except Exception as e:
            load_time = time.perf_counter() - load_start_time
            self.log(f"Error en carga adaptativa TRA (tiempo: {load_time:.2f}s): {str(e)}", "ERROR")
    
    def _check_tra_cache(self, cache_key):
        """Verifica si existe cache válido para los parámetros TRA dados"""
        try:
            import os
            import json
            from datetime import datetime, timedelta
            
            cache_file = f"tra_cache_{cache_key.replace('/', '_').replace(':', '_')}.json"
            cache_ttl = timedelta(hours=2)  # Cache válido por 2 horas
            
            if not os.path.exists(cache_file):
                return False
            
            # Verificar TTL
            mtime = datetime.fromtimestamp(os.path.getmtime(cache_file))
            if datetime.now() - mtime > cache_ttl:
                self.tra_debug_log(f"Cache expirado: {cache_file}")
                try:
                    os.remove(cache_file)
                except Exception:
                    pass
                return False
            
            # Cargar datos desde cache
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            
            if isinstance(cached_data, list) and len(cached_data) > 0:
                self.cached_ventas_tra = [tuple(item) for item in cached_data]
                self.log(f"[TRA] Cache cargado: {len(self.cached_ventas_tra)} registros", "INFO")
                return True
            
            return False
            
        except Exception as e:
            self.tra_debug_log(f"Error verificando cache TRA: {e}")
            return False
    
    def _save_tra_cache(self, cache_key, data):
        """Guarda datos TRA en cache local con TTL"""
        try:
            import json
            
            cache_file = f"tra_cache_{cache_key.replace('/', '_').replace(':', '_')}.json"
            
            # Convertir tuplas a listas para JSON
            json_data = [list(item) if isinstance(item, tuple) else item for item in data]
            
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, default=str)
            
            # Solo loggear si es un guardado significativo
            if len(data) > 100:
                self.tra_debug_log(
                    f"Cache guardado: {len(data)} registros",
                    level="DEBUG",
                    throttle_key="cache_save",
                    throttle_seconds=10.0
                )
            
        except Exception as e:
            self.tra_debug_log(f"Error guardando cache TRA: {e}")
    
    def _fetch_tra_chunk_optimized(self, fecha_inicio, fecha_fin, sede_codigo, start_row=1, fetch_size=500):
        """Consulta optimizada para chunks TRA con índices y selección minimal"""
        try:
            # Consulta optimizada: solo seleccionar columnas necesarias y usar índices
            # Usar el método correcto del DatabaseManager para evitar consultas duplicadas
            return self.db_manager.obtener_ventas_por_producto_chunk(
                fecha_inicio, fecha_fin, sede_codigo, start_row, fetch_size,
                exclude_depts=getattr(self, 'excluded_depts', [])
            )
            
        except Exception as e:
            self.tra_debug_log(f"Error en consulta TRA: {e}", level="ERROR", throttle_key="query_error", throttle_seconds=5.0)
            # Fallback a método original si la optimizada falla
            try:
                return self.db_manager.obtener_ventas_por_producto_chunk(
                    fecha_inicio, fecha_fin, sede_codigo, start_row, fetch_size,
                    exclude_depts=getattr(self, 'excluded_depts', [])
                )
            except Exception as e2:
                self.tra_debug_log(f"Error en fallback TRA: {e2}", level="ERROR")
                return []
    
    def _update_tra_ui_after_chunk(self, total_records, chunk_count, records_per_sec=0):
        """Actualiza UI TRA después de cargar un chunk en segundo plano con estadísticas de rendimiento"""
        try:
            # Evitar errores si la UI fue destruida/reconstruida
            if not hasattr(self, 'root') or not self.root.winfo_exists():
                return
            # Actualizar status con estadísticas de rendimiento
            if hasattr(self, 'api_status') and self.api_status.winfo_exists():
                status_text = f"RI: {total_records} registros"
                if records_per_sec > 0:
                    status_text += f" | {records_per_sec:.0f} reg/s"
                self.api_status.config(text=status_text, foreground="#004C97")
            
            # Actualizar paginación y vista si el tree existe
            if hasattr(self, 'tra_tree') and self.tra_tree and self.tra_tree.winfo_exists():
                if hasattr(self, 'tra_pagina_label') and self.tra_pagina_label.winfo_exists():
                    # Nota: solo actualizar etiqueta; el refresco completo se hace en aplicar_filtro_tra()
                    self.tra_pagina_label.config(text=f"{total_records} registros")
        except Exception as e:
            self.tra_debug_log(f"Error actualizando UI TRA: {e}", level="ERROR")
            if hasattr(self, 'aplicar_filtro_tra') and total_records > 10:
                self.aplicar_filtro_tra()
            
        except Exception as e:
            self.tra_debug_log(f"Error actualizando UI: {e}", level="ERROR", throttle_key="ui_update_error", throttle_seconds=5.0)
    
    def _background_load_ventas_tra_fast(self):
        """Versión super optimizada de carga TRA - primeros datos en <2 segundos"""
        if not hasattr(self, 'tra_fecha_inicio') or not hasattr(self, 'tra_fecha_fin') or not hasattr(self, 'tra_sede_codigo'):
            self.log("No hay parámetros TRA para carga rápida", "ERROR")
            return
        
        load_start_time = time.perf_counter()
        
        try:
            # Normalizar código de sede para persistencia (00/% -> ICH)
            sede_persistencia = self.tra_sede_codigo
            if sede_persistencia in ('00', '%', 'ALL'):
                sede_persistencia = 'ICH'

            # ATAJO: Verificar si hay datos persistidos frescos (Nodo Maestro)
            from pal.services.tra import get_persisted_rotation, save_rotation_persistence, clasificar_rotacion
            dias_context = (self.tra_fecha_fin - self.tra_fecha_inicio).days or 365
            persisted = get_persisted_rotation(self.db_manager, sede=sede_persistencia, dias_rango=dias_context)
            if persisted:
                self.log(f"🚀 TRA: Usando datos de rotación persistidos (Sede: {sede_persistencia}, Rango: {dias_context}d)", "SUCCESS")
                datos_persisted = self.db_manager.obtener_ventas_persisted_tra(sede=sede_persistencia, dias_rango=dias_context)
                if datos_persisted:
                    self.cached_ventas_tra = [tuple(r) for r in datos_persisted]
                    self.root.after(0, lambda: self._update_tra_phase(3, len(self.cached_ventas_tra), time.perf_counter() - load_start_time))
                    self.root.after(0, self._finalize_tra_loading)
                    return

            # FASE 1: Carga ultra rápida (primeros 50 registros) - <1 segundo
            self.log("🚀 TRA: Iniciando carga ultra rápida (50 registros)...", "INFO")
            
            # Obtener primeros datos rápidamente
            datos_ultra_rapidos = self.db_manager.obtener_ventas_completas_tra(
                self.tra_fecha_inicio, 
                self.tra_fecha_fin, 
                self.tra_sede_codigo, 
                limit=50,
                exclude_depts=getattr(self, 'excluded_depts', [])
            )
            
            if not datos_ultra_rapidos:
                self.root.after(0, self._show_no_data_tra)
                return
            
            # Clasificar y mostrar inmediatamente
            from pal.services.tra import clasificar_rotacion
            self.cached_ventas_tra = clasificar_rotacion(datos_ultra_rapidos)
            phase1_time = time.perf_counter() - load_start_time
            
            # Actualizar UI inmediatamente
            self.root.after(0, lambda: self._update_tra_phase(1, len(self.cached_ventas_tra), phase1_time))
            
            # FASE 2: Carga rápida (200 registros más) - a los 2-3 segundos
            time.sleep(0.5)  # Pausa mínima
            self.tra_debug_log("Iniciando fase 2: 200 registros adicionales")
            
            datos_rapidos = self.db_manager.obtener_ventas_completas_tra(
                self.tra_fecha_inicio, 
                self.tra_fecha_fin, 
                self.tra_sede_codigo, 
                limit=250,  # Total 250 (ya tenemos 50)
                exclude_depts=getattr(self, 'excluded_depts', [])
            )
            
            if datos_rapidos and len(datos_rapidos) > len(self.cached_ventas_tra):
                self.cached_ventas_tra = clasificar_rotacion(datos_rapidos)
                phase2_time = time.perf_counter() - load_start_time
                self.root.after(0, lambda: self._update_tra_phase(2, len(self.cached_ventas_tra), phase2_time))
            
            # FASE 3: Carga progresiva en chunks (resto de datos)
            time.sleep(1)  # Dar tiempo a la UI
            self.tra_debug_log("Iniciando fase 3: carga progresiva por chunks")
            
            chunk_size = 500
            start_row = 251  # Comenzar después de los primeros 250
            chunk_count = 0
            consecutive_failures = 0
            max_consecutive_failures = 2
            
            existentes = {r[0]: r for r in (self.cached_ventas_tra or [])}
            
            while consecutive_failures < max_consecutive_failures:
                chunk_start = time.perf_counter()
                
                # Usar función de chunk thread-safe
                chunk_data = self.db_manager.obtener_ventas_por_producto_chunk(
                    self.tra_fecha_inicio, 
                    self.tra_fecha_fin, 
                    self.tra_sede_codigo, 
                    start_row=start_row, 
                    fetch_size=chunk_size,
                    exclude_depts=getattr(self, 'excluded_depts', [])
                )
                
                if not chunk_data:
                    consecutive_failures += 1
                    self.tra_debug_log(f"Chunk vacío en posición {start_row} (fallo {consecutive_failures})")
                    start_row += chunk_size
                    continue
                
                consecutive_failures = 0
                chunk_count += 1
                new_records = 0
                
                # Añadir nuevos registros
                for r in chunk_data:
                    codigo_str = str(r[0])
                    if codigo_str not in existentes:
                        existentes[codigo_str] = r
                        new_records += 1
                
                # Actualizar cache con clasificación de rotación
                self.cached_ventas_tra = clasificar_rotacion(list(existentes.values()))
                
                chunk_time = time.perf_counter() - chunk_start
                total_time = time.perf_counter() - load_start_time
                
                self.tra_debug_log(
                    f"Chunk {chunk_count}: +{new_records} nuevos | "
                    f"Total: {len(existentes)} | "
                    f"Tiempo chunk: {chunk_time:.2f}s"
                )
                
                # Actualizar UI cada 2 chunks
                if chunk_count % 2 == 0:
                    self.root.after(0, lambda c=chunk_count, t=len(existentes): 
                                  self._update_tra_chunk_progress(c, t))
                
                start_row += len(chunk_data)
                
                # Pausa breve para no sobrecargar la BD
                time.sleep(0.1)
                
                # Si el chunk es más pequeño, probablemente sea el último
                if len(chunk_data) < chunk_size:
                    self.tra_debug_log(f"Último chunk detectado ({len(chunk_data)} < {chunk_size})")
                    break
            
            # Finalización
            total_time = time.perf_counter() - load_start_time
            final_count = len(existentes)
            
            self.log(
                f"✅ TRA: Carga completa finalizada - "
                f"{final_count} registros en {total_time:.2f}s | "
                f"{chunk_count} chunks procesados", 
                "SUCCESS"
            )

            # GUARDAR PERSISTENCIA (Actuar como Nodo Maestro) - En hilo separado para evitar freeze
            # IMPORTANTE: Si es reporte masivo, NO guardar en la base de datos (evita inflar tabla con ventas 0)
            if not getattr(self, 'tra_include_zero_sales', False):
                def _bg_save_persistence():
                    try:
                        from pal.services.tra import save_rotation_persistence
                        user_nodo = self.current_user['username'] if self.current_user else "SISTEMA"
                        dias = (self.tra_fecha_fin - self.tra_fecha_inicio).days or 365
                        
                        # Normalizar sede para persistencia (00/% -> ICH)
                        sede_save = self.tra_sede_codigo
                        if sede_save in ('00', '%', 'ALL'):
                            sede_save = 'ICH'
                            
                        save_rotation_persistence(self.db_manager, self.cached_ventas_tra, sede_save, dias, user_nodo)
                    except Exception as e:
                        self.log(f"Error guardando persistencia RI: {e}", "DEBUG")
                
                import threading
                threading.Thread(target=_bg_save_persistence, daemon=True, name="RI_Persistence_Fast").start()
            else:
                self.log("ℹ️ RI: Persistencia omitida por Reporte Masivo (Ventas 0)", "DEBUG")
            
            # Aplicar filtros finales
            self.root.after(0, self._finalize_tra_loading)
            
        except Exception as e:
            load_time = time.perf_counter() - load_start_time
            self.log(f"Error en carga rápida TRA (tiempo: {load_time:.2f}s): {str(e)}", "ERROR")
            self.root.after(0, lambda: self._show_tra_error(str(e)))
    
    def _update_tra_phase(self, phase, count, elapsed_time):
        """Actualiza UI tras cada fase de carga TRA"""
        try:
            phase_names = {1: "Ultra Rápida", 2: "Rápida", 3: "Completa"}
            phase_name = phase_names.get(phase, f"Fase {phase}")
            
            self.api_status.config(
                text=f"RI: {phase_name} - {count} registros ({elapsed_time:.1f}s)", 
                foreground="#004C97"
            )
            
            # Aplicar filtros para mostrar datos
            self.aplicar_filtro_tra()
            
            self.tra_debug_log(f"Fase {phase} completada: {count} registros en {elapsed_time:.1f}s")
            
        except Exception as e:
            self.tra_debug_log(f"Error actualizando fase {phase}: {e}")
    
    def _update_tra_chunk_progress(self, chunk_num, total_records):
        """Actualiza progreso durante carga por chunks"""
        try:
            self.api_status.config(
                text=f"RI: Chunk {chunk_num} - {total_records} registros", 
                foreground="#004C97"
            )
            
            # Reaplicar filtros para mostrar datos actualizados
            self.aplicar_filtro_tra()
            
        except Exception as e:
            self.tra_debug_log(f"Error actualizando chunk {chunk_num}: {e}")
    
    def _finalize_tra_loading(self):
        """Finaliza la carga TRA"""
        try:
            self.api_status.config(text="RI: Completo", foreground="green")
            self.aplicar_filtro_tra()
            self.log("🎉 TRA: Carga finalizada - datos listos", "SUCCESS")
        except Exception as e:
            self.tra_debug_log(f"Error finalizando carga TRA: {e}")
    
    def _show_no_data_tra(self):
        """Muestra mensaje cuando no hay datos TRA"""
        try:
            if hasattr(self, 'tra_tree'):
                self.tra_tree.delete(*self.tra_tree.get_children())
                self.tra_tree.insert("", tk.END, values=(
                    "", "No se encontraron ventas para este rango", "", "", "", "", "", ""
                ), tags=("no_data",))
            
            self.api_status.config(text="RI: Sin datos", foreground="orange")
            messagebox.showinfo("Sin resultados", "No se encontraron ventas para ese rango y sede.")
            
        except Exception as e:
            self.tra_debug_log(f"Error mostrando 'sin datos': {e}")
    
    def _show_tra_error(self, error_msg):
        """Muestra error en la UI TRA"""
        try:
            if hasattr(self, 'tra_tree'):
                self.tra_tree.delete(*self.tra_tree.get_children())
                self.tra_tree.insert("", tk.END, values=(
                    "ERROR", f"Error cargando datos: {error_msg[:50]}...", "", "", "", "", "", ""
                ), tags=("error",))
            
            self.api_status.config(text="RI: Error", foreground="red")
            
        except Exception as e:
            self.tra_debug_log(f"Error mostrando error: {e}")

    def mostrar_alertas_paginadas(self, datos):
        """Mostrar datos con estado de favoritos"""
        self.stock_tree.delete(*self.stock_tree.get_children())
        favoritos = self._get_favoritos_local()
    
        for codigo, desc, stock, nivel in datos:
            es_favorito = codigo in favoritos
            estado = "✓" if es_favorito else "☐"
            tags = ('favorito' if es_favorito else '', nivel.lower().replace('ítica', 'itica'))
        
            self.stock_tree.insert("", tk.END, 
                                values=(estado, codigo, desc, stock, nivel),
                                tags=tags)
        # Limpiar selección después de actualizar
        self.stock_tree.selection_remove(self.stock_tree.selection())

    def listar_hilos_activos(self):
            hilos_activos = threading.enumerate()
            self.log(f"Hilos activos: {[h.name for h in hilos_activos]}", "DEBUG")
        


    def monitorear_favoritos(self):
        """Monitorea favoritos y productos críticos de rotación alta/media"""
        self.log("[MONITOR] Hilo de monitoreo de favoritos iniciado", "INFO")
        wait_count = 0
        while True:
            try:
                if not self.db_manager.conn:
                    wait_count += 1
                    if wait_count % 6 == 0:  # Log cada 6 iteraciones (360s = 6 min)
                        self.log(f"[MONITOR] Esperando conexión BD... ({wait_count * 60}s)", "DEBUG")
                    time.sleep(60)
                    continue
                
                wait_count = 0  # Reset cuando hay conexión
            
                # Obtener favoritos desde el JSON
                favoritos = self._get_favoritos_local()
                if not favoritos:
                    time.sleep(300)
                    continue
            
                try:
                    self.monitorear_quiebres_stock()
                except Exception as e:
                    self.log(f"[MONITOR] Error en monitor de quiebres: {e}", "DEBUG")
                         
                time.sleep(300)  # 5 minutos
            
            except Exception as e:
                self.log(f"[MONITOR] Error monitoreo stock: {str(e)}", "ERROR")
                time.sleep(100)
    
    def monitorear_quiebres_stock(self):
        """
        Detecta quiebres de stock (Stock 0 en almacenes tratables para productos de alta rotación)
        usando la lógica directa de ventas después de última compra.
        """
        try:
            from pal.ui.popups import show_stock_break_popup
            
            sedes_config = self.config_manager.get_sedes_config()
            quiebres_consolidados = {}
            found_new_quiebre = False
            
            # Revisar por cada sede configurada
            for sede_name, config in sedes_config.items():
                depositos = config.get('almacenes_tratables', [])
                if not depositos:
                    continue
                
                # Mapear nombre de sede a código (ej. 'Cabudare' -> '0301') si es posible, 
                # o usar la sede_name como contexto de rotación.
                # Nota: El monitor usa 365 días por defecto para determinar importancia.
                quiebres = self.db_manager.obtener_quiebres_directos(depositos, solo_alta_rotacion=True, sede_context=sede_name, dias_context=365)
                
                if quiebres:
                    quiebres_consolidados[sede_name] = quiebres
                    
                    # Verificar si hay alguno nuevo para disparar el popup
                    for q in quiebres:
                        # La llave ahora incluye sede para notificar individualmente por ocurrencia
                        notif_key = f"QUIEBRE_{q['codigo']}_{sede_name}"
                        if notif_key not in self.ultimas_notificaciones:
                            self.log(f"⚠️ ¡QUIEBRE! {q['codigo']} ({q['descripcion']}) - Sede: {sede_name} | Perdidas: {int(q['unidades_perdidas'])}", "WARNING")
                            self.ultimas_notificaciones.add(notif_key)
                            # Mostrar notificación tipo toast individual
                            self.mostrar_notificacion_quiebre(q['codigo'], q['descripcion'], sede_name)
                            found_new_quiebre = True
                
        except Exception as e:
            self.log(f"Error en monitoreo de quiebres: {str(e)}", "DEBUG")
                
# Los métodos legacy _detectar_y_notificar_criticos y _mostrar_alerta_compras fueron eliminados.



    def mostrar_notificacion_quiebre(self, codigo, desc, sede_name):
        """Muestra un toast para quiebre de stock en una sede específica."""
        mensaje = f"Código: {codigo} | Sede: {sede_name}\nDescripción: {desc}"

        # Persistir en el Centro de Notificaciones
        try:
            if hasattr(self, 'notification_manager'):
                usuario = self.current_user.get('username') if self.current_user else None
                self.notification_manager.add(
                    title="Quiebre de Stock Detectado",
                    message=mensaje,
                    priority="urgent",
                    module="Stock",
                    modulo_ruta="stock",
                    accion_etiqueta="Ver Stock",
                    usuario=usuario,
                    datos={"codigo": codigo, "sede": sede_name},
                )
        except Exception as e:
            self.log(f"Error al registrar quiebre en notificaciones: {e}", "DEBUG")

        # Toast visual (mantener compatibilidad)
        try:
            if hasattr(self, 'toaster'):
                self.toaster.show_toast(
                    "QUIEBRE DE STOCK DETECTADO",
                    mensaje,
                    duration=10,
                    threaded=False
                )
        except Exception as e:
            self.log(f"Error en toast de quiebre: {e}", "DEBUG")


    
    def exportar_stock_excel(self):
        """Exporta datos de quiebres de stock en formato Excel con formato profesional - ASYNC"""
        import threading
        from tkinter import filedialog, messagebox
        from pal.services.exports import export_stock_excel
        from pal.services.stock import filter_alertas
        
        try:
            # Permisos: STOCK.exportar
            allowed = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    allowed = self.permissions.tiene_permiso(self.current_user['id'], 'STOCK', 'exportar')
                if self.current_user and self.current_user.get('username','').lower() == 'admin':
                    allowed = True
            except Exception:
                allowed = False
            
            if not allowed:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para exportar quiebres de stock")
                return

            # Verificar si hay datos
            if not hasattr(self, 'cached_alertas') or not self.cached_alertas:
                messagebox.showinfo("Sin datos", "No hay datos de quiebres para exportar.")
                return

            # Aplicar filtros actuales de la UI
            dept_desc = self.stock_dept_var.get() if hasattr(self, 'stock_dept_var') else 'Todos'
            group_desc = self.stock_group_var.get() if hasattr(self, 'stock_group_var') else 'Todos'
            subgroup_desc = self.stock_subgroup_var.get() if hasattr(self, 'stock_subgroup_var') else 'Todos'
            search_text = self.stock_search_var.get() if hasattr(self, 'stock_search_var') else ''
            
            dept_code = self.tra_dept_dict.get(dept_desc) if dept_desc != 'Todos' else None
            group_code = None
            if group_desc != 'Todos' and dept_code:
                group_code = self.tra_group_dict.get(dept_code, {}).get(group_desc)
            subgroup_code = None
            if subgroup_desc != 'Todos' and dept_code and group_code:
                key = f"{dept_code}|{group_code}"
                subgroup_code = self.tra_sub_dict.get(key, {}).get(subgroup_desc)

            # Usar vista efectiva (sin departamentos excluidos)
            alertas_base = getattr(self, 'cached_alertas_effective', self.cached_alertas)
            
            # Filtrar
            filtrados = filter_alertas(
                alertas=alertas_base,
                producto_jerarquia=getattr(self, 'all_jerarquia', {}),
                dept_code=dept_code,
                group_code=group_code,
                subgroup_code=subgroup_code,
                search_text=search_text,
                favoritos=getattr(self, 'favoritos', set())
            )

            if not filtrados:
                messagebox.showinfo("Sin resultados", "No hay quiebres que coincidan con los filtros actuales.")
                return

            # Diálogo para guardar
            filename = filedialog.asksaveasfilename(
                title="Guardar Quiebres de Stock",
                defaultextension=".xlsx",
                initialfile=f"quiebres_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filename:
                return

            # Obtener configuraciones para la exportación
            try:
                # Obtener todos los depósitos tratables para la columna de stock
                sedes_config = self.config_manager.get_sedes_config()
                seleccionadas = []
                # location_groups para el service
                location_groups_svc = {}
                
                for name, cfg in sedes_config.items():
                    deps = cfg.get('almacenes_tratables', [])
                    if deps:
                        # Extraer código de sede de manera consistente
                        cod_sede = cfg.get('codigo_sede')
                        if not cod_sede:
                            if " - " in name: cod_sede = name.split(" - ")[0]
                            else: cod_sede = deps[0] if deps else name
                        
                        if cod_sede not in ('00', 'ICH', '%', 'ALL'):
                            seleccionadas.extend(deps)
                            location_groups_svc[name] = deps
                
                seleccionadas = list(set(seleccionadas)) # Únicos
                
            except Exception as e:
                self.log(f"Error preparando parámetros de exportación: {e}", "ERROR")
                seleccionadas = []
                location_groups_svc = {}

            # Ejecutar en segundo plano
            def run_export():
                try:
                    self.log(f"Iniciando exportación de {len(filtrados)} quiebres...", "INFO")
                    
                    # El service espera: (codigo, descripcion, stock, nivel) 
                    # Nuestras tuplas son: (codigo, desc, sede, unid_perd, dias, comp, vent, dept, grp, sub, brand)
                    # Mapear a diccionarios para mayor claridad y compatibilidad
                    datos_svc = []
                    for q in filtrados:
                        # Estructura q: (codigo, descripcion, sede, unidades_perdidas, dias_quiebre, ultima_compra, ultima_venta)
                        # Pasamos los 7 elementos básicos, la jerarquía se resuelve en el service
                        if len(q) >= 7:
                            datos_svc.append((q[0], q[1], q[2], q[3], q[4], q[5], q[6]))
                        else:
                            # Fallback si por alguna razón la tupla fuese diferente
                            datos_svc.append(q)

                    count = export_stock_excel(
                        filename=filename,
                        datos_exportar=datos_svc,
                        seleccionadas=seleccionadas,
                        location_groups=location_groups_svc,
                        db_manager=self.db_manager,
                        current_localidad=getattr(self, 'current_sede_name', 'Cabudare'),
                        permissions_manager=self.permissions,
                        current_user_id=self.current_user['id'] if self.current_user else None
                    )
                    
                    self.log(f"✅ Exportación completada: {count} registros en {filename}", "SUCCESS")
                    self.root.after(0, lambda: messagebox.showinfo("Exportación Exitosa", f"Se han exportado {count} quiebres de stock exitosamente."))
                    
                except Exception as e:
                    self.log(f"Error exportando stock: {e}", "ERROR")
                    def _show_err(err=e):
                        messagebox.showerror("Error de Exportación", f"No se pudo completar la exportación:\n{str(err)}")
                    self.root.after(0, _show_err)

            threading.Thread(target=run_export, daemon=True).start()
            
        except Exception as e:
            self.log(f"Error al iniciar exportación de stock: {e}", "ERROR")
            messagebox.showerror("Error", f"Error inesperado al exportar: {e}")

    def exportar_tra_excel(self):
        """Exporta datos TRA en formato Excel con múltiples hojas y formato profesional - ASYNC"""
        import threading
        
        try:
            # Permisos: TRA.exportar
            allowed = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    allowed = self.permissions.tiene_permiso(self.current_user['id'], 'TRA', 'exportar')
                if self.current_user and self.current_user.get('username','').lower() == 'admin':
                    allowed = True
            except Exception:
                allowed = False
            if not allowed:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para exportar datos de RI")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENIED', usuario_id=self.current_user['id'], modulo='TRA', detalle='exportar')
                except Exception:
                    pass
                return
            # Verificar si hay datos para exportar
            if not hasattr(self, 'cached_ventas_tra') or not self.cached_ventas_tra:
                messagebox.showwarning("Sin datos", "No hay datos RI cargados para exportar")
                return
            
            # Verificar si openpyxl está disponible
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror(
                    "Módulo no encontrado", 
                    "Para exportar en Excel necesita instalar openpyxl:\n\n"
                    "pip install openpyxl\n\n"
                    "Use la exportación CSV como alternativa."
                )
                return
            
            # Obtener los mismos datos filtrados que se muestran en la interfaz
            if not hasattr(self, 'cached_ventas_tra') or not self.cached_ventas_tra:
                messagebox.showwarning("Sin datos", "No hay datos RI cargados para exportar")
                return
            
            # Aplicar los mismos filtros que se usan en la interfaz
            dept_cod = self.tra_dept_dict.get(self.tra_dept_var.get()) if hasattr(self, 'tra_dept_var') else None
            group_cod = None
            sub_cod = None
            
            if dept_cod and hasattr(self, 'tra_group_var'):
                group_desc = self.tra_group_var.get()
                group_cod = self.tra_group_dict.get(dept_cod, {}).get(group_desc)
                
                if group_cod and hasattr(self, 'tra_sub_var'):
                    sub_desc = self.tra_sub_var.get()
                    key = f"{dept_cod}|{group_cod}"
                    sub_cod = self.tra_sub_dict.get(key, {}).get(sub_desc)
            
            texto = self.tra_search_var.get() if hasattr(self, 'tra_search_var') else ''
            favoritos = self._get_favoritos_local()
            
            # Partir de los mismos datos base que la vista (incluyendo filtro por proveedor)
            datos_base = list(self.cached_ventas_tra)
            proveedor_cod = getattr(self, 'tra_proveedor_codigo', None)
            if proveedor_cod:
                codigos_prov = self._get_codigos_por_proveedor_cached(proveedor_cod)
                if codigos_prov:
                    datos_base = [r for r in datos_base if str(r[0]) in codigos_prov]
                else:
                    datos_base = []
            
            # Usar las mismas funciones de filtrado que la interfaz
            from pal.services.tra import filter_ventas_tra
            
            datos_exportar = filter_ventas_tra(
                ventas=datos_base,
                dept_code=dept_cod,
                group_code=group_cod,
                sub_code=sub_cod,
                search_text=texto,
                filter_rotacion='TODAS',
                favoritos=favoritos
            )
            # Exclusión global por departamento (para exportación RI)
            excluded_set = getattr(self, '_excluded_depts_set', set())
            if excluded_set:
                datos_exportar = [r for r in datos_exportar if len(r) > 2 and str(r[2]) not in excluded_set]
            
            # DEBUG: Log para diagnosticar discrepancias
            self.log(f"[EXPORT DEBUG] Datos originales en cache: {len(self.cached_ventas_tra)} registros", "DEBUG")
            self.log(f"[EXPORT DEBUG] Filtros aplicados - Dept: {dept_cod}, Group: {group_cod}, Sub: {sub_cod}, Texto: '{texto}'", "DEBUG")
            self.log(f"[EXPORT DEBUG] Datos para exportar: {len(datos_exportar)} registros", "DEBUG")
            if datos_exportar:
                primer_item = datos_exportar[0]
                self.log(f"[EXPORT DEBUG] Primer item: {primer_item}", "DEBUG")
                if len(primer_item) > 5:
                    neto_raw = primer_item[5]
                    self.log(f"[EXPORT DEBUG] Neto raw del primer item: {neto_raw} (tipo: {type(neto_raw)})", "DEBUG")
            
            if not datos_exportar:
                messagebox.showwarning("Sin datos", "No hay registros RI para exportar")
                return
            
            # Configurar progreso
            total_registros = len(datos_exportar)
            self.global_progress.pack(side=tk.RIGHT, padx=10)
            self.global_progress['value'] = 0
            self.global_progress['maximum'] = total_registros
            self.api_status.config(text="Exportando RI: 0%", foreground="#004C97")
            
            # Sugerir nombre por defecto y permitir al usuario elegir carpeta/archivo
            default_name = f"reporte_tra_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                from tkinter import filedialog
                filename = filedialog.asksaveasfilename(
                    parent=self.root if hasattr(self, 'root') else None,
                    title="Guardar reporte RI como...",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
                )
            except Exception as e:
                # Si falla el diálogo, caer al nombre por defecto en la carpeta actual
                self.log(f"[EXPORT] Error mostrando diálogo de guardar RI: {e} - usando ruta por defecto", "WARNING")
                filename = default_name

            if not filename:
                # Usuario canceló
                self.log("[EXPORT] Usuario canceló el diálogo de guardar RI", "INFO")
                self._cleanup_export_progress()
                return
            
            # Callback de progreso thread-safe
            def progress_cb(i, total):
                if i % max(1, total // 50) == 0 or i == total:
                    progreso = int((i / total) * 100)
                    # Usar after() para actualizar UI de forma segura desde otro hilo
                    self.root.after(0, lambda: self._update_export_progress(i, progreso, "TRA"))
            
            # Función para ejecutar la exportación en hilo separado
            def export_thread():
                try:
                    from pal.services.exports import export_tra_excel
                    
                    # Pasar permisos y usuario para verificar ver_costo_utilidad
                    permissions_mgr = getattr(self, 'permissions', None)
                    user_id = self.current_user.get('id') if self.current_user else None

                    # Etiqueta de proveedor (si hay filtro activo)
                    prov_cod = getattr(self, 'tra_proveedor_codigo', None)
                    prov_desc = getattr(self, 'tra_proveedor_descripcion', None)
                    prov_label = None
                    if prov_cod:
                        prov_label = prov_desc or prov_cod
                    
                    # Obtener código de sede
                    tra_sede_cod = getattr(self, 'tra_sede_codigo', None)
                    
                    # Obtener fechas para cálculo de estado
                    fecha_ini = getattr(self, 'tra_fecha_inicio', None)
                    fecha_fin = getattr(self, 'tra_fecha_fin', None)

                    total_registros = export_tra_excel(
                        filename=filename,
                        datos_tra=datos_exportar,
                        db_manager=self.db_manager,
                        progress_cb=progress_cb,
                        permissions_manager=permissions_mgr,
                        current_user_id=user_id,
                        provider_label=prov_label,
                        sede_codigo=tra_sede_cod,
                        fecha_inicio=fecha_ini,
                        fecha_fin=fecha_fin,
                    )
                    
                    # Notificar éxito en el hilo principal
                    self.root.after(0, lambda: self._export_success(
                        "RI",
                        total_registros,
                        filename,
                        "• Hojas incluidas: Datos principales, Resumen por rotación, Productos de baja rotación, Resumen jerárquico (con gráfico)\n"
                        "• Formato: Tablas con filtros, formatos condicionales y gráficos"
                    ))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='TRA', detalle=filename, exitoso=True)
                    except Exception:
                        pass
                    
                except Exception as e:
                    # Notificar error en el hilo principal
                    self.root.after(0, lambda err=str(e): self._export_error("RI", err))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='TRA', detalle=str(e), exitoso=False)
                    except Exception:
                        pass
            
            # Iniciar exportación en hilo separado
            thread = threading.Thread(target=export_thread, daemon=True, name="ExportTRA")
            thread.start()
            
        except Exception as e:
            self.log(f"Error iniciando exportación TRA: {str(e)}", "ERROR")
            self.api_status.config(text="API: Error", foreground="red")
            messagebox.showerror("Error en Exportación RI", f"Error durante la exportación:\n{str(e)}")
            self._cleanup_export_progress()
    
    def exportar_mbrp_excel(self):
        """Exporta datos MBRP en formato Excel con múltiples hojas y análisis de rentabilidad - ASYNC"""
        import threading
        
        try:
            # Permisos: MBRP.exportar
            allowed = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    allowed = self.permissions.tiene_permiso(self.current_user['id'], 'MBRP', 'exportar')
                if self.current_user and self.current_user.get('username','').lower() == 'admin':
                    allowed = True
            except Exception:
                allowed = False
            if not allowed:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para exportar datos de MBRP")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENEGADO', usuario_id=self.current_user['id'], modulo='MBRP', detalle='exportar')
                except Exception:
                    pass
                return
            # Verificar si hay datos para exportar
            if not hasattr(self, 'cached_ventas_mbrp') or not self.cached_ventas_mbrp:
                messagebox.showwarning("Sin datos", "No hay datos MBRP cargados para exportar")
                return
            
            # Verificar si openpyxl está disponible
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror(
                    "Módulo no encontrado", 
                    "Para exportar en Excel necesita instalar openpyxl:\n\n"
                    "pip install openpyxl\n\n"
                    "Use la exportación CSV como alternativa."
                )
                return
            
            # Preparar datos filtrados para exportar (replicando filtros activos de la vista MBRP)
            if not hasattr(self, 'cached_ventas_mbrp') or not self.cached_ventas_mbrp:
                messagebox.showwarning("Sin datos", "No hay datos MBRP cargados para exportar")
                return
            
            # 1) Partir de todos los datos en caché
            datos_base = list(self.cached_ventas_mbrp)
            
            # 2) Exclusión global por departamento (igual que en aplicar_filtro_mbrp)
            excluded_set = getattr(self, '_excluded_depts_set', set())
            if excluded_set:
                datos_base = [r for r in datos_base if len(r) > 2 and str(r[2]) not in excluded_set]
            
            # 3) Filtro por proveedor (si está seleccionado en MBRP)
            proveedor_cod = getattr(self, 'mbrp_proveedor_codigo', None)
            if proveedor_cod and datos_base:
                codigos_prov = self._get_codigos_por_proveedor_cached(proveedor_cod)
                if codigos_prov:
                    datos_base = [r for r in datos_base if str(r[0]) in codigos_prov]
                else:
                    datos_base = []
            
            # 4) Filtros jerárquicos y de búsqueda (usando el mismo helper que TRA)
            dept_cod = self.mbrp_dept_dict.get(self.mbrp_dept_var.get()) if hasattr(self, 'mbrp_dept_var') else None
            group_cod = None
            sub_cod = None
            if dept_cod and hasattr(self, 'mbrp_group_var'):
                group_desc = self.mbrp_group_var.get()
                group_cod = self.mbrp_group_dict.get(dept_cod, {}).get(group_desc)
                if group_cod and hasattr(self, 'mbrp_sub_var'):
                    sub_desc = self.mbrp_sub_var.get()
                    key = f"{dept_cod}|{group_cod}"
                    sub_cod = self.mbrp_sub_dict.get(key, {}).get(sub_desc)
            texto = self.mbrp_search_var.get() if hasattr(self, 'mbrp_search_var') else ''
            
            from pal.services.tra import filter_ventas_tra
            datos_exportar = filter_ventas_tra(
                ventas=datos_base,
                dept_code=dept_cod,
                group_code=group_cod,
                sub_code=sub_cod,
                search_text=texto,
                filter_rotacion='TODAS',
                favoritos=self._get_favoritos_local(),
            )
            
            if not datos_exportar:
                messagebox.showwarning("Sin datos", "No hay registros MBRP para exportar")
                return
            
            # Configurar progreso
            total_registros = len(datos_exportar)
            self.global_progress.pack(side=tk.RIGHT, padx=10)
            self.global_progress['value'] = 0
            self.global_progress['maximum'] = total_registros
            self.api_status.config(text="Exportando MBRP: 0%", foreground="#004C97")
            
            # Sugerir nombre por defecto y permitir al usuario elegir carpeta/archivo
            default_name = f"reporte_mbrp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                from tkinter import filedialog
                filename = filedialog.asksaveasfilename(
                    parent=self.root if hasattr(self, 'root') else None,
                    title="Guardar reporte MBRP como...",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
                )
            except Exception as e:
                # Si falla el diálogo, caer al nombre por defecto en la carpeta actual
                self.log(f"[EXPORT] Error mostrando diálogo de guardar MBRP: {e} - usando ruta por defecto", "WARNING")
                filename = default_name

            if not filename:
                # Usuario canceló
                self.log("[EXPORT] Usuario canceló el diálogo de guardar MBRP", "INFO")
                self._cleanup_export_progress()
                return
            
            # Callback de progreso thread-safe
            def progress_cb(i, total):
                if i % max(1, total // 50) == 0 or i == total:
                    progreso = int((i / total) * 100)
                    # Usar after() para actualizar UI de forma segura desde otro hilo
                    self.root.after(0, lambda: self._update_export_progress(i, progreso, "MBRP"))
            
            # Función para ejecutar la exportación en hilo separado
            def export_thread():
                try:
                    from pal.services.exports import export_mbrp_excel
                    
                    # Pasar permisos y usuario para verificar ver_costo_utilidad
                    permissions_mgr = getattr(self, 'permissions', None)
                    user_id = self.current_user.get('id') if self.current_user else None

                    # Etiqueta de proveedor (si hay filtro activo en MBRP)
                    prov_cod = getattr(self, 'mbrp_proveedor_codigo', None)
                    prov_desc = getattr(self, 'mbrp_proveedor_descripcion', None)
                    prov_label = None
                    if prov_cod:
                        prov_label = prov_desc or prov_cod
                    
                    total_registros = export_mbrp_excel(
                        filename=filename,
                        datos_mbrp=datos_exportar,
                        db_manager=self.db_manager,
                        progress_cb=progress_cb,
                        permissions_manager=permissions_mgr,
                        current_user_id=user_id,
                        provider_label=prov_label,
                        sede_codigo=self.mbrp_sede_codigo,
                        fecha_inicio=self.mbrp_fecha_inicio,
                        fecha_fin=self.mbrp_fecha_fin,
                    )
                    
                    # Notificar éxito en el hilo principal
                    self.root.after(0, lambda: self._export_success(
                        "MBRP",
                        total_registros,
                        filename,
                        "• Hojas incluidas: Datos principales, Resumen por rentabilidad, Productos críticos\n"
                        "• Formato: Tablas con filtros y formato condicional por margen"
                    ))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='MBRP', detalle=filename, exitoso=True)
                    except Exception:
                        pass
                    
                except Exception as e:
                    # Notificar error en el hilo principal
                    self.root.after(0, lambda err=str(e): self._export_error("MBRP", err))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='MBRP', detalle=str(e), exitoso=False)
                    except Exception:
                        pass
            
            # Iniciar exportación en hilo separado
            thread = threading.Thread(target=export_thread, daemon=True, name="ExportMBRP")
            thread.start()
            
        except Exception as e:
            self.log(f"Error iniciando exportación MBRP: {str(e)}", "ERROR")
            self.api_status.config(text="API: Error", foreground="red")
            messagebox.showerror("Error en Exportación MBRP", f"Error durante la exportación:\n{str(e)}")
            self._cleanup_export_progress()
    
    def exportar_excel(self):
        """Exporta datos de stock en formato Excel con múltiples hojas y formato avanzado - ASYNC"""
        import threading
        import time
        
        tiempo_inicio_total = time.time()
        self.log("[EXPORT TIMER] ⏱️ Iniciando proceso de exportación...", "INFO")
        
        try:
            # Permisos: STOCK.exportar
            allowed = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    allowed = self.permissions.tiene_permiso(self.current_user['id'], 'STOCK', 'exportar')
                if self.current_user and self.current_user.get('username','').lower() == 'admin':
                    allowed = True
            except Exception:
                allowed = False
            if not allowed:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para exportar datos de Stock")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENIED', usuario_id=self.current_user['id'], modulo='STOCK', detalle='exportar')
                except Exception:
                    pass
                return
            # Verificar si openpyxl está disponible
            try:
                import openpyxl
            except ImportError:
                messagebox.showerror(
                    "Módulo no encontrado", 
                    "Para exportar en Excel necesita instalar openpyxl:\n\n"
                    "pip install openpyxl\n\n"
                    "Use la exportación CSV como alternativa."
                )
                return
            
            # Verificar que existan localidades y depósitos configurados
            if not hasattr(self, 'stock_localidades') or not self.stock_localidades:
                messagebox.showwarning("Depósitos", "No hay depósitos configurados. Configure primero los depósitos.")
                return
            
            # Mostrar diálogo para seleccionar sedes/localidades a exportar
            tiempo_pre_dialogo = time.time()
            seleccionadas = self._seleccionar_sedes_para_exportar()
            tiempo_post_dialogo = time.time()
            self.log(f"[EXPORT TIMER] 📊 Diálogo de selección: {tiempo_post_dialogo - tiempo_pre_dialogo:.2f}s", "INFO")
            
            if not seleccionadas:
                return  # Usuario canceló o no seleccionó ninguna sede
            
            # Construir grupos por sede para exportar (usando selección marcada en el diálogo si está disponible)
            location_groups_dynamic = getattr(self, '_export_selected_location_groups', None)
            if not location_groups_dynamic:
                # Derivar de stock_localidades según intersección con seleccionadas
                location_groups_dynamic = {}
                locs = getattr(self, 'stock_localidades', {}) or {}
                for sede, deps in locs.items():
                    codes = [d['codigo'] for d in deps if d['codigo'] in seleccionadas]
                    if codes:
                        location_groups_dynamic[sede] = codes

            # 2. Obtener datos filtrados por los depósitos seleccionados para exportar
            tiempo_pre_filtrado = time.time()
            datos_exportar = self._obtener_datos_filtrados_por_depositos(seleccionadas)
            tiempo_post_filtrado = time.time()
            self.log(f"[EXPORT TIMER] 🔍 Filtrado de datos: {tiempo_post_filtrado - tiempo_pre_filtrado:.2f}s ({len(datos_exportar)} registros)", "INFO")
            
            if not datos_exportar:
                messagebox.showwarning("Sin datos", "No hay registros para exportar en las sedes seleccionadas")
                return

            # 3. Configurar progreso
            total_registros = len(datos_exportar)
            self.global_progress.pack(side=tk.RIGHT, padx=10)
            self.global_progress['value'] = 0
            self.global_progress['maximum'] = total_registros
            self.api_status.config(text="Creando Excel: 0%", foreground="#004C97")

            # Sugerir nombre por defecto y permitir al usuario elegir carpeta/archivo
            default_name = f"reporte_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            try:
                from tkinter import filedialog
                filename = filedialog.asksaveasfilename(
                    parent=self.root if hasattr(self, 'root') else None,
                    title="Guardar reporte de Stock como...",
                    defaultextension=".xlsx",
                    initialfile=default_name,
                    filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
                )
            except Exception as e:
                # Si falla el diálogo, caer al nombre por defecto en la carpeta actual
                self.log(f"[EXPORT] Error mostrando diálogo de guardar Stock: {e} - usando ruta por defecto", "WARNING")
                filename = default_name

            if not filename:
                # Usuario canceló
                self.log("[EXPORT] Usuario canceló el diálogo de guardar Stock", "INFO")
                self._cleanup_export_progress()
                return

            # Callback de progreso thread-safe
            def progress_cb(i, total):
                if i % max(1, total // 20) == 0 or i == total:
                    progreso = int((i / total) * 100)
                    # Usar after() para actualizar UI de forma segura desde otro hilo
                    self.root.after(0, lambda: self._update_export_progress(i, progreso, "Stock"))

            # Función para ejecutar la exportación en hilo separado
            def export_thread():
                tiempo_inicio_export = time.time()
                try:
                    from pal.services.exports import export_stock_excel
                    
                    self.root.after(0, lambda: self.log(f"[EXPORT TIMER] 📝 Iniciando escritura de Excel ({len(datos_exportar)} registros, {len(seleccionadas)} depósitos)...", "INFO"))
                    
                    total_registros = export_stock_excel(
                        filename=filename,
                        datos_exportar=datos_exportar,
                        seleccionadas=seleccionadas,
                        location_groups=location_groups_dynamic,
                        db_manager=self.db_manager,
                        progress_cb=progress_cb,
                        current_localidad=getattr(self, 'stock_localidad_actual', 'Cabudare'),
                    )
                    
                    tiempo_fin_export = time.time()
                    tiempo_total_export = tiempo_fin_export - tiempo_inicio_export
                    tiempo_total_proceso = tiempo_fin_export - tiempo_inicio_total
                    
                    self.root.after(0, lambda: self.log(
                        f"[EXPORT TIMER] ✅ Excel completado en {tiempo_total_export:.2f}s | Proceso total: {tiempo_total_proceso:.2f}s", 
                        "SUCCESS"
                    ))
                    
                    # Notificar éxito en el hilo principal
                    ubicaciones_info = f"🏢 Depósitos: {len(seleccionadas)}\n\n"
                    self.root.after(0, lambda: self._export_success(
                        "Stock",
                        total_registros,
                        filename,
                        f"{ubicaciones_info}"
                        "Características:\n"
                        "• 3 hojas: Datos, Resumen, Críticos\n"
                        "• Tablas con filtros automáticos\n"
                        "• Formato condicional por niveles\n"
                        "• Columnas auto-ajustadas"
                    ))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='STOCK', detalle=filename, exitoso=True)
                    except Exception:
                        pass
                    
                except Exception as e:
                    # Notificar error en el hilo principal
                    self.root.after(0, lambda err=str(e): self._export_error("Stock", err))
                    try:
                        if hasattr(self, 'audit_db') and self.current_user:
                            self.audit_db.log_action(
                                accion='EXPORT', usuario_id=self.current_user['id'], modulo='STOCK', detalle=str(e), exitoso=False)
                    except Exception:
                        pass
            
            # Iniciar exportación en hilo separado
            thread = threading.Thread(target=export_thread, daemon=True, name="ExportStock")
            thread.start()

        except Exception as e:
            self.log(f"Error iniciando exportación Excel: {str(e)}", "ERROR")
            self.api_status.config(text="API: Error", foreground="red")
            messagebox.showerror("Error en Exportación Excel", f"Error durante la exportación:\n{str(e)}")
            self._cleanup_export_progress()
    
    def _safe_update_api_status(self, text, color):
        """Actualiza el estado API de forma segura verificando que el widget exista"""
        try:
            if hasattr(self, 'api_status') and hasattr(self, 'root') and self.root.winfo_exists():
                self.api_status.config(text=text, foreground=color)
        except tk.TclError:
            pass  # Widget destruido, ignorar
                
            
            
    def _obtener_datos_filtrados_por_depositos(self, depositos_seleccionados):
        """Obtiene datos filtrados específicamente para los depósitos seleccionados en exportación"""
        try:
            # Si los depósitos seleccionados son los mismos que en configuración, usar cached_alertas
            depositos_configurados = set(getattr(self, 'stock_depositos_seleccionados', []) or [])
            depositos_exportar = set(depositos_seleccionados)
            
            # Si son exactamente los mismos, usar el cache
            if depositos_exportar == depositos_configurados:
                datos_base = list(self.cached_alertas)
            else:
                # Si son diferentes, recargar desde BD solo para los depósitos de exportación
                from pal.services.stock import get_alertas_stock
                
                # Obtener alertas para los depósitos específicos
                alertas_filtradas = get_alertas_stock(
                    self.db_manager, 
                    ubicaciones=list(depositos_seleccionados)
                )
                
                if not alertas_filtradas:
                    return []
                
                datos_base = list(alertas_filtradas)
            
            # Aplicar mismos filtros de UI que en aplicar_filtro_stock
            datos_filtrados = list(getattr(self, 'cached_alertas_effective', datos_base))

            # Exclusión por departamento (global)
            excluded_set = getattr(self, '_excluded_depts_set', set())
            if excluded_set:
                datos_filtrados = [
                    r for r in datos_filtrados
                    if not self._esta_excluido_por_departamento(r[0], excluded_set)
                ]
            
            # Filtro jerárquico
            dept_code = self.dept_dict.get(self.dept_var.get()) if hasattr(self, 'dept_dict') else None
            group_code = self.group_dict.get(self.group_var.get()) if hasattr(self, 'group_dict') else None
            sub_code = self.sub_dict.get(self.sub_var.get()) if hasattr(self, 'sub_dict') else None
        
            if any([dept_code, group_code, sub_code]):
                datos_filtrados = [
                    r for r in datos_filtrados 
                    if self._coincide_jerarquia(r[0], dept_code, group_code, sub_code)
                ]
        
            # Filtro texto
            texto_busqueda = self.search_var.get().strip().lower() if hasattr(self, 'search_var') else ''
            if texto_busqueda:
                datos_filtrados = [
                    r for r in datos_filtrados 
                    if texto_busqueda in (r[1].lower() + r[0].lower())
                ]
        
            # Filtro nivel
            filtro_nivel = self.filter_var.get().upper() if hasattr(self, 'filter_var') else 'TODAS'
            if filtro_nivel != 'TODAS':
                datos_filtrados = [r for r in datos_filtrados if r[3].upper() == filtro_nivel]
        
            # Ordenar por severidad local (CRÍTICA, MEDIA, LEVE) y por stock de la localidad activa
            favoritos = self._get_favoritos_local()
            def _norm_nivel(n):
                s = str(n or '').upper()
                for a,b in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U')]:
                    s = s.replace(a,b)
                return s
            def _rank(n):
                s = _norm_nivel(n)
                return 0 if s == 'CRITICA' else 1 if s == 'MEDIA' else 2 if s == 'LEVE' else 3
            def _fav(code):
                return 0 if str(code) in favoritos else 1
            return sorted(
                datos_filtrados,
                key=lambda r: (
                    _fav(r[0]),
                    _rank(r[3] if len(r) > 3 else ''),
                    int(r[2] or 0) if len(r) > 2 and r[2] is not None else 0,
                    str(r[0])
                )
            )
            
        except Exception as e:
            self.log(f"Error filtrando datos por depósitos: {e}", "ERROR")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}", "ERROR")
            return []
    
    def _obtener_datos_filtrados(self):
        """Réplica de la lógica de filtrado SIN paginación"""
        # Aplicar mismos filtros que en aplicar_filtro_stock
        datos_filtrados = list(getattr(self, 'cached_alertas_effective', self.cached_alertas))
    
        # Exclusión por departamento (global)
        excluded_set = getattr(self, '_excluded_depts_set', set())
        if excluded_set:
            datos_filtrados = [
                r for r in datos_filtrados
                if not self._esta_excluido_por_departamento(r[0], excluded_set)
            ]

        # Filtro jerárquico
        dept_code = self.dept_dict.get(self.dept_var.get())
        group_code = self.group_dict.get(self.group_var.get())
        sub_code = self.sub_var.get() and self.sub_dict.get(self.sub_var.get())
    
        if any([dept_code, group_code, sub_code]):
            datos_filtrados = [
                r for r in datos_filtrados 
                if self._coincide_jerarquia(r[0], dept_code, group_code, sub_code)
            ]
    
        # Filtro texto
        texto_busqueda = self.search_var.get().strip().lower()
        if texto_busqueda:
            datos_filtrados = [
                r for r in datos_filtrados 
                if texto_busqueda in (r[1].lower() + r[0].lower())
            ]
    
        # Filtro nivel
        filtro_nivel = self.filter_var.get().upper()
        if filtro_nivel != 'TODAS':
            datos_filtrados = [r for r in datos_filtrados if r[3].upper() == filtro_nivel]
    
        # Ordenar por severidad local (CRÍTICA, MEDIA, LEVE) y por stock de la localidad activa
        favoritos = self._get_favoritos_local()
        def _norm_nivel(n):
            s = str(n or '').upper()
            for a,b in [('Á','A'),('É','E'),('Í','I'),('Ó','O'),('Ú','U')]:
                s = s.replace(a,b)
            return s
        def _rank(n):
            s = _norm_nivel(n)
            return 0 if s == 'CRITICA' else 1 if s == 'MEDIA' else 2 if s == 'LEVE' else 3
        def _fav(code):
            return 0 if str(code) in favoritos else 1
        return sorted(
            datos_filtrados,
            key=lambda r: (
                _fav(r[0]),
                _rank(r[3] if len(r) > 3 else ''),
                int(r[2] or 0) if len(r) > 2 and r[2] is not None else 0,
                str(r[0])
            )
        )
    
    # === Funciones auxiliares para exportación asíncrona ===
    
    def _update_export_progress(self, current, percentage, module_name):
        """Actualiza la barra de progreso y el estado durante la exportación (thread-safe)"""
        try:
            if hasattr(self, 'global_progress') and self.global_progress.winfo_exists():
                self.global_progress['value'] = current
            if hasattr(self, 'api_status'):
                self.api_status.config(text=f"Exportando {module_name}: {percentage}%", foreground="#004C97")
        except tk.TclError:
            pass  # Widget destruido, ignorar
    
    def _export_success(self, module_name, total_registros, filename, extra_info=""):
        """Muestra mensaje de éxito tras completar la exportación (thread-safe)"""
        try:
            self._cleanup_export_progress()
            self.api_status.config(text="API: Lista", foreground="green")
            
            messagebox.showinfo(
                f"Exportación {module_name} Exitosa",
                f"Reporte {module_name} generado con éxito:\n\n"
                f"• Registros: {total_registros}\n"
                f"{extra_info}\n\n"
                f"📁 Ruta: {os.path.abspath(filename)}"
            )
            
            # Restaurar estado después de 3 segundos
            self.root.after(3000, lambda: self._safe_update_api_status("API: Lista", "green"))
            
        except Exception as e:
            self.log(f"Error mostrando mensaje de éxito: {e}", "ERROR")
    
    def _export_error(self, module_name, error_msg):
        """Muestra mensaje de error tras fallar la exportación (thread-safe)"""
        try:
            self._cleanup_export_progress()
            self.api_status.config(text="API: Error", foreground="red")
            self.log(f"Error en exportación {module_name}: {error_msg}", "ERROR")
            
            messagebox.showerror(
                f"Error en Exportación {module_name}",
                f"Error durante la exportación:\n{error_msg}"
            )
            
            # Restaurar estado después de 3 segundos
            self.root.after(3000, lambda: self._safe_update_api_status("API: Lista", "green"))
            
        except Exception as e:
            self.log(f"Error mostrando mensaje de error: {e}", "ERROR")
    
    def _cleanup_export_progress(self):
        """Limpia la barra de progreso tras finalizar exportación (thread-safe)"""
        try:
            if hasattr(self, 'global_progress') and self.global_progress.winfo_exists():
                self.global_progress.pack_forget()
        except tk.TclError:
            pass  # Widget ya fue destruido

    def buscar_por_fecha(self):
        # Obtener fechas como objetos datetime.datetime (incluyendo hora)
        fecha_inicio = datetime.combine(self.fecha_inicio.get_date(), datetime.min.time())
        fecha_fin = datetime.combine(self.fecha_fin.get_date(), datetime.max.time())

        query = "SELECT * FROM pal_envios_programados WHERE fecha_programada BETWEEN ? AND ?"
        params = (fecha_inicio, fecha_fin)  # Ahora son datetime.datetime

        records = self.db_manager.fetch_data(query, params)
        self.tree.delete(*self.tree.get_children())
        for row in records:
            self.tree.insert("", tk.END, values=row)

    def cargar_eventos_calendario(self):
        """Carga en el calendario los envíos PENDIENTES desde la BD y resalta las fechas."""
        try:
            if not hasattr(self, 'cal'):
                return
            # Limpiar eventos anteriores del calendario
            try:
                for ev_id in self.cal.get_calevents():
                    self.cal.calevent_remove(ev_id)
            except Exception:
                pass

            # Asegurar conexión y consultar envíos pendientes
            if not self.db_manager.ensure_connection():
                return
            rows = self.db_manager.fetch_data(
                "SELECT id, numero_cliente, fecha_programada, tipo_envio, codigo_producto, estado "
                "FROM pal_envios_programados WHERE estado = 'PENDIENTE'"
            )

            # Mapear eventos por fecha (YYYY-MM-DD)
            self.eventos_por_fecha = {}
            for id_envio, numero, fecha_prog, tipo, codigo, estado in (rows or []):
                try:
                    key = fecha_prog.strftime('%Y-%m-%d')
                except Exception:
                    key = str(fecha_prog)[:10]
                self.eventos_por_fecha.setdefault(key, []).append({
                    'id': id_envio,
                    'numero': numero,
                    'fecha': fecha_prog,
                    'tipo': (tipo or ''),
                    'codigo': codigo,
                    'estado': estado,
                })

            # Resaltar fechas en el calendario
            for key, lst in self.eventos_por_fecha.items():
                try:
                    y, m, d = [int(x) for x in key.split('-')]
                    dt = datetime(y, m, d)
                    self.cal.calevent_create(dt, f"{len(lst)} pendiente(s)", 'pendiente')
                except Exception:
                    continue
            try:
                self.cal.tag_config('pendiente', background='#FFB81C', foreground='#004C97')
            except Exception:
                pass

            # Mostrar detalles del día seleccionado
            self.mostrar_eventos_fecha(None)
        except Exception as e:
            try:
                self.log(f"Error cargando eventos de calendario: {e}", "ERROR")
            except Exception:
                pass

    def mostrar_eventos_fecha(self, event):
        """Muestra en el panel de detalles los envíos del día seleccionado."""
        try:
            if not hasattr(self, 'eventos_text') or not hasattr(self, 'cal'):
                return 0
            fecha_sel = self.cal.get_date()  # 'YYYY-MM-DD' por date_pattern
            key = str(fecha_sel)

            self.eventos_text.delete('1.0', tk.END)
            eventos = (getattr(self, 'eventos_por_fecha', {}) or {}).get(key, [])
            if not eventos:
                self.eventos_text.insert(tk.END, "Sin envíos pendientes para esta fecha.")
                return 0

            # Ordenar por hora si es posible
            try:
                eventos_sorted = sorted(eventos, key=lambda e: e['fecha'])
            except Exception:
                eventos_sorted = eventos

            for e in eventos_sorted:
                try:
                    hora = e['fecha'].strftime('%H:%M')
                except Exception:
                    hora = ''
                linea = (
                    f"#{e['id']} | {hora} | {e['tipo']} | Cliente: {e['numero']} | "
                    f"Producto: {e['codigo'] or '-'} | Estado: {e['estado']}\n"
                )
                self.eventos_text.insert(tk.END, linea)
            return 0
        except Exception as e:
            try:
                self.log(f"Error mostrando eventos: {e}", "ERROR")
            except Exception:
                pass
            return 0

    def programar_actualizaciones_stock(self):
        def actualizar():
            while True:
                if hasattr(self, 'last_refresh'):  # <-- Verificar atributo
                    self.actualizar_alertas_stock()
                time.sleep(3600)  # <-- Actualizar cada hora
        threading.Thread(target=actualizar, daemon=True).start()
                        


    def setup_bindings(self):
        """Configurar eventos del teclado y widgets"""
        # Doble click en la tabla (solo si existe)
        if hasattr(self, 'tree'):
            self.tree.bind("<Double-1>", lambda e: self.on_tree_double_click(e) or 0)
        
        # Validación en tiempo real del código de producto (solo si existe)
        if hasattr(self, 'cod_producto'):
            self.cod_producto.bind("<KeyRelease>", lambda e: self.buscar_descripcion(e) or 0)
        
        # Atajo de teclado para consola de debug (Ctrl+Shift+D)
        self.root.bind_all("<Control-Shift-D>", lambda e: (self.debug_console.toggle() or 0))

    def setup_modern_ui(self):   
        self.root.title("Casapro - Nexus")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 600)
        
        # create_header(self)  # Header eliminado para aprovechar espacio
        self.create_main_workspace()
        self.create_status_panel()

        style = ttk.Style()
        style.configure("TButton", background="#FFB81C", foreground="#004C97")
        style.map("TButton",
            background=[('active', '#e89f00')],
            foreground=[('active', '#003f7e')])
        
        # Sincronizar estados actuales tras crear los widgets
        if hasattr(self, 'db_manager') and self.db_manager.conn:
            self.update_status('connected', server=self.db_manager.server)
        
        if hasattr(self, 'api_state') and self.api_state == "active":
            self.update_status('api_connected')
        

    def load_tra_filters(self):
        """Carga departamentos, grupos y subgrupos para la pestaña T.R.A"""
        try:
            # Departamentos
            deps = self.db_manager.fetch_data(
                "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS"
            )
            self.tra_dept_dict = {desc: cod for cod, desc in deps if cod and desc}
            self.tra_dept_combo['values'] = ['Todos'] + list(self.tra_dept_dict.keys())
            self.tra_dept_var.set('Todos')
        except Exception as e:
            print("Error cargando departamentos del modulo tra:", e)

            # Inicializar grupos y subgrupos vacíos
            self.tra_group_dict = {}
            self.tra_group_combo['values'] = ['Todos']
            self.tra_group_var.set('Todos')

            self.tra_sub_dict = {}
            self.tra_sub_combo['values'] = ['Todos']
            self.tra_sub_var.set('Todos')

    def on_dept_selected(self):
        desc = self.dept_var.get()
        codigo = self.dept_dict.get(desc)
        if not codigo:
            self.group_combo['values'] = ['Todos']
            self.group_var.set('Todos')
            self.sub_combo['values'] = ['Todos']
            self.sub_var.set('Todos')
            self.aplicar_filtro_stock()
            return
        try:
            grupos = self.db_manager.fetch_data(
                "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WHERE C_DEPARTAMENTO = ?",
                (codigo,)
            )
        except Exception as e:
            print("Error cargando grupos:", e)
            grupos = []
        self.group_dict = {desc: cod for cod, desc in grupos if cod and desc}
        self.group_combo['values'] = ['Todos'] + list(self.group_dict.keys())
        self.group_var.set('Todos')
        self.sub_combo['values'] = ['Todos']
        self.sub_var.set('Todos')
        self.aplicar_filtro_stock()

    def on_tra_dept_selected(self, event=None):
        """Actualiza grupos cuando se selecciona departamento"""
        dept = self.tra_dept_var.get() if hasattr(self, 'tra_dept_var') else None
        dept_cod = self.tra_dept_dict.get(dept) if dept else None
        
        # DEBUG: Log de selección de departamento
        self.log(f"🔍 [TRA] Departamento seleccionado: '{dept}' (código: {dept_cod})", "DEBUG")
        self.log(f"🔍 [TRA] Diccionario tra_dept_dict tiene {len(self.tra_dept_dict)} departamentos", "DEBUG")
        self.log(f"🔍 [TRA] Diccionario tra_group_dict tiene {len(self.tra_group_dict)} claves", "DEBUG")

        # Resetear subgrupos
        if hasattr(self, 'tra_sub_combo'):
            self.tra_sub_combo['values'] = ['Todos']
            self.tra_sub_var.set('Todos')
    
        if dept_cod and hasattr(self, 'tra_group_combo'):
            grupos = list(self.tra_group_dict.get(dept_cod, {}).keys())
            self.tra_group_combo['values'] = ['Todos'] + grupos
            self.log(f"🔍 [TRA] Grupos cargados para dept {dept_cod}: {len(grupos)} grupos", "DEBUG")
        elif hasattr(self, 'tra_group_combo'):
            self.tra_group_combo['values'] = ['Todos']
            self.log(f"🔍 [TRA] No hay dept_cod, reseteando grupos a 'Todos'", "DEBUG")
    
        if hasattr(self, 'tra_group_var'):
            self.tra_group_var.set('Todos')
        
        # Resetear página actual y aplicar filtros
        self.tra_current_page = 1
        self.log(f"🔍 [TRA] Llamando aplicar_filtro_tra()...", "DEBUG")
        self.aplicar_filtro_tra()

    def on_group_selected(self):
        dept_desc = self.dept_var.get()
        grupo_desc = self.group_var.get()
        dept_codigo = self.dept_dict.get(dept_desc)
        grupo_codigo = self.group_dict.get(grupo_desc)
        if not (dept_codigo and grupo_codigo):
            self.sub_combo['values'] = ['Todos']
            self.sub_var.set('Todos')
            self.aplicar_filtro_stock()
            return
        try:
            subs = self.db_manager.fetch_data(
                "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WHERE C_IN_DEPARTAMENTO = ? AND C_IN_GRUPO = ?",
                (dept_codigo, grupo_codigo)
            )
        except Exception as e:
            print("Error cargando subgrupos:", e)
            subs = []
        self.sub_dict = {desc: cod for cod, desc in subs if cod and desc}
        self.sub_combo['values'] = ['Todos'] + list(self.sub_dict.keys())
        self.sub_var.set('Todos')
        def esperar_inicio():
            if self.db_manager.conn:
                self.aplicar_filtro_stock()
            else:
                self.root.after(100, esperar_inicio)

        esperar_inicio()

    def on_tra_group_selected(self, event=None):
        """Actualiza subgrupos cuando se selecciona grupo"""
        dept = self.tra_dept_var.get() if hasattr(self, 'tra_dept_var') else None
        group = self.tra_group_var.get() if hasattr(self, 'tra_group_var') else None
        dept_cod = self.tra_dept_dict.get(dept) if dept else None
        group_cod = self.tra_group_dict.get(dept_cod, {}).get(group) if dept_cod else None
        
        # DEBUG: Log de selección de grupo
        self.log(f"🔍 [TRA] Grupo seleccionado: '{group}' (dept: {dept}, dept_cod: {dept_cod}, group_cod: {group_cod})", "DEBUG")

        if hasattr(self, 'tra_sub_combo'):
            if dept_cod and group_cod:
                # Usar string como key (formato: "dept|group")
                key = f"{dept_cod}|{group_cod}"
                subgrupos = list(self.tra_sub_dict.get(key, {}).keys())
                self.tra_sub_combo['values'] = ['Todos'] + subgrupos
                self.log(f"🔍 [TRA] Subgrupos cargados para {key}: {len(subgrupos)} subgrupos", "DEBUG")
            else:
                self.tra_sub_combo['values'] = ['Todos']
                self.log(f"🔍 [TRA] No hay dept_cod/group_cod, reseteando subgrupos a 'Todos'", "DEBUG")
    
        if hasattr(self, 'tra_sub_var'):
            self.tra_sub_var.set('Todos')
        
        # Resetear página actual y aplicar filtros
        self.tra_current_page = 1
        self.log(f"🔍 [TRA] Llamando aplicar_filtro_tra()...", "DEBUG")
        self.aplicar_filtro_tra()

        
        
    
    def cargar_jerarquia_productos(self):
        """Filtra la jerarquía usando solo los códigos actualmente en alerta."""
        start = time.perf_counter()
        try:
            if not hasattr(self, 'all_jerarquia') or not self.all_jerarquia:
                self.log("Jerarquía vacía, iniciando carga completa", "WARNING")
                self._cargar_toda_jerarquia_productos()

            # Normalizar códigos antes de filtrar
            def _s(x):
                try:
                    return str(x).strip()
                except Exception:
                    return ""
            codigos_en_alerta = {_s(r[0]) for r in self.cached_alertas}
            all_jerarquia_norm = {_s(k): v for k, v in (self.all_jerarquia or {}).items()}
            self.producto_jerarquia = {cod: all_jerarquia_norm[cod] for cod in codigos_en_alerta if cod in all_jerarquia_norm}
            elapsed = time.perf_counter() - start
            self.log(f"🔍 Filtrado de jerarquía: {len(self.producto_jerarquia)} en {elapsed:.2f}s", "DEBUG")
        except Exception as e:
            self.log(f"Error filtrando jerarquía: {e}", "ERROR")
            self.producto_jerarquia = {}

    def _cargar_toda_jerarquia_productos(self):
        """Carga el mapeo completo de productos a jerarquía con caché local."""
        start = time.perf_counter()
        try:
            if os.path.exists(JERARQUIA_CACHE_FILE):
                mtime = datetime.fromtimestamp(os.path.getmtime(JERARQUIA_CACHE_FILE))
                if datetime.now() - mtime < JERARQUIA_CACHE_TTL:
                    self.log("Cargando jerarquía desde cache", "INFO")
                    with open(JERARQUIA_CACHE_FILE, 'r', encoding='utf-8') as f:
                        self.all_jerarquia = json.load(f)
                    elapsed = time.perf_counter() - start
                    self.log(f"✅ Jerarquía leída en {elapsed:.2f}s (productos: {len(self.all_jerarquia)})", "DEBUG")
                    return
                else:
                    self.log("Cache vencido, recargando jerarquía desde BD", "INFO")
            else:
                self.log("Cache no existe, cargando jerarquía desde BD", "INFO")

            filas = self.db_manager.fetch_data(
                "SELECT C_CODIGO, C_DEPARTAMENTO, C_GRUPO, C_SUBGRUPO FROM MA_PRODUCTOS"
            )
            # Normalizar y cargar incluyendo productos con campos vacíos
            def _s(x):
                try:
                    s = str(x).strip()
                    return s if s and s.lower() != 'none' else ""
                except Exception:
                    return ""
            self.all_jerarquia = {}
            for fila in filas or []:
                try:
                    if not fila:
                        continue
                    cod = _s(fila[0])
                    if not cod:
                        continue
                    dep = _s(fila[1]) if len(fila) > 1 else ""
                    grp = _s(fila[2]) if len(fila) > 2 else ""
                    sub = _s(fila[3]) if len(fila) > 3 else ""
                    self.all_jerarquia[cod] = (dep, grp, sub)
                except Exception:
                    continue
            with open(JERARQUIA_CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.all_jerarquia, f, ensure_ascii=False)
            elapsed = time.perf_counter() - start
            self.log(f"✅ Jerarquía cargada y cacheada ({len(self.all_jerarquia)} productos) en {elapsed:.2f}s", "SUCCESS")

        except Exception as e:
            self.log(f"Error cargando jerarquía: {e}", "ERROR")
            self.all_jerarquia = {}

    def _init_jerarquia_async(self):
        """Thread en background para cargar y filtrar jerarquía sin bloquear UI."""
        total_start = time.perf_counter()
        try:
            from pal.services.stock import load_all_jerarquia, build_producto_jerarquia
            # Cargar jerarquía (con caché)
            all_jerarquia = load_all_jerarquia(
                self.db_manager,
                JERARQUIA_CACHE_FILE,
                int(JERARQUIA_CACHE_TTL.total_seconds())
            )
            self.all_jerarquia = all_jerarquia or {}
            # Filtrar sólo códigos en alerta
            codigos_en_alerta = {r[0] for r in getattr(self, 'cached_alertas', [])}
            self.producto_jerarquia = build_producto_jerarquia(self.all_jerarquia, codigos_en_alerta)
        except Exception as e:
            self.log(f"Error inicializando jerarquía: {e}", "ERROR")
            self.producto_jerarquia = {}
        elapsed = time.perf_counter() - total_start
        self.log(f"🏁 Jerarquía inicializada en {elapsed:.2f}s", "INFO")
        # Actualizar UI en hilo principal
        try:
            self.root.after(0, self.aplicar_filtro_stock)
        except Exception:
            pass

    def load_stock_filters(self):
        """Sincroniza los combos de Stock con la jerarquía unificada de la aplicación."""
        if not hasattr(self, 'stock_dept_combo'): return
        
        # Si no hay datos cargados aún, no hacemos nada
        if not hasattr(self, 'tra_dept_dict') or not self.tra_dept_dict:
            return

        try:
            # Población directa desde el diccionario unificado (igual que RI/MBRP)
            depts = sorted(list(self.tra_dept_dict.keys()))
            self.stock_dept_combo['values'] = ['Todos'] + depts
            
            # Asegurar que el valor actual sea válido
            if self.stock_dept_var.get() not in self.stock_dept_combo['values']:
                self.stock_dept_var.set('Todos')
                # Si reseteamos a Todos, limpiar hijos
                self.stock_group_combo['values'] = ['Todos']
                self.stock_group_var.set('Todos')
                self.stock_subgroup_combo['values'] = ['Todos']
                self.stock_subgroup_var.set('Todos')
            
            self.log("✅ Filtros de stock sincronizados con jerarquía unificada", "DEBUG")
            
        except Exception as e:
            self.log(f"Error sincronizando filtros stock: {e}", "ERROR")

    def on_stock_dept_selected(self, event=None):
        """Maneja el cambio de departamento en stock usando la jerarquía unificada"""
        dept_desc = self.stock_dept_var.get()
        dept_cod = self.tra_dept_dict.get(dept_desc)
        
        if dept_desc == 'Todos' or not dept_cod:
            self.stock_group_combo['values'] = ['Todos']
            self.stock_group_var.set('Todos')
            self.stock_subgroup_combo['values'] = ['Todos']
            self.stock_subgroup_var.set('Todos')
        else:
            # Obtener grupos directamente del diccionario de TRA
            grupos_dict = self.tra_group_dict.get(dept_cod, {})
            self.stock_group_combo['values'] = ['Todos'] + sorted(list(grupos_dict.keys()))
            self.stock_group_var.set('Todos')
            self.stock_subgroup_combo['values'] = ['Todos']
            self.stock_subgroup_var.set('Todos')
            
        self.aplicar_filtro_stock()

    def on_stock_group_selected(self, event=None):
        """Maneja el cambio de grupo en stock usando la jerarquía unificada"""
        dept_desc = self.stock_dept_var.get()
        group_desc = self.stock_group_var.get()
        
        dept_cod = self.tra_dept_dict.get(dept_desc)
        group_cod = self.tra_group_dict.get(dept_cod, {}).get(group_desc) if dept_cod else None
        
        if group_desc == 'Todos' or not group_cod:
            self.stock_subgroup_combo['values'] = ['Todos']
            self.stock_subgroup_var.set('Todos')
        else:
            # Obtener subgrupos usando la misma clave que TRA (dept|group)
            key = f"{dept_cod}|{group_cod}"
            subs_dict = self.tra_sub_dict.get(key, {})
            self.stock_subgroup_combo['values'] = ['Todos'] + sorted(list(subs_dict.keys()))
            self.stock_subgroup_var.set('Todos')
            
        self.aplicar_filtro_stock()

    def cambiar_pagina_stock(self, delta):
        """Cambia la página actual de la vista de stock"""
        if not hasattr(self, 'stock_current_page'): self.stock_current_page = 1
        self.stock_current_page += delta
        self.aplicar_filtro_stock()

    def aplicar_filtro_stock(self):
        """Actualiza la vista del treeview de quiebres aplicando filtros y paginación"""
        if not hasattr(self, 'stock_tree') or not self.stock_tree.winfo_exists() or not hasattr(self, 'cached_alertas'):
            return
            
        try:
            from pal.services.stock import filter_alertas, paginate
            
            # Obtener valores de los nuevos filtros (y convertirlos a códigos si es necesario)
            dept_desc = self.stock_dept_var.get() if hasattr(self, 'stock_dept_var') else 'Todos'
            group_desc = self.stock_group_var.get() if hasattr(self, 'stock_group_var') else 'Todos'
            subgroup_desc = self.stock_subgroup_var.get() if hasattr(self, 'stock_subgroup_var') else 'Todos'
            search_text = self.stock_search_var.get() if hasattr(self, 'stock_search_var') else ''
            favoritos = getattr(self, 'favoritos', set())

            # Convertir descripciones a códigos para el filtro lógico usando diccionarios unificados
            dept_code = self.tra_dept_dict.get(dept_desc) if dept_desc != 'Todos' else None
            
            # El grupo y subgrupo dependen de sus padres
            group_code = None
            if group_desc != 'Todos' and dept_code:
                group_code = self.tra_group_dict.get(dept_code, {}).get(group_desc)
                
            subgroup_code = None
            if subgroup_desc != 'Todos' and dept_code and group_code:
                key = f"{dept_code}|{group_code}"
                subgroup_code = self.tra_sub_dict.get(key, {}).get(subgroup_desc)
            
            # Usar vista efectiva (sin departamentos excluidos)
            alertas_base = getattr(self, 'cached_alertas_effective', self.cached_alertas)
            
            # Filtrar
            filtrados = filter_alertas(
                alertas=alertas_base,
                producto_jerarquia=getattr(self, 'all_jerarquia', {}),
                dept_code=dept_code,
                group_code=group_code,
                subgroup_code=subgroup_code,
                search_text=search_text,
                favoritos=favoritos
            )
            
            # Paginación
            page_data, total_pages, current = paginate(
                filtrados, 
                getattr(self, 'stock_current_page', 1), 
                getattr(self, 'stock_page_size', 250)
            )
            
            self.stock_current_page = current
            if hasattr(self, 'stock_pagination_label'):
                self.stock_pagination_label.config(text=f"Página {current}/{total_pages}")
            
            # Limpiar y llenar treeview
            self.stock_tree.delete(*self.stock_tree.get_children())
            
            for q in page_data:
                # q: (codigo, descripcion, sede, unid_perd, dias, ult_compra, ult_venta)
                codigo = str(q[0])
                is_fav = "⭐" if codigo in favoritos else "☆"
                
                # Formatear unidades para visualización (entero redondeado)
                try:
                    u_display = int(round(float(q[3])))
                except (ValueError, TypeError):
                    u_display = 0
                
                # Crear nueva tupla para visualización con el valor redondeado
                row_values = (is_fav, q[0], q[1], q[2], u_display, q[4], q[5], q[6])
                
                if codigo in favoritos:
                    tags = ('favorito',)
                elif hasattr(self, 'new_stock_break_codes') and codigo in self.new_stock_break_codes:
                    tags = ('nuevo_quiebre',)
                else:
                    tags = ('quiebre',)
                    
                self.stock_tree.insert("", tk.END, values=row_values, tags=tags)
                
        except Exception as e:
            self.log(f"Error filtrando quiebres stock: {e}", "DEBUG")
        # Diagnóstico: distribución tras aplicar filtro seleccionado (solo en modo debug)
        # Comentado para reducir spam en logs
        # try:
        #     leves = medias = criticas = 0
        #     for _, _, stock_val, _ in filtrados:
        #         try:
        #             s = int(stock_val or 0)
        #         except Exception:
        #             s = 0
        #         if s >= 15:
        #             leves += 1
        #         elif s >= 8:
        #             medias += 1
        #         else:
        #             criticas += 1
        #     self.log(
        #         f"Filtro='{filtro_nivel}', total filtradas={len(filtrados)} | Leves:{leves} Medias:{medias} Críticas:{criticas}",
        #         "DEBUG"
        #     )
        # except Exception:
        #     pass

    def _coincide_jerarquia(self, codigo, dept_code, group_code, sub_code):
        """Helper function para filtro jerárquico optimizado"""
        jerarquia = self.producto_jerarquia.get(codigo)
        if not jerarquia:
            return False
    
        dep, grp, sub = jerarquia
        return  (not dept_code or dep == dept_code) and \
                (not group_code or grp == group_code) and \
                (not sub_code or sub == sub_code)
    
    def _coincide_jerarquia_tra(self, codigo, tra_dept_code, tra_group_code, tra_sub_code):
        """Helper function para filtro jerárquico optimizado"""
        jerarquia = self.producto_jerarquia.get(codigo)
        if not jerarquia:
            return False
    
        dep, grp, sub = jerarquia
        return  (not tra_dept_code or dep == tra_dept_code) and \
                (not tra_group_code or grp == tra_group_code) and \
                (not tra_sub_code or sub == tra_sub_code)

    def _esta_excluido_por_departamento(self, codigo, excluded_depts_set):
        """Devuelve True si el producto pertenece a un departamento excluido."""
        try:
            jerarquia = self.producto_jerarquia.get(codigo)
            if not jerarquia:
                return False
            dep = str(jerarquia[0])
            return dep in excluded_depts_set
        except Exception:
            return False

    def _update_excluded_set(self):
        """Construye y cachea el set de departamentos excluidos para uso rápido."""
        try:
            self._excluded_depts_set = set(str(x) for x in (getattr(self, 'excluded_depts', []) or []))
        except Exception:
            self._excluded_depts_set = set()

    def _rebuild_effective_views(self):
        """Reconstruye vistas 'efectivas' sin departamentos excluidos para evitar filtrar en cada refresco."""
        try:
            # Asegurar set cacheado
            self._update_excluded_set()
            excluded = getattr(self, '_excluded_depts_set', set())

            # Stock: excluir por códigos mapeados a depto
            try:
                if hasattr(self, 'cached_alertas') and self.cached_alertas:
                    if hasattr(self, 'producto_jerarquia') and self.producto_jerarquia and excluded:
                        # Construir set de códigos excluidos una sola vez
                        excl_codes = {str(code) for code, (dep, *_rest) in self.producto_jerarquia.items() if str(dep) in excluded}
                        self.cached_alertas_effective = [r for r in self.cached_alertas if str(r[0]) not in excl_codes]
                    else:
                        self.cached_alertas_effective = list(self.cached_alertas)
                else:
                    self.cached_alertas_effective = []
            except Exception:
                self.cached_alertas_effective = list(getattr(self, 'cached_alertas', []) or [])

            # TRA: índice 2 es depto
            try:
                if hasattr(self, 'cached_ventas_tra') and self.cached_ventas_tra:
                    if excluded:
                        self.cached_ventas_tra_effective = [r for r in self.cached_ventas_tra if len(r) > 2 and str(r[2]) not in excluded]
                    else:
                        self.cached_ventas_tra_effective = list(self.cached_ventas_tra)
                else:
                    self.cached_ventas_tra_effective = []
            except Exception:
                self.cached_ventas_tra_effective = list(getattr(self, 'cached_ventas_tra', []) or [])

            # MBRP: índice 2 es depto
            try:
                if hasattr(self, 'cached_ventas_mbrp') and self.cached_ventas_mbrp:
                    if excluded:
                        self.cached_ventas_mbrp_effective = [r for r in self.cached_ventas_mbrp if len(r) > 2 and str(r[2]) not in excluded]
                    else:
                        self.cached_ventas_mbrp_effective = list(self.cached_ventas_mbrp)
                else:
                    self.cached_ventas_mbrp_effective = []
            except Exception:
                self.cached_ventas_mbrp_effective = list(getattr(self, 'cached_ventas_mbrp', []) or [])
            
            # Precargar mapeo producto->proveedor para estadísticas coherentes.
            # IMPORTANTE: esto puede implicar consultas pesadas a BD, por lo que se
            # ejecuta en un hilo separado para no bloquear la UI.
            try:
                threading.Thread(
                    target=self._preload_productos_proveedores,
                    daemon=True,
                    name="PreloadProdProv"
                ).start()
            except Exception:
                # Si por alguna razón no se puede crear el hilo, intentar al menos
                # ejecutar de forma síncrona sin romper la aplicación.
                self._preload_productos_proveedores()
            
        except Exception:
            pass
        
    def actualizar_controles_paginacion(self, total_paginas):
        """Actualiza los controles de paginación"""
        if not hasattr(self, 'pagination_label'):
            return
        self.pagination_label.config(text=f"Página {self.current_page}/{total_paginas}")
        if hasattr(self, 'btn_prev'):
            self.btn_prev['state'] = 'normal' if self.current_page > 1 else 'disabled'
        if hasattr(self, 'btn_next'):
            self.btn_next['state'] = 'normal' if self.current_page < total_paginas else 'disabled'
    
    def load_depositos_stock(self):
        """Carga lista de depósitos desde la base de datos y aplica sedes personalizadas (opcional)"""
        try:
            depositos = self.db_manager.obtener_depositos()
            if not depositos:
                self.log("No se encontraron depósitos en la BD", "WARNING")
                return
            
            # Agrupar por localidad según prefijo (base por defecto)
            localidades = {}
            for cod, desc in depositos:
                cod = str(cod).strip()
                desc = str(desc).strip()
                
                # Determinar localidad por prefijo
                if cod.startswith('03'):
                    localidad = 'Cabudare'
                elif cod.startswith('01'):
                    localidad = 'Barinas'
                elif cod.startswith('04'):
                    localidad = 'Guanare'
                else:
                    localidad = 'Otra'
                
                if localidad not in localidades:
                    localidades[localidad] = []
                localidades[localidad].append({'codigo': cod, 'descripcion': desc})
            
            # Map rápido código->descripción
            self.stock_depositos_por_codigo = {str(c).strip(): str(d).strip() for c, d in depositos}

            # Aplicar sedes personalizadas si existen en preferencias
            sedes_custom = getattr(self, 'stock_localidades_custom', {}) or {}
            if sedes_custom:
                # Remover códigos personalizados de sus sedes originales
                custom_all_codes = set()
                for sede, codes in sedes_custom.items():
                    for cod in (codes or []):
                        custom_all_codes.add(str(cod).strip())
                if custom_all_codes:
                    for sede_base, lst in list(localidades.items()):
                        localidades[sede_base] = [d for d in lst if d['codigo'] not in custom_all_codes]
                # Agregar sedes personalizadas con sus depósitos
                for sede, codes in sedes_custom.items():
                    sede = str(sede).strip()
                    if sede not in localidades:
                        localidades[sede] = []
                    for cod in (codes or []):
                        c = str(cod).strip()
                        desc = self.stock_depositos_por_codigo.get(c, c)
                        if all(d['codigo'] != c for d in localidades[sede]):
                            localidades[sede].append({'codigo': c, 'descripcion': desc})
            
            # Guardar en atributos
            self.stock_localidades = localidades
            
            # Inicializar seleccionados (default a Cabudare)
            if not hasattr(self, 'stock_depositos_seleccionados'):
                self.stock_depositos_seleccionados = [d['codigo'] for d in localidades.get('Cabudare', [])]
            
            # Actualizar label de depósitos y columnas del tree si existe
            self._update_stock_depositos_label()
            if hasattr(self, '_rebuild_stock_tree_columns') and hasattr(self, 'stock_tree') and self.stock_tree:
                try:
                    self._rebuild_stock_tree_columns()
                except Exception:
                    pass
            
            self.log(f"✅ Depósitos cargados: {len(depositos)} depósitos en {len(localidades)} localidades (incluye personalizadas: {len(sedes_custom)})", "SUCCESS")
            
        except Exception as e:
            self.log(f"Error cargando depósitos: {e}", "ERROR")
    
    def _update_stock_depositos_label(self):
        """Actualiza el label con los depósitos seleccionados"""
        try:
            if not hasattr(self, 'stock_depositos_seleccionados'):
                self.stock_depositos_seleccionados = ['0301']
            
            if hasattr(self, 'stock_depositos_label'):
                # Crear texto descriptivo
                descs = [self.stock_depositos_por_codigo.get(cod, cod) 
                        for cod in sorted(self.stock_depositos_seleccionados)]
                texto = ', '.join(descs) if descs else '[Sin seleccionar]'
                self.stock_depositos_label.config(text=f"📦 {texto}")
        except Exception as e:
            self.log(f"Error actualizando label de depósitos: {e}", "DEBUG")
    
    def abrir_menu_depositos_stock(self):
        """Abre diálogo para configurar localidad activa y depósitos por cada sede (persistentes)."""
        try:
            import tkinter as tk
            from tkinter import ttk
            
            if not hasattr(self, 'stock_localidades') or not self.stock_localidades:
                self.log("Cargando depósitos...", "INFO")
                return
            
            ventana = tk.Toplevel(self.root)
            ventana.title("Configurar Localidad y Depósitos")
            ventana.geometry("900x680")
            ventana.resizable(False, False)
            ventana.grab_set()

            # Localidad activa
            frame_loc = ttk.LabelFrame(ventana, text="1. Localidad activa", padding=10)
            frame_loc.pack(fill=tk.X, padx=10, pady=(10,5))
            ttk.Label(frame_loc, text="Localidad:" ).pack(side=tk.LEFT)
            loc_var = tk.StringVar(value=getattr(self, 'stock_localidad_actual', 'Cabudare'))
            ttk.Combobox(frame_loc, textvariable=loc_var, state='readonly',
                         values=sorted(self.stock_localidades.keys()), width=20).pack(side=tk.LEFT, padx=8)

            # Gestión de sedes personalizadas
            frame_custom = ttk.LabelFrame(ventana, text="2. Sedes personalizadas (agrupaciones manuales)", padding=10)
            frame_custom.pack(fill=tk.X, padx=10, pady=5)

            # Fila 1: Crear sede personalizada
            row_add = ttk.Frame(frame_custom)
            row_add.pack(fill=tk.X, pady=3)
            new_sede_var = tk.StringVar()
            ttk.Label(row_add, text="Nueva sede:").pack(side=tk.LEFT)
            entry_new_sede = ttk.Entry(row_add, textvariable=new_sede_var, width=22)
            entry_new_sede.pack(side=tk.LEFT, padx=6)
            def add_sede():
                nombre = (new_sede_var.get() or '').strip()
                if not nombre:
                    return
                if not hasattr(self, 'stock_localidades_custom') or self.stock_localidades_custom is None:
                    self.stock_localidades_custom = {}
                if nombre in (self.stock_localidades_custom or {}):
                    messagebox.showinfo("Info", f"La sede '{nombre}' ya existe")
                    return
                self.stock_localidades_custom[nombre] = []
                # Reaplicar y reabrir diálogo para refrescar UI
                self.load_depositos_stock()
                ventana.destroy()
                self.abrir_menu_depositos_stock()
            ttk.Button(row_add, text="➕ Agregar sede", command=add_sede).pack(side=tk.LEFT, padx=6)

            # Fila 2: Eliminar sede personalizada
            row_del = ttk.Frame(frame_custom)
            row_del.pack(fill=tk.X, pady=3)
            ttk.Label(row_del, text="Eliminar sede:").pack(side=tk.LEFT)
            del_sede_var = tk.StringVar()
            def _custom_sedes_list():
                sc = getattr(self, 'stock_localidades_custom', {}) or {}
                return sorted([s for s in sc.keys() if s not in ['Cabudare','Barinas','Guanare','Otra']])
            cb_del_sede = ttk.Combobox(row_del, textvariable=del_sede_var, width=22, values=_custom_sedes_list(), state='readonly')
            cb_del_sede.pack(side=tk.LEFT, padx=6)
            def eliminar_sede():
                sede = (del_sede_var.get() or '').strip()
                if not sede:
                    return
                if messagebox.askyesno("Confirmar", f"¿Eliminar la sede personalizada '{sede}'?\nLos depósitos volverán a su sede base.", parent=ventana):
                    try:
                        if hasattr(self, 'stock_localidades_custom') and sede in (self.stock_localidades_custom or {}):
                            self.stock_localidades_custom.pop(sede, None)
                        if hasattr(self, 'stock_depositos_por_sede') and isinstance(self.stock_depositos_por_sede, dict):
                            self.stock_depositos_por_sede.pop(sede, None)
                    except Exception:
                        pass
                    self.load_depositos_stock()
                    ventana.destroy()
                    self.abrir_menu_depositos_stock()
            ttk.Button(row_del, text="🗑️ Eliminar", command=eliminar_sede).pack(side=tk.LEFT, padx=6)

            # Fila 3: Mover depósito a sede
            row_move = ttk.Frame(frame_custom)
            row_move.pack(fill=tk.X, pady=3)
            ttk.Label(row_move, text="Mover depósito:").pack(side=tk.LEFT)
            all_deps = sorted([(c, self.stock_depositos_por_codigo.get(c, c)) for c in self.stock_depositos_por_codigo.keys()])
            move_dep_var = tk.StringVar()
            cb_dep = ttk.Combobox(row_move, textvariable=move_dep_var, width=14, values=[c for c,_ in all_deps])
            cb_dep.pack(side=tk.LEFT, padx=6)
            ttk.Label(row_move, text="a sede:").pack(side=tk.LEFT)
            move_sede_var = tk.StringVar()
            sedes_names = sorted((self.stock_localidades or {}).keys())
            cb_sede = ttk.Combobox(row_move, textvariable=move_sede_var, width=20, values=sedes_names, state='readonly')
            cb_sede.pack(side=tk.LEFT, padx=6)
            def mover_dep():
                dep = (move_dep_var.get() or '').strip()
                sede_dest = (move_sede_var.get() or '').strip()
                if not dep or not sede_dest:
                    return
                if not hasattr(self, 'stock_localidades_custom') or self.stock_localidades_custom is None:
                    self.stock_localidades_custom = {}
                # Remover de cualquier sede personalizada previa
                for s in list(self.stock_localidades_custom.keys()):
                    lst = self.stock_localidades_custom.get(s) or []
                    if dep in lst:
                        lst = [x for x in lst if x != dep]
                        self.stock_localidades_custom[s] = lst
                # Si el destino es una sede personalizada, asignar ahí; si es base, no lo incluimos en custom (volverá a su sede por prefijo)
                if sede_dest not in ['Cabudare','Barinas','Guanare','Otra']:
                    self.stock_localidades_custom.setdefault(sede_dest, [])
                    if dep not in self.stock_localidades_custom[sede_dest]:
                        self.stock_localidades_custom[sede_dest].append(dep)
                # Reaplicar y recargar UI
                self.load_depositos_stock()
                ventana.destroy()
                self.abrir_menu_depositos_stock()
            ttk.Button(row_move, text="➡️ Mover", command=mover_dep).pack(side=tk.LEFT, padx=6)

            # Depósitos por sede
            frame_deps = ttk.LabelFrame(ventana, text="3. Seleccionar Depósitos por sede (visibilidad)", padding=10)
            frame_deps.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

            canvas = tk.Canvas(frame_deps)
            vsb = ttk.Scrollbar(frame_deps, orient="vertical", command=canvas.yview)
            inner = ttk.Frame(canvas)
            inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0,0), window=inner, anchor="nw")
            canvas.configure(yscrollcommand=vsb.set)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)

            # Estado inicial: usar mapa por sede si existe, si no derivar de lista combinada
            if not hasattr(self, 'stock_depositos_por_sede') or not self.stock_depositos_por_sede:
                self.stock_depositos_por_sede = {}
                seleccionados = set(getattr(self, 'stock_depositos_seleccionados', []) or [])
                for sede, deps in (self.stock_localidades or {}).items():
                    self.stock_depositos_por_sede[sede] = [d['codigo'] for d in deps if d['codigo'] in seleccionados]
            vars_por_sede = {}

            # Construir columnas por sede
            for sede in sorted(self.stock_localidades.keys()):
                lf = ttk.LabelFrame(inner, text=f"🏢 {sede}", padding=8)
                lf.pack(fill=tk.X, padx=6, pady=6)
                vars_por_sede[sede] = {}

                # Botones rápidos
                quick = ttk.Frame(lf)
                quick.pack(fill=tk.X, pady=(0,6))
                def _select_all(s=sede):
                    for v in vars_por_sede[s].values(): v.set(True)
                def _clear_all(s=sede):
                    for v in vars_por_sede[s].values(): v.set(False)
                ttk.Button(quick, text="Seleccionar todo", command=_select_all).pack(side=tk.LEFT)
                ttk.Button(quick, text="Limpiar", command=_clear_all).pack(side=tk.LEFT, padx=5)

                # Lista de depósitos
                for dep in self.stock_localidades[sede]:
                    cod = dep['codigo']; desc = dep['descripcion']
                    var = tk.BooleanVar(value=cod in (self.stock_depositos_por_sede.get(sede, []) or []))
                    vars_por_sede[sede][cod] = var
                    ttk.Checkbutton(lf, text=f"{cod} - {desc}", variable=var).pack(anchor=tk.W, padx=10)

            # Guardar
            frame_btns = ttk.Frame(ventana)
            frame_btns.pack(fill=tk.X, padx=10, pady=10)
            
            def guardar():
                self.stock_localidad_actual = loc_var.get() or 'Cabudare'
                # Recolectar por sede y combinado (visibilidad)
                self.stock_depositos_por_sede = {sede: [cod for cod, v in vars_por_sede[sede].items() if v.get()] for sede in vars_por_sede}
                self.stock_depositos_seleccionados = []
                for lst in self.stock_depositos_por_sede.values():
                    self.stock_depositos_seleccionados.extend(lst)
                # Actualizar columnas dinámicas y guardar preferencias (incluye sedes_custom)
                self._update_stock_depositos_label()
                self._save_stock_preferences()
                # Recargar
                self.current_page = 1
                self.recargar_stock()
                ventana.destroy()
                self.log(f"✅ Localidad: {self.stock_localidad_actual} | Depósitos seleccionados: {', '.join(self.stock_depositos_seleccionados)}", "SUCCESS")

            ttk.Button(frame_btns, text="✅ Guardar", command=guardar).pack(side=tk.LEFT, padx=5)
            ttk.Button(frame_btns, text="❌ Cancelar", command=ventana.destroy).pack(side=tk.LEFT, padx=5)

        except Exception as e:
            self.log(f"Error abriendo menú de depósitos: {e}", "ERROR")
    
    def _seleccionar_sedes_para_exportar(self):
        """Muestra diálogo para que el usuario seleccione qué sedes/localidades exportar"""
        import tkinter as tk
        from tkinter import ttk
        
        resultado = []
        
        try:
            # Crear ventana emergente
            ventana = tk.Toplevel(self.root)
            ventana.title("Seleccionar Sedes para Exportar")
            ventana.geometry("600x450")
            ventana.resizable(False, False)
            ventana.grab_set()
            
            # Título
            ttk.Label(
                ventana, 
                text="Seleccione las sedes/localidades que desea incluir en la exportación:",
                font=('Arial', 10, 'bold')
            ).pack(pady=15, padx=10)
            
            # Frame principal con scrollbar
            frame_main = ttk.Frame(ventana)
            frame_main.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            
            canvas = tk.Canvas(frame_main)
            scrollbar = ttk.Scrollbar(frame_main, orient="vertical", command=canvas.yview)
            frame_sedes = ttk.Frame(canvas)
            
            frame_sedes.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            canvas.create_window((0, 0), window=frame_sedes, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Variables para cada sede
            vars_sedes = {}
            depositos_por_sede = {}
            
            # Crear checkboxes por cada localidad/sede (incluye sedes personalizadas)
            for localidad in sorted(self.stock_localidades.keys()):
                depositos = self.stock_localidades[localidad]
                # Preferir selección por sede; fallback a selección combinada
                conf_por_sede = getattr(self, 'stock_depositos_por_sede', {}) or {}
                if localidad in conf_por_sede and conf_por_sede[localidad]:
                    depositos_filtrados = [d for d in depositos if d['codigo'] in conf_por_sede[localidad]]
                else:
                    seleccionados = set(getattr(self, 'stock_depositos_seleccionados', []) or [])
                    depositos_filtrados = [d for d in depositos if d['codigo'] in seleccionados]
                
                # Si no hay depósitos configurados para esta sede, omitirla
                if not depositos_filtrados:
                    continue
                
                depositos_codigos = [d['codigo'] for d in depositos_filtrados]
                depositos_por_sede[localidad] = depositos_codigos
                
                # Frame para cada sede
                frame_sede = ttk.LabelFrame(frame_sedes, text=f"🏢 {localidad}", padding=10)
                frame_sede.pack(fill=tk.X, padx=10, pady=5)
                
                # Variable de checkbox para la sede
                var_sede = tk.BooleanVar(value=False)
                vars_sedes[localidad] = var_sede
                
                # Checkbox principal de la sede
                chk_sede = ttk.Checkbutton(
                    frame_sede,
                    text=f"Incluir toda la sede ({len(depositos_filtrados)} depósito{'s' if len(depositos_filtrados) > 1 else ''})",
                    variable=var_sede,
                    style='Bold.TCheckbutton'
                )
                chk_sede.pack(anchor=tk.W, pady=(0, 5))
                
                # Mostrar depósitos de esta sede
                frame_deps = ttk.Frame(frame_sede)
                frame_deps.pack(fill=tk.X, padx=20)
                
                for deposito in depositos_filtrados:
                    ttk.Label(
                        frame_deps,
                        text=f"  • {deposito['codigo']} - {deposito['descripcion']}",
                        foreground="#666"
                    ).pack(anchor=tk.W, pady=2)
            
            # Botones de selección rápida
            frame_quick = ttk.Frame(ventana)
            frame_quick.pack(fill=tk.X, padx=10, pady=5)
            
            def seleccionar_todas():
                for var in vars_sedes.values():
                    var.set(True)
            
            def deseleccionar_todas():
                for var in vars_sedes.values():
                    var.set(False)
            
            ttk.Button(frame_quick, text="✅ Seleccionar Todas", command=seleccionar_todas).pack(side=tk.LEFT, padx=5)
            ttk.Button(frame_quick, text="❌ Deseleccionar Todas", command=deseleccionar_todas).pack(side=tk.LEFT, padx=5)
            
            # Frame de botones principales
            frame_botones = ttk.Frame(ventana)
            frame_botones.pack(fill=tk.X, padx=10, pady=10)
            
            def confirmar_exportacion():
                """Confirma la selección y cierra el diálogo"""
                # Recopilar depósitos de las sedes seleccionadas
                depositos_seleccionados = []
                sedes_seleccionadas = []
                export_groups = {}
                
                for localidad, var in vars_sedes.items():
                    if var.get():
                        sedes_seleccionadas.append(localidad)
                        sel_codes = depositos_por_sede[localidad]
                        export_groups[localidad] = sel_codes
                        depositos_seleccionados.extend(sel_codes)
                
                if not depositos_seleccionados:
                    messagebox.showwarning(
                        "Sin Selección",
                        "Debe seleccionar al menos una sede para exportar",
                        parent=ventana
                    )
                    return
                
                # Guardar grupos seleccionados para la exportación actual
                self._export_selected_location_groups = export_groups
                resultado.extend(depositos_seleccionados)
                self.log(f"✅ Sedes seleccionadas para exportar: {', '.join(sedes_seleccionadas)}", "INFO")
                ventana.destroy()
            
            def cancelar_exportacion():
                """Cancela la exportación"""
                ventana.destroy()
            
            ttk.Button(
                frame_botones, 
                text="✅ Exportar Selección", 
                command=confirmar_exportacion,
                style='Accent.TButton'
            ).pack(side=tk.LEFT, padx=5)
            ttk.Button(
                frame_botones, 
                text="❌ Cancelar", 
                command=cancelar_exportacion
            ).pack(side=tk.LEFT, padx=5)
            
            # Centrar ventana
            ventana.update_idletasks()
            x = (ventana.winfo_screenwidth() // 2) - (ventana.winfo_width() // 2)
            y = (ventana.winfo_screenheight() // 2) - (ventana.winfo_height() // 2)
            ventana.geometry(f"+{x}+{y}")
            
            # Esperar a que se cierre la ventana
            ventana.wait_window()
            
            return resultado
        except Exception as e:
            self.log(f"Error seleccionando sedes: {e}", "ERROR")
            return []
        return resultado
    
    def _save_stock_preferences(self):
        """Guarda las preferencias de stock en un archivo local."""
        try:
            import json
            pref_file = "stock_depositos_preference.json"
            pref_data = {
                'localidad': getattr(self, 'stock_localidad_actual', 'Cabudare'),
                'depositos': self.stock_depositos_seleccionados,
                'depositos_por_sede': getattr(self, 'stock_depositos_por_sede', {}),
                'sedes_custom': getattr(self, 'stock_localidades_custom', {}),
                'timestamp': str(time.time())
            }
            with open(pref_file, 'w', encoding='utf-8') as f:
                json.dump(pref_data, f, ensure_ascii=False)
            self.log(f"💾 Preferencia de stock guardada", "DEBUG")
        except Exception as e:
            self.log(f"Error guardando preferencia de stock: {e}", "DEBUG")

    def _save_global_settings(self):
        """Guarda configuraciones globales en la base de datos."""
        try:
            if not self.db_manager.conn:
                return

            # Guardar exclusiones de departamentos
            excluded_list = list(getattr(self, 'excluded_depts', []) or [])
            json_value = json.dumps(excluded_list)
            
            query = """
            MERGE pal_global_settings AS target
            USING (SELECT 'excluded_depts' AS setting_key) AS source
            ON (target.setting_key = source.setting_key)
            WHEN MATCHED THEN
                UPDATE SET setting_value = ?, last_modified = GETDATE()
            WHEN NOT MATCHED THEN
                INSERT (setting_key, setting_value, description, last_modified)
                VALUES ('excluded_depts', ?, 'Lista de codigos de departamento excluidos globalmente de los reportes.', GETDATE());
            """
            self.db_manager.execute_query(query, (json_value, json_value))
            self.log(f"🌍 Guardadas {len(excluded_list)} exclusiones globales en la BD", "INFO")

        except Exception as e:
            self.log(f"Error guardando configuraciones globales: {e}", "ERROR")
    
    def _load_global_settings(self):
        """Carga configuraciones globales desde la base de datos."""
        try:
            if not self.db_manager.conn:
                return

            # Cargar exclusiones de departamentos
            query = "SELECT setting_value FROM pal_global_settings WHERE setting_key = 'excluded_depts'"
            result = self.db_manager.fetch_data(query)
            
            if result and result[0][0]:
                excluded_list = json.loads(result[0][0])
                self.excluded_depts = set(str(x) for x in excluded_list)
            else:
                self.excluded_depts = set()
            
            self._update_excluded_set()
            self.log(f"🌍 Cargadas {len(self.excluded_depts)} exclusiones globales desde la BD", "INFO")

        except Exception as e:
            self.log(f"Error cargando configuraciones globales: {e}", "WARNING")
            self.excluded_depts = set()

    def _start_module_services(self):
        """Inicializa servicios de fondo según los módulos habilitados post-login."""
        try:
            self.log("🚀 Iniciando servicios de módulos...", "INFO")
            
            # Envio de Mensajes
            if self.modules_enabled.get("envio_mensajes", False):
                from pal.services.envios import ProgramadorEnvios, EnvioProgramado
                self.programador = ProgramadorEnvios(self.db_manager, self)
                self.envios_programados = EnvioProgramado(self.db_manager)
                self.log("  ✓ Programador de mensajes activo", "DEBUG")

            # Stock
            if self.modules_enabled.get("stock", False):
                from win10toast import ToastNotifier
                # Inicializar el notificador antes de iniciar el hilo que lo usa
                self.toaster = ToastNotifier()
                self.monitor_thread = threading.Thread(target=self.monitorear_favoritos, daemon=True, name="MonitorStock")
                self.monitor_thread.start()
                self.programar_actualizaciones_stock()
                self.log("  ✓ Monitor de stock y quiebres activo", "DEBUG")

            # TRA / MBRP (Inicialización de estructuras si no existen)
            if self.modules_enabled.get("tra", False):
                self.tra_dept_dict = {}
                self.tra_group_dict = {}
                self.tra_sub_dict = {}
            
            if self.modules_enabled.get("mbrp", False):
                self.mbrp_dept_dict = {}
                self.mbrp_group_dict = {}
                self.mbrp_sub_dict = {}
                self.cached_ventas_mbrp = []
                self.mbrp_loader_thread = None
                self.mbrp_page_size = 500
                self.mbrp_current_page = 1

        except Exception as e:
            self.log(f"Error iniciando servicios de módulos: {e}", "ERROR")

    # _load_global_module_settings eliminada según requerimiento del usuario (usar pal_usuarios_modulos exclusivamente)
    
    def on_deposito_stock_selected(self):
        """Handler cuando cambian los depósitos seleccionados"""
        try:
            self.current_page = 1
            self.recargar_stock()
        except Exception as e:
            self.log(f"Error en cambio de depósito: {e}", "ERROR")

    def hover_row(self, event):
        item = self.stock_tree.identify_row(event.y)
        if item:  # Solo añadir tag si se identificó un item
            try:
                self.stock_tree.tk.call(self.stock_tree, 'tag', 'add', 'hover', item)
            except Exception as e:
                print(f"Error en hover_row: {str(e)}")
        return 0  # Valor entero apropiado para WNDPROC en Windows

    def leave_row(self, event):
        item = self.stock_tree.identify_row(event.y)
        if item:  # Solo remover tag si se identificó un item
            try:
                self.stock_tree.tk.call(self.stock_tree, 'tag', 'remove', 'hover', item)
            except Exception as e:
                print(f"Error en leave_row: {str(e)}")
        return 0  # Valor entero apropiado para WNDPROC en Windows

    def select_row(self, event):
        selected_items = self.stock_tree.selection()
        if not selected_items:  # Verificar si hay selección
            return 0  # Valor entero apropiado para WNDPROC en Windows
        item = selected_items[0]
        try:
            # Limpiar tags previos
            self.stock_tree.tk.call(self.stock_tree, 'tag', 'remove', 'selected', '')
            # Aplicar tag a nuevo item seleccionado
            self.stock_tree.tk.call(self.stock_tree, 'tag', 'add', 'selected', item)
        except Exception as e:
            print(f"Error en select_row: {str(e)}")
        return 0  # Valor entero apropiado para WNDPROC en Windows

    

    def on_favorito_click(self, event):
        """Manejar clic en la columna de favoritos usando cache local"""
        if not hasattr(self, 'stock_tree'):
            return
        try:
            region = self.stock_tree.identify_region(event.x, event.y)
            if region == "cell":
                col = self.stock_tree.identify_column(event.x)
                item = self.stock_tree.identify_row(event.y)
                if item and col == "#1":
                    values = self.stock_tree.item(item, 'values')
                    if values and len(values) > 1:
                        codigo = values[1]
                        if self._toggle_favorito_local(codigo):
                            self.aplicar_filtro_stock()
        except Exception as e:
            print(f"Error en on_favorito_click: {e}")
        return 0
        

    def _update_stock_extra_columns_headings(self):
        """Ya no es necesario con las columnas estáticas por sede"""
        pass

    def _build_existencias_sedes_map(self, codigos):
        """Devuelve map {codigo: {sede_name: total}} usando SOLO los depósitos seleccionados por sede (visibilidad)."""
        try:
            if not codigos:
                return {}
            locs = getattr(self, 'stock_localidades', {}) or {}
            sel_por_sede = getattr(self, 'stock_depositos_por_sede', {}) or {}
            seleccionados_global = set(getattr(self, 'stock_depositos_seleccionados', []) or [])

            # Construir lista de depósitos visibles por sede (prioriza 'por_sede'; fallback al set global; luego todos)
            dep_to_sede = {}
            all_deps = []
            for sede, deps in locs.items():
                # Preferir selección por sede si existe
                codes_sel = sel_por_sede.get(sede)
                if codes_sel:
                    codes_visibles = set(str(c) for c in codes_sel)
                elif seleccionados_global:
                    codes_visibles = {str(d['codigo']) for d in deps if str(d['codigo']) in seleccionados_global}
                else:
                    codes_visibles = {str(d['codigo']) for d in deps}
                for d in deps:
                    cod_dep = str(d['codigo'])
                    if cod_dep in codes_visibles:
                        dep_to_sede[cod_dep] = sede
                        all_deps.append(cod_dep)
            all_deps = list(dict.fromkeys(all_deps))  # únicos
            if not all_deps:
                return {c: {s:0 for s in locs.keys()} for c in codigos}

            # Consulta por artículos y depósitos visibles con chunking para evitar límite de parámetros
            all_rows = []
            max_params_per_query = 2000  # Margen seguro bajo el límite de 2100
            
            # Calcular tamaño de chunk para códigos basado en número de depósitos
            if len(all_deps) > 0:
                max_codigos_por_chunk = max(1, max_params_per_query // len(all_deps))
            else:
                max_codigos_por_chunk = 100  # Valor por defecto
            
            for i in range(0, len(codigos), max_codigos_por_chunk):
                chunk_codigos = codigos[i:i + max_codigos_por_chunk]
                codigo_placeholders = ','.join('?' * len(chunk_codigos))
                deposito_placeholders = ','.join('?' * len(all_deps))
                sql = (
                    f"SELECT c_codarticulo, c_coddeposito, ISNULL(SUM(n_cantidad),0) AS total "
                    f"FROM MA_DEPOPROD WITH (NOLOCK) "
                    f"WHERE c_codarticulo IN ({codigo_placeholders}) AND c_coddeposito IN ({deposito_placeholders}) "
                    f"GROUP BY c_codarticulo, c_coddeposito"
                )
                params = chunk_codigos + all_deps
                
                try:
                    chunk_rows = self.db_manager.fetch_data(sql, params) or []
                    all_rows.extend(chunk_rows)
                except Exception as e:
                    self.stock_debug_log(f"Error cargando stock chunk {i//max_codigos_por_chunk + 1}: {e}", "ERROR")
                    continue
            
            rows = all_rows
            # Acumular por sede dinámica
            res = {c: {sede: 0 for sede in locs.keys()} for c in codigos}
            for codigo, deposito, total in rows:
                try:
                    sede = dep_to_sede.get(str(deposito))
                    if sede and codigo in res:
                        res[codigo][sede] += int(total or 0)
                except Exception:
                    continue
            return res
        except Exception as e:
            self.stock_debug_log(f"Error construyendo existencias por sede: {e}", "ERROR")
            return {}

    def mostrar_alertas_paginadas(self, datos):
        """Mostrar datos con estado de favoritos y filas alternadas, incluyendo columnas por sede"""
        if not hasattr(self, 'stock_tree') or not self.stock_tree.winfo_exists():
            return
        self.stock_tree.delete(*self.stock_tree.get_children())
        favoritos = self._get_favoritos_local()
        
        # Crear mapeo de códigos de depósito a nombres de sede
        sede_name_map = {}
        try:
            sedes_config = self.config_manager.get_sedes_config()
            for sede_name, config in sedes_config.items():
                almacenes = config.get('almacenes_tratables', [])
                for almacen in almacenes:
                    sede_name_map[almacen] = sede_name
        except Exception:
            # Fallback a mapeo manual si falla la configuración
            sede_name_map = {
                '0301': 'Cabudare',
                '0401': 'Guanare',
                '0101': 'Barinas'
            }
        
        for idx, (codigo, desc, sede_codigo, unid_perd, dias, ult_compra, ult_venta) in enumerate(datos):
            es_favorito = codigo in favoritos
            estado = "✓" if es_favorito else "☐"

            # Tags
            if es_favorito:
                tags = ('favorito',)
            else:
                tags = ('quiebre',) if idx % 2 == 0 else ()
            
            # Formatear fechas
            def fmt_date(d):
                if not d or str(d).lower() == 'none': return "N/A"
                if isinstance(d, datetime): return d.strftime("%d/%m/%Y")
                return str(d)

            compra_str = fmt_date(ult_compra)
            venta_str = fmt_date(ult_venta)
            
            # Convertir código de sede a nombre
            sede_nombre = sede_name_map.get(sede_codigo, sede_codigo)

            row_values = [estado, codigo, desc, sede_nombre, int(unid_perd), int(dias), compra_str, venta_str]
            self.stock_tree.insert(
                "", tk.END, 
                values=tuple(row_values),
                tags=tags)
        
    def aplicar_filtro(self):
        self.current_filter = self.filter_var.get()
        self.current_page = 1
        # Limpiar selección antes de actualizar
        if hasattr(self, 'stock_tree'):
            try:
                self.stock_tree.selection_remove(self.stock_tree.selection())
            except Exception:
                pass
        self.actualizar_alertas_stock()

    def cambiar_pagina(self, delta):
        self.current_page += delta
        self.aplicar_filtro_stock()
        



    def create_main_workspace(self):

        self.main_notebook = ttk.Notebook(self.root)
        self.main_notebook.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        
        # Pestaña de Dashboard (Pantalla Principal) - siempre visible
        self.dashboard_tab = ttk.Frame(self.main_notebook)
        self.main_notebook.add(self.dashboard_tab, text="🏠 Inicio")
        from pal.ui.tabs.dashboard import setup_dashboard_tab
        setup_dashboard_tab(self)
        
        # Pestaña de Registros (solo si el módulo de mensajes está activo)
        if self.modules_enabled.get("envio_mensajes", False):
            self.records_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.records_tab, text="Registros")
            from pal.ui.tabs.records import setup_records_tab as setup_records_tab_ui
            setup_records_tab_ui(self)


        # Pestaña de Mensajería
        if self.modules_enabled.get("envio_mensajes", False):
            self.messaging_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.messaging_tab, text="Mensajería")
            from pal.ui.tabs.messaging import setup_messaging_tab as setup_messaging_tab_ui
            setup_messaging_tab_ui(self)

        # Pestaña de Estadísticas
        if self.modules_enabled.get("estadisticas", False):
            self.stats_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.stats_tab, text="📊 Estadísticas")
            from pal.ui.tabs.stats import setup_stats_tab as setup_stats_tab_ui
            setup_stats_tab_ui(self)  

        # Pestaña de Calendario
        if self.modules_enabled.get("calendario", False):
            self.calendar_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.calendar_tab, text="📅 Calendario")
            from pal.ui.tabs.calendar import setup_calendar_tab as setup_calendar_tab_ui
            setup_calendar_tab_ui(self)

        # Pestaña T.R.A (Rotación de Ventas)
        if self.modules_enabled.get("tra", False):
            self.tra_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.tra_tab, text="📈 RI")
            from pal.ui.tabs.tra import setup_tra_tab as setup_tra_tab_ui
            setup_tra_tab_ui(self)
            self.root.after(500, self._update_hierarchy_combos)
        # Pe staña MBRP (Movimiento de Baja Rotación)
        if self.modules_enabled.get("mbrp", False):
            self.mbrp_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.mbrp_tab, text="📉 MBRP")
            from pal.ui.tabs.mbrp import setup_mbrp_tab as setup_mbrp_tab_ui
            setup_mbrp_tab_ui(self)
            # Actualizar combos MBRP inmediatamente si ya hay datos cargados
            if hasattr(self, 'mbrp_dept_dict') and self.mbrp_dept_dict:
                self.mbrp_dept_combo['values'] = ['Todos'] + list(self.mbrp_dept_dict.keys())
                self.mbrp_dept_var.set('Todos')
            # Programar actualización adicional en caso de que se carguen después
            self.root.after(500, self._update_hierarchy_combos)
            self.root.after(500, self._update_hierarchy_combos)

        # Pestaña de Clientes (Menú de Submódulos)
        if self.modules_enabled.get("clientes", True):
            self.clientes_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.clientes_tab, text="👥 Clientes")
            
            # Crear instancias de las subvistas
            self.clientes_menu_view = ClientesMenu(self.clientes_tab, controller=self)
            self.clientes_reportes_view = None # Se creará bajo demanda

            # Mostrar el menú inicial por defecto
            self.show_clientes_sub_view('menu')

        # Alerta de stock Supervisores
        if self.modules_enabled.get("stock", False):
            self.stock_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.stock_tab, text="⚠️ Quiebre de Stock")
            from pal.ui.tabs.stock import setup_stock_tab as setup_stock_tab_ui
            setup_stock_tab_ui(self)

        # Pestaña de Logística (Abastecimiento)
        if self.modules_enabled.get("logistica", False):
            self.logistica_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.logistica_tab, text="🚚 Logística")
            from pal.ui.tabs.abastecimiento import AbastecimientoTab
            self.abastecimiento_view = AbastecimientoTab(self.logistica_tab, self)
            self.abastecimiento_view.pack(expand=True, fill=tk.BOTH)
        
        # Pestaña de Administración (Configuraciones Globales)
        if self.modules_enabled.get("admin", False):
            self.admin_tab = ttk.Frame(self.main_notebook)
            self.main_notebook.add(self.admin_tab, text="⚙️ Config. Global")

            from pal.ui.admin_menu import AdminMenu
            self.admin_menu_view = AdminMenu(self.admin_tab, controller=self)
            
            # Inicializar los frames de las subvistas como None
            self.sedes_servidores_view = None
            self.sedes_almacenes_view = None
            self.admin_users_view = None
            self.admin_roles_view = None
            self.admin_exclusions_view = None
            self.admin_audit_view = None
            self.admin_sedes_tradicional_view = None

            # Mostrar el menú inicial
            self.show_admin_sub_view('menu')
    
    def mostrar_tra_filtrado(self, datos_filtrados):
        """Muestra datos filtrados TRA en el Treeview con colores, stock actual e ideal, y días restantes"""
        self.tra_tree.delete(*self.tra_tree.get_children())
        
        # Actualizar variable para paginación
        self.tra_ventas_datos_filtrados = datos_filtrados
        
        # Solo mostrar una página de datos
        items_por_pagina = min(self.tra_page_size, len(datos_filtrados))
        page_rows = datos_filtrados[:items_por_pagina]
        codigos = [r[0] for r in page_rows]
        sede = (self.sede_var.get() or '').split(' - ')[0]
        stock_map = self.obtener_stock_actual_bulk(codigos, sede)
        fecha_inicio = self.fecha_inicio_entry.get_date()
        fecha_fin = self.fecha_fin_entry.get_date()
        
        for fila in page_rows:
            codigo, desc, _, _, _, neto, rotacion = fila
            stock_actual = int(stock_map.get(codigo, 0) or 0)
            
            # Calcular stock ideal y días restantes
            stock_ideal = self.calcular_stock_ideal_producto(neto)
            dias_restantes = self.calcular_dias_restantes(stock_actual, neto, fecha_inicio, fecha_fin)
            
            # Determinar tag de color según rotación
            tag_rotacion = rotacion.lower()
            
            # Verificar si producto tiene alerta de stock y es de rotación ALTA o MEDIA
            alerta_stock = self._check_product_stock_alert(codigo)
            es_rotacion_critica = rotacion.upper() in ['ALTA', 'MEDIA']
            tiene_alerta_critica = alerta_stock and es_rotacion_critica
            
            # Si tiene alerta crítica (alta/media rotación + stock bajo), agregar indicador
            desc_mostrada = desc
            tags_mostrados = (tag_rotacion,)
            if tiene_alerta_critica:
                # Agregar indicador visual prominente al inicio de la descripción
                desc_mostrada = f"🔴 {desc}"
                # Agregar tag de alerta además del tag de rotación
                tags_mostrados = (tag_rotacion, 'stock_alert')
            
            # Formatear neto según el modo de display (unidades o dólares con IVA)
            neto_valor = float(neto or 0)
            mostrar_dolares = getattr(self, 'tra_mostrar_dolares_var', tk.BooleanVar()).get()
            
            if mostrar_dolares:
                # Convertir a dólares usando precio con IVA
                precio_unitario = self._obtener_precio_producto(codigo)
                ventas_dolares = neto_valor * precio_unitario
                neto_formateado = f"${ventas_dolares:,.2f}"
            else:
                # Mostrar como unidades (valor original)
                neto_formateado = int(neto_valor) if neto_valor == int(neto_valor) else round(neto_valor, 2)
            
            self.tra_tree.insert(
                "", tk.END,
                values=(codigo, desc_mostrada, rotacion, neto_formateado, stock_actual, stock_ideal, dias_restantes),
                tags=tags_mostrados
            )
    
    def cambiar_pagina_tra(self, delta):
        """Cambia de página en el módulo TRA"""
        self.tra_current_page += delta
        self.aplicar_filtro_tra()  # Esto recalculará la paginación y mostrará los datos
    
    def mostrar_pagina_tra(self):
        """Muestra la página actual de resultados"""
        if not hasattr(self, 'tra_ventas_datos_filtrados'):
            return
    
        total_items = len(self.tra_ventas_datos_filtrados)
        total_pages = max(1, (total_items + self.tra_page_size - 1) // self.tra_page_size)
    
        # Asegurar página válida
        self.tra_current_page = max(1, min(self.tra_current_page, total_pages))
    
        # Calcular rango de items a mostrar
        start_idx = (self.tra_current_page - 1) * self.tra_page_size
        end_idx = min(start_idx + self.tra_page_size, total_items)
        page_data = self.tra_ventas_datos_filtrados[start_idx:end_idx]
    
        # Actualizar Treeview con colores, stock actual, stock ideal y días restantes
        self.tra_tree.delete(*self.tra_tree.get_children())
        codigos = [r[0] for r in page_data]
        sede = (self.sede_var.get() or '').split(' - ')[0]
        stock_map = self.obtener_stock_actual_bulk(codigos, sede)
        fecha_inicio = self.fecha_inicio_entry.get_date()
        fecha_fin = self.fecha_fin_entry.get_date()
        for fila in page_data:
            codigo, desc, _, _, _, neto, rotacion = fila
            stock_actual = int(stock_map.get(codigo, 0) or 0)
            stock_ideal = self.calcular_stock_ideal_producto(neto)
            dias_restantes = self.calcular_dias_restantes(stock_actual, neto, fecha_inicio, fecha_fin)
            tag_rotacion = rotacion.lower()
            # Formatear neto según el modo de display (unidades o dólares con IVA)
            neto_valor = float(neto or 0)
            mostrar_dolares = getattr(self, 'tra_mostrar_dolares_var', tk.BooleanVar()).get()
            
            if mostrar_dolares:
                # Convertir a dólares usando precio con IVA
                precio_unitario = self._obtener_precio_producto(codigo)
                ventas_dolares = neto_valor * precio_unitario
                neto_formateado = f"${ventas_dolares:,.2f}"
            else:
                # Mostrar como unidades (valor original)
                neto_formateado = int(neto_valor) if neto_valor == int(neto_valor) else round(neto_valor, 2)
            
            # Verificar si producto tiene alerta de stock y es de rotación ALTA o MEDIA
            alerta_stock = self._check_product_stock_alert(codigo)
            es_rotacion_critica = rotacion.upper() in ['ALTA', 'MEDIA']
            tiene_alerta_critica = alerta_stock and es_rotacion_critica
            
            # Si tiene alerta crítica (alta/media rotación + stock bajo), agregar indicador
            desc_mostrada = desc
            tags_mostrados = (tag_rotacion,)
            if tiene_alerta_critica:
                # Agregar indicador visual prominente al inicio de la descripción
                desc_mostrada = f"🔴 {desc}"
                # Agregar tag de alerta además del tag de rotación
                tags_mostrados = (tag_rotacion, 'stock_alert')
            
            self.tra_tree.insert(
                "", "end", 
                values=(codigo, desc_mostrada, rotacion, neto_formateado, stock_actual, stock_ideal, dias_restantes),
                tags=tags_mostrados
            )
    
        # Actualizar controles
        self.tra_pagina_label.config(text=f"Página {self.tra_current_page}/{total_pages}")
        self.tra_btn_prev['state'] = 'normal' if self.tra_current_page > 1 else 'disabled'
        self.tra_btn_next['state'] = 'normal' if self.tra_current_page < total_pages else 'disabled'
        
    def tra_debug_log(self, message: str, level: str = "DEBUG", throttle_key: str = None, throttle_seconds: float = 2.0):
        """Log DEBUG inteligente para módulo TRA con throttling para evitar spam.
        
        Args:
            message: Mensaje a loggear
            level: Nivel del log (DEBUG, INFO, etc.)
            throttle_key: Clave para agrupar mensajes similares (ej: "chunk_load")
            throttle_seconds: Segundos mínimos entre logs del mismo throttle_key
        """
        if not getattr(self, 'tra_debug', False):
            return
            
        try:
            # Inicializar throttle cache si no existe
            if not hasattr(self, '_tra_debug_throttle'):
                self._tra_debug_throttle = {}
            
            current_time = time.time()
            
            # Si hay throttle_key, verificar si debemos loggear
            if throttle_key:
                last_time = self._tra_debug_throttle.get(throttle_key, 0)
                if current_time - last_time < throttle_seconds:
                    return  # Suprimir este log por throttling
                self._tra_debug_throttle[throttle_key] = current_time
            
            self.log(f"[TRA DEBUG] {message}", level)
        except Exception:
            pass

    def stock_debug_log(self, message: str, level: str = "DEBUG", throttle_key: str = None, throttle_seconds: float = 2.0):
        """Log DEBUG inteligente para módulo Stock con throttling para evitar spam.
        
        Args:
            message: Mensaje a loggear
            level: Nivel del log (DEBUG, INFO, etc.)
            throttle_key: Clave para agrupar mensajes similares
            throttle_seconds: Segundos mínimos entre logs del mismo throttle_key
        """
        if not getattr(self, 'stock_debug', False):
            return
            
        try:
            # Inicializar throttle cache si no existe
            if not hasattr(self, '_stock_debug_throttle'):
                self._stock_debug_throttle = {}
            
            current_time = time.time()
            
            # Si hay throttle_key, verificar si debemos loggear
            if throttle_key:
                last_time = self._stock_debug_throttle.get(throttle_key, 0)
                if current_time - last_time < throttle_seconds:
                    return  # Suprimir este log por throttling
                self._stock_debug_throttle[throttle_key] = current_time
            
            self.log(f"[STOCK DEBUG] {message}", level)
        except Exception:
            pass

    def _get_codigos_por_proveedor_cached(self, cod_proveedor: str):
        """Obtiene códigos de producto por proveedor usando un pequeño caché en memoria."""
        try:
            if not cod_proveedor:
                return set()
            if not hasattr(self, '_codigos_por_proveedor_cache'):
                self._codigos_por_proveedor_cache = {}
            cod_proveedor = str(cod_proveedor).strip()
            if cod_proveedor in self._codigos_por_proveedor_cache:
                return self._codigos_por_proveedor_cache[cod_proveedor]
            rows = self.db_manager.obtener_codigos_por_proveedor(cod_proveedor)
            codigos = set(str(c) for c in rows) if rows else set()
            self._codigos_por_proveedor_cache[cod_proveedor] = codigos
            return codigos
        except Exception as e:
            self.log(f"Error obteniendo códigos por proveedor (cache): {e}", "ERROR")
            return set()
    
    def _preload_productos_proveedores(self):
        """Precarga mapeo productos->proveedores para TODOS los productos en cached_ventas_tra.
        
        Esto garantiza que las estadísticas por proveedor usen exactamente los mismos datos
        que las estadísticas por departamento (mismo universo de datos precargados en RI).
        """
        try:
            # Solo ejecutar si hay datos TRA cargados
            if not hasattr(self, 'cached_ventas_tra') or not self.cached_ventas_tra:
                return
            
            # Extraer todos los códigos de productos del universo RI
            codigos_universo = set()
            for r in self.cached_ventas_tra:
                if r and len(r) > 0:
                    try:
                        codigo = str(r[0]).strip()
                        if codigo:
                            codigos_universo.add(codigo)
                    except Exception:
                        continue
            
            if not codigos_universo:
                self.log("No hay códigos de productos para precargar proveedores", "DEBUG")
                return
            
            self.log(f"Precargando proveedores para {len(codigos_universo)} productos del universo RI...", "INFO")
            
            # Consultar MA_PRODXPROV en batches para obtener relaciones producto-proveedor
            cached_proveedor_por_codigo = {}
            codigos_list = sorted(codigos_universo)
            
            try:
                BATCH_SIZE = 1800  # Margen para límite de parámetros SQL Server (~2100)
                total_relaciones = 0
                
                for i in range(0, len(codigos_list), BATCH_SIZE):
                    batch = codigos_list[i:i + BATCH_SIZE]
                    placeholders = ','.join('?' * len(batch))
                    query = f"""
                        SELECT px.c_codigo,
                               px.c_codprovee,
                               COALESCE(pr.c_descripcio, px.c_codprovee) AS descrip
                        FROM MA_PRODXPROV px WITH (NOLOCK)
                        LEFT JOIN MA_PROVEEDORES pr WITH (NOLOCK)
                            ON px.c_codprovee = pr.c_codproveed
                        WHERE px.c_codigo IN ({placeholders})
                    """
                    rows = self.db_manager.fetch_data(query, batch) or []
                    
                    for cod_prod, cod_prov, desc_prov in rows:
                        try:
                            cod_prod_str = str(cod_prod).strip()
                            cod_prov_str = str(cod_prov).strip()
                            desc_prov_str = str(desc_prov) if desc_prov else cod_prov_str
                        except Exception:
                            continue
                        
                        if not cod_prod_str or not cod_prov_str:
                            continue
                        
                        # Guardar la PRIMERA relación encontrada (algunos productos tienen múltiples proveedores)
                        if cod_prod_str not in cached_proveedor_por_codigo:
                            cached_proveedor_por_codigo[cod_prod_str] = (cod_prov_str, desc_prov_str)
                            total_relaciones += 1
                
                # Guardar en atributo de instancia para uso en estadísticas
                self.cached_proveedor_por_codigo = cached_proveedor_por_codigo
                
                productos_sin_prov = len(codigos_universo) - len(cached_proveedor_por_codigo)
                self.log(
                    f"✅ Proveedores precargados: {total_relaciones} relaciones producto-proveedor | "
                    f"{productos_sin_prov} productos sin proveedor asignado",
                    "SUCCESS"
                )
                
            except Exception as e:
                self.log(f"Error consultando MA_PRODXPROV para precarga: {e}", "ERROR")
                # En caso de error, crear cache vacío para no bloquear estadísticas
                self.cached_proveedor_por_codigo = {}
                
        except Exception as e:
            self.log(f"Error precargando productos-proveedores: {e}", "ERROR")
            # Asegurar que exista el atributo aunque sea vacío
            self.cached_proveedor_por_codigo = {}

    def abrir_selector_proveedor_tra(self):
        """Abre un selector de proveedor para filtrar TRA."""
        self._abrir_selector_proveedor(contexto="TRA")

    def abrir_selector_proveedor_mbrp(self):
        """Abre un selector de proveedor para filtrar MBRP."""
        self._abrir_selector_proveedor(contexto="MBRP")

    def _abrir_selector_proveedor(self, contexto: str = "TRA"):
        """Ventana emergente para seleccionar proveedor desde MA_PROVEEDORES.

        contexto: "TRA" o "MBRP" para saber qué filtro actualizar.
        """
        try:
            import tkinter as tk
            from tkinter import ttk

            ventana = tk.Toplevel(self.root)
            ventana.title(f"Seleccionar Proveedor ({contexto})")
            ventana.geometry("800x550")
            ventana.resizable(False, False)
            ventana.grab_set()

            ttk.Label(
                ventana,
                text="Seleccione un proveedor para filtrar los productos:" ,
                font=('Arial', 10, 'bold')
            ).pack(pady=10, padx=10, anchor=tk.W)

            # Marco de búsqueda y modo
            search_frame = ttk.Frame(ventana)
            search_frame.pack(fill=tk.X, padx=10, pady=5)

            ttk.Label(search_frame, text="Modo:").pack(side=tk.LEFT, padx=(0, 5))
            mode_var = tk.StringVar(value="Proveedor")
            mode_combo = ttk.Combobox(
                search_frame,
                textvariable=mode_var,
                state='readonly',
                width=12,
                values=("Proveedor", "Producto"),
            )
            mode_combo.pack(side=tk.LEFT, padx=(0, 10))

            ttk.Label(search_frame, text="Buscar:").pack(side=tk.LEFT)
            search_var = tk.StringVar()
            search_entry = ttk.Entry(search_frame, textvariable=search_var)
            search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

            # Tabla de proveedores
            tree_frame = ttk.Frame(ventana)
            tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
            columns = ("Código", "Descripción")
            tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=12)

            # Estilo de selección similar a TRA/MBRP (azul oscuro)
            style = ttk.Style(tree)
            style.configure('Proveedor.Treeview', rowheight=22)
            style.map('Proveedor.Treeview',
                      background=[('selected', '#0D47A1')],
                      foreground=[('selected', '#FFFFFF')])
            tree.configure(style='Proveedor.Treeview')

            vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)
            tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            tree_frame.grid_rowconfigure(0, weight=1)
            tree_frame.grid_columnconfigure(0, weight=1)

            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=120 if col == "Código" else 400, anchor='w')

            # Cargar proveedores base (modo Proveedor)
            proveedores_cache = self.db_manager.obtener_proveedores()

            def refrescar_lista(*args):
                texto = search_var.get().strip()
                modo = (mode_var.get() or "Proveedor").strip()
                tree.delete(*tree.get_children())

                if modo == "Proveedor":
                    if not proveedores_cache:
                        return
                    for cod, desc in proveedores_cache:
                        if texto:
                            t = texto.lower()
                            if t not in str(cod).lower() and t not in str(desc).lower():
                                continue
                        tree.insert("", tk.END, values=(cod, desc))
                else:
                    # Modo Producto: buscar proveedores asociados a códigos de producto
                    if not texto:
                        return
                    try:
                        rows = self.db_manager.obtener_proveedores_por_producto(texto)
                    except Exception as e:
                        self.log(f"Error obteniendo proveedores por producto: {e}", "ERROR")
                        rows = []
                    for cod, desc in rows:
                        tree.insert("", tk.END, values=(cod, desc))

            search_var.trace_add('write', lambda *args: refrescar_lista())
            mode_var.trace_add('write', lambda *args: refrescar_lista())
            refrescar_lista()

            seleccionado = {"codigo": None, "descripcion": None}

            # Comportamiento de selección visual mejorado (similar a TRA/MBRP)
            def on_tree_click(event=None):
                try:
                    region = tree.identify_region(event.x, event.y)
                    if region in ('cell', 'tree'):
                        item = tree.identify_row(event.y)
                        if item:
                            tree.selection_set(item)
                            tree.focus(item)
                            tree.see(item)
                except Exception:
                    pass

            def on_double_click(event=None):
                item = tree.focus()
                if not item:
                    return
                values = tree.item(item, 'values')
                if not values:
                    return
                seleccionado["codigo"] = str(values[0]).strip()
                seleccionado["descripcion"] = str(values[1]).strip()
                ventana.destroy()

            tree.bind('<Button-1>', on_tree_click)
            tree.bind("<Double-1>", on_double_click)

            # Botones inferiores
            btn_frame = ttk.Frame(ventana)
            btn_frame.pack(fill=tk.X, padx=10, pady=10)

            def aplicar_seleccion():
                item = tree.focus()
                if not item:
                    ventana.destroy()
                    return
                values = tree.item(item, 'values')
                if not values:
                    ventana.destroy()
                    return
                seleccionado["codigo"] = str(values[0]).strip()
                seleccionado["descripcion"] = str(values[1]).strip()
                ventana.destroy()

            def limpiar_filtro():
                seleccionado["codigo"] = None
                seleccionado["descripcion"] = None
                ventana.destroy()

            ttk.Button(btn_frame, text="✅ Seleccionar", command=aplicar_seleccion).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="🧹 Quitar filtro", command=limpiar_filtro).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="❌ Cancelar", command=ventana.destroy).pack(side=tk.RIGHT, padx=5)

            # Centrar
            ventana.update_idletasks()
            x = (ventana.winfo_screenwidth() // 2) - (ventana.winfo_width() // 2)
            y = (ventana.winfo_screenheight() // 2) - (ventana.winfo_height() // 2)
            ventana.geometry(f"+{x}+{y}")

            ventana.wait_window()

            # Aplicar resultado
            cod_sel = seleccionado["codigo"]
            desc_sel = seleccionado["descripcion"]
            if contexto.upper() == "TRA":
                self.tra_proveedor_codigo = cod_sel
                # No es obligatorio mostrar descripción en UI, pero se guarda por si se necesita
                self.tra_proveedor_descripcion = desc_sel
                # Actualizar etiqueta de proveedor seleccionado en pestaña TRA
                try:
                    label = getattr(self, 'tra_proveedor_label', None)
                    if hasattr(self, 'tra_proveedor_label_var') and label is not None:
                        if cod_sel:
                            texto = desc_sel or cod_sel
                            self.tra_proveedor_label_var.set(texto)
                            # Mostrar la etiqueta solo cuando hay proveedor
                            if not label.winfo_manager():
                                label.pack(side=tk.LEFT, padx=(5, 0))
                        else:
                            # Filtro limpiado: ocultar etiqueta
                            self.tra_proveedor_label_var.set("")
                            if label.winfo_manager():
                                label.pack_forget()
                except Exception:
                    pass
                self.aplicar_filtro_tra()
            elif contexto.upper() == "MBRP":
                self.mbrp_proveedor_codigo = cod_sel
                self.mbrp_proveedor_descripcion = desc_sel
                # Actualizar etiqueta de proveedor seleccionado en pestaña MBRP
                try:
                    label = getattr(self, 'mbrp_proveedor_label', None)
                    if hasattr(self, 'mbrp_proveedor_label_var') and label is not None:
                        if cod_sel:
                            texto = desc_sel or cod_sel
                            self.mbrp_proveedor_label_var.set(texto)
                            if not label.winfo_manager():
                                label.pack(side=tk.LEFT, padx=(5, 0))
                        else:
                            self.mbrp_proveedor_label_var.set("")
                            if label.winfo_manager():
                                label.pack_forget()
                except Exception:
                    pass
                self.aplicar_filtro_mbrp()

        except Exception as e:
            self.log(f"Error en selector de proveedor ({contexto}): {e}", "ERROR")

    def mbrp_debug_log(self, message: str, level: str = "DEBUG", throttle_key: str = None, throttle_seconds: float = 2.0):
        """Log DEBUG inteligente para módulo MBRP con throttling para evitar spam.
        
        Args:
            message: Mensaje a loggear
            level: Nivel del log (DEBUG, INFO, etc.)
            throttle_key: Clave para agrupar mensajes similares
            throttle_seconds: Segundos mínimos entre logs del mismo throttle_key
        """
        if not getattr(self, 'mbrp_debug', False):
            return
            
        try:
            # Inicializar throttle cache si no existe
            if not hasattr(self, '_mbrp_debug_throttle'):
                self._mbrp_debug_throttle = {}
            
            current_time = time.time()
            
            # Si hay throttle_key, verificar si debemos loggear
            if throttle_key:
                last_time = self._mbrp_debug_throttle.get(throttle_key, 0)
                if current_time - last_time < throttle_seconds:
                    return  # Suprimir este log por throttling
                self._mbrp_debug_throttle[throttle_key] = current_time
            
            self.log(f"[MBRP DEBUG] {message}", level)
        except Exception:
            pass

    def aplicar_filtro_tra(self):
        """Aplica filtros jerárquicos, proveedor y texto a los datos TRA."""
        if not hasattr(self, 'cached_ventas_tra') or not self.cached_ventas_tra:
            # Solo aviso en warning, no debug
            self.log("No hay datos TRA cacheados para filtrar", "WARNING")
            return

        # Filtro por proveedor (si está seleccionado)
        datos_base = list(self.cached_ventas_tra)
        proveedor_cod = getattr(self, 'tra_proveedor_codigo', None)
        if proveedor_cod:
            codigos_prov = self._get_codigos_por_proveedor_cached(proveedor_cod)
            if codigos_prov:
                datos_base = [r for r in datos_base if str(r[0]) in codigos_prov]
            else:
                datos_base = []

        # Obtener códigos seleccionados
        dept_cod = self.tra_dept_dict.get(self.tra_dept_var.get()) if hasattr(self, 'tra_dept_var') else None
        group_cod = None
        sub_cod = None
        
        # DEBUG: Log de valores de filtro
        dept_desc = self.tra_dept_var.get() if hasattr(self, 'tra_dept_var') else 'N/A'
        self.log(f"🔍 [FILTRO TRA] Iniciando filtrado...", "DEBUG")
        self.log(f"🔍 [FILTRO TRA] Dept seleccionado: '{dept_desc}' -> código: {dept_cod}", "DEBUG")
        self.log(f"🔍 [FILTRO TRA] Datos base: {len(datos_base)} registros", "DEBUG")

        if dept_cod and hasattr(self, 'tra_group_var'):
            group_desc = self.tra_group_var.get()
            group_cod = self.tra_group_dict.get(dept_cod, {}).get(group_desc)
            self.log(f"🔍 [FILTRO TRA] Grupo seleccionado: '{group_desc}' -> código: {group_cod}", "DEBUG")
        
            if group_cod and hasattr(self, 'tra_sub_var'):
                sub_desc = self.tra_sub_var.get()
                # Usar string como key (formato: "dept|group")
                key = f"{dept_cod}|{group_cod}"
                sub_cod = self.tra_sub_dict.get(key, {}).get(sub_desc)
    
        # Obtener texto de búsqueda
        texto = self.tra_search_var.get() if hasattr(self, 'tra_search_var') else ''
        
        # Obtener favoritos
        favoritos = self._get_favoritos_local()
    
        # Usar las nuevas funciones de filtrado
        from pal.services.tra import filter_ventas_tra, paginate_tra
        
        # Debug: verificar datos antes del filtro (con throttling para evitar spam)
        self.tra_debug_log(
            f"Aplicando filtros - Datos: {len(self.cached_ventas_tra)} | "
            f"Dept: {dept_cod}, Group: {group_cod}, Sub: {sub_cod}, Texto: '{texto}'",
            throttle_key="filter_input",
            throttle_seconds=2.0
        )
        
        # Obtener alertas de stock si están disponibles
        alertas_para_filtro = getattr(self, 'cached_alertas', [])
        
        # Filtrar y ordenar
        datos_filtrados = filter_ventas_tra(
            ventas=datos_base,
            dept_code=dept_cod,
            group_code=group_cod,
            sub_code=sub_cod,
            search_text=texto,
            filter_rotacion='TODAS',  # Por ahora no implementamos filtro por rotación en UI
            favoritos=favoritos,
            alertas_stock=alertas_para_filtro  # Pasar alertas para ordenamiento prioritario
        )
        # Exclusión global por departamento en RI
        excluded_set = getattr(self, '_excluded_depts_set', set())
        if excluded_set:
            datos_filtrados = [r for r in datos_filtrados if len(r) > 2 and str(r[2]) not in excluded_set]
        
        self.tra_debug_log(
            f"Resultado filtrado: {len(datos_filtrados)} registros",
            throttle_key="filter_result",
            throttle_seconds=2.0
        )
        
        # Asegurar que la página actual sea válida
        if not hasattr(self, 'tra_current_page') or self.tra_current_page < 1:
            self.tra_current_page = 1
            
        # Paginación
        datos_pagina, total_paginas, self.tra_current_page = paginate_tra(
            datos_filtrados, self.tra_current_page, self.tra_page_size
        )
        
        self.tra_debug_log(
            f"Paginación: página {self.tra_current_page}/{total_paginas} ({len(datos_pagina)} registros)",
            throttle_key="pagination",
            throttle_seconds=1.5
        )
        
        # Actualizar vista
        self.mostrar_tra_paginado(datos_pagina)
        self.actualizar_controles_paginacion_tra(total_paginas)
    
    def mostrar_tra_paginado(self, datos):
        """Muestra datos TRA paginados en el Treeview con colores, stock actual e ideal, días restantes y porcentajes"""
        if not hasattr(self, 'tra_tree'):
            self.tra_debug_log("Error: tra_tree no existe")  # Usar función condicional
            return
            
        self.tra_tree.delete(*self.tra_tree.get_children())
        
        if not datos:
            # Solo loggear una vez para evitar spam
            self.tra_debug_log(
                "No hay datos para mostrar en esta página",
                level="DEBUG",
                throttle_key="empty_page",
                throttle_seconds=5.0
            )
            return
            
        # Optimización: solo calcular porcentajes si no están en cache o si hay cambios significativos
        cached_data = getattr(self, 'cached_ventas_tra', [])
        
        # Cache de porcentajes para evitar recalcular constantemente
        if not hasattr(self, 'tra_porcentajes_map') or not hasattr(self, '_tra_last_porcentaje_count'):
            self.tra_porcentajes_map = {}
            self._tra_last_porcentaje_count = 0
        
        # Solo recalcular si hay cambios significativos (>10% de diferencia)
        current_count = len(cached_data)
        count_diff = abs(current_count - self._tra_last_porcentaje_count)
        
        if count_diff > max(10, current_count * 0.1):  # Más de 10 registros o 10% de diferencia
            from pal.services.tra import calcular_porcentajes_representacion
            
            # Verificar si estamos en modo dólares para precargar precios
            mostrar_dolares = getattr(self, 'tra_mostrar_dolares_var', tk.BooleanVar()).get()
            precios_cache = None
            
            if mostrar_dolares:
                # Precargar precios para todos los productos en cache
                codigos_precios = [str(item[0]) for item in cached_data]
                if codigos_precios:
                    self._cargar_precios_bulk(codigos_precios)
                    # Usar cache de precios existente
                    precios_cache = getattr(self, '_precios_cache', {})
            
            self.tra_porcentajes_map = calcular_porcentajes_representacion(
                cached_data, 
                mostrar_dolares=mostrar_dolares,
                precios_cache=precios_cache
            )
            self._tra_last_porcentaje_count = current_count
            self.tra_debug_log(
                f"Porcentajes recalculados para {current_count} productos (modo: {'dólares' if mostrar_dolares else 'unidades'})",
                throttle_key="percentage_calc",
                throttle_seconds=5.0
            )
            
        # Optimización: cache de stock para evitar consultas repetidas
        codigos = [r[0] for r in datos]
        sede = self.tra_sede_codigo or '0301'
        
        # Cache de stock con TTL de 30 segundos
        # Cache de stock con TTL de 30 segundos
        if not hasattr(self, '_stock_cache') or not hasattr(self, '_stock_cache_time') or not hasattr(self, '_stock_cache_sede'):
            self._stock_cache = {}
            self._stock_cache_time = 0
            self._stock_cache_sede = None
            
        # Si la sede cambió desde la última vez, invalidar cache para evitar mostrar stock de otra sede
        if self._stock_cache_sede != sede:
            self._stock_cache = {}
            self._stock_cache_sede = sede
            self._stock_cache_time = 0
        
        current_time = time.time()
        cache_ttl = 30  # 30 segundos
        
        # Identificar qué códigos necesitamos consultar
        codigos_faltantes = [
            cod for cod in codigos 
            if cod not in self._stock_cache or (current_time - self._stock_cache_time) > cache_ttl
        ]
        
        # Solo consultar los códigos que faltan en cache
        if codigos_faltantes:
            nuevos_stocks = self.obtener_stock_actual_bulk(codigos_faltantes, sede)
            self._stock_cache.update(nuevos_stocks)
            self._stock_cache_time = current_time
            self.tra_debug_log(
                f"Stock consultado: {len(codigos_faltantes)} códigos nuevos",
                throttle_key="stock_query",
                throttle_seconds=3.0
            )
        
        # Usar el cache para obtener el stock
        stock_map = {cod: self._stock_cache.get(cod, 0) for cod in codigos}
        
        for idx, fila in enumerate(datos):
            try:
                # Extraer valores de la fila
                codigo = fila[0]
                desc = fila[1]
                dept_cod = str(fila[2]) if len(fila) > 2 else ""
                group_cod = str(fila[3]) if len(fila) > 3 else ""
                sub_cod = str(fila[4]) if len(fila) > 4 else ""
                neto = fila[5] if len(fila) > 5 else 0
                rotacion = fila[6] if len(fila) > 6 else None
                precio = fila[7] if len(fila) > 7 else 0
                impuesto1 = fila[8] if len(fila) > 8 else 0
                
                # Obtener stock
                stock_actual = float(stock_map.get(codigo, 0) or 0)
                porcentaje = getattr(self, 'tra_porcentajes_map', {}).get(str(codigo), 0.0)

                # Calcular stock ideal y días restantes
                stock_ideal = self.calcular_stock_ideal_producto(neto)
                dias_restantes = self.calcular_dias_restantes(
                    stock_actual, neto, 
                    self.tra_fecha_inicio or datetime.now(),
                    self.tra_fecha_fin or datetime.now()
                )

                # Calcular Estado Stock basado en Días Restantes
                estado_stock = "N/A"
                if dias_restantes is not None:
                    try:
                        dr = float(dias_restantes)
                        if dr < 25: estado_stock = "Posible quiebre"
                        elif 25 <= dr <= 59: estado_stock = "Alerta Compra"
                        elif 60 <= dr <= 90: estado_stock = "Optimo"
                        elif 91 <= dr <= 119: estado_stock = "Critico"
                        else: estado_stock = "Sobre Stock"
                    except: pass

                # Determinar tag de color (solo por rotación)
                tag_base = (str(rotacion).lower() if rotacion else "sin_clasificar")
                tag_rotacion = tag_base if idx % 2 == 0 else f"{tag_base}_alt"

                # Formatear valores finales (sin información de quiebre de stock)
                neto_valor = float(neto or 0)
                mostrar_dolares = getattr(self, 'tra_mostrar_dolares_var', tk.BooleanVar()).get()
                precio_real = float(precio or 0) * (1 + float(impuesto1 or 0) / 100)
                ventas_display = f"{neto_valor * precio_real:.2f}" if mostrar_dolares and precio_real > 0 else f"{neto_valor:.0f}"
                
                values = (
                    codigo, desc, rotacion or "SIN CLASIFICAR",
                    ventas_display, f"{porcentaje:.2f}%",
                    int(stock_actual), int(stock_ideal),
                    f"{dias_restantes:.0f}" if dias_restantes is not None else "N/A",
                    estado_stock
                )
                
                # Insertar en el treeview con tag de rotación únicamente
                item_id = self.tra_tree.insert("", tk.END, values=values, tags=(tag_rotacion,))
            
            except Exception as e:
                self.log(f"Error procesando fila TRA {fila}: {str(e)}", "ERROR")
                continue
   
    def actualizar_controles_paginacion_tra(self, total_paginas):
        """Actualiza los controles de paginación TRA"""
        if hasattr(self, 'tra_pagina_label'):
            self.tra_pagina_label.config(text=f"Página {self.tra_current_page}/{total_paginas}")
        
        if hasattr(self, 'tra_btn_prev'):
            self.tra_btn_prev['state'] = 'normal' if self.tra_current_page > 1 else 'disabled'
        
        if hasattr(self, 'tra_btn_next'):
            self.tra_btn_next['state'] = 'normal' if self.tra_current_page < total_paginas else 'disabled'
    
    def actualizar_display_ventas_tra(self):
        """Actualiza el display de ventas entre unidades y dólares para TRA con clean reload"""
        try:
            # Actualizar encabezado de la columna
            if hasattr(self, 'tra_tree'):
                mostrar_dolares = getattr(self, 'tra_mostrar_dolares_var', tk.BooleanVar()).get()
                nuevo_encabezado = "Ventas ($)" if mostrar_dolares else "Ventas"
                
                # Actualizar el encabezado de la columna "Ventas" (índice 3)
                self.tra_tree.heading("Ventas", text=nuevo_encabezado)
                
                # Clean reload: recargar la página actual para asegurar consistencia
                if hasattr(self, 'tra_ventas_datos_filtrados') and self.tra_ventas_datos_filtrados:
                    # Recargar página actual para asegurar datos consistentes
                    self.mostrar_pagina_tra()
                else:
                    # Si no hay datos filtrados, actualizar items existentes
                    if mostrar_dolares:
                        # Cargar precios en bulk solo cuando se necesita mostrar en dólares
                        codigos_visibles = []
                        for item in self.tra_tree.get_children():
                            values = list(self.tra_tree.item(item)['values'])
                            if len(values) >= 1:
                                codigos_visibles.append(values[0])
                        
                        if codigos_visibles:
                            self._cargar_precios_bulk(codigos_visibles)
                    
                    # Actualizar datos existentes en el treeview
                    for item in self.tra_tree.get_children():
                        values = list(self.tra_tree.item(item)['values'])
                        if len(values) >= 4:  # Asegurarse que hay suficientes columnas
                            # Preservar formato original del código
                            codigo = values[0] if isinstance(values[0], str) else str(values[0])
                            ventas_actuales = values[3]
                            
                            # Obtener el valor original de unidades desde los datos cacheados
                            unidades_originales = 0
                            for fila in getattr(self, 'cached_ventas_tra', []):
                                if str(fila[0]) == str(codigo):
                                    unidades_originales = float(fila[5] if len(fila) > 5 else 0)  # neto está en índice 5
                                    break
                            
                            if mostrar_dolares:
                                # Convertir unidades a dólares usando precio cacheado
                                if unidades_originales > 0:
                                    # Usar precio desde cache de precios
                                    precio_unitario = self._obtener_precio_producto(codigo)
                                    ventas_dolares = unidades_originales * precio_unitario
                                    values[3] = f"${ventas_dolares:,.2f}"
                                else:
                                    values[3] = "$0.00"
                            else:
                                # Mostrar como unidades (valor original desde cache)
                                values[3] = int(unidades_originales) if unidades_originales == int(unidades_originales) else unidades_originales
                            
                            # Actualizar el item
                            self.tra_tree.item(item, values=values)
                        
        except Exception as e:
            self.log(f"Error actualizando display ventas TRA: {e}", "ERROR")
    
    def actualizar_display_ventas_mbrp(self):
        """Actualiza el display de ventas entre unidades y dólares para MBRP con clean reload"""
        try:
            # Actualizar encabezado de la columna
            if hasattr(self, 'mbrp_tree'):
                mostrar_dolares = getattr(self, 'mbrp_mostrar_dolares_var', tk.BooleanVar()).get()
                nuevo_encabezado = "Ventas ($)" if mostrar_dolares else "Ventas"
                
                # Actualizar el encabezado de la columna "Ventas" (índice 3)
                self.mbrp_tree.heading("Ventas", text=nuevo_encabezado)
                
                # Clean reload: recargar la página actual para asegurar consistencia
                if hasattr(self, 'mbrp_ventas_datos_filtrados') and self.mbrp_ventas_datos_filtrados:
                    # Recargar página actual para asegurar datos consistentes
                    self.cambiar_pagina_mbrp(0)  # 0 significa recargar página actual
                else:
                    # Si no hay datos filtrados, actualizar items existentes
                    if mostrar_dolares:
                        # Cargar precios en bulk solo cuando se necesita mostrar en dólares
                        codigos_visibles = []
                        for item in self.mbrp_tree.get_children():
                            values = list(self.mbrp_tree.item(item)['values'])
                            if len(values) >= 1:
                                codigos_visibles.append(values[0])
                        
                        if codigos_visibles:
                            self._cargar_precios_bulk(codigos_visibles)
                    
                    # Actualizar datos existentes en el treeview
                    for item in self.mbrp_tree.get_children():
                        values = list(self.mbrp_tree.item(item)['values'])
                        if len(values) >= 4:  # Asegurarse que hay suficientes columnas
                            # Preservar formato original del código
                            codigo = values[0] if isinstance(values[0], str) else str(values[0])
                            ventas_actuales = values[3]
                            
                            # Obtener el valor original de unidades y precio desde los datos cacheados
                            unidades_originales = 0
                            precio_unitario = 0
                            for fila in getattr(self, 'cached_ventas_mbrp', []):
                                if str(fila[0]) == str(codigo):
                                    unidades_originales = float(fila[5] if len(fila) > 5 else 0)  # neto está en índice 5
                                    # Después de clasificación MBRP, precio está en índice 7
                                    precio_unitario = float(fila[7] if len(fila) > 7 else 0)
                                    break
                            
                            if mostrar_dolares:
                                # Convertir unidades a dólares usando precio cacheado
                                if unidades_originales > 0:
                                    # Usar precio desde cache de precios
                                    precio_unitario = self._obtener_precio_producto(codigo)
                                    ventas_dolares = unidades_originales * precio_unitario
                                    values[3] = f"${ventas_dolares:,.2f}"
                                else:
                                    values[3] = "$0.00"
                            else:
                                # Mostrar como unidades (valor original desde cache)
                                values[3] = int(unidades_originales) if unidades_originales == int(unidades_originales) else unidades_originales
                            
                            # Actualizar el item
                            self.mbrp_tree.item(item, values=values)
                        
        except Exception as e:
            self.log(f"Error actualizando display ventas MBRP: {e}", "ERROR")
    
    def _convertir_unidades_a_dolares(self, codigo_producto, unidades):
        """Convierte unidades vendidas a dólares usando el precio cacheado del producto"""
        try:
            if not codigo_producto or unidades <= 0:
                return 0.0
            
            # Usar el precio desde cache de productos (más eficiente y confiable)
            precio_unitario = self._obtener_precio_producto(codigo_producto)
            
            return unidades * precio_unitario
                
        except Exception as e:
            self.log(f"Error convirtiendo unidades a dólares para {codigo_producto}: {e}", "ERROR")
            return 0.0
            
            # Obtener precio del producto desde la base de datos
            precio = self._obtener_precio_producto(codigo_producto)
            if precio > 0:
                return unidades * precio
            else:
                return 0.0
                
        except Exception as e:
            self.log(f"Error convirtiendo unidades a dólares para {codigo_producto}: {e}", "ERROR")
            return 0.0
    
    def _cargar_precios_bulk(self, codigos):
        """Carga precios en bulk para una lista de códigos de producto"""
        try:
            if not codigos:
                return
            
            # Inicializar cache si no existe
            if not hasattr(self, '_precios_cache'):
                self._precios_cache = {}
                self._precios_cache_time = 0
            
            import time
            current_time = time.time()
            cache_ttl = 300  # 5 minutos de cache
            
            # Limpiar cache si expiró
            if current_time - self._precios_cache_time > cache_ttl:
                self._precios_cache = {}
                self._precios_cache_time = current_time
            
            # Filtrar códigos que no están en cache
            codigos_faltantes = [str(c).strip() for c in codigos if str(c) not in self._precios_cache]
            
            if not codigos_faltantes:
                return  # Ya tenemos todos los precios en cache
            
            # Consultar precios en chunk para evitar límite de 2100 parámetros de SQL Server
            chunk_size = 2000  # Dejamos margen bajo el límite de 2100
            
            self.log(f"Cargando precios para {len(codigos_faltantes)} productos en chunks de {chunk_size}...", "INFO")
            all_results = []
            
            for i in range(0, len(codigos_faltantes), chunk_size):
                chunk = codigos_faltantes[i:i + chunk_size]
                placeholders = ','.join(['?' for _ in chunk])
                query = f"""
                    SELECT C_CODIGO, COALESCE(n_precio1, 0) AS precio, COALESCE(n_impuesto1, 0) AS impuesto1
                    FROM MA_PRODUCTOS WITH (NOLOCK)
                    WHERE C_CODIGO IN ({placeholders})
                """
                
                try:
                    chunk_results = self.db_manager.fetch_data(query, chunk)
                    all_results.extend(chunk_results)
                except Exception as e:
                    self.log(f"Error cargando chunk {i//chunk_size + 1} de precios: {e}", "ERROR")
                    continue
            
            results = all_results
            
            # Actualizar cache con precio + IVA
            for row in results:
                if row and len(row) >= 3:
                    codigo = str(row[0]).strip()
                    precio_base = float(row[1] or 0)
                    impuesto_pct = float(row[2] or 0)
                    # Calcular precio con IVA incluido
                    precio_con_iva = precio_base * (1.0 + (impuesto_pct / 100.0))
                    self._precios_cache[codigo] = precio_con_iva
                    # Debug: log para verificar caché
                    if codigo == '016208':  # Debug específico para el código mencionado
                        self.tra_debug_log(
                            f"CACHE PRECIO {codigo}: base={precio_base}, iva={impuesto_pct}%, con_iva={precio_con_iva}",
                            throttle_key="cache_price_debug",
                            throttle_seconds=5.0
                        )
            
            # Para códigos no encontrados, guardar precio 0
            for codigo in codigos_faltantes:
                if codigo not in self._precios_cache:
                    self._precios_cache[codigo] = 0.0
            
            self.log(f"Precios cargados: {len(self._precios_cache)} productos en cache ({len(results)} recuperados de BD)", "SUCCESS")
            
        except Exception as e:
            self.log(f"Error cargando precios bulk: {e}", "ERROR")

    def _obtener_precio_producto(self, codigo_producto):
        """Obtiene el precio de un producto con IVA incluido desde la base de datos con cacheo"""
        try:
            if not codigo_producto:
                return 0.0
            
            # Inicializar cache si no existe
            if not hasattr(self, '_precios_cache'):
                self._precios_cache = {}
                self._precios_cache_time = 0
            
            import time
            current_time = time.time()
            cache_ttl = 300  # 5 minutos de cache
            
            # Limpiar cache si expiró
            if current_time - self._precios_cache_time > cache_ttl:
                self._precios_cache = {}
                self._precios_cache_time = current_time
            
            # Verificar si ya está en cache
            codigo_str = str(codigo_producto).strip()
            if codigo_str in self._precios_cache:
                return self._precios_cache[codigo_str]
            
            # Debug: log del código que se está consultando
            self.tra_debug_log(
                f"Consultando precio para código: '{codigo_producto}' (tipo: {type(codigo_producto)})",
                throttle_key="price_query",
                throttle_seconds=5.0
            )
            
            # Asegurarse que el código sea string para evitar problemas de conversión
            codigo_str = str(codigo_producto).strip()
            
            # Consultar precio con IVA a la base de datos - evitar cualquier conversión implícita
            query = """
                SELECT TOP 1 COALESCE(n_precio1, 0) AS precio, COALESCE(n_impuesto1, 0) AS impuesto1
                FROM MA_PRODUCTOS WITH (NOLOCK)
                WHERE RTRIM(LTRIM(C_CODIGO)) = RTRIM(LTRIM(?))
            """
            
            result = self.db_manager.fetch_data(query, (codigo_str,))
            if result and len(result) > 0:
                precio_base = float(result[0][0] or 0)
                impuesto_pct = float(result[0][1] or 0)
                # Calcular precio con IVA incluido
                precio_con_iva = precio_base * (1.0 + (impuesto_pct / 100.0))
                self._precios_cache[codigo_producto] = precio_con_iva
                # Debug: log para verificar cálculo
                self.tra_debug_log(
                    f"Precio para {codigo_producto}: base={precio_base}, iva={impuesto_pct}%, con_iva={precio_con_iva}",
                    throttle_key="price_calc",
                    throttle_seconds=5.0
                )
                return precio_con_iva
            else:
                self._precios_cache[codigo_producto] = 0.0
                return 0.0
                
        except Exception as e:
            self.log(f"Error obteniendo precio para {codigo_producto}: {e}", "ERROR")
            # En caso de error, devolver 0.0 y cachear para evitar repetir el error
            if hasattr(self, '_precios_cache'):
                self._precios_cache[codigo_producto] = 0.0
            return 0.0
    
    def _check_jerarquia_cache(self):
        """Verifica si existe cache válido de jerarquías"""
        try:
            import os
            import json
            from datetime import datetime, timedelta
            
            cache_file = "jerarquia_cache.json"
            cache_ttl = timedelta(hours=12)  # Cache por 12 horas
            
            if not os.path.exists(cache_file):
                return None
            
            # Verificar TTL
            mtime = datetime.fromtimestamp(os.path.getmtime(cache_file))
            if datetime.now() - mtime > cache_ttl:
                self.log("Cache de jerarquía expirado, recargando...", "INFO")
                try:
                    os.remove(cache_file)
                except Exception:
                    pass
                return None
            
            # Cargar datos desde cache
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            
            self.log("⚙️ Cache de jerarquía cargado exitosamente", "SUCCESS")
            return cached_data
            
        except Exception as e:
            self.log(f"Error verificando cache de jerarquía: {e}", "ERROR")
            return None
    
    def _save_jerarquia_cache(self, data):
        """Guarda jerarquías en cache local"""
        try:
            import json
            
            cache_file = "jerarquia_cache.json"
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            self.log("💾 Cache de jerarquía guardado", "DEBUG")
            
        except Exception as e:
            self.log(f"Error guardando cache de jerarquía: {e}", "ERROR")
    
    def cargar_jerarquia_unificada(self):
        """Carga jerarquías TRA y MBRP con una sola consulta optimizada JOIN"""
        import time
        # Evitar cargas duplicadas si ya fue cargado por otro hilo
        if getattr(self, 'jerarquias_unificadas_cargadas', False):
            # IMPORTANTE: Verificar que los diccionarios realmente tengan datos
            # Si el flag está en True pero los diccionarios están vacíos, hay que recargar
            if (hasattr(self, 'tra_dept_dict') and self.tra_dept_dict and 
                hasattr(self, 'tra_group_dict') and self.tra_group_dict):
                # Asegurar que los combos estén actualizados incluso si omitimos la carga
                self._update_hierarchy_combos()
                self.log("Jerarquía unificada ya cargada — se omite carga duplicada", "DEBUG")
                return
            else:
                # Flag está en True pero diccionarios vacíos - resetear flag y recargar
                self.log("⚠️ Flag de jerarquía en True pero diccionarios vacíos - recargando...", "WARNING")
                self.jerarquias_unificadas_cargadas = False
        start_time = time.perf_counter()
        
        # Verificar cache primero
        cached_data = self._check_jerarquia_cache()
        if cached_data and 'tra' in cached_data and 'mbrp' in cached_data:
            # Cargar TRA desde cache
            tra_data = cached_data['tra']
            self.tra_dept_dict = tra_data.get('departments', {})
            self.tra_group_dict = tra_data.get('groups', {})
            self.tra_sub_dict = tra_data.get('subgroups', {})
            
            # Cargar MBRP desde cache
            mbrp_data = cached_data['mbrp']
            self.mbrp_dept_dict = mbrp_data.get('departments', {})
            self.mbrp_group_dict = mbrp_data.get('groups', {})
            self.mbrp_sub_dict = mbrp_data.get('subgroups', {})
            
            # Actualizar combos si existen
            self._update_hierarchy_combos()
            
            total_items = (len(self.tra_dept_dict) + sum(len(v) for v in self.tra_group_dict.values()) + 
                          sum(len(v) for v in self.tra_sub_dict.values()))
            
            load_time = time.perf_counter() - start_time
            self.jerarquias_unificadas_cargadas = True
            self.log(f"⚡ Jerarquía UNIFICADA cargada desde cache en {load_time:.3f}s - {total_items} elementos", "SUCCESS")
            
            # Programar actualización adicional por si los combos se crean después
            if hasattr(self, 'root'):
                self.root.after(200, self._update_hierarchy_combos)
            return
        
        # Si no hay cache válido, cargar desde BD con consulta optimizada
        if not self.db_manager or not self.db_manager.ensure_connection():
            self.log("No hay conexión válida para cargar jerarquía unificada", "WARNING")
            return
        
        try:
            # Consulta JOIN optimizada: UNA SOLA consulta en lugar de 6
            query = """
            SELECT DISTINCT
                d.C_CODIGO as dept_cod, d.C_DESCRIPCIO as dept_desc,
                g.C_CODIGO as group_cod, g.C_DESCRIPCIO as group_desc,
                s.C_CODIGO as sub_cod, s.C_DESCRIPCIO as sub_desc
            FROM MA_DEPARTAMENTOS d
            LEFT JOIN MA_GRUPOS g ON d.C_CODIGO = g.C_DEPARTAMENTO
            LEFT JOIN MA_SUBGRUPOS s ON g.C_DEPARTAMENTO = s.C_IN_DEPARTAMENTO 
                AND g.C_CODIGO = s.C_IN_GRUPO
            WHERE d.C_CODIGO IS NOT NULL AND d.C_DESCRIPCIO IS NOT NULL
            ORDER BY d.C_CODIGO, g.C_CODIGO, s.C_CODIGO
            """
            
            data = self.db_manager.fetch_data(query)
            
            # Procesar resultados de forma eficiente O(n) en lugar de O(n²)
            tra_dept_dict = {}
            tra_group_dict = {}
            tra_sub_dict = {}
            
            for row in data:
                dept_cod, dept_desc, group_cod, group_desc, sub_cod, sub_desc = row
                
                # Departamentos (evitar duplicados)
                if dept_cod and dept_desc and dept_desc.strip() not in tra_dept_dict:
                    tra_dept_dict[dept_desc.strip()] = dept_cod.strip()
                    
                # Grupos por departamento  
                if group_cod and group_desc and dept_cod:
                    dept_key = dept_cod.strip()
                    if dept_key not in tra_group_dict:
                        tra_group_dict[dept_key] = {}
                    tra_group_dict[dept_key][group_desc.strip()] = group_cod.strip()
                    
                # Subgrupos por departamento y grupo
                if sub_cod and sub_desc and dept_cod and group_cod:
                    # Usar string como key en lugar de tupla para compatibilidad con JSON
                    key = f"{dept_cod.strip()}|{group_cod.strip()}"
                    if key not in tra_sub_dict:
                        tra_sub_dict[key] = {}
                    tra_sub_dict[key][sub_desc.strip()] = sub_cod.strip()
            # Asignar a ambos módulos (TRA y MBRP comparten la misma jerarquía)
            self.tra_dept_dict = tra_dept_dict
            self.tra_group_dict = tra_group_dict
            self.tra_sub_dict = tra_sub_dict
            
            # DEBUG: Verificar asignación
            self.log(f"🔍 [JERARQUIA] Asignados a TRA: {len(self.tra_dept_dict)} depts, {len(self.tra_group_dict)} groups", "DEBUG")
            
            self.mbrp_dept_dict = tra_dept_dict.copy()
            self.mbrp_group_dict = tra_group_dict.copy()
            self.mbrp_sub_dict = tra_sub_dict.copy()
            
            # DEBUG: Verificar asignación MBRP
            self.log(f"🔍 [JERARQUIA] Asignados a MBRP: {len(self.mbrp_dept_dict)} depts, {len(self.mbrp_group_dict)} groups", "DEBUG")
            
            # Calcular totales
            total_items = len(tra_dept_dict) + sum(len(v) for v in tra_group_dict.values()) + sum(len(v) for v in tra_sub_dict.values())
            load_time = time.perf_counter() - start_time

            if total_items == 0:
                # Fallback explícito si la consulta no devolvió datos útiles
                self.log("Jerar quía unificada vacía, aplicando fallback por módulos", "WARNING")
                self.cargar_jerarquia_tra()
                self.cargar_jerarquia_mbrp()
                return
            
            # Actualizar combos MBRP inmediatamente si la pestaña ya existe
            if hasattr(self, 'mbrp_dept_combo') and self.mbrp_dept_dict:
                try:
                    self.mbrp_dept_combo['values'] = ['Todos'] + list(self.mbrp_dept_dict.keys())
                    if hasattr(self, 'mbrp_dept_var'):
                        self.mbrp_dept_var.set('Todos')
                except Exception:
                    pass
            
            # Actualizar combos si están disponibles
            self._update_hierarchy_combos()
            self._update_hierarchy_combos()
            
            # Marcar como cargado para evitar duplicados
            self.jerarquias_unificadas_cargadas = True
            
            # Guardar en cache unificado
            cache_data = {
                'tra': {
                    'departments': self.tra_dept_dict,
                    'groups': self.tra_group_dict,
                    'subgroups': self.tra_sub_dict
                },
                'mbrp': {
                    'departments': self.mbrp_dept_dict,
                    'groups': self.mbrp_group_dict,
                    'subgroups': self.mbrp_sub_dict
                }
            }
            self._save_jerarquia_cache(cache_data)
            
            self.log(f"⚙️ Jerarquía UNIFICADA cargada y cacheada en {load_time:.3f}s - {total_items} elementos", "SUCCESS")
            
            # Programar actualización adicional después de 200ms por si los combos se crean tarde
            if hasattr(self, 'root'):
                self.root.after(200, self._update_hierarchy_combos)
            
        except Exception as e:
            self.log(f"Error cargando jerarquía unificada: {e}", "ERROR")
            # Fallback a métodos individuales
            self.cargar_jerarquia_tra()
            self.cargar_jerarquia_mbrp()
    
        # Vincular evento de cambio de pestaña para automatizaciones
        self.main_notebook.bind("<<NotebookTabChanged>>", self._on_main_tab_changed)
        
        # Carga inicial automática de quiebres (silenciosa al inicio)
        if self.modules_enabled.get("stock", False):
            self.root.after(3000, lambda: self.actualizar_alertas_stock(force_refresh=False))

    def _on_main_tab_changed(self, event):
        """Maneja la automatización al cambiar de pestaña"""
        try:
            selected_tab = self.main_notebook.select()
            tab_text = self.main_notebook.tab(selected_tab, "text")
            
            # Si entra a Quiebre de Stock, actualizar automáticamente
            if "Quiebre de Stock" in tab_text:
                self.actualizar_alertas_stock(force_refresh=False)
        except Exception as e:
            self.log(f"Error en cambio de pestaña: {e}", "DEBUG")

    def _update_hierarchy_combos(self):
        """Actualiza los combos de jerar quía para ambos módulos"""
        try:
            tra_actualizado = False
            mbrp_actualizado = False
            
            # Debug: verificar estado
            has_tra_combo = hasattr(self, 'tra_dept_combo')
            has_mbrp_combo = hasattr(self, 'mbrp_dept_combo')
            has_tra_dict = hasattr(self, 'tra_dept_dict') and self.tra_dept_dict
            has_mbrp_dict = hasattr(self, 'mbrp_dept_dict') and self.mbrp_dept_dict
            
            # TRA combos - con validación exhaustiva del widget
            if (has_tra_combo and has_tra_dict):
                try:
                    # Validar que el widget existe y es accesible
                    if self.tra_dept_combo.winfo_exists():
                        valores_tra = ['Todos'] + list(self.tra_dept_dict.keys())
                        self.tra_dept_combo['values'] = valores_tra
                        if hasattr(self, 'tra_dept_var'):
                            self.tra_dept_var.set('Todos')
                        
                        # NUEVO: Inicializar combos de grupos y subgrupos con TODAS las opciones disponibles
                        # Esto asegura que los filtros sean funcionales inmediatamente
                        if hasattr(self, 'tra_group_combo') and hasattr(self, 'tra_group_dict'):
                            try:
                                # Recopilar TODOS los grupos de TODOS los departamentos
                                all_groups = set()
                                for dept_groups in self.tra_group_dict.values():
                                    all_groups.update(dept_groups.keys())
                                
                                if all_groups:
                                    self.tra_group_combo['values'] = ['Todos'] + sorted(list(all_groups))
                                    if hasattr(self, 'tra_group_var'):
                                        self.tra_group_var.set('Todos')
                            except Exception:
                                pass
                        
                        if hasattr(self, 'tra_sub_combo') and hasattr(self, 'tra_sub_dict'):
                            try:
                                # Recopilar TODOS los subgrupos de TODOS los grupos
                                all_subs = set()
                                for sub_groups in self.tra_sub_dict.values():
                                    all_subs.update(sub_groups.keys())
                                
                                if all_subs:
                                    self.tra_sub_combo['values'] = ['Todos'] + sorted(list(all_subs))
                                    if hasattr(self, 'tra_sub_var'):
                                        self.tra_sub_var.set('Todos')
                            except Exception:
                                pass
                        
                        tra_actualizado = True
                except (tk.TclError, AttributeError):
                    # Widget destruido o no accesible
                    pass
                
            # MBRP combos - con validación exhaustiva del widget
            if (has_mbrp_combo and has_mbrp_dict):
                try:
                    # Validar que el widget existe y es accesible
                    if self.mbrp_dept_combo.winfo_exists():
                        valores_mbrp = ['Todos'] + list(self.mbrp_dept_dict.keys())
                        self.mbrp_dept_combo['values'] = valores_mbrp
                        if hasattr(self, 'mbrp_dept_var'):
                            self.mbrp_dept_var.set('Todos')
                        
                        # NUEVO: Inicializar combos de grupos y subgrupos para MBRP también
                        if hasattr(self, 'mbrp_group_combo') and hasattr(self, 'mbrp_group_dict'):
                            try:
                                all_groups = set()
                                for dept_groups in self.mbrp_group_dict.values():
                                    all_groups.update(dept_groups.keys())
                                
                                if all_groups:
                                    self.mbrp_group_combo['values'] = ['Todos'] + sorted(list(all_groups))
                                    if hasattr(self, 'mbrp_group_var'):
                                        self.mbrp_group_var.set('Todos')
                            except Exception:
                                pass
                        
                        if hasattr(self, 'mbrp_sub_combo') and hasattr(self, 'mbrp_sub_dict'):
                            try:
                                all_subs = set()
                                for sub_groups in self.mbrp_sub_dict.values():
                                    all_subs.update(sub_groups.keys())
                                
                                if all_subs:
                                    self.mbrp_sub_combo['values'] = ['Todos'] + sorted(list(all_subs))
                                    if hasattr(self, 'mbrp_sub_var'):
                                        self.mbrp_sub_var.set('Todos')
                            except Exception:
                                pass
                        
                        mbrp_actualizado = True
                except (tk.TclError, AttributeError):
                    # Widget destruido o no accesible
                    pass

            # STOCK combos - Nuevo bloque para sincronizar quiebres de stock
            if hasattr(self, 'stock_dept_combo') and hasattr(self, 'stock_dept_var'):
                try:
                    if self.stock_dept_combo.winfo_exists():
                        # Llamar a load_stock_filters que ahora es eficiente
                        self.load_stock_filters()
                except Exception:
                    pass
            
            # Log consolidado - solo una línea
            if tra_actualizado or mbrp_actualizado:
                modulos = []
                if tra_actualizado:
                    modulos.append(f"TRA({len(self.tra_dept_dict)} depts)")
                if mbrp_actualizado:
                    modulos.append(f"MBRP({len(self.mbrp_dept_dict)} depts)")
                self.log(f"✅ Filtros actualizados: {', '.join(modulos)}", "SUCCESS")
            elif has_mbrp_dict and not mbrp_actualizado:
                # Debug: MBRP tiene datos pero combo no se actualizó
                self.log(f"⚠️ MBRP tiene {len(self.mbrp_dept_dict)} depts pero combo aún no creado (has_combo={has_mbrp_combo})", "DEBUG")
                
        except Exception as e:
            self.log(f"Error actualizando combos de jerar quía: {e}", "ERROR")
    
    def _inicializar_modulos_paralelo(self):
        """Inicializa módulos en paralelo real usando ThreadPoolExecutor"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import time
        
        start_time = time.perf_counter()
        self.log("🚀 Iniciando carga paralela de módulos...", "INFO")
        
        # Definir tareas a ejecutar en paralelo
        tareas = []
        
        # Stock: filtros + alertas iniciales
        if self.modules_enabled.get("stock", False):
            tareas.append(("stock_filters", self._load_stock_parallel))
            tareas.append(("stock_alerts", self._load_stock_alerts_parallel))
        
        # Jerarquías TRA/MBRP unificadas
        if (self.modules_enabled.get("tra", False) or self.modules_enabled.get("mbrp", False)):
            tareas.append(("jerarquias", self._load_hierarchies_parallel))
        
        # Jerarquía de productos (para stock)
        if self.modules_enabled.get("stock", False):
            tareas.append(("producto_jerarquia", self._init_jerarquia_async))
        
        # Ejecutar tareas en paralelo con máximo 4 hilos
        max_workers = min(4, len(tareas)) if tareas else 1
        resultados = {}
        
        try:
            with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="ModuleInit") as executor:
                # Enviar todas las tareas
                future_to_task = {
                    executor.submit(task_func): task_name 
                    for task_name, task_func in tareas
                }
                
                # Procesar resultados conforme van completando
                for future in as_completed(future_to_task, timeout=30):
                    task_name = future_to_task[future]
                    try:
                        resultado = future.result(timeout=5)
                        resultados[task_name] = {"status": "success", "result": resultado}
                        self.log(f"✅ Módulo {task_name} cargado exitosamente", "SUCCESS")
                    except Exception as e:
                        resultados[task_name] = {"status": "error", "error": str(e)}
                        self.log(f"❌ Error en módulo {task_name}: {e}", "ERROR")
                        
        except Exception as e:
            self.log(f"Error en carga paralela: {e}", "ERROR")
            # Evitar fallback si ya hay módulos críticos listos (por ejemplo, jerarquías)
            if not getattr(self, 'jerarquias_unificadas_cargadas', False):
                self._fallback_inicializacion_secuencial()
            else:
                self.log("Omitiendo fallback: jerarquías ya cargadas", "INFO")
            return
        
        # Estadísticas finales
        total_time = time.perf_counter() - start_time
        exitosos = sum(1 for r in resultados.values() if r["status"] == "success")
        fallidos = len(resultados) - exitosos
        
        self.log(
            f"🎆 Inicialización paralela completada en {total_time:.3f}s | "
            f"Módulos: {exitosos} exitosos, {fallidos} fallidos", 
            "SUCCESS" if fallidos == 0 else "WARNING"
        )
        
        # Ejecutar tareas post-carga en hilo principal
        self.root.after(100, self._post_init_tasks, resultados)
    
    def _load_stock_parallel(self):
        """Carga filtros de stock en hilo paralelo"""
        try:
            self.load_stock_filters()
            return {"filters_loaded": True}
        except Exception as e:
            raise Exception(f"Error cargando filtros stock: {e}")
    
    def _load_stock_alerts_parallel(self):
        """Carga quiebres iniciales de stock en hilo paralelo"""
        try:
            # Ensure connection is valid before loading alerts
            if not self.db_manager.ensure_connection():
                raise Exception("No hay conexión válida a la base de datos")
            
            # Get configured warehouses from all sedes
            sedes_config = self.config_manager.get_sedes_config()
            all_warehouses = []
            for sede_name, config in sedes_config.items():
                depositos = config.get('almacenes_tratables', [])
                all_warehouses.extend(depositos)
            
            if not all_warehouses:
                self.log("No hay almacenes configurados para stock", "WARNING")
                return {"alerts_loaded": False, "error": "No warehouses configured"}
            
            # Use new quiebre detection logic
            quiebres = self.db_manager.obtener_quiebres_directos(all_warehouses)
            if quiebres:
                # Convert dict format to tuple format for compatibility
                self.cached_alertas = [
                    (q['codigo'], q['descripcion'], q['sede'], q['unidades_perdidas'], 
                     q['dias_quiebre'], q['ultima_compra'], q['ultima_venta'])
                    for q in quiebres
                ]
                self.last_refresh = datetime.now()
                self.log(f"Quiebres stock cargados: {len(quiebres)} registros", "SUCCESS")
            else:
                self.cached_alertas = []
                self.last_refresh = datetime.now()
            
            return {"alerts_loaded": True, "count": len(quiebres) if quiebres else 0}
        except Exception as e:
            self.log(f"Error cargando quiebres stock: {e}", "ERROR")
            # Don't raise - allow app to continue without initial alerts
            return {"alerts_loaded": False, "error": str(e)}
    
    def _load_hierarchies_parallel(self):
        """Carga jerarquías unificadas en hilo paralelo"""
        try:
            self.cargar_jerarquia_unificada()
            # Programar actualización de combos en hilo principal después de cargar
            self.root.after(100, self._update_hierarchy_combos)
            return {"hierarchies_loaded": True}
        except Exception as e:
            # Fallback a métodos individuales
            try:
                if self.modules_enabled.get("tra", False):
                    self.cargar_jerarquia_tra()
                if self.modules_enabled.get("mbrp", False):
                    self.cargar_jerarquia_mbrp()
                # Programar actualización de combos incluso en fallback
                self.root.after(100, self._update_hierarchy_combos)
                return {"hierarchies_loaded": True, "fallback": True}
            except Exception as e2:
                raise Exception(f"Error cargando jerarquías: {e} (fallback: {e2})")
    
    def _fallback_inicializacion_secuencial(self):
        """Fallback a inicialización secuencial si falla la paralela"""
        self.log("⚠️ Fallback: carga secuencial de módulos", "WARNING")
        
        try:
            if self.modules_enabled.get("stock", False):
                self.load_stock_filters()
                self.actualizar_alertas_stock(force_refresh=True)
                threading.Thread(target=self._init_jerarquia_async, daemon=True).start()
            
            if (self.modules_enabled.get("tra", False) or self.modules_enabled.get("mbrp", False)) and not getattr(self, 'jerarquias_unificadas_cargadas', False):
                self.cargar_jerarquia_unificada()
                
        except Exception as e:
            self.log(f"Error en fallback secuencial: {e}", "ERROR")
    
    def _post_init_tasks(self, resultados):
        """Tareas post-inicialización en hilo principal"""
        try:
            # Realizar búsqueda inicial si stock está habilitado y fue exitoso
            if (self.modules_enabled.get("stock", False) and 
                resultados.get("stock_filters", {}).get("status") == "success"):
                # Esto se ejecuta en hilo principal para actualizar UI
                pass  # search_records ya se ejecutó antes
                
        except Exception as e:
            self.log(f"Error en tareas post-inicialización: {e}", "ERROR")
    
    def cargar_jerarquia_tra(self):
        """Carga diccionarios para departamentos, grupos y subgrupos con cache optimizado"""
        # Intentar cargar desde cache primero
        cached_data = self._check_jerarquia_cache()
        if cached_data and 'tra' in cached_data:
            try:
                tra_data = cached_data['tra']
                self.tra_dept_dict = tra_data.get('departments', {})
                self.tra_group_dict = tra_data.get('groups', {})
                self.tra_sub_dict = tra_data.get('subgroups', {})
                
                # Actualizar combo de departamentos si existe
                if hasattr(self, 'tra_dept_combo') and self.tra_dept_dict:
                    self.tra_dept_combo['values'] = ['Todos'] + list(self.tra_dept_dict.keys())
                    self.tra_dept_var.set('Todos')
                
                total_items = len(self.tra_dept_dict) + sum(len(v) for v in self.tra_group_dict.values()) + sum(len(v) for v in self.tra_sub_dict.values())
                self.log(f"⚙️ Jerarquía TRA cargada desde cache - {total_items} elementos", "SUCCESS")
                return
            except Exception as e:
                self.log(f"Error usando cache TRA: {e}", "ERROR")
        
        # Verificar conexión a la base de datos
        if not self.db_manager or not self.db_manager.ensure_connection():
            self.log("No hay conexión válida a la base de datos para cargar jerarquía TRA", "WARNING")
            return
        
        # Inicializar diccionarios vacíos por si hay errores
        self.tra_dept_dict = {}
        self.tra_group_dict = {}
        self.tra_sub_dict = {}
        
        try:
            # Departamentos con manejo de errores específico
            try:
                deps = self.db_manager.fetch_data(
                    "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                self.tra_dept_dict = {desc.strip(): cod.strip() for cod, desc in deps if cod and desc and str(cod).strip() and str(desc).strip()}
                
                # Actualizar combo de departamentos si existe
                if hasattr(self, 'tra_dept_combo') and self.tra_dept_dict:
                    self.tra_dept_combo['values'] = ['Todos'] + list(self.tra_dept_dict.keys())
                    self.tra_dept_var.set('Todos')
                    
                self.log(f"Departamentos TRA cargados: {len(self.tra_dept_dict)}", "DEBUG")
            except Exception as e:
                self.log(f"Error cargando departamentos TRA: {str(e)}", "ERROR")
        
            # Grupos por departamento con manejo de errores específico
            try:
                grupos = self.db_manager.fetch_data(
                    "SELECT C_DEPARTAMENTO, C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WHERE C_DEPARTAMENTO IS NOT NULL AND C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                self.tra_group_dict = {}
                for dept, cod, desc in grupos:
                    if dept and cod and desc:
                        dept_clean = str(dept).strip()
                        cod_clean = str(cod).strip()
                        desc_clean = str(desc).strip()
                        
                        if dept_clean not in self.tra_group_dict:
                            self.tra_group_dict[dept_clean] = {}
                        self.tra_group_dict[dept_clean][desc_clean] = cod_clean
                        
                self.log(f"Grupos TRA cargados: {sum(len(v) for v in self.tra_group_dict.values())}", "DEBUG")
            except Exception as e:
                self.log(f"Error cargando grupos TRA: {str(e)}", "ERROR")
        
            # Subgrupos por departamento y grupo con manejo de errores específico
            try:
                subs = self.db_manager.fetch_data(
                    "SELECT C_IN_DEPARTAMENTO, C_IN_GRUPO, C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WHERE C_IN_DEPARTAMENTO IS NOT NULL AND C_IN_GRUPO IS NOT NULL AND C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                self.tra_sub_dict = {}
                for dept, grp, cod, desc in subs:
                    if dept and grp and cod and desc:
                        dept_clean = str(dept).strip()
                        grp_clean = str(grp).strip()
                        cod_clean = str(cod).strip()
                        desc_clean = str(desc).strip()
                        
                        # Usar string como key en lugar de tupla para compatibilidad con JSON
                        key = f"{dept_clean}|{grp_clean}"
                        if key not in self.tra_sub_dict:
                            self.tra_sub_dict[key] = {}
                        self.tra_sub_dict[key][desc_clean] = cod_clean
                        
                self.log(f"Subgrupos TRA cargados: {sum(len(v) for v in self.tra_sub_dict.values())}", "DEBUG")
            except Exception as e:
                self.log(f"Error cargando subgrupos TRA: {str(e)}", "ERROR")
            
            # Verificar si se cargaron datos
            total_items = len(self.tra_dept_dict) + sum(len(v) for v in self.tra_group_dict.values()) + sum(len(v) for v in self.tra_sub_dict.values())
            if total_items > 0:
                self.log(f"✅ Jerarquía TRA cargada correctamente - Total elementos: {total_items}", "SUCCESS")
                
                # Guardar en cache para próximas cargas
                try:
                    cache_data = {
                        'tra': {
                            'departments': self.tra_dept_dict,
                            'groups': self.tra_group_dict,
                            'subgroups': self.tra_sub_dict
                        }
                    }
                    
                    # Si ya existe cache con MBRP, combinar
                    existing_cache = self._check_jerarquia_cache()
                    if existing_cache:
                        cache_data.update(existing_cache)
                    
                    self._save_jerarquia_cache(cache_data)
                except Exception as e:
                    self.log(f"Error guardando cache TRA: {e}", "DEBUG")
            else:
                self.log("⚠️ Jerarquía TRA cargada pero sin datos disponibles", "WARNING")
        
        except Exception as e:
            self.log(f"Error general cargando jerarquía TRA: {str(e)}", "ERROR")
            # Asegurar que los diccionarios estén inicializados incluso en caso de error
            if not hasattr(self, 'tra_dept_dict'):
                self.tra_dept_dict = {}
            if not hasattr(self, 'tra_group_dict'):
                self.tra_group_dict = {}
            if not hasattr(self, 'tra_sub_dict'):
                self.tra_sub_dict = {}

    def cargar_jerarquia_mbrp(self):
        """Carga diccionarios para departamentos, grupos y subgrupos MBRP con cache optimizado"""
        # Intentar cargar desde cache primero
        cached_data = self._check_jerarquia_cache()
        if cached_data and 'mbrp' in cached_data:
            try:
                mbrp_data = cached_data['mbrp']
                self.mbrp_dept_dict = mbrp_data.get('departments', {})
                self.mbrp_group_dict = mbrp_data.get('groups', {})
                self.mbrp_sub_dict = mbrp_data.get('subgroups', {})
                
                # Actualizar combo de departamentos si existe
                if hasattr(self, 'mbrp_dept_combo') and self.mbrp_dept_dict:
                    self.mbrp_dept_combo['values'] = ['Todos'] + list(self.mbrp_dept_dict.keys())
                    self.mbrp_dept_var.set('Todos')
                
                self.log("⚙️ Jerarquía MBRP cargada desde cache", "SUCCESS")
                return
            except Exception as e:
                self.log(f"Error usando cache MBRP: {e}", "ERROR")
        
        if not getattr(self.db_manager, 'conn', None):
            return
        try:
            # Departamentos
            deps = self.db_manager.fetch_data(
                "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS"
            )
            self.mbrp_dept_dict = {desc: cod for cod, desc in deps if cod and desc}
            
            # Actualizar combo de departamentos si existe
            if hasattr(self, 'mbrp_dept_combo'):
                self.mbrp_dept_combo['values'] = ['Todos'] + list(self.mbrp_dept_dict.keys())
                self.mbrp_dept_var.set('Todos')
        
            # Grupos por departamento
            grupos = self.db_manager.fetch_data(
                "SELECT C_DEPARTAMENTO, C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS"
            )
            self.mbrp_group_dict = {}
            for dept, cod, desc in grupos:
                if dept not in self.mbrp_group_dict:
                    self.mbrp_group_dict[dept] = {}
                if desc and cod:
                    self.mbrp_group_dict[dept][desc] = cod
        
            # Subgrupos por departamento y grupo
            subs = self.db_manager.fetch_data(
                "SELECT C_IN_DEPARTAMENTO, C_IN_GRUPO, C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS"
            )
            self.mbrp_sub_dict = {}
            for dept, grp, cod, desc in subs:
                # Usar string como key en lugar de tupla para compatibilidad con JSON
                key = f"{dept}|{grp}"
                if key not in self.mbrp_sub_dict:
                    self.mbrp_sub_dict[key] = {}
                if desc and cod:
                    self.mbrp_sub_dict[key][desc] = cod
            
            self.log("Jerarquía MBRP cargada correctamente", "SUCCESS")
            
            # Guardar en cache para próximas cargas
            try:
                cache_data = {
                    'mbrp': {
                        'departments': self.mbrp_dept_dict,
                        'groups': self.mbrp_group_dict,
                        'subgroups': self.mbrp_sub_dict
                    }
                }
                
                # Si ya existe cache con TRA, combinar
                existing_cache = self._check_jerarquia_cache()
                if existing_cache:
                    cache_data.update(existing_cache)
                
                self._save_jerarquia_cache(cache_data)
            except Exception as e:
                self.log(f"Error guardando cache MBRP: {e}", "DEBUG")
        
        except Exception as e:
            self.log(f"Error cargando jerarquía MBRP: {str(e)}", "ERROR")
            self.mbrp_dept_dict = {}
            self.mbrp_group_dict = {}
            self.mbrp_sub_dict = {}

    
    def cargar_tra_base(self):
        """Carga datos TRA de forma instantánea - delegando trabajo pesado al hilo paralelo"""
        try:
            # Validación rápida del rango de fechas
            fecha_inicio = self.fecha_inicio_entry.get_date()
            fecha_fin = self.fecha_fin_entry.get_date()
            
            dias_diferencia = (fecha_fin - fecha_inicio).days
            # Se ha eliminado la restricción de 1 año; ahora se permite cualquier rango 
            # pero con previo aviso visual en la UI.
            if dias_diferencia < 0:
                messagebox.showerror("Error", "La fecha de fin no puede ser anterior a la de inicio.")
                return
            
            sede = self.sede_var.get().split(" - ")[0]
            # Mapear ICH a consulta global
            if sede == '00':
                sede = '%'
            
            # Si ya hay una carga en curso, evitar lanzar otra simultáneamente
            try:
                if getattr(self, 'tra_loader_thread', None) is not None and self.tra_loader_thread.is_alive():
                    self.log("TRA: Ya hay una carga en curso; ignorando clic adicional", "WARNING")
                    try:
                        self.api_status.config(text="TRA: Carga en curso...", foreground="#004C97")
                    except Exception:
                        pass
                    return
            except Exception:
                pass

            # Limpiar datos previos y preparar parámetros
            self.cached_ventas_tra = []
            self.tra_fecha_inicio = fecha_inicio
            self.tra_fecha_fin = fecha_fin
            self.tra_sede_codigo = sede
            
            # Capturar estado del reporte masivo (con fallback seguro)
            self.tra_include_zero_sales = False
            if hasattr(self, 'tra_masivo_var'):
                self.tra_include_zero_sales = self.tra_masivo_var.get()
            
            # Mostrar estado de carga inmediatamente
            try:
                self.api_status.config(text="TRA: Iniciando carga...", foreground="#004C97")
                self.global_progress.pack(side=tk.RIGHT, padx=10)
                self.global_progress.config(mode="indeterminate")
                self.global_progress.start(5)
            except Exception:
                pass
            
            # Log inicial
            self.log(f"🚀 TRA: Iniciando carga desde {fecha_inicio} hasta {fecha_fin} para sede {sede}", "INFO")
            
            # Mostrar mensaje temporal en la tabla
            if hasattr(self, 'tra_tree'):
                self.tra_tree.delete(*self.tra_tree.get_children())
                # Insertar mensaje de carga
                self.tra_tree.insert("", tk.END, values=(
                    "...", "Cargando datos TRA...", "...", "...", "...", "...", "...", "..."
                ), tags=("loading",))
            
            # Iniciar carga paralela inmediatamente (sin carga inicial bloqueante)
            from threading import Thread
            self.tra_loader_thread = Thread(target=self._background_load_ventas_tra, daemon=True, name="tra_loader")
            self.tra_loader_thread.start()
            self.log("✅ TRA: Carga paralela iniciada - UI lista para uso", "SUCCESS")
            
            # Ocultar progreso después de un momento
            self.root.after(2000, self._hide_tra_progress)
            
        except Exception as e:
            error_msg = f"Error iniciando carga TRA: {str(e)}"
            self.log(error_msg, "ERROR")
            try:
                self.api_status.config(text="TRA: Error", foreground="red")
                self.global_progress.stop()
                self.global_progress.pack_forget()
            except Exception:
                pass
    
    def _hide_tra_progress(self):
        """Oculta el progreso TRA después de iniciada la carga"""
        try:
            self.global_progress.stop()
            self.global_progress.pack_forget()
        except Exception:
            pass
    
    def on_tra_row_selected(self, event=None):
        """Log detallado al seleccionar una fila del TRA Treeview (solo en modo debug)."""
        try:
            if not getattr(self, 'tra_debug', False):
                return
            if not hasattr(self, 'tra_tree'):
                return
            sel = self.tra_tree.selection()
            if not sel:
                return
            item = self.tra_tree.item(sel[0])
            vals = item.get('values') or []
            if len(vals) < 7:
                return
            codigo = str(vals[0])
            desc = str(vals[1])
            rotacion = str(vals[2])
            try:
                neto = float(vals[3])
            except Exception:
                neto = 0.0
            try:
                stock_actual = int(vals[5])
            except Exception:
                stock_actual = 0
            try:
                stock_ideal = int(vals[6])
            except Exception:
                stock_ideal = 0
            
            estado_stock = str(vals[8]) if len(vals) > 8 else "N/A"

            # Intervalo de fechas desde el UI
            try:
                fi = self.fecha_inicio_entry.get_date()
                ff = self.fecha_fin_entry.get_date()
            except Exception:
                from datetime import datetime
                fi = ff = datetime.now().date()

            msg = (
                f"Selección TRA -> Código: {codigo} | Desc: {desc} | Rotación: {rotacion}\n"
                f"Stock: {stock_actual} | Ideal: {stock_ideal} | Estado: {estado_stock} | Ventas: {neto}"
            )
            self.tra_debug_log(msg)
        except Exception as e:
            # No queremos romper selección por logs
            self.tra_debug_log(f"Error en on_tra_row_selected: {e}")
    
    def _on_tra_mouse_motion(self, event):
        """Maneja el movimiento del mouse para mostrar tooltips TRA"""
        try:
            if not hasattr(self, 'tra_tooltip') or not hasattr(self, 'tra_tree'):
                self.tra_debug_log("Tooltip o tree no disponible")
                return
            
            # Obtener item bajo el cursor
            item = self.tra_tree.identify_row(event.y)
            if not item:
                self.tra_tooltip.hide_tooltip()
                return
                
            # Obtener datos del item
            values = self.tra_tree.item(item, 'values')
            if not values:
                self.tra_debug_log("No hay valores en el item")
                return
                
            self.tra_debug_log(f"Valores del item: {values} (longitud: {len(values)})")
                
            # Ajustar índices según la nueva estructura de columnas
            # Columnas: Código, Descripción, Rotación, Neto, Representación %, Stock Actual, Stock Ideal, Días restantes
            if len(values) < 8:
                self.tra_debug_log(f"Item con pocos valores: {len(values)}")
                return
                
            codigo = str(values[0])
            descripcion = str(values[1])
            rotacion = str(values[2])
            
            try:
                neto = float(values[3])
            except (ValueError, IndexError):
                neto = 0.0
                
            # Saltar la columna de Representación % (index 4)
            try:
                stock_actual = int(values[5])  # Stock Actual
            except (ValueError, IndexError):
                stock_actual = 0
                
            try:
                stock_ideal = int(values[6])  # Stock Ideal
            except (ValueError, IndexError):
                stock_ideal = 0
            
            self.tra_debug_log(f"Mostrando tooltip para: {codigo} - {descripcion[:20]}...")
            
            # Mostrar tooltip con información completa
            self.tra_tooltip.show_tooltip(
                event, codigo, descripcion, neto, rotacion, stock_actual, stock_ideal
            )
            
        except Exception as e:
            self.tra_debug_log(f"Error en tooltip motion: {str(e)}")

    def calcular_stock_ideal_producto(self, neto_ventas):
        """Calcula el stock ideal para un producto basado en sus ventas netas"""
        try:
            fecha_inicio = self.fecha_inicio_entry.get_date()
            fecha_fin = self.fecha_fin_entry.get_date()
            dias_periodo = (fecha_fin - fecha_inicio).days or 1
            promedio_diario = neto_ventas / dias_periodo if dias_periodo else 0
            ciclo_reposicion_dias = 30
            factor_seguridad = 1.25
            factor_crecimiento = 1.15
            stock_ideal = promedio_diario * ciclo_reposicion_dias * factor_seguridad * factor_crecimiento
            return max(0, int(round(stock_ideal, 0)))
        except Exception as e:
            self.log(f"Error calculando stock ideal: {str(e)}", "ERROR")
            return 0
    
    def recalcular_stock_ideal_tra(self):
        """Recalcula el stock ideal para todos los productos TRA mostrados"""
        try:
            # Verificar tanto tra_ventas_datos (método antiguo) como cached_ventas_tra (nuevo método)
            if (not hasattr(self, 'tra_ventas_datos') or not self.tra_ventas_datos) and \
               (not hasattr(self, 'cached_ventas_tra') or not self.cached_ventas_tra):
                messagebox.showwarning("Sin datos", "Primero debe cargar los datos TRA")
                return
            
            self.log("Recalculando stock ideal para productos TRA...", "INFO")
            
            # Usar los datos cacheados y reaplicar los filtros para refrescar la vista
            if hasattr(self, 'cached_ventas_tra') and self.cached_ventas_tra:
                self.aplicar_filtro_tra()  # Esto recalculará y mostrará todo
            elif hasattr(self, 'tra_ventas_datos_filtrados') and self.tra_ventas_datos_filtrados:
                self.mostrar_pagina_tra()  # Método legacy
            else:
                # Si no hay datos filtrados, usar el método legacy
                datos_a_usar = getattr(self, 'tra_ventas_datos', [])
                if datos_a_usar:
                    self.tra_tree.delete(*self.tra_tree.get_children())
                    for fila in datos_a_usar[:self.tra_page_size]:  # Mostrar solo primera página
                        try:
                            if len(fila) >= 7:
                                codigo, desc, _, _, _, neto, rotacion = fila
                            elif len(fila) >= 6:
                                codigo, desc, _, _, _, neto = fila
                                rotacion = "SIN CLASIFICAR"
                            else:
                                continue
                                
                            # Calcular stock ideal
                            stock_ideal = self.calcular_stock_ideal_producto(neto)
                            
                            # Determinar tag de color según rotación
                            tag_rotacion = rotacion.lower() if rotacion else "sin_clasificar"
                            
                            self.tra_tree.insert(
                                "", tk.END,
                                values=(codigo, desc, rotacion, round(neto, 2), stock_ideal),
                                tags=(tag_rotacion,)
                            )
                        except Exception as e:
                            self.log(f"Error procesando fila en recalcular: {fila} -> {e}", "ERROR")
                            continue
            
            self.log("Stock ideal recalculado correctamente", "SUCCESS")
            
        except Exception as e:
            self.log(f"Error recalculando stock ideal: {str(e)}", "ERROR")
            messagebox.showerror("Error", f"Error al recalcular stock ideal: {str(e)}")
    
    def calcular_dias_restantes(self, stock_actual: int, neto_ventas: float, fecha_inicio, fecha_fin) -> int:
        """Calcula los días restantes antes de romper stock según promedio diario de ventas"""
        try:
            dias_periodo = (fecha_fin - fecha_inicio).days or 1
            promedio_diario = neto_ventas / dias_periodo if dias_periodo else 0
            if promedio_diario <= 0:
                return 0  # Sin consumo o regresos netos, no decrementa stock
            return max(0, int(stock_actual // promedio_diario))
        except Exception as e:
            self.log(f"Error calculando días restantes: {str(e)}", "ERROR")
            return 0
    
    def obtener_stock_actual_bulk(self, codigos: list, deposito: str) -> dict:
        """Obtiene stock actual por código filtrando por los almacenes tratables de la sede seleccionada."""
        try:
            if not codigos:
                return {}
            
            # Obtener los almacenes vendibles (tratables) para esta sede/contexto
            # Esto implementa el "nuevo motor" de segregación
            tratables = self.config_manager.get_tratables_by_sede(deposito)
            if not tratables:
                # Si no hay configuración, fallback al comportamiento anterior pero logueando advertencia
                self.log(f"Advertencia: No se encontraron almacenes tratables para {deposito}", "WARNING")
                return {str(c): 0 for c in codigos}

            # Evitar IN () muy grande dividiendo en chunks si es necesario
            MAX_IN = 900  # límite seguro para SQL Server
            resultado = {}
            
            # Preparar placeholders para los depósitos
            dep_placeholders = ",".join(["?"] * len(tratables))
            
            for i in range(0, len(codigos), MAX_IN):
                chunk = codigos[i:i+MAX_IN]
                art_placeholders = ",".join(["?"] * len(chunk))
                
                sql = (
                    f"SELECT c_codarticulo, SUM(n_cantidad) "
                    f"FROM MA_DEPOPROD WITH (NOLOCK) "
                    f"WHERE c_coddeposito IN ({dep_placeholders}) "
                    f"  AND c_codarticulo IN ({art_placeholders}) "
                    f"GROUP BY c_codarticulo"
                )
                
                params = list(tratables) + list(chunk)
                rows = self.db_manager.fetch_data(sql, params)
                
                for cod, sum_qty in (rows or []):
                    try:
                        resultado[str(cod)] = int(sum_qty or 0)
                    except Exception:
                        resultado[str(cod)] = 0
                
                # asegurar claves para todos los códigos del chunk
                for cod in chunk:
                    resultado.setdefault(str(cod), 0)
                    
            return resultado
        except Exception as e:
            self.log(f"Error obteniendo stock actual bulk: {str(e)}", "ERROR")
            return {}
    
    def calcular_stock_ideal(self, fecha_inicio, fecha_fin):
        """Método legacy para compatibilidad con DataFrames"""
        if not hasattr(self, 'tra_ventas_df') or self.tra_ventas_df.empty:
            return

        df = self.tra_ventas_df.copy()
    
        dias_periodo = (fecha_fin - fecha_inicio).days or 1  # evitar división por cero
        promedio_diario = df['neto'] / dias_periodo
    
        # Parámetros para el cálculo
        ciclo_reposicion_dias = 30
        stock_ideal = promedio_diario * ciclo_reposicion_dias * 1.25 * 1.15
    
        df['stock_ideal'] = stock_ideal.round(0).astype(int)

        self.tra_ventas_df = df






    def navigate_to_module(self, modulo_ruta: str):
        """
        Navega a la pestaña correspondiente al módulo indicado.
        Llamado por NotificationBell al hacer clic en "Tratar".

        modulo_ruta puede ser:
          'stock', 'tra', 'mbrp', 'clientes', 'admin', 'mensajes',
          'estadisticas', 'calendario', 'registros'
        """
        try:
            ruta = (modulo_ruta or '').lower().strip()
            tab_map = {
                'stock':        'stock_tab',
                'tra':          'tra_tab',
                'mbrp':         'mbrp_tab',
                'clientes':     'clientes_tab',
                'admin':        'admin_tab',
                'mensajes':     'messaging_tab',
                'estadisticas': 'stats_tab',
                'calendario':   'calendar_tab',
                'registros':    'records_tab',
                'inicio':       'dashboard_tab',
            }
            attr = tab_map.get(ruta)
            if attr and hasattr(self, attr):
                tab = getattr(self, attr)
                if tab and tab.winfo_exists():
                    self.main_notebook.select(tab)
                    return
            self.log(f"[navigate_to_module] Módulo desconocido o no disponible: '{modulo_ruta}'", "WARNING")
        except Exception as e:
            self.log(f"[navigate_to_module] Error navegando a '{modulo_ruta}': {e}", "ERROR")

    def show_records_view(self):
        self.main_notebook.select(self.records_tab)

    def show_admin_view(self):
        """Muestra la pestaña de administración."""
        if hasattr(self, 'admin_tab'):
            self.main_notebook.select(self.admin_tab)

    def show_messaging_view(self):
        self.main_notebook.select(self.messaging_tab)







    def log(self, message: str, level: str = 'INFO'):
        """Registrar mensaje con prefijo estandarizado para identificar origen en consola.
        Si el panel de logs aún no existe, hace un print de respaldo con prefijo [PAL][APP].
        """
        levels = ['DEBUG', 'INFO', 'WARNING', 'ERROR', 'SUCCESS']
        if level not in levels:
            level = 'INFO'

        # Asegurar contadores por defecto
        if not hasattr(self, 'log_counter'):
            self.log_counter = 0
        if not hasattr(self, 'max_logs'):
            self.max_logs = 200

        # Construir prefijo con ubicación del caller y thread
        try:
            caller = inspect.stack()[1]
            module = inspect.getmodule(caller.frame)
            mod_name = getattr(module, '__name__', 'unknown')
            func_name = caller.function
            line_no = caller.lineno
        except Exception:
            mod_name, func_name, line_no = 'unknown', 'unknown', 0
        thread_name = threading.current_thread().name
        timestamp = time.strftime("%H:%M:%S")
        prefix = f"[PAL][APP][{level}][{thread_name}][{mod_name}.{func_name}:{line_no}]"
        console_entry = f"{prefix} {message}\n"

        # SIEMPRE escribir en consola de debug flotante
        if hasattr(self, 'debug_console'):
            try:
                self.debug_console.write(console_entry.strip(), level)
            except Exception:
                pass

        # Rotar logs si excede el máximo
        try:
            if self.log_counter >= self.max_logs:
                self.limpiar_logs()
        except Exception:
            pass

        # Insertar en panel de logs si existe; usar solo mensaje como contenido y nivel como tag
        if hasattr(self, 'logs_text') and self.logs_text:
            try:
                self.logs_text.configure(state='normal')
                # Mostrar también el prefijo dentro del panel para consistencia
                self.logs_text.insert(tk.END, console_entry, level)
                self.logs_text.see(tk.END)  # Auto-scroll
                self.logs_text.configure(state='disabled')
                self.log_counter += 1
                return
            except Exception:
                pass

        # Respaldo a consola si no hay UI de logs disponible
        try:
            print(console_entry.strip())
        except Exception:
            pass

    def limpiar_logs(self):
        """Limpia el panel de logs y reinicia el contador de manera segura."""
        # Reiniciar contador siempre
        self.log_counter = 0
        # Limpiar el widget si existe
        if hasattr(self, 'logs_text') and self.logs_text:
            try:
                self.logs_text.configure(state='normal')
                self.logs_text.delete('1.0', tk.END)
                self.logs_text.configure(state='disabled')
            except Exception:
                # No bloquear si el widget no está listo
                pass

    
    def load_connection_settings(self):
        """Cargar configuración de conexión desde archivo, desencriptando valores."""
        if not os.path.exists(CONFIG_FILE):
            return None

        config = configparser.ConfigParser()
        config.read(CONFIG_FILE)

        if 'Database' not in config:
            return None

        try:
        # Obtener los valores encriptados
            server_enc = config['Database'].get('server', '')
            database_enc = config['Database'].get('database', '')
            user_enc = config['Database'].get('user', '')

            # Desencriptar solo si existen datos
            server = self.cred_manager.decrypt(server_enc) if server_enc else ''
            database = self.cred_manager.decrypt(database_enc) if database_enc else ''
            user = self.cred_manager.decrypt(user_enc) if user_enc else ''

            return {
                'server': server,
                'database': database,
                'user': user
            }
        except Exception as e:
            # Mostrar error usando ErrorCode en la interfaz
            self.notification_manager.show_error("Error", f"No se pudo cargar la configuración: {str(e)}")
            return None
    

    def _ensure_admin_user(self):
        """Crea el usuario 'admin' si no existe con contraseña predeterminada '123' y habilita módulos."""
        try:
            rows = self.db_manager.fetch_data("SELECT id FROM pal_usuarios WHERE username = ?", ("admin",))
            if rows:
                return  # Ya existe
            # Crear con contraseña por defecto '123'
            salt = bcrypt.gensalt(rounds=12)
            pwd_hash = bcrypt.hashpw(b"123", salt).decode('utf-8')
            self.db_manager.execute_query(
                "INSERT INTO pal_usuarios (username, password_hash, nombre_completo, email, activo) VALUES (?, ?, ?, ?, 1)",
                ("admin", pwd_hash, "Administrador del Sistema", None)
            )
            user_id = int(self.db_manager.fetch_data("SELECT id FROM pal_usuarios WHERE username = ?", ("admin",))[0][0])
            for modulo_db in DB_MODULE_TO_FLAG.keys():
                self.db_manager.execute_query(
                    """
                    IF NOT EXISTS (SELECT 1 FROM pal_usuarios_modulos WHERE usuario_id = ? AND modulo = ?)
                        INSERT INTO pal_usuarios_modulos (usuario_id, modulo, habilitado) VALUES (?, ?, 1)
                    ELSE
                        UPDATE pal_usuarios_modulos SET habilitado = 1 WHERE usuario_id = ? AND modulo = ?
                    """,
                    (user_id, modulo_db, user_id, modulo_db, user_id, modulo_db)
                )
            self.log("Usuario 'admin' creado con contraseña predeterminada y módulos habilitados", "SUCCESS")
        except Exception as e:
            self.log(f"No se pudo crear usuario admin: {e}", "ERROR")

    def _ensure_admin_password_interactive(self):
        """Verifica el hash del usuario admin y permite definir una contraseña si es placeholder o inválido."""
        try:
            rows = self.db_manager.fetch_data("SELECT id, password_hash FROM pal_usuarios WHERE username = ?", ("admin",))
            if not rows:
                return
            user_id, pwd_hash = rows[0]
            pwd_hash = str(pwd_hash) if pwd_hash is not None else ""
            needs_set = (not pwd_hash) or ("PLACEHOLDER" in pwd_hash) or (not pwd_hash.startswith("$2")) or (len(pwd_hash) < 55)
            if not needs_set:
                return
            # Pedir nueva contraseña
            while True:
                p1 = simpledialog.askstring("Configurar contraseña", "Ingrese nueva contraseña para admin:", show='*', parent=self.root)
                if p1 is None:
                    # Cancelado
                    break
                p2 = simpledialog.askstring("Confirmar contraseña", "Confirme la contraseña:", show='*', parent=self.root)
                if p2 is None:
                    break
                if p1 != p2:
                    messagebox.showerror("Error", "Las contraseñas no coinciden")
                    continue
                # Generar hash y guardar
                salt = bcrypt.gensalt(rounds=12)
                new_hash = bcrypt.hashpw(p1.encode("utf-8"), salt).decode("utf-8")
                self.db_manager.execute_query("UPDATE pal_usuarios SET password_hash = ? WHERE id = ?", (new_hash, int(user_id)))
                self.log("Contraseña de admin actualizada", "SUCCESS")
                messagebox.showinfo("Listo", "Se actualizó la contraseña de admin")
                break
        except Exception as e:
            self.log(f"No se pudo configurar contraseña de admin: {e}", "WARNING")

    def attempt_auto_connect(self):
        """Intentar conexión automática y habilitar login integrado en el splash."""
        try:
            settings = self.load_connection_settings()
            if not settings:
                self.root.after(0, self.show_settings)
                return

            server = settings.get('server')
            database = settings.get('database')
            user = settings.get('user')
            password = self.cred_manager.get_temp_password()
            api_token = self.cred_manager.get_whatsapp_token()

            if server and database:
                total_start = time.perf_counter()
                # Reportar progreso: 70% - conectando a BD
                try:
                    self.splash.set_progress(0.70)
                except:
                    pass
                
                if self.db_manager.connect(server, database, user, password):
                    # Reportar progreso: 85% - BD conectada
                    try:
                        self.splash.set_progress(0.85)
                    except:
                        pass
                    
                    self.update_status('connected', server=server, api_token=api_token)
                    self.log("Conexión a BD exitosa", "SUCCESS")
                    # Inicializar servicios de seguridad y auth
                    try:
                        from pal.core.audit_db import AuditDB
                        self.audit_db = AuditDB(self.db_manager)
                    except Exception:
                        self.audit_db = None
                    self.auth = AuthManager(self.db_manager)
                    # Crear admin si no existe (password predeterminada '123')
                    self._ensure_admin_user()
                    
                    # Reportar progreso: 95% - habilitando login
                    try:
                        self.splash.set_progress(0.95)
                    except:
                        pass
                    
                    # AHORA SÍ habilitar login (después de que BD esté lista)
                    try:
                        if hasattr(self, 'splash') and self.splash:
                            self.splash.enable_login(self._splash_login_submit)
                            self.splash.login_status.config(
                                text="Ingrese sus credenciales", 
                                foreground="#004C97"
                            )
                    except Exception:
                        pass

                    # (Opcional) Verificar/crear esquema de seguridad luego del login
                    # Se hará tras login exitoso para evitar prompts durante el splash

                    # Continuar con carga del resto de módulos en paralelo
                    # Carga inicial de módulos se hará tras el login exitoso

                    total_elapsed = time.perf_counter() - total_start
                    self.log(f"🏁 App inicializada (parcial) en {total_elapsed:.2f}s", "INFO")
                return

            self.root.after(0, self.show_settings)
        except Exception:
            self.audit_log.log_event(
                "AUTO_CONNECT_FAILED",
                os.getlogin(),
                "FAILED",
                ErrorCode.DB_CONNECTION_FAILED
            )
            self.root.after(0, self.show_settings)





    def create_status_panel(self):
        status_frame = ttk.Frame(self.root, style="Status.TFrame")
        status_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Indicadores de estado
        self.db_status = ttk.Label(status_frame, text="BD: Desconectado", foreground="red")
        self.db_status.pack(side=tk.LEFT, padx=10)
        
        self.api_status = ttk.Label(status_frame, text="API: Inactiva", foreground="orange")
        self.api_status.pack(side=tk.LEFT)
        
        # ── Campana de Notificaciones (Centro de Notificaciones) ──────
        # Debe estar ANTES de la alerta para aparecer a la izquierda
        self.notification_bell = None
        try:
            navigate_fn = getattr(self, "navigate_to_module", None)
            usuario = None
            if self.current_user:
                usuario = self.current_user.get("username")
            bell = NotificationBell(
                parent=status_frame,
                notification_manager=self.notification_manager,
                navigate_fn=navigate_fn,
                usuario_actual=usuario,
            )
            bell.pack(side=tk.LEFT, padx=8)
            self.notification_bell = bell
        except Exception as _bell_err:
            self.log(f"[create_status_panel] Error creando NotificationBell: {_bell_err}", "ERROR")
        
        # Anuncio de Alerta (Parpadeante) - Display de alertas basada en notificaciones
        self.alerts_display_label = tk.Label(
            status_frame,
            text="",
            font=("Segoe UI", 9, "bold"),
            foreground="#D32F2F",
            cursor="hand2",
            padx=10
        )
        # Oculto por defecto
        self.alerts_display_label.pack_forget()
        self.alerts_display_label.bind("<Button-1>", lambda e: self._handle_alert_click())

        # Menú de usuario (Cerrar sesión, Configuración solo para admin)
        self.user_menu = ttk.Menubutton(status_frame, text="Usuario")
        user_menu_popup = tk.Menu(self.user_menu, tearoff=0)
        user_menu_popup.add_command(label="Cerrar sesión", command=self.logout)
        
        # Agregar opción de configuración solo si el usuario actual es admin
        is_admin = self.current_user and self.current_user.get('username', '').lower() == 'admin'
        if is_admin:
            user_menu_popup.add_separator()
            user_menu_popup.add_command(label="⚙️ Configuración Avanzada", command=self.show_settings)
        
        self.user_menu['menu'] = user_menu_popup
        self.user_menu.pack(side=tk.RIGHT, padx=10)

        # Barra de progreso global
        self.global_progress = ttk.Progressbar(
        status_frame, 
        orient="horizontal",
        length=200,
        mode="determinate"
        )
        self.global_progress.pack(side=tk.RIGHT, padx=10)
        self.global_progress.pack_forget()  # Ocultar inicialmente


    def setup_tooltips(self):
        if hasattr(self, 'num_cliente'):
            self.help_tooltips.add_tooltip(self.num_cliente, "Número de cliente (1-11 dígitos)")
        if hasattr(self, 'cod_producto'):
            self.help_tooltips.add_tooltip(self.cod_producto, "Código de producto (buscar en base de datos)")
        if hasattr(self, 'user_menu'):
            self.help_tooltips.add_tooltip(self.user_menu, "Menú de usuario con opciones de configuración")

    def show_settings(self):
    # Si ya existe una ventana, la traemos al frente
        if hasattr(self, 'settings_window') and self.settings_window and self.settings_window.winfo_exists():
            self.settings_window.lift()
            return

        self.settings_window = tk.Toplevel(self.root)
        self.settings_window.title("Configuración Avanzada")
        self.settings_window.geometry("600x400")
    
        # Configurar protocolo de cierre
        self.settings_window.protocol("WM_DELETE_WINDOW", self.on_settings_close)
    
        notebook = ttk.Notebook(self.settings_window)
        notebook.pack(fill=tk.BOTH, expand=True)

        # Pestaña de Conexión
        conn_frame = ttk.Frame(notebook)
        notebook.add(conn_frame, text="Conexión BD")
        self.create_connection_tab(conn_frame)
        
        # Pestaña de API
        api_frame = ttk.Frame(notebook)
        notebook.add(api_frame, text="API WhatsApp")
        self.create_api_tab(api_frame)
    
        # Pestaña de Módulos: gestionar usuarios y privilegios
        modules_frame = ttk.Frame(notebook)
        notebook.add(modules_frame, text="Gestión de Usuarios")
        self._create_user_management_tab(modules_frame)

        # Pestaña de Actualizaciones
        update_frame = ttk.Frame(notebook)
        notebook.add(update_frame, text="Actualizaciones")
        self.create_update_tab(update_frame)
        
        # Pestaña de Depuración
        debug_frame = ttk.Frame(notebook)
        notebook.add(debug_frame, text="Depuración")

        self.debug_vars = {}
        for idx, (key, label) in enumerate([
            ("tra", "Habilitar logs de depuración de T.R.A"),
            ("stock", "Habilitar logs de depuración de Stock"),
            ("mbrp", "Habilitar logs de depuración de MBRP"),
            ("db", "Habilitar logs de depuración de Base de Datos"),
        ]):
            var = tk.BooleanVar(value=self.debug_flags.get(key, False) if hasattr(self, 'debug_flags') else False)
            cb = ttk.Checkbutton(debug_frame, text=label, variable=var)
            cb.grid(row=idx, column=0, sticky="w", padx=10, pady=5)
            self.debug_vars[key] = var

        ttk.Button(
            debug_frame,
            text="Guardar Depuración",
            command=self._save_debug_config
        ).grid(row=4, column=0, sticky="e", pady=10, padx=10)
        
        # Pestaña de Caché y Limpieza
        cache_frame = ttk.Frame(notebook)
        notebook.add(cache_frame, text="Caché y Limpieza")
        self._create_cache_cleanup_tab(cache_frame)

    def _create_cache_cleanup_tab(self, parent):
        """Pestaña para gestionar y limpiar cachés."""
        frame = ttk.Frame(parent, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Gestionar Cachés del Sistema", font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=10)
        
        # Información de cachés
        info_frame = ttk.LabelFrame(frame, text="Cachés Disponibles", padding=10)
        info_frame.pack(fill=tk.X, padx=5, pady=10)
        
        import os
        caches = [
            ("productos_jerarquia_cache.json", "Jerarquía de Productos"),
            ("jerarquia_cache.json", "Jerarquía TRA/MBRP"),
            ("stock_depositos_preference.json", "Preferencias de Depósitos"),
        ]
        
        cache_info = ttk.Frame(info_frame)
        cache_info.pack(fill=tk.BOTH, expand=True)
        
        for cache_file, desc in caches:
            cache_row = ttk.Frame(cache_info)
            cache_row.pack(fill=tk.X, pady=5)
            
            exists = os.path.exists(cache_file)
            status = "✔️" if exists else "❌"
            size = f"{os.path.getsize(cache_file) / 1024:.1f} KB" if exists else "No existe"
            
            ttk.Label(cache_row, text=f"{status} {desc}").pack(side=tk.LEFT, padx=5)
            ttk.Label(cache_row, text=size, foreground="gray").pack(side=tk.RIGHT, padx=5)
        
        # Botones de limpieza
        action_frame = ttk.LabelFrame(frame, text="Limpiar Caché", padding=10)
        action_frame.pack(fill=tk.X, padx=5, pady=10)
        
        btn1 = ttk.Button(
            action_frame,
            text="🗑️ Limpiar Jerarquía de Productos",
            command=lambda: self._clear_cache_file("productos_jerarquia_cache.json")
        )
        btn1.pack(fill=tk.X, pady=5)
        
        btn2 = ttk.Button(
            action_frame,
            text="🗑️ Limpiar Caché TRA/MBRP",
            command=lambda: self._clear_cache_file("jerarquia_cache.json")
        )
        btn2.pack(fill=tk.X, pady=5)
        
        btn3 = ttk.Button(
            action_frame,
            text="🗑️ Limpiar Preferencias de Depósitos",
            command=lambda: self._clear_cache_file("stock_depositos_preference.json")
        )
        btn3.pack(fill=tk.X, pady=5)
        
        # Botón para limpiar TODO
        btn_all = ttk.Button(
            action_frame,
            text="🗑️ LIMPIAR TODO",
            command=self._clear_all_caches
        )
        btn_all.pack(fill=tk.X, pady=10)
        
        # Nota informativa
        info_text = ttk.Frame(frame)
        info_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=10)
        
        ttk.Label(
            info_text,
            text="⚠️ Nota: Limpiar el caché forzará que el sistema recargue los datos desde la base de datos.\n"
                 "Esto puede tomar más tiempo la próxima vez que cargues los módulos.",
            foreground="orange",
            wraplength=500,
            justify=tk.LEFT
        ).pack(fill=tk.BOTH, expand=True)
    
    def _clear_cache_file(self, filename):
        """Limpia un archivo de caché específíco."""
        try:
            import os
            if os.path.exists(filename):
                os.remove(filename)
                self.log(f"✅ Caché '{filename}' eliminado", "SUCCESS")
                messagebox.showinfo("Listo", f"✅ Caché '{filename}' eliminado exitosamente")
            else:
                messagebox.showinfo("Info", f"ℹ️ El archivo '{filename}' no existe")
        except Exception as e:
            self.log(f"❌ Error eliminando caché: {e}", "ERROR")
            messagebox.showerror("Error", f"❌ No se pudo eliminar el caché:\n{e}")
    
    def _clear_all_caches(self):
        """Limpia todos los cachés del sistema."""
        if not messagebox.askyesno("Confirmar", 
                                   "🗑️ ¿Está seguro de que desea limpiar TODO el caché?\n\n"
                                   "Esto incluirá:\n"
                                   "• Jerarquía de Productos\n"
                                   "• Caché TRA/MBRP\n"
                                   "• Preferencias de Depósitos"):
            return
        
        import os
        caches = [
            "productos_jerarquia_cache.json",
            "jerarquia_cache.json",
            "stock_depositos_preference.json",
        ]
        
        count = 0
        for cache_file in caches:
            try:
                if os.path.exists(cache_file):
                    os.remove(cache_file)
                    count += 1
            except Exception as e:
                self.log(f"❌ Error eliminando {cache_file}: {e}", "ERROR")
        
        self.log(f"✅ {count} archivo(s) de caché eliminados", "SUCCESS")
        messagebox.showinfo("Listo", f"✅ Se eliminaron {count} archivo(s) de caché exitosamente")
    
    def _create_user_management_tab(self, parent):
        """Pestaña para gestionar usuarios y sus módulos (sólo admin)."""
        frame = ttk.Frame(parent, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Permiso requerido: ADMIN.usuarios (fallback: usuario 'admin')
        allowed = False
        try:
            if hasattr(self, 'permissions') and self.current_user:
                allowed = self.permissions.tiene_permiso(self.current_user['id'], 'ADMIN', 'usuarios')
            if self.current_user and self.current_user.get('username', '').lower() == 'admin':
                allowed = True
        except Exception:
            allowed = self.current_user and self.current_user.get('username', '').lower() == 'admin'
        
        if not allowed:
            ttk.Label(frame, text="No tienes permiso para gestionar usuarios.", foreground="orange").pack(pady=20)
            return
        
        ttk.Label(frame, text="Gestionar módulos de usuarios", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=10)
        
        # Selector de usuario
        sel_frame = ttk.Frame(frame)
        sel_frame.pack(fill=tk.X, padx=5, pady=10)
        ttk.Label(sel_frame, text="Seleccionar usuario:").pack(side=tk.LEFT, padx=5)
        
        try:
            rows = self.db_manager.fetch_data("SELECT id, username FROM pal_usuarios WHERE activo = 1 ORDER BY username")
            self.usuarios_list = [(int(r[0]), str(r[1])) for r in (rows or [])]
        except Exception:
            self.usuarios_list = []
        
        self.user_combo = ttk.Combobox(sel_frame, values=[u[1] for u in self.usuarios_list], state="readonly", width=25)
        self.user_combo.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        if self.usuarios_list:
            self.user_combo.current(0)
            self.user_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_modules_list())
        
        # Frame para módulos
        modules_frame = ttk.LabelFrame(frame, text="Módulos habilitados", padding=10)
        modules_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=10)
        
        self.user_mod_vars = {}
        for idx, modulo_db in enumerate(sorted(DB_MODULE_TO_FLAG.keys())):
            var = tk.BooleanVar(value=False)
            cb = ttk.Checkbutton(modules_frame, text=f"{modulo_db}", variable=var)
            cb.grid(row=idx, column=0, sticky="w", padx=10, pady=5)
            self.user_mod_vars[modulo_db] = var
        
        # Botones de acción
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=10)
        ttk.Button(btn_frame, text="Guardar Módulos", command=self._save_user_modules_admin).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Recargar", command=self._refresh_modules_list).pack(side=tk.LEFT, padx=5)
        
        # Cargar módulos del primer usuario
        if self.usuarios_list:
            self._refresh_modules_list()
    
    def _refresh_modules_list(self):
        """Recarga los módulos del usuario seleccionado desde la BD."""
        try:
            sel_username = self.user_combo.get()
            user_id = next((u[0] for u in self.usuarios_list if u[1] == sel_username), None)
            if not user_id:
                return
            
            # Obtener módulos habilitados del usuario
            rows = self.db_manager.fetch_data(
                "SELECT modulo FROM pal_usuarios_modulos WHERE usuario_id = ? AND habilitado = 1",
                (user_id,)
            )
            enabled = {str(r[0]) for r in (rows or [])}
            
            # Actualizar checkboxes
            for modulo_db, var in self.user_mod_vars.items():
                var.set(modulo_db in enabled)
        except Exception as e:
            self.log(f"Error recargando módulos: {e}", "WARNING")
    
    def _save_user_modules_admin(self):
        """Guarda los módulos del usuario seleccionado en la BD."""
        try:
            # Permiso requerido
            can_manage = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    can_manage = self.permissions.tiene_permiso(self.current_user['id'], 'ADMIN', 'usuarios')
                if self.current_user and self.current_user.get('username', '').lower() == 'admin':
                    can_manage = True
            except Exception:
                can_manage = False
            if not can_manage:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para modificar módulos de usuarios")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENIED', usuario_id=self.current_user['id'], modulo='ADMIN', detalle='modificar_modulos')
                except Exception:
                    pass
                return
            
            sel_username = self.user_combo.get()
            user_id = next((u[0] for u in self.usuarios_list if u[1] == sel_username), None)
            if not user_id:
                messagebox.showwarning("Error", "Selecciona un usuario")
                return
            
            # Guardar cada módulo
            for modulo_db, var in self.user_mod_vars.items():
                habilitado = 1 if var.get() else 0
                # Upsert
                self.db_manager.execute_query(
                    """
                    IF NOT EXISTS (SELECT 1 FROM pal_usuarios_modulos WHERE usuario_id = ? AND modulo = ?)
                        INSERT INTO pal_usuarios_modulos (usuario_id, modulo, habilitado) VALUES (?, ?, ?)
                    ELSE
                        UPDATE pal_usuarios_modulos SET habilitado = ? WHERE usuario_id = ? AND modulo = ?
                    """,
                    (user_id, modulo_db, user_id, modulo_db, habilitado, habilitado, user_id, modulo_db)
                )
            
            messagebox.showinfo("Listo", f"Módulos de '{sel_username}' guardados")
            self.log(f"Módulos de '{sel_username}' actualizados", "SUCCESS")
            try:
                if hasattr(self, 'audit_db'):
                    self.audit_db.log_action(
                        accion='USER_MODULES_UPDATE', usuario_id=self.current_user['id'], modulo='ADMIN', detalle=f"user={sel_username}")
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
            self.log(f"Error guardando módulos: {e}", "ERROR")
    
    def _create_admin_users_tab(self, parent):
        """Pestaña de administración de usuarios con crear/editar/eliminar."""
        main_frame = ttk.Frame(parent, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Secciones: Left (listado) y Right (formulario)
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # LEFT: Listado de usuarios
        ttk.Label(left_frame, text="Usuarios del Sistema", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=5)
        
        # Treeview con usuarios
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.admin_users_tree = ttk.Treeview(
            tree_frame,
            columns=("id", "username", "nombre", "activo"),
            height=12,
            yscrollcommand=scrollbar.set
        )
        scrollbar.config(command=self.admin_users_tree.yview)
        
        self.admin_users_tree.column("#0", width=0)
        self.admin_users_tree.column("id", width=30)
        self.admin_users_tree.column("username", width=100)
        self.admin_users_tree.column("nombre", width=150)
        self.admin_users_tree.column("activo", width=60)
        
        self.admin_users_tree.heading("#0", text="")
        self.admin_users_tree.heading("id", text="ID")
        self.admin_users_tree.heading("username", text="Usuario")
        self.admin_users_tree.heading("nombre", text="Nombre Completo")
        self.admin_users_tree.heading("activo", text="Activo")
        
        self.admin_users_tree.pack(fill=tk.BOTH, expand=True)
        self.admin_users_tree.bind("<<TreeviewSelect>>", lambda e: self._load_user_details())
        
        # Botones de acción
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="Recargar", command=self._reload_users_list).pack(side=tk.LEFT, padx=5)
        
        # RIGHT: Formulario de usuario
        ttk.Label(right_frame, text="Detalles del Usuario", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=5)
        
        form_frame = ttk.LabelFrame(right_frame, text="Información", padding=10)
        form_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        ttk.Label(form_frame, text="Usuario:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.admin_user_username = ttk.Entry(form_frame, width=25)
        self.admin_user_username.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(form_frame, text="Nombre Completo:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.admin_user_nombre = ttk.Entry(form_frame, width=25)
        self.admin_user_nombre.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(form_frame, text="Email:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.admin_user_email = ttk.Entry(form_frame, width=25)
        self.admin_user_email.grid(row=2, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(form_frame, text="Contraseña:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.admin_user_password = ttk.Entry(form_frame, width=25, show="*")
        self.admin_user_password.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        
        ttk.Label(form_frame, text="(dejar en blanco para no cambiar)", foreground="gray").grid(row=4, column=1, sticky="w", padx=5)
        
        ttk.Label(form_frame, text="Activo:").grid(row=5, column=0, sticky="w", padx=5, pady=5)
        self.admin_user_activo = tk.BooleanVar(value=True)
        ttk.Checkbutton(form_frame, variable=self.admin_user_activo).grid(row=5, column=1, sticky="w", padx=5, pady=5)
        
        form_frame.columnconfigure(1, weight=1)
        
        # Módulos
        modules_frame = ttk.LabelFrame(right_frame, text="Módulos Habilitados", padding=10)
        modules_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.admin_user_mod_vars = {}
        for idx, modulo_db in enumerate(sorted(DB_MODULE_TO_FLAG.keys())):
            var = tk.BooleanVar(value=False)
            cb = ttk.Checkbutton(modules_frame, text=modulo_db, variable=var)
            cb.grid(row=idx, column=0, sticky="w", padx=5, pady=2)
            self.admin_user_mod_vars[modulo_db] = var
        
        # Botones de acción
        action_frame = ttk.Frame(right_frame)
        action_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(action_frame, text="Nuevo Usuario", command=self._new_admin_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Guardar", command=self._save_admin_user).pack(side=tk.LEFT, padx=5)
        ttk.Button(action_frame, text="Eliminar", command=self._delete_admin_user).pack(side=tk.LEFT, padx=5)
        
        # Cargar lista inicial
        self._reload_users_list()
    
    def _reload_users_list(self):
        """Recarga la lista de usuarios desde la BD."""
        try:
            self.admin_users_tree.delete(*self.admin_users_tree.get_children())
            rows = self.db_manager.fetch_data("SELECT id, username, nombre_completo, activo FROM pal_usuarios ORDER BY username")
            for r in (rows or []):
                user_id, username, nombre, activo = r
                activo_str = "✓" if activo else "❌"
                self.admin_users_tree.insert("", tk.END, values=(user_id, username, nombre, activo_str))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar usuarios: {e}")
    
    def _load_user_details(self):
        """Carga detalles del usuario seleccionado en el formulario."""
        try:
            sel = self.admin_users_tree.selection()
            if not sel:
                self._clear_user_form()
                return
            
            item = sel[0]
            values = self.admin_users_tree.item(item, "values")
            user_id = int(values[0])
            
            # Obtener datos del usuario
            rows = self.db_manager.fetch_data(
                "SELECT username, nombre_completo, email, activo FROM pal_usuarios WHERE id = ?",
                (user_id,)
            )
            if rows:
                username, nombre, email, activo = rows[0]
                self.admin_user_username.delete(0, tk.END)
                self.admin_user_username.insert(0, username)
                self.admin_user_nombre.delete(0, tk.END)
                self.admin_user_nombre.insert(0, nombre or "")
                self.admin_user_email.delete(0, tk.END)
                self.admin_user_email.insert(0, email or "")
                self.admin_user_password.delete(0, tk.END)  # No cargar password
                self.admin_user_activo.set(bool(activo))
                
                # Cargar módulos
                mod_rows = self.db_manager.fetch_data(
                    "SELECT modulo FROM pal_usuarios_modulos WHERE usuario_id = ? AND habilitado = 1",
                    (user_id,)
                )
                enabled_mods = {str(r[0]) for r in (mod_rows or [])}
                for modulo_db, var in self.admin_user_mod_vars.items():
                    var.set(modulo_db in enabled_mods)
                
                self.current_admin_user_id = user_id
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar detalles: {e}")
    
    def _clear_user_form(self):
        """Limpia el formulario de usuario."""
        self.admin_user_username.delete(0, tk.END)
        self.admin_user_nombre.delete(0, tk.END)
        self.admin_user_email.delete(0, tk.END)
        self.admin_user_password.delete(0, tk.END)
        self.admin_user_activo.set(True)
        for var in self.admin_user_mod_vars.values():
            var.set(False)
        self.current_admin_user_id = None
    
    def _new_admin_user(self):
        """Prepara formulario para crear nuevo usuario."""
        self._clear_user_form()
        self.admin_users_tree.selection_remove(self.admin_users_tree.selection())
        self.admin_user_username.focus()
    
    def _save_admin_user(self):
        """Guarda o crea usuario según lo indicado en el formulario."""
        try:
            # Permiso requerido
            can_manage = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    can_manage = self.permissions.tiene_permiso(self.current_user['id'], 'ADMIN', 'usuarios')
                if self.current_user and self.current_user.get('username', '').lower() == 'admin':
                    can_manage = True
            except Exception:
                can_manage = False
            if not can_manage:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para crear/editar usuarios")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENIED', usuario_id=self.current_user['id'], modulo='ADMIN', detalle='guardar_usuario')
                except Exception:
                    pass
                return
            
            # PROTECCIÓN: Solo el usuario 'admin' puede modificar su propia cuenta
            username = self.admin_user_username.get().strip()
            is_editing_admin = (hasattr(self, 'current_admin_user_id') and self.current_admin_user_id is not None and
                               username.lower() == 'admin')
            
            if is_editing_admin and self.current_user and self.current_user.get('username', '').lower() != 'admin':
                usuario_actual = self.current_user.get('username', 'DESCONOCIDO')
                
                messagebox.showerror(
                    "Acceso denegado",
                    "⚠️ PROTECCIÓN DE SEGURIDAD\n\nSolo el usuario 'admin' puede modificar su propia cuenta.\n\n" +
                    "Por seguridad, el acceso al usuario 'admin' está restringido."
                )
                
                self.log(
                    f"[SECURITY] Intento de modificar usuario admin desde: {usuario_actual}",
                    "ERROR"
                )
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='SECURITY_VIOLATION_ADMIN_MODIFY', usuario_id=self.current_user['id'], modulo='ADMIN',
                            detalle=f'Usuario {usuario_actual} intentó modificar cuenta admin', exitoso=False)
                except Exception:
                    pass
                return
            
            nombre = self.admin_user_nombre.get().strip()
            email = self.admin_user_email.get().strip() or None
            password = self.admin_user_password.get().strip()
            activo = 1 if self.admin_user_activo.get() else 0
            
            if not username or not nombre:
                messagebox.showwarning("Error", "Usuario y nombre son obligatorios")
                return
            
            is_new = not hasattr(self, 'current_admin_user_id') or self.current_admin_user_id is None
            user_id = None
            
            if is_new:
                # Crear nuevo usuario
                if not password:
                    messagebox.showwarning("Error", "La contraseña es obligatoria para nuevos usuarios")
                    return
                from pal.core.user_management import UserManager
                um = UserManager(self.db_manager)
                try:
                    user_id = um.crear_usuario(username, password, nombre, email)
                    messagebox.showinfo("Listo", f"Usuario '{username}' creado exitosamente")
                    try:
                        if hasattr(self, 'audit_db'):
                            self.audit_db.log_action(
                                accion='USER_CREATE', usuario_id=self.current_user['id'], modulo='ADMIN', detalle=f"username={username}")
                    except Exception:
                        pass
                except Exception as create_err:
                    messagebox.showerror("Error", f"No se pudo crear usuario: {str(create_err)}")
                    self.log(f"Error creando usuario: {create_err}", "ERROR")
                    return
            else:
                # Actualizar usuario existente
                user_id = self.current_admin_user_id
                self.db_manager.execute_query(
                    "UPDATE pal_usuarios SET nombre_completo = ?, email = ?, activo = ? WHERE id = ?",
                    (nombre, email, activo, user_id)
                )
                if password:
                    # Actualizar contraseña
                    salt = bcrypt.gensalt(rounds=12)
                    pwd_hash = bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')
                    self.db_manager.execute_query(
                        "UPDATE pal_usuarios SET password_hash = ? WHERE id = ?",
                        (pwd_hash, user_id)
                    )
                messagebox.showinfo("Listo", f"Usuario '{username}' actualizado exitosamente")
                try:
                    if hasattr(self, 'audit_db'):
                        self.audit_db.log_action(
                            accion='USER_UPDATE', usuario_id=self.current_user['id'], modulo='ADMIN', detalle=f"id={user_id}")
                except Exception:
                    pass
            
            # Guardar módulos (solo si user_id es válido)
            if user_id:
                for modulo_db, var in self.admin_user_mod_vars.items():
                    habilitado = 1 if var.get() else 0
                    self.db_manager.execute_query(
                        """
                        IF NOT EXISTS (SELECT 1 FROM pal_usuarios_modulos WHERE usuario_id = ? AND modulo = ?)
                            INSERT INTO pal_usuarios_modulos (usuario_id, modulo, habilitado) VALUES (?, ?, ?)
                        ELSE
                            UPDATE pal_usuarios_modulos SET habilitado = ? WHERE usuario_id = ? AND modulo = ?
                        """,
                        (user_id, modulo_db, user_id, modulo_db, habilitado, habilitado, user_id, modulo_db)
                    )
                
                self.log(f"Usuario '{username}' guardado", "SUCCESS")
                self._reload_users_list()
                self._clear_user_form()
            else:
                messagebox.showerror("Error", "No se asignó un ID válido al usuario")
                self.log(f"Error: user_id inválido para '{username}'", "ERROR")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar usuario: {e}")
            self.log(f"Error guardando usuario: {e}", "ERROR")
    
    def _delete_admin_user(self):
        """Elimina (desactiva) el usuario seleccionado."""
        try:
            # Permiso requerido
            can_manage = False
            try:
                if hasattr(self, 'permissions') and self.current_user:
                    can_manage = self.permissions.tiene_permiso(self.current_user['id'], 'ADMIN', 'usuarios')
                if self.current_user and self.current_user.get('username', '').lower() == 'admin':
                    can_manage = True
            except Exception:
                can_manage = False
            if not can_manage:
                messagebox.showwarning("Permiso denegado", "No tienes permiso para desactivar usuarios")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='PERMISSION_DENIED', usuario_id=self.current_user['id'], modulo='ADMIN', detalle='desactivar_usuario')
                except Exception:
                    pass
                return
            
            if not hasattr(self, 'current_admin_user_id') or self.current_admin_user_id is None:
                messagebox.showwarning("Error", "Selecciona un usuario para eliminar")
                return
            
            username = self.admin_user_username.get()
            
            # PROTECCIÓN: Impedir que se desactive el usuario 'admin'
            if username.lower() == 'admin':
                usuario_actual = self.current_user.get('username', 'DESCONOCIDO') if self.current_user else 'DESCONOCIDO'
                
                messagebox.showerror(
                    "Acceso denegado",
                    "⚠️ PROTECCIÓN DE SEGURIDAD\n\nNo se puede desactivar el usuario 'admin'.\n\n" +
                    "Este usuario es esencial para el sistema."
                )
                
                self.log(
                    f"[SECURITY] Intento de desactivar usuario admin desde: {usuario_actual}",
                    "ERROR"
                )
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='SECURITY_VIOLATION_ADMIN_DELETE', usuario_id=self.current_user['id'], modulo='ADMIN',
                            detalle=f'Usuario {usuario_actual} intentó desactivar cuenta admin', exitoso=False)
                except Exception:
                    pass
                return
            if messagebox.askyesno("Confirmar", f"¿Desactivar usuario '{username}'?"):
                self.db_manager.execute_query(
                    "UPDATE pal_usuarios SET activo = 0 WHERE id = ?",
                    (self.current_admin_user_id,)
                )
                messagebox.showinfo("Listo", f"Usuario '{username}' desactivado")
                self.log(f"Usuario '{username}' desactivado", "INFO")
                try:
                    if hasattr(self, 'audit_db'):
                        self.audit_db.log_action(
                            accion='USER_DEACTIVATE', usuario_id=self.current_user['id'], modulo='ADMIN', detalle=f"user={username}")
                except Exception:
                    pass
                self._reload_users_list()
                self._clear_user_form()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo desactivar usuario: {e}")
    
    # =========================
    # Roles y Permisos (Fase 5)
    # =========================
    def _create_roles_permissions_tab(self, parent):
        container = ttk.Frame(parent, padding=10)
        container.pack(fill=tk.BOTH, expand=True)

        # Split left (roles) / right (details + permisos + asignación a usuarios)
        left = ttk.Frame(container)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0,10))
        right = ttk.Frame(container)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # LEFT: Roles list
        ttk.Label(left, text="Roles", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        tree_frame = ttk.Frame(left)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.roles_tree = ttk.Treeview(
            tree_frame,
            columns=("id","nombre","sistema"),
            height=14,
            yscrollcommand=scrollbar.set
        )
        scrollbar.config(command=self.roles_tree.yview)
        self.roles_tree.column("#0", width=0)
        self.roles_tree.column("id", width=40)
        self.roles_tree.column("nombre", width=160)
        self.roles_tree.column("sistema", width=70)
        self.roles_tree.heading("#0", text="")
        self.roles_tree.heading("id", text="ID")
        self.roles_tree.heading("nombre", text="Nombre")
        self.roles_tree.heading("sistema", text="Sistema")
        self.roles_tree.pack(fill=tk.BOTH, expand=True)
        self.roles_tree.bind("<<TreeviewSelect>>", lambda e: self._on_role_select())

        btns = ttk.Frame(left)
        btns.pack(fill=tk.X, pady=5)
        ttk.Button(btns, text="Nuevo", command=self._new_role).pack(side=tk.LEFT, padx=2)
        ttk.Button(btns, text="Guardar", command=self._save_role).pack(side=tk.LEFT, padx=2)
        ttk.Button(btns, text="Eliminar", command=self._delete_role).pack(side=tk.LEFT, padx=2)
        ttk.Button(btns, text="Recargar", command=self._reload_roles_list).pack(side=tk.LEFT, padx=2)

        # RIGHT: Details and permissions
        details = ttk.LabelFrame(right, text="Detalle del Rol", padding=10)
        details.pack(fill=tk.X)
        ttk.Label(details, text="Nombre:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.role_name_entry = ttk.Entry(details, width=30)
        self.role_name_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        ttk.Label(details, text="Descripción:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.role_desc_entry = ttk.Entry(details, width=50)
        self.role_desc_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        details.columnconfigure(1, weight=1)

        # Permissions catalog (by module)
        self.perm_vars_by_id = {}
        self.perm_catalog_by_module = {}
        perms_rows = self.db_manager.fetch_data("SELECT id, codigo, modulo, descripcion FROM pal_permisos ORDER BY modulo, codigo") or []
        for pid, codigo, modulo, desc in perms_rows:
            m = str(modulo)
            self.perm_catalog_by_module.setdefault(m, []).append((int(pid), str(codigo), str(desc or "")))

        # Scrollable permissions container
        perms_container = ttk.LabelFrame(right, text="Permisos por módulo", padding=5)
        perms_container.pack(fill=tk.BOTH, expand=True, pady=10)

        canvas = tk.Canvas(perms_container, borderwidth=0, highlightthickness=0)
        vscroll = ttk.Scrollbar(perms_container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        perms_inner = ttk.Frame(canvas)
        inner_window = canvas.create_window((0, 0), window=perms_inner, anchor="nw")

        def _on_perms_inner_configure(event=None):
            try:
                canvas.configure(scrollregion=canvas.bbox("all"))
                canvas.itemconfigure(inner_window, width=canvas.winfo_width())
            except Exception:
                pass
        def _on_canvas_configure(event=None):
            try:
                canvas.itemconfigure(inner_window, width=canvas.winfo_width())
            except Exception:
                pass
        perms_inner.bind("<Configure>", _on_perms_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)

        # Optional: mouse wheel scrolling
        def _on_mousewheel(event):
            try:
                delta = -1 * (event.delta // 120)
                canvas.yview_scroll(delta, "units")
            except Exception:
                pass
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Build permission sections inside scrollable frame
        self.perm_section_frames = {}
        row_idx = 0
        for modulo, plist in sorted(self.perm_catalog_by_module.items()):
            sect = ttk.LabelFrame(perms_inner, text=modulo, padding=8)
            sect.grid(row=row_idx, column=0, sticky="ew", padx=5, pady=5)
            row_idx += 1
            # Checkboxes for this module
            col = 0
            for pid, codigo, desc in plist:
                var = tk.BooleanVar(value=False)
                cb = ttk.Checkbutton(sect, text=codigo.split('.')[-1], variable=var)
                cb.grid(row=0, column=col, sticky="w", padx=6, pady=4)
                self.perm_vars_by_id[pid] = var
                col += 1
        for i in range(row_idx):
            perms_inner.grid_rowconfigure(i, weight=0)
        perms_inner.grid_columnconfigure(0, weight=1)

        # Fixed action bar for saving permissions
        perm_actions = ttk.Frame(right)
        perm_actions.pack(fill=tk.X, pady=(0,5))
        ttk.Button(perm_actions, text="Guardar Permisos del Rol", command=self._save_role_permissions).pack(side=tk.RIGHT)

        # User-role assignment
        assign = ttk.LabelFrame(right, text="Asignación de Roles a Usuario", padding=10)
        assign.pack(fill=tk.BOTH, expand=False, pady=10)
        ttk.Label(assign, text="Usuario:").grid(row=0, column=0, sticky="w", padx=5)
        self.assign_user_combo = ttk.Combobox(assign, state="readonly", width=25)
        self.assign_user_combo.grid(row=0, column=1, sticky="w", padx=5)
        self.assign_user_combo.bind("<<ComboboxSelected>>", lambda e: self._load_user_roles_selector())

        # Roles checkboxes for selected user
        self.user_role_vars = {}
        self.user_roles_frame = ttk.Frame(assign)
        self.user_roles_frame.grid(row=1, column=0, columnspan=2, sticky="w", padx=5, pady=5)
        ttk.Button(assign, text="Guardar Roles de Usuario", command=self._save_user_roles_assignment).grid(row=2, column=1, sticky="e", padx=5, pady=5)

        # Initialize data
        self.current_role_id = None
        self._reload_roles_list()
        self._reload_users_combo_for_assignment()

    def _reload_roles_list(self):
        try:
            rows = self.db_manager.fetch_data("SELECT id, nombre, es_sistema FROM pal_roles ORDER BY nombre") or []
            # Clear
            for item in self.roles_tree.get_children():
                self.roles_tree.delete(item)
            for rid, nombre, es_sistema in rows:
                self.roles_tree.insert("", tk.END, values=(int(rid), str(nombre), "Sí" if bool(es_sistema) else "No"))
        except Exception as e:
            self.log(f"Error cargando roles: {e}", "ERROR")

    def _on_role_select(self):
        try:
            sel = self.roles_tree.selection()
            if not sel:
                return
            values = self.roles_tree.item(sel[0], "values")
            role_id = int(values[0])
            self.current_role_id = role_id
            # Load details
            row = self.db_manager.fetch_data("SELECT nombre, descripcion FROM pal_roles WHERE id = ?", (role_id,))
            if row:
                self.role_name_entry.delete(0, tk.END)
                self.role_name_entry.insert(0, str(row[0][0] or ""))
                self.role_desc_entry.delete(0, tk.END)
                self.role_desc_entry.insert(0, str(row[0][1] or ""))
            # Load permissions for role
            for var in self.perm_vars_by_id.values():
                var.set(False)
            perm_rows = self.db_manager.fetch_data("SELECT permiso_id FROM pal_roles_permisos WHERE rol_id = ?", (role_id,)) or []
            for (pid,) in perm_rows:
                if int(pid) in self.perm_vars_by_id:
                    self.perm_vars_by_id[int(pid)].set(True)
        except Exception as e:
            self.log(f"Error cargando rol: {e}", "ERROR")

    def _new_role(self):
        try:
            self.current_role_id = None
            self.role_name_entry.delete(0, tk.END)
            self.role_desc_entry.delete(0, tk.END)
            for var in self.perm_vars_by_id.values():
                var.set(False)
        except Exception:
            pass

    def _save_role(self):
        try:
            nombre = self.role_name_entry.get().strip()
            descripcion = self.role_desc_entry.get().strip() or None
            if not nombre:
                messagebox.showwarning("Error", "El nombre del rol es obligatorio")
                return
            if self.current_role_id is None:
                # Crear
                self.db_manager.execute_query(
                    "INSERT INTO pal_roles (nombre, descripcion, es_sistema) VALUES (?, ?, 0)",
                    (nombre, descripcion)
                )
                self.current_role_id = int(self.db_manager.fetch_data("SELECT id FROM pal_roles WHERE nombre = ?", (nombre,))[0][0])
            else:
                # Actualizar
                self.db_manager.execute_query(
                    "UPDATE pal_roles SET nombre = ?, descripcion = ? WHERE id = ?",
                    (nombre, descripcion, self.current_role_id)
                )
            self._reload_roles_list()
            messagebox.showinfo("Listo", "Rol guardado")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el rol: {e}")

    def _delete_role(self):
        try:
            if not self.current_role_id:
                messagebox.showwarning("Error", "Selecciona un rol")
                return
            row = self.db_manager.fetch_data("SELECT es_sistema, nombre FROM pal_roles WHERE id = ?", (self.current_role_id,))
            if not row:
                return
            es_sistema, nombre = bool(row[0][0]), str(row[0][1])
            if es_sistema:
                messagebox.showwarning("No permitido", "No se puede eliminar un rol del sistema")
                return
            if not messagebox.askyesno("Confirmar", f"¿Eliminar rol '{nombre}'?"):
                return
            # Eliminar asignaciones y rol
            self.db_manager.execute_query("DELETE FROM pal_roles_permisos WHERE rol_id = ?", (self.current_role_id,))
            self.db_manager.execute_query("DELETE FROM pal_usuarios_roles WHERE rol_id = ?", (self.current_role_id,))
            self.db_manager.execute_query("DELETE FROM pal_roles WHERE id = ?", (self.current_role_id,))
            self.current_role_id = None
            self._reload_roles_list()
            self._new_role()
            messagebox.showinfo("Listo", "Rol eliminado")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el rol: {e}")

    def _save_role_permissions(self):
        try:
            if not self.current_role_id:
                messagebox.showwarning("Error", "Selecciona o crea un rol antes de guardar permisos")
                return
            # Reemplazar conjunto de permisos del rol
            self.db_manager.execute_query("DELETE FROM pal_roles_permisos WHERE rol_id = ?", (self.current_role_id,))
            selected_perm_ids = [pid for pid, var in self.perm_vars_by_id.items() if var.get()]
            for pid in selected_perm_ids:
                self.db_manager.execute_query(
                    "IF NOT EXISTS (SELECT 1 FROM pal_roles_permisos WHERE rol_id = ? AND permiso_id = ?)\n"
                    "INSERT INTO pal_roles_permisos (rol_id, permiso_id) VALUES (?, ?)\n",
                    (self.current_role_id, pid, self.current_role_id, pid)
                )
            messagebox.showinfo("Listo", "Permisos del rol guardados")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron guardar permisos: {e}")

    def _reload_users_combo_for_assignment(self):
        try:
            rows = self.db_manager.fetch_data("SELECT id, username FROM pal_usuarios WHERE activo = 1 ORDER BY username") or []
            self.assign_users = [(int(r[0]), str(r[1])) for r in rows]
            self.assign_user_combo['values'] = [u[1] for u in self.assign_users]
            if self.assign_users:
                self.assign_user_combo.current(0)
                self._load_user_roles_selector()
        except Exception as e:
            self.log(f"Error cargando usuarios: {e}", "WARNING")

    def _load_user_roles_selector(self):
        try:
            # Clear existing
            for child in self.user_roles_frame.winfo_children():
                child.destroy()
            self.user_role_vars = {}
            sel_username = self.assign_user_combo.get()
            user_id = next((u[0] for u in getattr(self, 'assign_users', []) if u[1] == sel_username), None)
            if not user_id:
                return
            # Load roles and current assignments
            roles = self.db_manager.fetch_data("SELECT id, nombre FROM pal_roles ORDER BY nombre") or []
            assigned = self.db_manager.fetch_data("SELECT rol_id FROM pal_usuarios_roles WHERE usuario_id = ?", (user_id,)) or []
            assigned_set = {int(r[0]) for r in assigned}
            # Create checkboxes
            col = 0
            row = 0
            for rid, nombre in roles:
                var = tk.BooleanVar(value=int(rid) in assigned_set)
                cb = ttk.Checkbutton(self.user_roles_frame, text=str(nombre), variable=var)
                cb.grid(row=row, column=col, sticky="w", padx=6, pady=4)
                self.user_role_vars[int(rid)] = var
                col += 1
                if col >= 3:
                    col = 0
                    row += 1
        except Exception as e:
            self.log(f"Error cargando roles de usuario: {e}", "WARNING")

    def _save_user_roles_assignment(self):
        try:
            sel_username = self.assign_user_combo.get()
            user_id = next((u[0] for u in getattr(self, 'assign_users', []) if u[1] == sel_username), None)
            if not user_id:
                messagebox.showwarning("Error", "Selecciona un usuario")
                return
            # Replace role set
            self.db_manager.execute_query("DELETE FROM pal_usuarios_roles WHERE usuario_id = ?", (user_id,))
            selected_role_ids = [rid for rid, var in self.user_role_vars.items() if var.get()]
            for rid in selected_role_ids:
                self.db_manager.execute_query(
                    "IF NOT EXISTS (SELECT 1 FROM pal_usuarios_roles WHERE usuario_id = ? AND rol_id = ?)\n"
                    "INSERT INTO pal_usuarios_roles (usuario_id, rol_id) VALUES (?, ?)\n",
                    (user_id, rid, user_id, rid)
                )
            # Limpiar cache permisos del usuario si el servicio está disponible
            try:
                if hasattr(self, 'permissions'):
                    self.permissions.limpiar_cache_usuario(user_id)
            except Exception:
                pass
            messagebox.showinfo("Listo", f"Roles de '{sel_username}' actualizados")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron guardar roles del usuario: {e}")
    
    def _save_modules_config(self):
        """Deprecated: módulos ahora se controlan por BD (pal_usuarios_modulos)."""
        pass

    # =========================
    # Temas (Configuración de UI)
    # =========================
    def _create_temas_tab(self, parent):
        from pal.ui.themes import THEMES, apply_theme, get_current_theme
        
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="🎨 Temas de Interfaz", font=("Segoe UI", 14, "bold")).pack(pady=(0, 20))
        
        current_theme = get_current_theme(self)
        
        themes_frame = ttk.Frame(frame)
        themes_frame.pack(fill=tk.BOTH, expand=True)
        
        for idx, (theme_key, theme_data) in enumerate(THEMES.items()):
            card = ttk.Frame(themes_frame, relief="solid", borderwidth=1)
            card.grid(row=idx // 3, column=idx % 3, padx=10, pady=10, sticky="nsew")
            
            is_selected = (theme_key == current_theme)
            
            bg_color = theme_data["colors"]["bg_main"]
            accent_color = theme_data["colors"]["accent"]
            
            preview = tk.Frame(card, bg=bg_color, width=80, height=50)
            preview.pack(pady=(10, 5))
            preview.pack_propagate(False)
            tk.Label(preview, text=" preview ", bg=accent_color, fg="white", font=("Segoe UI", 8)).pack(expand=True)
            
            ttk.Label(card, text=theme_data["name"], font=("Segoe UI", 10, "bold")).pack()
            ttk.Label(card, text=theme_data["description"], font=("Segoe UI", 8)).pack()
            
            if is_selected:
                ttk.Label(card, text="✓ ACTIVO", foreground="green", font=("Segoe UI", 9, "bold")).pack(pady=5)
                btn_text = "Aplicar de nuevo"
            else:
                ttk.Label(card, text="   ", font=("Segoe UI", 8)).pack(pady=5)
                btn_text = "Aplicar"
            
            btn = ttk.Button(card, text=btn_text, command=lambda tk=theme_key: self._apply_theme(tk))
            btn.pack(pady=5)
        
        themes_frame.columnconfigure((0, 1, 2), weight=1)
        themes_frame.rowconfigure((0, 1), weight=1)
        
        ttk.Label(frame, text="Nota: Algunos cambios pueden requerir reiniciar la aplicación.", 
                  font=("Segoe UI", 8), foreground="gray").pack(pady=10)
    
    def _apply_theme(self, theme_key):
        from pal.ui.themes import apply_theme
        apply_theme(self, theme_key)
        messagebox.showinfo("Tema aplicado", f"El tema '{theme_key}' ha sido aplicado.\n\nNota: Los cambios se aplicarán completamente al reiniciar la aplicación.")

    # =========================
    # Auditoría (Fase 6)
    # =========================
    def _create_audit_tab(self, parent):
        frame = ttk.Frame(parent, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)

        # Filtros
        filters = ttk.Frame(frame)
        filters.pack(fill=tk.X)
        ttk.Label(filters, text="Usuario:").pack(side=tk.LEFT)
        self.audit_user_combo = ttk.Combobox(filters, state="readonly", width=25)
        self.audit_user_combo.pack(side=tk.LEFT, padx=5)
        ttk.Button(filters, text="Recargar", command=self._reload_audit_logs).pack(side=tk.LEFT, padx=5)

        # Cargar usuarios para filtro
        try:
            rows = self.db_manager.fetch_data("SELECT id, username FROM pal_usuarios ORDER BY username") or []
            self.audit_users = [(int(r[0]), str(r[1])) for r in rows]
            self.audit_user_combo['values'] = [u[1] for u in self.audit_users]
        except Exception:
            self.audit_users = []

        # Tabla
        table_frame = ttk.Frame(frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        cols = ("Fecha", "Usuario", "Acción", "Módulo", "Éxito", "IP", "Detalle")
        self.audit_tree = ttk.Treeview(table_frame, columns=cols, show='headings', height=12)
        
        # Configurar estilos de colores para filas intercaladas
        style = ttk.Style()
        style.configure('Treeview', rowheight=25)
        self.audit_tree.tag_configure('oddrow', background='#F0F0F0', foreground='#000000')
        self.audit_tree.tag_configure('evenrow', background='#FFFFFF', foreground='#000000')
        self.audit_tree.tag_configure('error_row', background='#FFE6E6', foreground='#CC0000')
        self.audit_tree.tag_configure('success_row', background='#E6F3E6', foreground='#006600')
        
        for c in cols:
            self.audit_tree.heading(c, text=c)
            self.audit_tree.column(c, width=120 if c != "Detalle" else 300, anchor='w')
        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=self.audit_tree.yview)
        self.audit_tree.configure(yscrollcommand=vsb.set)
        self.audit_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        self._reload_audit_logs()

    def _reload_audit_logs(self):
        try:
            # Limpiar
            for item in self.audit_tree.get_children():
                self.audit_tree.delete(item)
            # Filtro de usuario
            sel_user = self.audit_user_combo.get() if hasattr(self, 'audit_user_combo') else ''
            user_id = next((u[0] for u in getattr(self, 'audit_users', []) if u[1] == sel_user), None)
            if user_id:
                query = (
                    "SELECT TOP 500 a.fecha, u.username, a.accion, a.modulo, a.exitoso, a.ip_address, a.detalle "
                    "FROM pal_auditoria_accesos a LEFT JOIN pal_usuarios u ON u.id = a.usuario_id "
                    "WHERE a.usuario_id = ? ORDER BY a.fecha DESC"
                )
                rows = self.db_manager.fetch_data(query, (user_id,)) or []
            else:
                query = (
                    "SELECT TOP 500 a.fecha, u.username, a.accion, a.modulo, a.exitoso, a.ip_address, a.detalle "
                    "FROM pal_auditoria_accesos a LEFT JOIN pal_usuarios u ON u.id = a.usuario_id "
                    "ORDER BY a.fecha DESC"
                )
                rows = self.db_manager.fetch_data(query) or []
            
            # Insertar filas con colores intercalados
            for idx, r in enumerate(rows):
                fecha, username, accion, modulo, exitoso, ip, detalle = r
                
                # Determinar tag de color
                # Prioridad: error_row > success_row > oddrow/evenrow
                tags = ()
                if not exitoso:  # Filas de error en rojo
                    tags = ('error_row',)
                elif exitoso:  # Filas de éxito en verde
                    tags = ('success_row',)
                else:  # Colores intercalados por defecto
                    tags = ('oddrow' if idx % 2 == 0 else 'evenrow',)
                
                self.audit_tree.insert("", tk.END, values=(
                    str(fecha), str(username or ''), str(accion or ''), str(modulo or ''),
                    'Sí' if bool(exitoso) else 'No', str(ip or ''), str(detalle or '')
                ), tags=tags)
        except Exception as e:
            try:
                self.log(f"Error cargando auditoría: {e}", "ERROR")
            except Exception:
                pass

    def show_clientes_sub_view(self, view_name):
        """Gestiona la navegación entre subvistas dentro de la pestaña Clientes."""
        # Ocultar todas las subvistas de clientes
        if hasattr(self, 'clientes_menu_view') and self.clientes_menu_view.winfo_exists():
            self.clientes_menu_view.pack_forget()
        
        # La vista de reportes ahora se maneja dentro de esta función
        if hasattr(self, 'clientes_reportes_view') and self.clientes_reportes_view and self.clientes_reportes_view.winfo_exists():
            self.clientes_reportes_view.pack_forget()
        
        # Vista de estadísticas
        if hasattr(self, 'clientes_estadisticas_view') and self.clientes_estadisticas_view and self.clientes_estadisticas_view.winfo_exists():
            self.clientes_estadisticas_view.pack_forget()
        
        # Mostrar la vista solicitada
        if view_name == 'menu':
            if hasattr(self, 'clientes_menu_view'):
                self.clientes_menu_view.pack(fill=tk.BOTH, expand=True)
        elif view_name == 'reportes':
            # Crear la vista de reportes si no existe
            if not hasattr(self, 'clientes_reportes_view') or not self.clientes_reportes_view or not self.clientes_reportes_view.winfo_exists():
                from pal.ui.tabs.clientes_reportes import ClientesReportesTab
                self.clientes_reportes_view = ClientesReportesTab(self.clientes_tab, self)
            self.clientes_reportes_view.pack(fill=tk.BOTH, expand=True)
        elif view_name == 'estadisticas':
            # Crear la vista de estadísticas si no existe
            if not hasattr(self, 'clientes_estadisticas_view') or not self.clientes_estadisticas_view or not self.clientes_estadisticas_view.winfo_exists():
                from pal.ui.tabs.clientes_estadisticas import ClientesEstadisticasTab
                self.clientes_estadisticas_view = ClientesEstadisticasTab(self.clientes_tab, self)
            self.clientes_estadisticas_view.pack(fill=tk.BOTH, expand=True)

    def show_admin_sub_view(self, view_name):
        """Gestiona la navegación entre subvistas dentro de la pestaña de Administración (Config. Global)."""
        # Ocultar todas las subvistas
        views = [
            'admin_menu_view', 'sedes_servidores_view', 'sedes_almacenes_view',
            'admin_users_view', 'admin_roles_view', 'admin_exclusions_view',
            'admin_audit_view', 'admin_temas_view'
        ]
        for v in views:
            if hasattr(self, v) and getattr(self, v) and getattr(self, v).winfo_exists():
                getattr(self, v).pack_forget()

        # Mostrar la vista solicitada
        if view_name == 'menu':
            self.admin_menu_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'sedes_servidores':
            if not self.sedes_servidores_view or not self.sedes_servidores_view.winfo_exists():
                from pal.ui.tabs.sedes_servidores import SedesServidoresTab
                self.sedes_servidores_view = SedesServidoresTab(self.admin_tab, self)
            self.sedes_servidores_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'sedes_almacenes':
            if not self.sedes_almacenes_view or not self.sedes_almacenes_view.winfo_exists():
                from pal.ui.tabs.sedes_config import SedesAlmacenesTab
                self.sedes_almacenes_view = SedesAlmacenesTab(self.admin_tab, self)
            self.sedes_almacenes_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'exclusiones':
            if not self.admin_exclusions_view or not self.admin_exclusions_view.winfo_exists():
                self.admin_exclusions_view = ttk.Frame(self.admin_tab)
                self._create_admin_exclusions_tab(self.admin_exclusions_view)
            self.admin_exclusions_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'usuarios':
            if not self.admin_users_view or not self.admin_users_view.winfo_exists():
                self.admin_users_view = ttk.Frame(self.admin_tab)
                self._create_admin_users_tab(self.admin_users_view)
            self.admin_users_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'roles':
            if not self.admin_roles_view or not self.admin_roles_view.winfo_exists():
                self.admin_roles_view = ttk.Frame(self.admin_tab)
                self._create_roles_permissions_tab(self.admin_roles_view)
            self.admin_roles_view.pack(fill=tk.BOTH, expand=True)
            
        elif view_name == 'auditoria':
            if not self.admin_audit_view or not self.admin_audit_view.winfo_exists():
                self.admin_audit_view = ttk.Frame(self.admin_tab)
                self._create_audit_tab(self.admin_audit_view)
            self.admin_audit_view.pack(fill=tk.BOTH, expand=True)
        
        elif view_name == 'temas':
            if not hasattr(self, 'admin_temas_view') or not self.admin_temas_view or not self.admin_temas_view.winfo_exists():
                self.admin_temas_view = ttk.Frame(self.admin_tab)
                self._create_temas_tab(self.admin_temas_view)
            self.admin_temas_view.pack(fill=tk.BOTH, expand=True)

        # Botón de volver al menú (excepto si ya estamos en el menú)
        if view_name != 'menu':
            # Solo agregar si no existe
            if not hasattr(self, 'btn_admin_back') or not self.btn_admin_back or not self.btn_admin_back.winfo_exists():
                self.btn_admin_back = ttk.Button(self.admin_tab, text="◀ Volver al Config. Global", 
                                               command=lambda: self.show_admin_sub_view('menu'))
            self.btn_admin_back.pack(side=tk.BOTTOM, pady=10, anchor="e", padx=20)
        else:
            if hasattr(self, 'btn_admin_back') and self.btn_admin_back and self.btn_admin_back.winfo_exists():
                self.btn_admin_back.pack_forget()


    def _create_admin_sedes_tab(self, parent):
        """Crea la pestaña de configuración de sedes y stock."""
        from pal.ui.tabs.sedes_config import create_sedes_almacenes_tab
        create_sedes_almacenes_tab(parent, self)

    def _save_debug_config(self):
        try:
            flags = {k: v.get() for k, v in self.debug_vars.items()}
            save_debug_config(flags)
            # Actualizar flags en runtime
            self.debug_flags = flags
            self.tra_debug = flags.get('tra', False)
            self.stock_debug = flags.get('stock', False)
            try:
                setattr(self.db_manager, 'debug_enabled', flags.get('db', False))
            except Exception:
                pass
            self.log("Configuración de depuración guardada", "SUCCESS")
        except Exception as e:
            self.log(f"Error guardando depuración: {e}", "ERROR")

    def on_settings_close(self):
        try:
            if hasattr(self, 'settings_window') and self.settings_window and self.settings_window.winfo_exists():
                self.settings_window.destroy()
        except Exception:
            pass
        finally:
            self.settings_window = None

    def connect_to_database(self):
        """Alias for connect_db to maintain compatibility"""
        self.show_settings()

    def _require_login(self) -> bool:
        try:
            if not self.auth:
                return False
            # Mostrar diálogo de login hasta éxito o cancelación
            from pal.ui.login import LoginDialog
            dlg = LoginDialog(self.root)
            self.root.wait_window(dlg.top)
            if not dlg.result:
                self.update_status('error', message="Login cancelado")
                return False
            username, password = dlg.result
            resp = self.auth.login(username, password, ip_address=None)
            if not resp.get('success'):
                from tkinter import messagebox
                messagebox.showerror("Login fallido", resp.get('message', 'Error'))
                return self._require_login()
            self.session_token = resp['token']
            self.current_user = resp['user']
            self.log(f"Sesión iniciada como {self.current_user['username']}", "INFO")
            return True
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Error", f"Error durante login: {e}")
            return False

    def _reset_session_on_db_change(self):
        """Reset de sesión cuando se cambia de base de datos."""
        try:
            self.log("[DB CHANGE] Reseteando sesión...", "INFO")
            # Cerrar sesión actual
            if hasattr(self, 'session_token') and self.session_token and hasattr(self, 'auth'):
                try:
                    self.auth.logout(self.session_token)
                except Exception:
                    pass
            # Resetear variables de sesión
            self.session_token = None
            self.current_user = None
            self.auth = None
            self.permissions = None
            # Limpiar caché
            self.cached_ventas_tra = []
            self.cached_ventas_mbrp = []
            self.cached_alertas = []
            self.log("[DB CHANGE] Sesión reseteada correctamente", "SUCCESS")
        except Exception as e:
            self.log(f"[DB CHANGE] Error durante reset: {e}", "WARNING")

    def _ensure_security_schema_interactive(self):
        """Verifica tablas pal_* y ofrece crearlas si faltan."""
        try:
            missing = []
            try:
                missing = self.db_manager.check_security_schema()
            except Exception as e:
                self.log(f"No se pudo verificar esquema de seguridad: {e}", "ERROR")
                return

            if missing:
                from tkinter import messagebox
                msg = (
                    "¡Oh no!\n\n" 
                    "No existen las tablas necesarias para el funcionamiento normal:\n\n"
                    + "\n".join(f"• {t}" for t in missing)
                    + "\n\n¿Quieres crearlas ahora?"
                )
                if messagebox.askyesno("Crear tablas necesarias", msg):
                    try:
                        self.update_status('action', message="Creando tablas PAL...")
                        self.db_manager.ensure_security_tables()
                        self.update_status('connected')
                        messagebox.showinfo("Listo", "Tablas creadas correctamente.\n\nAhora inicia sesión con:\nUsuario: admin\nContraseña: 123")
                        self.log("Esquema pal_* creado/actualizado", "SUCCESS")
                    except Exception as e:
                        self.update_status('error', message=str(e))
                        messagebox.showerror(
                            "Error creando tablas",
                            f"No se pudieron crear las tablas necesarias.\n\nDetalle: {e}"
                        )
                        self.log(f"Error creando esquema pal_*: {e}", "ERROR")
        except Exception as e:
            self.log(f"Error en verificación de esquema: {e}", "ERROR")

    def connect_db(self):
        try:
            
            server = self.server_entry.get().strip()
            database = self.db_entry.get().strip()
            user = self.user_entry.get().strip()
            password = self.pwd_entry.get().strip()
            token = self.token_entry.get().strip()

            if not server:
                messagebox.showwarning("Error", "El campo Servidor es obligatorio")
                return
    
            if not password:
                password = self.cred_manager.get_temp_password() or ''
    
            if password:
                self.cred_manager.store_temp_password(password)

            # Detectar cambio de base de datos
            old_settings = self.load_connection_settings()
            db_changed = (
                old_settings and (
                    old_settings.get('server') != server or 
                    old_settings.get('database') != database or 
                    old_settings.get('user') != user
                )
            )
            
            if db_changed:
                self.log(f"[DB CHANGE] Cambio de BD: {old_settings.get('database')} → {database}", "INFO")
                self._reset_session_on_db_change()

            if self.db_manager.connect(server, database, user, password):
                self.save_connection_settings(server, database, user, token)
                self.update_status('connected', server=server, api_token=token)
                # Inicializar servicios de seguridad
                from pal.core.audit_db import AuditDB
                self.audit_db = AuditDB(self.db_manager)
                self.auth = AuthManager(self.db_manager)
                self._ensure_security_schema_interactive()
                # Requerir login
                if not self._require_login():
                    return
                self.settings_window.destroy()
                self.log("Conexión a BD exitosa", "SUCCESS")
                try:
                    if hasattr(self, 'audit_db'):
                        self.audit_db.log_action(accion='DB_CONNECTED', usuario_id=self.current_user['id'] if self.current_user else None, detalle=server, modulo='ADMIN')
                except Exception:
                    pass
                self.show_temp_notification("Conexión exitosa")
                self.search_records()
            
        except Exception as e:
            self.log(f"Error de conexión: {str(e)}", "ERROR")
            error_msg = str(e)
            self.audit_log.log_event(
            "DB_CONNECTION_ERROR",
            os.getlogin(),
            "FAILED",
            ErrorCode.DB_CONNECTION_FAILED
        )
            self.update_status('error', message=error_msg)
            try:
                if hasattr(self, 'audit_db'):
                    self.audit_db.log_action(
                        accion='DB_CONNECTION_ERROR', usuario_id=self.current_user['id'] if self.current_user else None, detalle=error_msg, exitoso=False, modulo='ADMIN')
            except Exception:
                pass
            self.show_temp_notification("Error de conexión", duration=5000)

    def create_connection_tab(self, parent):
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Servidor:").grid(row=0, column=0, sticky="w")
        self.server_entry = ttk.Entry(frame)
        self.server_entry.grid(row=0, column=1, sticky="ew", pady=5)
        
        ttk.Label(frame, text="Base de Datos:").grid(row=1, column=0, sticky="w")
        self.db_entry = ttk.Entry(frame)
        self.db_entry.grid(row=1, column=1, sticky="ew", pady=5)
        
        ttk.Label(frame, text="Usuario:").grid(row=2, column=0, sticky="w")
        self.user_entry = ttk.Entry(frame)
        self.user_entry.grid(row=2, column=1, sticky="ew", pady=5)
        
        ttk.Label(frame, text="Contraseña:").grid(row=3, column=0, sticky="w")
        self.pwd_entry = ttk.Entry(frame, show="*")
        self.pwd_entry.grid(row=3, column=1, sticky="ew", pady=5)

        # Botones en la parte inferior
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20, sticky="e")
    
        ttk.Button(btn_frame, text="Guardar", 
            command=self.connect_db).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Cancelar", 
            command=lambda: (self.settings_window.destroy() if self.settings_window else None)).pack(side=tk.RIGHT)
        
        # Cargar configuración existente
        settings = self.load_connection_settings()
        if settings:
            self.server_entry.insert(0, settings['server'])
            self.db_entry.insert(0, settings['database'])
            self.user_entry.insert(0, settings['user'])
    
    def update_status(self, status_type: str, message: str = "", server: str = "", api_token: Optional[str] = None):
        """Actualizar la barra de estado principal Y el dashboard"""
        status_config = {
            'connected': {
                'text': f"BD: Conectado" if server else "BD: Conectado",
                'color': "green",
                'dashboard_text': "Conectado ✓",
                'dashboard_color': "#10B981"
            },
            'error': {
                'text': f"Error: {message[:50]}..." if len(message) > 50 else f"Error: {message}",
                'color': "red",
                'dashboard_text': "Error ✗",
                'dashboard_color': "#EF4444"
            },
            'action': {
                'text': message,
                'color': "blue",
                'dashboard_text': "Procesando...",
                'dashboard_color': "#3B82F6"
            },
            'disconnected': {
                'text': "BD: Desconectado",
                'color': "orange",
                'dashboard_text': "Desconectado ✗",
                'dashboard_color': "#EF4444"
            }
        }
        
        config = status_config.get(status_type, {
            'text': "Estado desconocido", 
            'color': "gray",
            'dashboard_text': "Desconocido",
            'dashboard_color': "#6B7280"
        })
        
        # Actualizar footer status (Solo si el widget ya existe y no ha sido destruido)
        try:
            if hasattr(self, 'db_status') and self.db_status and self.db_status.winfo_exists():
                self.db_status.config(text=config['text'], foreground=config['color'])
        except (AttributeError, tk.TclError):
            pass
        
        # Actualizar dashboard status (si existe)
        try:
            if hasattr(self, 'dashboard_connection_status') and self.dashboard_connection_status.winfo_exists():
                self.dashboard_connection_status.config(
                    text=config['dashboard_text'], 
                    foreground=config['dashboard_color']
                )
        except Exception:
            pass  # Dashboard aún no está inicializado
        
        # Mantener estado API independiente
        if not hasattr(self, 'api_state'):
            self.api_state = "inactive"
    
        # Actualizar solo si hay cambio explícito y el widget existe
        try:
            if status_type == 'api_connected':
                self.api_state = "active"
                if hasattr(self, 'api_status') and self.api_status and self.api_status.winfo_exists():
                    self.api_status.config(text="API: Lista", foreground="green")
            elif status_type == 'api_error':
                self.api_state = "error"
                if hasattr(self, 'api_status') and self.api_status and self.api_status.winfo_exists():
                    self.api_status.config(text="API: Error", foreground="red")
        except (AttributeError, tk.TclError):
            pass

        # Habilitar/Deshabilitar UI dependiente de BD
        try:
            if status_type == 'connected':
                self._set_ui_connected(True)
            elif status_type in ('disconnected', 'error'):
                self._set_ui_connected(False)
        except Exception:
            pass

    def show_temp_notification(self, message: str, duration: int = 3000):
        """Mostrar notificación temporal en la esquina inferior derecha"""
        notification = tk.Toplevel(self.root)
        notification.wm_overrideredirect(True)
        notification.geometry(f"300x60+{self.root.winfo_x()+700}+{self.root.winfo_y()+600}")
    
        ttk.Label(notification, text=message, wraplength=280).pack(pady=10)
        notification.after(duration, notification.destroy)
    
    def save_api_settings(self):
        try:
            new_token = self.token_entry.get().strip()
        # Si el usuario ingresó algo, se guarda el nuevo token; de lo contrario se conserva el anterior
            if new_token:
                self.cred_manager.store_whatsapp_token(new_token)
                self.show_temp_notification("Token guardado")
                self.cred_manager.store_whatsapp_token(new_token)
                # Actualizar estado inmediatamente
                self.update_status('connected', api_token=new_token)
            else:
                self.show_temp_notification("No se detectaron cambios en el token")
            self.settings_window.destroy()
        except Exception as e:
            self.notification_manager.show_error("Error", f"Error guardando token: {str(e)}")

    def create_api_tab(self, parent):
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Token WhatsApp:").grid(row=0, column=0, sticky="w")
        self.token_entry = ttk.Entry(frame, show="*")
        self.token_entry.grid(row=0, column=1, sticky="ew", pady=5)

        # Botones
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=1, column=0, columnspan=2, pady=20, sticky="e")
    
        ttk.Button(btn_frame, text="Guardar", 
            command=self.save_api_settings).pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, 
          text="Cancelar", 
          command=lambda: (self.settings_window.destroy() if self.settings_window else None)).pack(side=tk.RIGHT)
        
        # Cargar token existente
        try:
            token = self.cred_manager.get_whatsapp_token()
            if token:
                self.token_entry.insert(0, token)
        except Exception as e:
            self.notification_manager.show_error("Error cargando token")

    def create_update_tab(self, parent):
        """Crea la pestaña de configuración de actualizaciones."""
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Información de versión actual
        version_label = ttk.Label(frame, text=f"Versión actual: {APP_VERSION}", font=("Arial", 10, "bold"))
        version_label.grid(row=0, column=0, columnspan=2, sticky="w", pady=10)
        
        # URL de actualizaciones
        ttk.Label(frame, text="URL de actualizaciones:").grid(row=1, column=0, sticky="w", pady=5)
        self.update_url_entry = ttk.Entry(frame, width=50)
        self.update_url_entry.grid(row=1, column=1, sticky="ew", pady=5, padx=5)
        # Cargar URL desde configuración
        try:
            current_url = load_update_url()
        except Exception:
            current_url = UPDATE_URL_DEFAULT
        self.update_url_entry.insert(0, current_url)
        
        # Estado de verificación periódica
        self.update_status_label = ttk.Label(frame, text="Estado: Verificación periódica activa", foreground="green")
        self.update_status_label.grid(row=2, column=0, columnspan=2, sticky="w", pady=5)
        
        # Botones
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=20, sticky="ew")
        
        ttk.Button(btn_frame, text="Verificar actualizaciones ahora", 
            command=self.check_updates_manual).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Actualizar e instalar", 
            command=self.update_and_install).pack(side=tk.LEFT, padx=5)
        
        frame.columnconfigure(1, weight=1)
    
    def _on_update_available(self, has_update: bool):
        """Callback cuando se encuentra una actualización disponible."""
        if has_update and hasattr(self, 'update_manager') and self.update_manager:
            try:
                latest_info = self.update_manager.get_latest_version_info()
                if latest_info:
                    message = f"Hay una nueva versión disponible:\n\n"
                    message += f"Versión actual: {APP_VERSION}\n"
                    message += f"Nueva versión: {latest_info.get('version', 'Desconocida')}\n\n"
                    if latest_info.get('changelog'):
                        message += f"Cambios:\n{latest_info['changelog']}\n\n"
                    message += "¿Deseas actualizar ahora?"
                    
                    response = messagebox.askyesno("Actualización disponible", message)
                    if response:
                        # Ejecutar en un hilo para no bloquear el messagebox
                        threading.Thread(
                            target=lambda: self.update_and_install(progress_callback=self.splash.set_progress),
                            daemon=True
                        ).start()
            except Exception as e:
                print(f"[ERROR] Error al procesar actualización disponible: {e}", flush=True)
    
    def save_update_url_config(self):
        """Guarda la URL de actualizaciones desde la UI."""
        new_url = self.update_url_entry.get().strip()
        if not new_url:
            messagebox.showerror("Error", "La URL no puede estar vacía.")
            return
        
        # Validar formato básico de URL
        if not (new_url.startswith('http://') or new_url.startswith('https://')):
            messagebox.showerror("Error", "La URL debe comenzar con http:// o https://")
            return
        
        # Guardar en configuración
        save_update_url(new_url)
        
        # Actualizar el UpdateManager si existe
        if hasattr(self, 'update_manager') and self.update_manager:
            self.update_manager.update_url = new_url
            # Reiniciar verificación periódica con nueva URL
            self.update_manager.stop_periodic_check()
            self.update_manager.start_periodic_check(
                callback=lambda has_update: self._on_update_available(has_update)
            )
        
        messagebox.showinfo("Éxito", f"URL de actualizaciones guardada:\n{new_url}")
        self.update_status_label.config(text=f"URL actualizada: {new_url}", foreground="green")
    
    def check_updates_manual(self):
        """Verifica actualizaciones manualmente desde la UI."""
        if not hasattr(self, 'update_manager') or not self.update_manager:
            messagebox.showerror("Error", "El gestor de actualizaciones no está disponible.")
            return
        
        try:
            # Obtener URL actual del campo de texto (puede haber cambiado)
            current_url = self.update_url_entry.get().strip()
            if current_url and current_url != self.update_manager.update_url:
                # Actualizar URL si cambió
                self.update_manager.update_url = current_url
            
            # Mostrar mensaje de verificación
            self.update_status_label.config(text="Verificando actualizaciones...", foreground="blue")
            self.settings_window.update()
            
            # Verificar actualizaciones
            has_update = self.update_manager.check_for_updates(show_no_update_message=True)
            
            if has_update:
                latest_info = self.update_manager.get_latest_version_info()
                if latest_info:
                    self.update_status_label.config(
                        text=f"Actualización disponible: {latest_info.get('version', 'Desconocida')}", 
                        foreground="orange"
                    )
                else:
                    self.update_status_label.config(text="Actualización disponible", foreground="orange")
            else:
                self.update_status_label.config(text="Ya tienes la versión más reciente", foreground="green")
        except Exception as e:
            self.update_status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"No se pudo verificar actualizaciones: {str(e)}")
    
    def update_and_install(self, progress_callback: Optional[Callable[[float], None]] = None):
        """Descarga e instala la actualización disponible."""
        if not hasattr(self, 'update_manager') or not self.update_manager:
            messagebox.showerror("Error", "El gestor de actualizaciones no está disponible.")
            return
        
        try:
            # Verificar primero si hay actualización
            if not self.update_manager.check_for_updates():
                messagebox.showinfo("Información", "No hay actualizaciones disponibles.")
                return
            
            # Mostrar progreso si la UI de settings está disponible
            if hasattr(self, 'settings_window') and self.settings_window and hasattr(self, 'update_status_label'):
                self.update_status_label.config(text="Descargando actualización...", foreground="blue")
                self.settings_window.update()
            
            # Descargar con el callback de progreso
            if self.update_manager.download_update(progress_callback=progress_callback):
                if hasattr(self, 'settings_window') and self.settings_window and hasattr(self, 'update_status_label'):
                    self.update_status_label.config(text="Instalando actualización...", foreground="blue")
                    self.settings_window.update()
                
                # Instalar (esto reiniciará la aplicación)
                if self.update_manager.install_update(restart_callback=self.shutdown):
                    messagebox.showinfo("Éxito", "La actualización se instalará al reiniciar la aplicación.")
                    # El gestor de actualizaciones manejará el reinicio automáticamente
                else:
                    if hasattr(self, 'settings_window') and self.settings_window and hasattr(self, 'update_status_label'):
                        self.update_status_label.config(text="Error al instalar actualización", foreground="red")
            else:
                if hasattr(self, 'settings_window') and self.settings_window and hasattr(self, 'update_status_label'):
                    self.update_status_label.config(text="Error al descargar actualización", foreground="red")
        except Exception as e:
            if hasattr(self, 'settings_window') and self.settings_window and hasattr(self, 'update_status_label'):
                self.update_status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"No se pudo completar la actualización: {str(e)}")

    def _initialize_post_login_components(self):
        """Inicializa componentes no críticos después del login para acelerar el inicio."""
        try:
            # Setup tooltips
            self.setup_tooltips()
            
            # Inicializar gestor de actualizaciones (30s después del login)
            def _init_updates():
                try:
                    update_url = load_update_url()
                    self.update_manager = UpdateManager(
                        app_name="Casapro Nexus",
                        current_version=APP_VERSION,
                        update_url=update_url,
                        update_check_interval=3600,
                        auto_download=False,
                        auto_install=False
                    )
                    self.update_manager.start_periodic_check(
                        callback=lambda has_update: self._on_update_available(has_update)
                    )
                    print(f"[DEBUG] Gestor de actualizaciones inicializado (versión {APP_VERSION})", flush=True)
                except Exception as e:
                    print(f"[WARNING] Error al inicializar actualizaciones: {e}", flush=True)
                    self.update_manager = None
            
            # Diferir inicialización de updates 30 segundos
            threading.Timer(30.0, _init_updates).start()
        except Exception as e:
            print(f"[WARNING] Error en inicialización post-login: {e}", flush=True)

    def _splash_login_submit(self, username: str, password: str) -> tuple[bool, str]:
        """
        Validar credenciales desde el splash screen e iniciar verificaciones críticas.
        """
        try:
            if not self.auth or not self.db_manager.conn:
                return False, "Base de datos no conectada"
            
            # 1. Verificación de actualizaciones obligatorias (cada 12h)
            try:
                last_check = load_last_update_check()
                now = datetime.now()
                
                # Si nunca se ha verificado o pasaron 12 horas
                if last_check is None or (now - last_check) > timedelta(hours=12):
                    print(f"[DEBUG] Ejecutando comprobación obligatoria de actualizaciones...", flush=True)
                    
                    # Asegurar que tenemos UpdateManager
                    if not hasattr(self, 'update_manager') or self.update_manager is None:
                        update_url = load_update_url()
                        self.update_manager = UpdateManager(
                            app_name="Casapro Nexus",
                            current_version=APP_VERSION,
                            update_url=update_url
                        )
                    
                    # Verificar si hay actualización
                    if self.update_manager.check_for_updates():
                        latest_info = self.update_manager.get_latest_version_info()
                        if latest_info:
                            # Programar diálogo mandatorio en el hilo principal
                            self.root.after(100, lambda: self.show_mandatory_update_dialog(latest_info))
                            return False, "Actualización obligatoria disponible"
                    
                    # Guardar fecha de comprobación exitosa
                    save_last_update_check(now)
                    
            except Exception as e:
                print(f"[WARNING] Error en comprobación obligatoria: {e}", flush=True)

            # 2. Validar credenciales
            resp = self.auth.login(username, password, ip_address=None)
            if not resp.get('success'):
                try:
                    if hasattr(self, 'audit_db'):
                        self.audit_db.log_access(
                            accion='LOGIN', usuario_id=None, exitoso=False, detalle=f"user={username}"
                        )
                except Exception:
                    pass
                return False, resp.get('message', 'Credenciales inválidas')
            
            # Guardar sesión
            self.session_token = resp['token']
            self.current_user = resp['user']
            
            self.log(f"✅ Sesión iniciada como {username}", "SUCCESS")
            try:
                if hasattr(self, 'audit_db'):
                    self.audit_db.log_access(accion='LOGIN', usuario_id=self.current_user['id'], exitoso=True)
            except Exception:
                pass
            
            # Forzar cambio de contraseña si admin usó la predeterminada (en hilo principal)
            if username.lower() == 'admin' and password == '123':
                # Programar en hilo principal después de que la UI esté lista
                self.root.after(500, self._prompt_change_admin_password)
            
            
            def _post_login_setup():
                try:
                    # Cargar tema guardado
                    try:
                        from pal.ui.themes import load_saved_theme
                        load_saved_theme(self)
                    except Exception as e:
                        print(f"[WARN] Error cargando tema: {e}")
                    
                    # CONFIGURAR COMPONENTES UI AHORA (Diferido desde la inicialización temprana)
                    try:
                        self.setup_modern_ui()
                    except Exception as e:
                        print(f"[ERROR] No se pudo inicializar UI moderna: {e}")

                    # Verificar y crear esquema si es necesario
                    try:
                        missing = self.db_manager.check_security_schema()
                        if missing:
                            self.db_manager.ensure_security_tables()
                        # Asegurar tablas y permisos de Logística
                        self.db_manager.ensure_logistica_tables()
                    except Exception as e:
                        self.log(f"Nota: {e}", "INFO")
                    
                    # 1. Cargar permisos del usuario
                    from pal.core.permissions import PermissionsManager
                    self.permissions = PermissionsManager(self.db_manager)
                    user_id = self.current_user['id']
                    db_mods = self.permissions.obtener_modulos_disponibles(user_id) or []
                    
                    # 2. Resetear y poblar módulos habilitados basado en la tabla del usuario
                    self.modules_enabled = {k: False for k in FLAG_TO_DB_MODULE.keys()}
                    
                    # 3. Mapear nombres de BD a flags de app
                    for m in db_mods:
                        if m in DB_MODULE_TO_FLAG:
                            flag = DB_MODULE_TO_FLAG[m]
                            self.modules_enabled[flag] = True
                    
                    self.log(f"Módulos disponibles para usuario {self.current_user['username']}: {', '.join(db_mods) or 'ninguno'}", "INFO")
                    
                    # 4. Verificar autorizaciones pendientes de logística
                    if "logistica" in self.modules_enabled and self.modules_enabled["logistica"]:
                        self.root.after(2000, self._check_pending_authorizations)

                    # Recrear workspace con módulos actualizados
                    try:
                        # Limpiar dashboard_tab específicamente antes de destruir
                        if hasattr(self, 'dashboard_tab') and self.dashboard_tab.winfo_exists():
                            for widget in self.dashboard_tab.winfo_children():
                                widget.destroy()
                        
                        if hasattr(self, 'main_notebook') and self.main_notebook.winfo_exists():
                            self.main_notebook.destroy()
                        
                        # Forzar actualización de la UI
                        self.root.update_idletasks()
                        
                        self.create_main_workspace()
                        
                        # Cargar lista de registros al iniciar si la pestaña existe
                        try:
                            if hasattr(self, 'search_records'):
                                self.search_records()
                        except Exception:
                            pass
                    except Exception as e:
                        self.log(f"Error recreando workspace: {e}", "WARNING")
                    
                    # ── Integración post-login del sistema de notificaciones ──────────
                    try:
                        from pal.infrastructure.notification_db_backend import PyodbcNotificationBackend
                        from pal.services.notifications import NotificationManager as CentralNotificationManager

                        # Guardar referencia al manager anterior para migrar notificaciones en memoria
                        old_manager = getattr(self, 'notification_manager', None)

                        # Crear nuevo manager con backend persistente
                        backend = PyodbcNotificationBackend(self.db_manager)
                        new_manager = CentralNotificationManager(db_backend=backend)

                        # Migrar notificaciones en memoria del manager anterior (si las hay)
                        if old_manager is not None and hasattr(old_manager, 'notifications'):
                            try:
                                for n in list(old_manager.notifications):
                                    if n.id not in {x.id for x in new_manager.notifications}:
                                        new_manager.notifications.append(n)
                            except Exception:
                                pass

                        # Reemplazar el manager
                        self.notification_manager = new_manager

                        # Cargar notificaciones activas del usuario desde la BD
                        usuario = self.current_user.get('username') if self.current_user else None
                        self.notification_manager.load_from_db(usuario=usuario)

                        # Actualizar la campana si ya existe en el status panel
                        if hasattr(self, 'notification_bell') and self.notification_bell:
                            # Desconectar observador del manager viejo (si existe)
                            if old_manager is not None and hasattr(old_manager, 'remove_observer'):
                                try:
                                    old_manager.remove_observer(
                                        self.notification_bell._on_notifications_changed
                                    )
                                except Exception:
                                    pass

                            # Conectar campana al nuevo manager
                            self.notification_bell._mgr = self.notification_manager
                            self.notification_bell.set_usuario(usuario)
                            self.notification_manager.add_observer(
                                self.notification_bell._on_notifications_changed
                            )
                            self.notification_bell._refresh_badge()

                        self.log("✅ Sistema de notificaciones persistente activado", "SUCCESS")
                    except Exception as _notif_err:
                        self.log(f"[Notificaciones] Error activando backend persistente: {_notif_err}", "WARNING")
                    # ─────────────────────────────────────────────────────────────────

                    # Inicializar servicios de fondo según módulos habilitados
                    self._start_module_services()
                    
                    # Inicializar módulos en segundo plano
                    threading.Thread(target=self._inicializar_modulos_paralelo, daemon=True).start()
                except Exception as e:
                    self.log(f"Error en setup post-login: {e}", "ERROR")
            
            # Ejecutar setup después de 150ms (asegurar que el splash y tabs se procesen)
            self.root.after(150, _post_login_setup)
            
            return True, "Login exitoso"
            
        except Exception as e:
            self.log(f"Error en login: {e}", "ERROR")
            return False, f"Error: {str(e)[:50]}"
    
    def _check_pending_authorizations(self):
        """Verifica si hay autorizaciones de logística pendientes y notifica al usuario."""
        try:
            if not self.db_manager or not self.current_user:
                return

            # Solo notificar a usuarios con permiso de autorizar
            if not self.permissions.tiene_permiso(self.current_user['id'], 'LOGISTICA', 'autorizar'):
                return

            # Contar pendientes
            sql = "SELECT COUNT(*) FROM pal_sugerencias_transferencia WHERE requiere_autorizacion = 1 AND fue_autorizada = 0 AND estado = 'pendiente'"
            res = self.db_manager.fetch_data(sql)
            count = res[0][0] if res and res[0] else 0

            if count > 0:
                self.log(f"📢 Hay {count} autorizaciones de abastecimiento pendientes", "WARNING")
                if hasattr(self, 'notification_manager'):
                    self.notification_manager.show_banner(
                        title="Autorizaciones Pendientes",
                        message=f"Atención: Hay {count} transferencias pendientes de autorización en Logística.",
                        bg="#f39c12", # Orange
                        duration=10000
                    )
        except Exception as e:
            self.log(f"Error verificando autorizaciones: {e}", "ERROR")
    
    def _prompt_change_admin_password(self):
        """
        Solicita cambio de contraseña para admin en el hilo principal.
        Ejecutado mediante root.after() para evitar problemas de threading.
        """
        try:
            while True:
                p1 = simpledialog.askstring(
                    "Cambiar contraseña", 
                    "Nueva contraseña para admin (dejado en blanco cancela):", 
                    show='*', 
                    parent=self.root
                )
                if p1 is None or p1 == '':
                    self.log("Cambio de contraseña cancelado", "INFO")
                    return
                
                p2 = simpledialog.askstring(
                    "Cambiar contraseña", 
                    "Confirmar contraseña:", 
                    show='*', 
                    parent=self.root
                )
                if p2 is None or p2 == '':
                    self.log("Cambio de contraseña cancelado", "INFO")
                    return
                
                if p1 != p2:
                    messagebox.showerror("Error", "Las contraseñas no coinciden. Intente nuevamente.")
                    continue
                
                if len(p1) < 4:
                    messagebox.showerror("Error", "La contraseña debe tener al menos 4 caracteres")
                    continue
                
                # Intentar actualizar contraseña
                if self.auth and self.current_user:
                    ok = self.auth.cambiar_password(self.current_user['id'], '123', p1)
                    if ok:
                        messagebox.showinfo("Éxito", "Contraseña actualizada exitosamente")
                        self.log("✅ Contraseña admin actualizada", "SUCCESS")
                        return
                    else:
                        messagebox.showerror("Error", "No se pudo actualizar la contraseña. Intente nuevamente.")
                        continue
                else:
                    messagebox.showerror("Error", "Sistema no listo para cambiar contraseña")
                    return
                    
        except Exception as e:
            self.log(f"Error en cambio de contraseña: {e}", "ERROR")
            messagebox.showerror("Error", f"Error al cambiar contraseña: {str(e)[:100]}")
    
    def show_mandatory_update_dialog(self, info: dict):
        """Muestra una ventana modal que obliga al usuario a actualizar."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Actualización Obligatoria")
        dialog.geometry("500x450")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Centrar diálogo
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() - 500) // 2
        y = (dialog.winfo_screenheight() - 450) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Contenido
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="Nueva versión disponible", font=("Segoe UI", 14, "bold")).pack(pady=(0, 10))
        ttk.Label(frame, text=f"Versión actual: {APP_VERSION}", font=("Segoe UI", 10)).pack(anchor="w")
        ttk.Label(frame, text=f"Versión requerida: {info.get('version', 'Desconocida')}", font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 15))
        
        ttk.Label(frame, text="Notas de la versión:", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        
        # Área de texto para changelog (con scroll)
        text_frame = ttk.Frame(frame)
        text_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        changelog_text = tk.Text(text_frame, height=10, font=("Segoe UI", 9), wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=changelog_text.yview)
        changelog_text.configure(yscrollcommand=scrollbar.set)
        
        changelog_text.insert(tk.END, info.get('changelog', 'Sin notas detalladas.'))
        changelog_text.config(state=tk.DISABLED)
        
        changelog_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        ttk.Label(frame, text="Es necesario actualizar para continuar usando la aplicación.", 
                  foreground="red", wraplength=450, font=("Segoe UI", 10, "italic")).pack(pady=15)
        
        # Botones
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill="x", side="bottom")
        
        def _do_update():
            # Deshabilitar botones
            btn_update.config(state=tk.DISABLED)
            btn_exit.config(state=tk.DISABLED)
            
            # Cambiar label de estado
            status_label.config(text="Iniciando descarga...", foreground="blue")
            
            # Ejecutar en hilo
            def _thread_update():
                try:
                    # Usar update_and_install existente de DatabaseApp
                    self.update_and_install(progress_callback=lambda p: self.root.after(0, lambda: status_label.config(text=f"Descargando: {int(p*100)}%")))
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror("Error", f"No se pudo completar la actualización: {e}"))
                    self.root.after(0, lambda: [btn_update.config(state=tk.NORMAL), btn_exit.config(state=tk.NORMAL)])

            threading.Thread(target=_thread_update, daemon=True).start()

        btn_update = ttk.Button(btn_frame, text="Actualizar ahora", command=_do_update, style="Accent.TButton")
        btn_update.pack(side=tk.RIGHT, padx=5)
        
        btn_exit = ttk.Button(btn_frame, text="Salir", command=self.shutdown)
        btn_exit.pack(side=tk.RIGHT)
        
        status_label = ttk.Label(frame, text="", font=("Segoe UI", 9))
        status_label.pack(side="bottom", pady=5)
        
        # Prevenir cierre accidental
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)
    
    
    def logout(self):
        """Cierra la sesión actual y muestra pantalla de login."""
        try:
            # Cerrar sesión en BD
            if self.auth and self.session_token:
                self.auth.logout(self.session_token)
            
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_access(accion='LOGOUT', usuario_id=self.current_user['id'], exitoso=True)
            except Exception:
                pass
            
            self.session_token = None
            self.current_user = None
            self.permissions = None
            
            # Resetear flags de jerarquías para el próximo usuario
            self.jerarquias_unificadas_cargadas = False
            
            self.log("Sesión cerrada", "INFO")
            
            # Ocultar ventana principal
            self.root.withdraw()
            
            # Crear pantalla de login similar al splash
            self._show_login_screen()
            
        except Exception as e:
            from tkinter import messagebox
            self.log(f"Error durante logout: {e}", "ERROR")
            messagebox.showerror("Error", f"No se pudo cerrar sesión: {e}")
            self.root.deiconify()  # Mostrar ventana de nuevo en caso de error
    
    def _show_login_screen(self):
        """Muestra una pantalla de login independiente después de logout."""
        from pal.ui.splash import SplashScreen
        
        # Crear splash de login
        login_splash = SplashScreen(self.root)
        login_splash.title("Iniciar Sesión")
        
        # Configurar como pantalla de login (sin progreso ni loops)
        login_splash.progress.pack_forget()  # Ocultar barra de progreso
        login_splash.progress_value = 100  # Evitar que la animación se reinicie
        
        # Resetear eventos para que NO se cierre automáticamente
        login_splash.minimum_time_elapsed.clear()
        login_splash.app_initialized.clear()
        login_splash.login_success.clear()
        
        # Habilitar login inmediatamente
        def _handle_login(username, password):
            success, message = self._splash_login_submit(username, password)
            if success:
                # Inicializar componentes diferidos después del login exitoso
                self._initialize_post_login_components()
                # Marcar login exitoso y cerrar
                login_splash.login_success.set()
                login_splash.after(100, lambda: [
                    login_splash.destroy(),
                    self.root.deiconify()
                ])
            return success, message
        
        login_splash.enable_login(_handle_login)
        
        # Esperar cierre del splash
        self.root.wait_window(login_splash)
        
        # Si el usuario cerró la ventana sin hacer login, cerrar app
        if not self.current_user:
            self.log("Login cancelado - cerrando aplicación", "INFO")
            self.root.quit()

    class NotificationManager:
        def __init__(self, root):
            self.root = root
        
        def show_success(self, message):
            self._show_notification("✓ Éxito", message, "#d4edda")
        
        def show_error(self, message, details=None):
            # Support both old style (message only) and new style (message, details)
            if details:
                full_message = f"{message}: {details}"
            else:
                full_message = message
            self._show_notification("⚠ Error", full_message, "#f8d7da")
        
        def _show_notification(self, title, message, color):
            notification = tk.Toplevel(self.root)
            notification.wm_overrideredirect(True)
            notification.attributes('-topmost', True)
            notification.geometry(f"+{self.root.winfo_rootx()+self.root.winfo_width()-300}+{self.root.winfo_rooty()+50}")
        
            frame = ttk.Frame(notification, relief="solid", borderwidth=1)
            frame.pack(padx=10, pady=10)
        
            ttk.Label(frame, text=title, foreground="#155724", font=("Segoe UI", 10, "bold")).pack()
            ttk.Label(frame, text=message).pack()
        
            notification.after(3000, notification.destroy)
        
        def show_banner(self, message, bg="#FFB81C", fg="white", duration=5000):
            """Muestra una banda superior no modal y auto-ocultable"""
            try:
                banner = tk.Toplevel(self.root)
                banner.wm_overrideredirect(True)
                banner.attributes('-topmost', True)
                # Dimensiones
                width = min(max(480, int(self.root.winfo_width()*0.6)), 900)
                height = 60
                # Posición centrada arriba
                x = self.root.winfo_rootx() + (self.root.winfo_width() - width)//2
                y = self.root.winfo_rooty() + 10
                banner.geometry(f"{width}x{height}+{x}+{y}")
                # Contenido
                frame = tk.Frame(banner, bg=bg, bd=2, relief="ridge")
                frame.pack(fill=tk.BOTH, expand=True)
                label = tk.Label(
                    frame,
                    text=message,
                    bg=bg,
                    fg=fg,
                    font=("Segoe UI", 12, "bold"),
                    padx=16,
                    pady=10,
                    wraplength=width-40,
                    justify="center"
                )
                label.pack(fill=tk.BOTH, expand=True)
                # Auto-ocultar
                banner.after(max(1500, int(duration)), banner.destroy)
            except Exception:
                # Fallback a messagebox no modal-like (evitar bloqueo): usar _safe_update_api_status
                try:
                    from tkinter import messagebox
                    messagebox.showinfo("Aviso", message)
                except Exception:
                    pass

    class HelpTooltips:
        def __init__(self, root):
            self.root = root
            self.tooltip_window = None
        
        def add_tooltip(self, widget, text):
            widget.bind("<Enter>", lambda e: self.show_tooltip(widget, text) or 0)  # Return 0 for WNDPROC
            widget.bind("<Leave>", lambda e: self.hide_tooltip() or 0)  # Return 0 for WNDPROC
        
        def show_tooltip(self, widget, text):
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + 20
        
            self.tooltip_window = tk.Toplevel(widget)
            self.tooltip_window.wm_overrideredirect(True)
            self.tooltip_window.wm_geometry(f"+{x}+{y}")
        
            label = ttk.Label(self.tooltip_window, text=text, background="#ffffe0",
                        relief="solid", borderwidth=1, padding=5)
            label.pack()
            return 0  # Return integer for WNDPROC
        
        def hide_tooltip(self):
            if self.tooltip_window:
                self.tooltip_window.destroy()
            return 0  # Return integer for WNDPROC



    def toggle_buttons(self, estado: str):
        """Habilitar/deshabilitar botones principales"""
        if not hasattr(self, 'buttons') or not self.buttons:
            return  # Prevenir error si no hay botones registrados
    
        for nombre, boton in self.buttons.items():
            try:
                boton.config(state=estado)
                # Cambiar color para mejor feedback visual
                boton.config(style="TButton" if estado == 'normal' else "Disabled.TButton")
            except Exception:
                pass

    def _set_ui_connected(self, connected: bool):
        """Habilita/deshabilita elementos de UI dependientes de BD y módulos."""
        if not hasattr(self, 'buttons'):
            self.buttons = {}
        # Reglas por módulo requerido
        required_module = {
            'btn_envio_masivo': 'envio_mensajes',
            'btn_programar_envio': 'envio_mensajes',
            'btn_export_csv': 'stock',
            'btn_stock_reload': 'stock',
            'btn_tra_cargar': 'tra',
            'btn_calendar_refresh': 'calendario',
            # CRUD registros no dependen de módulo, solo de BD
            'btn_buscar': None,
            'btn_guardar': None,
            'btn_actualizar': None,
            'btn_eliminar': None,
        }
        for key, btn in self.buttons.items():
            try:
                mod = required_module.get(key)
                enable = connected and (mod is None or self.modules_enabled.get(mod, False))
                btn.config(state='normal' if enable else 'disabled')
                btn.config(style='TButton' if enable else 'Disabled.TButton')
            except Exception:
                continue

        

    def toggle_password(self):
        if self.show_pwd_var.get():
            self.pwd_entry.config(show="")
        else:
            self.pwd_entry.config(show="*")

    def save_connection_settings(self, server: str, database: str, user: str, token: str = None):
        """Guarda la configuración en el archivo .ini"""
        try:
            config = configparser.ConfigParser()
            
            # Si existe el archivo, cargamos su contenido primero
            if os.path.exists(CONFIG_FILE):
                config.read(CONFIG_FILE)
            
            # Crear sección si no existe
            if 'Database' not in config:
                config.add_section('Database')
                
            # Actualizar valores
            config['Database']['server'] = self.cred_manager.encrypt(server)
            config['Database']['database'] = self.cred_manager.encrypt(database)
            config['Database']['user'] = self.cred_manager.encrypt(user)
            
            # Escribir archivo
            with open(CONFIG_FILE, 'w') as configfile:
                config.write(configfile)
            
            # Guardar token si existe
            if token:
                self.cred_manager.store_whatsapp_token(token)
                
            self.show_temp_notification("Configuración guardada ✅")
            
        except Exception as e:
            self.audit_log.log_event(
                "SAVE_CONFIG_FAILED",
                os.getlogin(),
                "FAILED",
                ErrorCode.INVALID_CONFIG
            )
            try:
                if hasattr(self, 'audit_db'):
                    self.audit_db.log_action(
                        accion='SAVE_CONFIG_FAILED', usuario_id=self.current_user['id'] if self.current_user else None, modulo='ADMIN', detalle=str(e), exitoso=False)
            except Exception:
                pass
            messagebox.showerror("Error", f"No se pudo guardar: {str(e)}")
    
    def validate_input(self):
        num = self.num_cliente.get().strip()
        cod = self.cod_producto.get().strip()
        inputs = {
            "Numero": num,
            "Codigo": cod
        }

        if not num or not cod:
            self.audit_log.log_event(
                "VALIDATION_ERROR", os.getlogin(), "FAILED", 
                ErrorCode.INVALID_CLIENT_NUMBER
            )
            messagebox.showwarning("Error", "Complete todos los campos")

            return False

        #Validar formato numerico

        if not re.match(r'^\d{1,11}$', num):
            self.audit_log.log_event(
                "INVALID_CLIENT_NUMBER", os.getlogin(), "FAILED",
                ErrorCode.INVALID_CLIENT_NUMBER
            )
            messagebox.showwarning("Error", str(ErrorCode.INVALID_CLIENT_NUMBER))
            return False
        
        dangerous_patterns = [r";.*--", r"/\*.*\*/", r"xp_", r"exec\(", r"union.*select"]
        for field, value in inputs.items():
            if any(re.search(pattern, value, re.IGNORECASE) for pattern in dangerous_patterns):
                self.audit_log.log_event(
                    f"DANGEROUS_INPUT_{field}", os.getlogin(), "FAILED",
                    ErrorCode.DANGEROUS_INPUT
                )
                self.notification_manager.show_warning("Error", str(ErrorCode.DANGEROUS_INPUT))
                return False
              
        return True 

    def clear_inputs(self):
        self.num_cliente.delete(0, tk.END)
        self.cod_producto.delete(0, tk.END)
        self.actualizar_descripcion("")

    def save_record(self):
        if not self.validate_input():
            return

        try:
            self.db_manager.execute_query(
                "INSERT INTO pal_clientes (numero_cliente, C_CODIGO) VALUES (?, ?)",
                (self.num_cliente.get(), self.cod_producto.get())
            )
            
            self.show_temp_notification("¡Guardado exitosamente!")
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='RECORD_CREATE', usuario_id=self.current_user['id'], modulo='REGISTROS', 
                        detalle=f"numero={self.num_cliente.get()} codigo={self.cod_producto.get()}")
            except Exception:
                pass

            # Restablecer el estado a 'Conectado' después de 3 segundos
            self.root.after(3000, lambda: self.update_status('connected'))

            self.search_records()
            self.clear_inputs()
        except Exception as e:
            self.notification_manager.show_success("Error", str(e))
            self.show_temp_notification("Error al guardar", duration=5000)
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='RECORD_CREATE', usuario_id=self.current_user['id'], modulo='REGISTROS', detalle=str(e), exitoso=False)
            except Exception:
                pass

    def search_records(self):
        # PRIMERO: Verificar que el módulo MENSAJES esté habilitado (donde está REGISTROS)
        # Si no está habilitado, retornar silenciosamente (no es un error del usuario)
        if not self.modules_enabled.get("envio_mensajes", False):
            return
        
        if not self.db_manager.conn:
            self.notification_manager.show_error("Error", "No hay conexión a la base de datos")
            self.show_settings()
            return
        
        # Validar que el widget tree exista y sea accesible
        if not hasattr(self, 'tree') or not self.tree or not self.tree.winfo_exists():
            return
        
        try:
            self.tree.delete(*self.tree.get_children())
            num = self.num_cliente.get().strip()
            cod = self.cod_producto.get().strip()

            query = "SELECT id, numero_cliente, C_CODIGO FROM pal_clientes"
            conditions = []
            params = []
            
            if num:
                conditions.append("numero_cliente LIKE ?")
                params.append(f"%{num}%")
            
            if cod:
                conditions.append("C_CODIGO LIKE ?")
                params.append(f"%{cod}%")

            if conditions:
                query += " WHERE " + " AND ".join(conditions)

            records = self.db_manager.fetch_data(query, params)
            for row in records:
                self.tree.insert("", tk.END, values=row)
                            
            
                
        except Exception as e:
            self.notification_manager.show_error("Error", str(e))
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='RECORD_SEARCH_ERROR', usuario_id=self.current_user['id'], modulo='REGISTROS', detalle=str(e), exitoso=False)
            except Exception:
                pass

    def enviar_a_todos(self):
        self.log("Iniciando proceso de envío masivo...", "INFO")
        if self.enviando: return
        self.toggle_buttons('disabled')
            
    
        try:
            records = self.db_manager.fetch_data("SELECT numero_cliente, C_CODIGO FROM pal_clientes")
            
            # Agrupar clientes
            clientes_con_error = set()

            clientes_dict = {}
            for numero, codigo in records:
                if numero not in clientes_dict:
                    clientes_dict[numero] = []
                clientes_dict[numero].append(codigo)
        
            self.log(f"Enviando a {numero}", "DEBUG")
            self.clientes_lista = list(clientes_dict.items())
            self.total = len(self.clientes_lista)
        
            if not self.clientes_lista:
                messagebox.showinfo("Info", "No hay clientes para notificar")
                return

            self.actual = 0
            self.enviando = True
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='MSG_MASS_START', usuario_id=self.current_user['id'], modulo='MENSAJES')
            except Exception:
                pass
        
            # Configurar UI de progreso
            self.progress = ttk.Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
            self.progress.pack(pady=10)
            self.lbl_progreso = ttk.Label(self.root, text="Preparando...")
            self.lbl_progreso.pack()
            self.progress["maximum"] = self.total

            self.root.after(1000, self.procesar_envio_masivo)
            
        except Exception as e:
            self.log(f"Error en envío masivo: {str(e)}", "ERROR")
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='MSG_MASS_ERROR', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=str(e), exitoso=False)
            except Exception:
                pass
            self.toggle_buttons('normal')
            messagebox.showerror("Error", f"Error obteniendo clientes: {str(e)}")
            self.enviando = False
        
        
        
        
        
    def _background_load_alertas_stock(self):
        """Mantiene compatibilidad pero deshabilitado ya que quiebres se cargan síncronamente."""
        try:
            self.stock_debug_log("Carga paralela de quiebres saltada (datos ya consolidados)")
        except Exception as e:
            self.stock_debug_log(f"Error en monitoreo: {e}")
        finally:
            self._stock_loading_in_progress = False
            self.stock_full_loading_started = False
    
    def _update_ui_after_chunk(self, total_records, chunk_number):
        """Actualiza la UI después de cargar un chunk de datos"""
        try:
            # Evitar errores si la UI fue destruida
            if not hasattr(self, 'root') or not self.root.winfo_exists():
                return
            if not hasattr(self, 'stock_tree') or not self.stock_tree or not self.stock_tree.winfo_exists():
                return
            # Actualizar filtros si es necesario
            self.aplicar_filtro_stock()
            
            # Log de progreso en UI
            estimated_pages = math.ceil(total_records / self.page_size)
            self.log(f"📈 [UI] Chunk {chunk_number}: {total_records} registros | ~{estimated_pages} páginas estimadas", "DEBUG")
            
        except Exception as e:
            self.log(f"Error actualizando UI después del chunk: {e}", "ERROR")
    
    def _finalize_stock_loading(self, final_count):
        """Finaliza el proceso de carga de stock y actualiza estadísticas finales"""
        try:
            # Reconstruir vistas efectivas antes de aplicar filtros
            try:
                self._rebuild_effective_views()
            except Exception:
                pass
            
            # Aplicar filtros finales
            self.aplicar_filtro_stock()
            
            # Calcular páginas totales
            total_pages = math.ceil(final_count / self.page_size)
            
            # Log final con estadísticas
            self.log(
                f"🏆 [FINAL] Carga completa: {final_count} registros | "
                f"{total_pages} páginas totales | "
                f"Tamaño de página: {self.page_size}", 
                "SUCCESS"
            )
            
            # Actualizar controles de paginación si existen
            if hasattr(self, 'pagination_label'):
                current_total_pages = math.ceil(len(self.cached_alertas) / self.page_size)
                self.pagination_label.config(text=f"Página {self.current_page}/{current_total_pages}")
            
        except Exception as e:
            self.log(f"Error finalizando carga de stock: {e}", "ERROR")
    
    def _get_total_stock_count(self):
        """Obtiene el total de registros de alertas disponibles en la BD para debug"""
        try:
            # Obtener depósitos tratables para el conteo de debug
            sedes_config = self.config_manager.get_sedes_config()
            all_deps = []
            for cfg in sedes_config.values():
                all_deps.extend(cfg.get('almacenes_tratables', []))
            
            if not all_deps:
                self.log("No hay depósitos para conteo de debug", "DEBUG")
                return 0
            
            all_deps = list(dict.fromkeys(all_deps))
            placeholders = ','.join(['?' for _ in all_deps])
            
            query = f"""
                SELECT COUNT(*) as total
                FROM (
                    SELECT c_codarticulo
                    FROM MA_DEPOPROD d WITH (NOLOCK)
                        INNER JOIN MA_PRODUCTOS p ON d.c_codarticulo = p.C_CODIGO
                        WHERE c_coddeposito IN ({placeholders})
                        GROUP BY c_codarticulo
                        HAVING SUM(n_cantidad) < 21
                ) as subquery
            """
            result = self.db_manager.fetch_data(query, all_deps)
            return result[0][0] if result and result[0] else 0
        except Exception as e:
            self.log(f"Error obteniendo total de registros: {e}", "ERROR")
            return 0
    
    def _retry_stock_loading(self):
        """Reintenta la carga de stock automáticamente tras un fallo"""
        try:
            # Evitar múltiples reintentos simultáneos
            if getattr(self, '_stock_loading_in_progress', False):
                self.log("🔄 [AUTO-RETRY] Carga ya en progreso, omitiendo reintento", "DEBUG")
                return
                
            self.log("🔄 [AUTO-RETRY] Reintentando carga de stock automáticamente...", "INFO")
            
            # Resetear estado para permitir nueva carga
            self.stock_full_loading_started = False
            self._stock_loading_in_progress = False  # Resetear también el flag de bloqueo
            
            # Limpiar datos parciales
            if len(self.cached_alertas) < 1000:
                self.cached_alertas = []
            
            # Forzar recarga
            self.actualizar_alertas_stock(force_refresh=True)
            
        except Exception as e:
            self.log(f"❌ [AUTO-RETRY] Error en reintento automático: {e}", "ERROR")
    
    def _background_load_ventas_tra(self):
        """Carga ventas TRA en segundo plano con soporte para Nodo Maestro (ICH incluido)"""
        load_start_time = time.perf_counter()
        try:
            # Normalizar sede para persistencia (00/% -> ICH)
            sede_p = self.tra_sede_codigo
            if sede_p in ('00', '%', 'ALL'):
                sede_p = 'ICH'
                
            # ATAJO CRÍTICO: Si ya hay persistencia, no calcular nada
            from pal.services.tra import get_persisted_rotation
            dias_context = (self.tra_fecha_fin - self.tra_fecha_inicio).days
            
            # Verificar si existe persistencia y si está fresca (TTL de 1 hora)
            # Nota: get_persisted_rotation devuelve el objeto con el timestamp para verificar frescura
            # IMPORTANTE: Si es reporte masivo, NO usar persistencia (solo tiene rotación)
            persisted_info = get_persisted_rotation(self.db_manager, sede=sede_p, dias_rango=dias_context) if not getattr(self, 'tra_include_zero_sales', False) else None
            if persisted_info:
                self.log(f"🚀 TRA: Usando datos de rotación persistidos para '{sede_p}' (Rango: {dias_context}d)", "SUCCESS")
                datos = self.db_manager.obtener_ventas_persisted_tra(sede=sede_p, dias_rango=dias_context)
                if datos:
                    self.cached_ventas_tra = [tuple(r) for r in datos]
                    # Simular resumen para _ui_finish
                    total_neto = sum(float(r[5] or 0) for r in self.cached_ventas_tra)
                    self.tra_total_neto_scaneado = total_neto
                    
                    def _ui_fast_finish():
                        self._rebuild_effective_views()
                        self.aplicar_filtro_tra()
                        self.log(f"Carga instantánea TRA completada: {len(self.cached_ventas_tra)} registros", "SUCCESS")
                        self.api_status.config(text="API: Lista", foreground="green")
                        self.global_progress.stop()
                        self.global_progress.pack_forget()
                    
                    self.root.after(0, _ui_fast_finish)
                    self.tra_debug_log("🚀 [TRA] Datos persistidos cargados con éxito, finalizando hilo de carga.")
                    return # SALIR DE LA FUNCIÓN, NO CALCULAR
            
            # Si no hay persistencia o no está fresca, proceder con el cálculo paralelo
            self.log(f"🔄 TRA: Iniciando cálculo paralelo para '{sede_p}' (Rango: {dias_context}d)...", "INFO")
            
            # Parámetros de chunk adaptativo
            chunk_size = 1000
            min_chunk = 200
            max_chunk = 5000
            target_ms = 150.0

            # Acumulador de neto escaneado para debug
            total_neto_scaneado = 0.0

            start = 1
            processed = 0
            ui_last_refreshed_at = 0
            chunk_index = 0
            first_refresh_done = False

            # Garantizar contenedor de datos
            if not hasattr(self, 'cached_ventas_tra') or self.cached_ventas_tra is None:
                self.cached_ventas_tra = []
            # Conjunto para evitar duplicados por código
            try:
                seen_codes = set(str(r[0]) for r in self.cached_ventas_tra)
            except Exception:
                seen_codes = set()

            self.tra_debug_log("Iniciando carga completa de datos TRA en background (adaptativa)...")

            # Obtener lista de almacenes tratables (Nuevo Motor)
            sedes_list = self.config_manager.get_tratables_by_sede(self.tra_sede_codigo)

            while True:
                t0 = time.perf_counter()
                rows = self.db_manager.obtener_ventas_por_producto_chunk(
                    fecha_inicio=self.tra_fecha_inicio,
                    fecha_fin=self.tra_fecha_fin,
                    sede_codigo=sedes_list,
                    start_row=start,
                    fetch_size=int(chunk_size),
                    include_zero_sales=getattr(self, 'tra_include_zero_sales', False),
                    exclude_depts=getattr(self, 'excluded_depts', [])
                )
                dt_ms = (time.perf_counter() - t0) * 1000.0

                if not rows:
                    break

                # Agregar filas sin clasificar (6 campos), evitando duplicados por código
                try:
                    nuevos = []
                    sum_neto_nuevos = 0.0
                    for r in rows:
                        try:
                            cod = str(r[0])
                        except Exception:
                            cod = None
                        if not cod or cod in seen_codes:
                            continue
                        seen_codes.add(cod)
                        nuevos.append(r)
                        # Sumar neto solo de nuevos registros
                        try:
                            sum_neto_nuevos += float(r[5] or 0)
                        except Exception:
                            pass
                    if nuevos:
                        self.cached_ventas_tra.extend(nuevos)
                        total_neto_scaneado += sum_neto_nuevos
                except Exception:
                    # Si falla por concurrencia (raro), recrear lista con filtro de duplicados
                    cleaned = []
                    sum_neto_cleaned = 0.0
                    for r in rows:
                        try:
                            cod = str(r[0])
                        except Exception:
                            cod = None
                        if not cod or cod in seen_codes:
                            continue
                        seen_codes.add(cod)
                        cleaned.append(r)
                        try:
                            sum_neto_cleaned += float(r[5] or 0)
                        except Exception:
                            pass
                    self.cached_ventas_tra = list(self.cached_ventas_tra) + cleaned
                    total_neto_scaneado += sum_neto_cleaned

                processed += len(rows)
                chunk_index += 1
                start += int(chunk_size)

                # Ajustar chunk de forma proporcional, limitado a rangos
                try:
                    factor = target_ms / max(1.0, dt_ms)
                    factor = max(0.5, min(2.0, factor))
                    chunk_size = int(max(min_chunk, min(max_chunk, chunk_size * factor)))
                except Exception:
                    pass

                # Actualización mínima de UI: siempre al primer chunk, luego cada ~3000 filas
                def _ui_update_slim(total=processed, last= len(rows), next_size=int(chunk_size)):
                    try:
                        self.tra_debug_log(f"+{last} (total {total}) en {dt_ms:.0f}ms, next_chunk={next_size}")
                        # Actualizar estado textual
                        try:
                            self.api_status.config(text=f"RI: Cargando {total} filas...", foreground="#004C97")
                        except Exception:
                            pass
                        # Primer chunk: refrescar vistas para feedback temprano
                        nonlocal first_refresh_done
                        if not first_refresh_done:
                            first_refresh_done = True
                            self.aplicar_filtro_tra()
                        else:
                            # Refrescar de forma más espaciada
                            if total - _ui_update_slim.last_shown >= 3000:
                                _ui_update_slim.last_shown = total
                                self.aplicar_filtro_tra()
                    except Exception:
                        pass
                # stateful attr on function
                if not hasattr(_ui_update_slim, 'last_shown'):
                    _ui_update_slim.last_shown = 0
                self.root.after(0, _ui_update_slim)

                # Pequeña pausa cooperativa (evitar saturar BD/CPU)
                time.sleep(0.05)

            # Clasificación de rotación al finalizar (fuera de UI)
            if self.cached_ventas_tra:
                try:
                    from pal.services.tra import clasificar_rotacion
                    raw_snapshot = list(self.cached_ventas_tra)
                    self.tra_debug_log(f"Clasificando rotación para {len(raw_snapshot)} registros...")
                    classified = clasificar_rotacion(raw_snapshot)
                    self.cached_ventas_tra = classified
                except Exception as e:
                    self.log(f"Error clasificando rotación TRA: {e}", "ERROR")

            # Guardar total de neto escaneado en atributo para inspección/debug
            try:
                self.tra_total_neto_scaneado = float(total_neto_scaneado)
            except Exception:
                self.tra_total_neto_scaneado = 0.0

            def _ui_finish(total=len(self.cached_ventas_tra) if self.cached_ventas_tra else 0, neto=self.tra_total_neto_scaneado, scanned=processed):
                try:
                    # Reconstruir vistas efectivas para TRA
                    try:
                        self._rebuild_effective_views()
                    except Exception:
                        pass
                    self.aplicar_filtro_tra()
                    # Log de cierre con resumen de escaneo
                    self.tra_debug_log(f"TRA: Escaneadas {scanned} filas | Neto total escaneado: {neto:.2f}")
                    self.log(f"Carga paralela de ventas TRA completada: {total} registros | Neto total escaneado: {neto:.2f}", "SUCCESS")
                    
                    # GUARDAR PERSISTENCIA (Actuar como Nodo Maestro) - En hilo separado para evitar freeze
                    # IMPORTANTE: Si es reporte masivo, NO guardar en la base de datos (evita inflar tabla con ventas 0)
                    if not getattr(self, 'tra_include_zero_sales', False):
                        def _bg_save_persistence_parallel():
                            try:
                                from pal.services.tra import save_rotation_persistence
                                user_nodo = self.current_user['username'] if self.current_user else "SISTEMA"
                                # Obtener fechas del objeto si están disponibles
                                f_inicio = getattr(self, 'tra_fecha_inicio', datetime.now())
                                f_fin = getattr(self, 'tra_fecha_fin', datetime.now())
                                dias = (f_fin - f_inicio).days or 365
                                
                                # Normalizar sede para persistencia (00/% -> ICH)
                                sede_save = self.tra_sede_codigo
                                if sede_save in ('00', '%', 'ALL'):
                                    sede_save = 'ICH'
                                    
                                save_rotation_persistence(self.db_manager, self.cached_ventas_tra, sede_save, dias, user_nodo)
                            except Exception as e:
                                self.log(f"Error guardando persistencia RI: {e}", "DEBUG")
                        
                        import threading
                        threading.Thread(target=_bg_save_persistence_parallel, daemon=True, name="RI_Persistence_Parallel").start()
                    else:
                        self.log("ℹ️ RI: Persistencia omitida por Reporte Masivo (Ventas 0)", "DEBUG")

                    try:
                        self.api_status.config(text="API: Lista", foreground="green")
                    except Exception:
                        pass
                    try:
                        self.global_progress.stop()
                        self.global_progress.pack_forget()
                    except Exception:
                        pass
                except Exception:
                    pass
            self.root.after(0, _ui_finish)
        except Exception as e:
            self.log(f"Error en carga paralela de ventas TRA: {e}", "ERROR")
            try:
                self.root.after(0, lambda: (self.global_progress.stop(), self.global_progress.pack_forget()))
            except Exception:
                pass
    def delete_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Seleccione un registro para eliminar")
            return

        if messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este registro?"):
            try:
 
                record_id_raw = self.tree.item(selected[0])['values'][0]
                print(f"Valor obtenido del Treeview: {record_id_raw} - Tipo: {type(record_id_raw)}")

                

                record_id_clean = re.sub(r"[^\d]", "", str(record_id_raw))  # Elimina todo excepto números

                if not record_id_clean.isdigit():

                   raise ValueError(f"El ID no es un número válido: {record_id_raw}")

                record_id = int(record_id_clean)  # Convertimos a entero limpio
                print(f"ID después de limpieza: {record_id}")

                self.audit_log.log_event(
                    action=f"DELETE_ID:{record_id}",
                    user=os.getlogin(),
                    status="SUCCESS" 
                )

                # Restablecer el estado a 'Conectado' después de 3 segundos
                self.root.after(3000, lambda: self.update_status('connected'))

                self.db_manager.execute_query("DELETE FROM pal_clientes WHERE id = ?", (record_id,))
                self.update_status('action', message="Registro eliminado")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='RECORD_DELETE', usuario_id=self.current_user['id'], modulo='REGISTROS', detalle=f"id={record_id}")
                except Exception:
                    pass
                self.search_records()
                self.clear_inputs() 
            except Exception as e:
                messagebox.showerror("Error", str(e))
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='RECORD_DELETE', usuario_id=self.current_user['id'], modulo='REGISTROS', detalle=str(e), exitoso=False)
                except Exception:
                    pass

    def crear_botones_masivos(self):
        frame_masivo = ttk.Frame(self.root)
        frame_masivo.pack(pady=10)
    
        self.btn_masivo = ttk.Button(frame_masivo, 
                               text="▶ Iniciar envío masivo", 
                               command=self.toggle_envio_masivo)
        self.btn_masivo.pack(side=tk.LEFT, padx=5)
    
        self.lbl_progreso = ttk.Label(frame_masivo, text="Envíos: 0/0")
        self.lbl_progreso.pack(side=tk.LEFT, padx=5)
    
        self.progress = ttk.Progressbar(frame_masivo, 
                                  orient=tk.HORIZONTAL, 
                                  length=200, 
                                  mode='determinate')

    def update_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Error", "Seleccione un registro para actualizar")
            return

        if not self.validate_input():
            return

        try:
            record_id_raw = self.tree.item(selected[0])['values'][0]
            print(f"valor obtenido del Treeview: {record_id_raw} - Tipo: {type(record_id_raw)}")

            record_id_clean = re.sub(r"[^\d]", "", str(record_id_raw))  # Elimina todo excepto números
            
            if not record_id_clean.isdigit():
                raise ValueError(f"El ID no es un número válido: {record_id_raw}")
            
            record_id = int(record_id_clean)  # Convertimos a entero limpio
            print(f"ID después de limpieza: {record_id}")
        

            self.db_manager.execute_query(
                "UPDATE pal_clientes SET numero_cliente = ?, C_CODIGO = ? WHERE id = ?",
                (self.num_cliente.get(), self.cod_producto.get(), record_id)
            )
            self.update_status('action', message="Registro actualizado")
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='RECORD_UPDATE', usuario_id=self.current_user['id'], modulo='REGISTROS', 
                        detalle=f"id={record_id}")
            except Exception:
                pass

            # Restablecer el estado a 'Conectado' después de 3 segundos
            self.root.after(3000, lambda: self.update_status('connected'))

            self.search_records()
            self.clear_inputs()
        except Exception as e:
            messagebox.showerror("Error", str(e))
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='RECORD_UPDATE', usuario_id=self.current_user['id'], modulo='REGISTROS', detalle=str(e), exitoso=False)
            except Exception:
                pass

    def buscar_descripcion(self, event=None):
        """Obtiene la descripción y actualiza el campo correspondiente"""
        # Ensure return value for WNDPROC
        codigo = self.cod_producto.get().strip()
        clean_codigo = re.sub(r'\D', '', codigo)  # Limpieza del código
    
        try:
            if not clean_codigo:
                self.actualizar_descripcion("Código no ingresado")
                return
            # Actualizar campo de código con versión limpia
            if clean_codigo != codigo:
                self.cod_producto.delete(0, tk.END)
                self.cod_producto.insert(0, clean_codigo)
        
            # Usar método reutilizable
            descripcion = self.obtener_descripcion_producto(clean_codigo)
            if descripcion:
                self.actualizar_descripcion(descripcion)
                print(f"Descripción: {descripcion}")
            else:
                self.actualizar_descripcion("Descripción no encontrada")

            # Validar stock usando el método reutilizable
            if self.validar_stock_producto(clean_codigo):
                stock_result = self.db_manager.fetch_data(
                    "SELECT n_cantidad FROM MA_DEPOPROD WHERE c_codarticulo = ? AND c_coddeposito = '0301'",
                    (clean_codigo,)
                )
                cantidad = int(stock_result[0][0]) if stock_result else 0
                self.show_temp_notification(f"Stock disponible: {cantidad} unidades ✅")
            else:
                self.show_temp_notification("No disponible o igual a 0 ❌")
            
        except Exception as e:
            self.actualizar_descripcion("Error en consulta")
            messagebox.showerror("Error", f"Error obteniendo datos: {str(e)}")
        return 0  # Return integer for WNDPROC
    
                       
    def _records_on_click(self, event):
        """Selecciona fila en el Treeview de Registros con un solo click"""
        try:
            if not hasattr(self, 'tree') or not self.tree or not self.tree.winfo_exists():
                return 0
            region = self.tree.identify_region(event.x, event.y)
            if region in ('cell', 'tree'):
                item = self.tree.identify_row(event.y)
                if item:
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
        except Exception:
            pass
        return 0

    def on_tree_double_click(self, event):
        """Manejo de doble click en la tabla con limpieza de datos"""
        # Return 0 at end for WNDPROC
        selected = self.tree.selection()
        if selected:
            try:
                # Obtener y limpiar datos del registro seleccionado
                values = self.tree.item(selected[0])['values']
                
                # Limpiar número de cliente
                clean_numero = re.sub(r'\D', '', str(values[1]))
                self.num_cliente.delete(0, tk.END)
                self.num_cliente.insert(0, clean_numero)
                
                # Limpiar código de producto
                clean_codigo = re.sub(r'\D', '', str(values[2]))
                self.cod_producto.delete(0, tk.END)
                self.cod_producto.insert(0, clean_codigo)
                
                # Actualizar descripción automáticamente
                self.buscar_descripcion()
                
            except IndexError:
                messagebox.showerror("Error", "El registro seleccionado es inválido")
            except Exception as e:
                messagebox.showerror("Error", f"Error al cargar datos: {str(e)}")
        return 0  # Return integer for WNDPROC
    

    def mostrar_ventana_programacion(self):
        try:
            win = tk.Toplevel(self.root)
            win.title("Programar Envío")
            win.transient(self.root)
            win.grab_set()
            win.resizable(False, False)
            frm = ttk.Frame(win, padding=10)
            frm.pack(fill=tk.BOTH, expand=True)

            # Número de cliente
            ttk.Label(frm, text="Número Cliente:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
            num_entry = ttk.Entry(frm, width=20)
            num_entry.grid(row=0, column=1, sticky='ew', padx=5, pady=5)

            # Fecha (Calendar) + Hora/Minuto (Spinbox)
            ttk.Label(frm, text="Fecha:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
            from tkcalendar import Calendar
            from datetime import datetime, timedelta
            default_date = datetime.now() + timedelta(minutes=2)
            cal = Calendar(frm, selectmode='day', year=default_date.year, month=default_date.month, day=default_date.day)
            cal.grid(row=1, column=1, sticky='w', padx=5, pady=5)

            ttk.Label(frm, text="Hora:").grid(row=1, column=2, sticky='e', padx=(15,5))
            spin_hora = ttk.Spinbox(frm, from_=0, to=23, width=4)
            spin_hora.insert(0, f"{default_date.hour:02d}")
            spin_hora.grid(row=1, column=3, sticky='w')
            ttk.Label(frm, text=":").grid(row=1, column=4, sticky='w')
            spin_minuto = ttk.Spinbox(frm, from_=0, to=59, width=4)
            spin_minuto.insert(0, f"{default_date.minute:02d}")
            spin_minuto.grid(row=1, column=5, sticky='w', padx=(0,5))

            # Tipo de envío
            ttk.Label(frm, text="Tipo:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
            tipo_var = tk.StringVar(value='DISPONIBILIDAD')
            tipo_combo = ttk.Combobox(frm, textvariable=tipo_var, values=['DISPONIBILIDAD', 'ENTREGA'], state='readonly', width=18)
            tipo_combo.grid(row=2, column=1, sticky='w', padx=5, pady=5)

            # Código de producto (solo para DISPONIBILIDAD)
            ttk.Label(frm, text="Código Producto:").grid(row=3, column=0, sticky='w', padx=5, pady=5)
            cod_entry = ttk.Entry(frm, width=20)
            cod_entry.grid(row=3, column=1, sticky='w', padx=5, pady=5)

            def _toggle_codigo(*args):
                if tipo_var.get() == 'DISPONIBILIDAD':
                    cod_entry.config(state='normal')
                else:
                    cod_entry.delete(0, tk.END)
                    cod_entry.config(state='disabled')
            try:
                tipo_var.trace_add('write', lambda *a: _toggle_codigo())
            except Exception:
                pass
            _toggle_codigo()

            def confirmar():
                from datetime import datetime
                numero = num_entry.get().strip()
                if not numero:
                    messagebox.showwarning("Error", "Ingrese el número de cliente")
                    return
                try:
                    # Obtener fecha del Calendar de forma robusta
                    try:
                        d = cal.selection_get()  # datetime.date
                    except Exception:
                        d = cal.get_date()
                        d = datetime.fromisoformat(str(d)).date()
                    h = int(spin_hora.get() or 0)
                    m = int(spin_minuto.get() or 0)
                    if not (0 <= h <= 23 and 0 <= m <= 59):
                        raise ValueError
                    fecha = datetime(d.year, d.month, d.day, h, m)
                except Exception:
                    messagebox.showwarning("Error", "Fecha/Hora inválida")
                    return

                tipo = tipo_var.get().strip().upper() or 'DISPONIBILIDAD'
                codigo = (cod_entry.get().strip() or None)
                if tipo == 'DISPONIBILIDAD' and not codigo:
                    messagebox.showwarning("Error", "Ingrese el código de producto para DISPONIBILIDAD")
                    return

                # Programar
                try:
                    if not hasattr(self, 'envios_programados'):
                        from pal.services.envios import EnvioProgramado
                        self.envios_programados = EnvioProgramado(self.db_manager)
                    ok = self.envios_programados.programar_envio(numero, fecha, tipo_envio=tipo, codigo_producto=codigo)
                    if ok:
                        messagebox.showinfo("Listo", "Envío programado")
                        try:
                            if hasattr(self, 'audit_db') and self.current_user:
                                self.audit_db.log_action(accion='MSG_SCHEDULE', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=f"cliente={numero} fecha={fecha} tipo={tipo}")
                        except Exception:
                            pass
                        win.destroy()
                    else:
                        messagebox.showerror("Error", "No se pudo programar el envío")
                except Exception as e:
                    messagebox.showerror("Error", f"Fallo programando: {e}")

            btns = ttk.Frame(frm)
            btns.grid(row=4, column=0, columnspan=6, sticky='e', pady=10)
            ttk.Button(btns, text="Programar", command=confirmar).pack(side=tk.RIGHT, padx=5)
            ttk.Button(btns, text="Cancelar", command=win.destroy).pack(side=tk.RIGHT)
            frm.columnconfigure(1, weight=1)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el programador: {e}")

    def notificar(self):

        try:
            codigo = self.cod_producto.get().strip()
            clean_codigo = re.sub(r'\D', '', codigo)  # Limpieza del código
        
            if not clean_codigo:
                messagebox.showwarning("Error", "Código no ingresado")
                return
        
            # Actualizar campo de código con versión limpia
            if clean_codigo != codigo:
                self.cod_producto.delete(0, tk.END)
                self.cod_producto.insert(0, clean_codigo)
    
            # Consultar cantidad
            query_cantidad = "SELECT n_cantidad FROM dbo.MA_DEPOPROD WHERE c_codarticulo = ? AND c_coddeposito = '0301'"
            result_cantidad = self.db_manager.fetch_data(query_cantidad, (clean_codigo,))
            if not result_cantidad or result_cantidad[0][0] <= 0:
                self.show_temp_notification("Cantidad no disponible o igual a 0")
                return
            
            # Consultar descripción
            query_descripcion = "SELECT COALESCE(cu_descripcion_corta, 'SIN DESCRIPCIÓN') FROM dbo.MA_PRODUCTOS WHERE C_CODIGO = ?"
            result_descripcion = self.db_manager.fetch_data(query_descripcion, (clean_codigo,))
            if not result_descripcion:
                messagebox.showinfo("Error", "Descripción no encontrada")
                return
        
            descripcion = result_descripcion[0][0]
    
            # Consultar número de cliente
            query_numero_cliente = "SELECT numero_cliente FROM pal_clientes WHERE C_CODIGO = ?"
            result_numero_cliente = self.db_manager.fetch_data(query_numero_cliente, (clean_codigo,))
            if not result_numero_cliente:
                messagebox.showinfo("Error", "Número de cliente no encontrado")
                return
        
            numero_cliente = result_numero_cliente[0][0]

            self.enviar_mensaje_whatsapp(numero_cliente, [descripcion])
            self.show_temp_notification("Enviado Exitosamente")
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='MSG_SEND', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=f"cliente={numero_cliente}")
            except Exception:
                pass
    
        except Exception as e:
            messagebox.showerror("Error", f"Error obteniendo descripción o cantidad: {str(e)}")
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='MSG_SEND', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=str(e), exitoso=False)
            except Exception:
                pass

    def procesar_envio_programado(self, id_envio, numero_cliente):
        """Envía un mensaje programado usando plantillas de WhatsApp y actualiza el estado"""
        try:
            # 1. Obtener datos extendidos del envío
            envio_data = self.db_manager.fetch_data(
                "SELECT tipo_envio, codigo_producto FROM pal_envios_programados WHERE id = ?", 
                (id_envio,)
            )
            
            if not envio_data:
                self.log(f"Envío {id_envio}: no encontrado", "ERROR")
                return False
            else:
                tipo_envio, codigo_producto = envio_data[0]

            tipo_envio, codigo_producto = envio_data[0]
            productos = None

            # 2. Configurar plantillas según tipo de envío
            if tipo_envio == "DISPONIBILIDAD": 

                # Validar datos del producto
                if not codigo_producto:
                    self.log(f"Envío {id_envio}: Sin código de producto", "ERROR")
                    return False
                
                # Obtener descripción del producto
                descripcion = self.db_manager.fetch_data(
                    "SELECT COALESCE(cu_descripcion_corta, 'SIN DESCRIPCIÓN') FROM MA_PRODUCTOS WHERE C_CODIGO = ?",
                    (codigo_producto,)
                )
            
                if not descripcion:
                    self.log(f"Producto {codigo_producto} no encontrado", "ERROR")
                    return False

                productos = [descripcion[0][0]]  # Adaptado al parámetro "productos" del método

            elif tipo_envio == "ENTREGA":
                pass  # Nombre exacto de tu plantilla en Meta

            # 3. Envío para tipos conocidos
            if self.enviar_mensaje_whatsapp(
                numero_cliente=numero_cliente,
                productos=productos,
                tipo_envio=tipo_envio
            ):
                self.db_manager.execute_query(
                    "UPDATE pal_envios_programados SET estado = 'ENVIADO' WHERE id = ?",
                    (id_envio,)
                )
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='MSG_SCHEDULED_SENT', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=f"id={id_envio}")
                except Exception:
                    pass
                return True

            return False

        except Exception as e:
            self.log(f"Error en envío {id_envio}: {str(e)}", "ERROR")
            self.audit_log.log_event(
                "ENVIO_PROGRAMADO_ERROR",
                os.getlogin(),
                "FAILED",
                error_code=ErrorCode.WHATSAPP_API_FAILURE
            )
            try:
                if hasattr(self, 'audit_db') and self.current_user:
                    self.audit_db.log_action(
                        accion='MSG_SCHEDULED_ERROR', usuario_id=self.current_user['id'], modulo='MENSAJES', detalle=str(e), exitoso=False)
            except Exception:
                pass
            return False
        finally:
            self.log(f"Finalizado el procesamiento del envío {id_envio} para el cliente {numero_cliente}", "INFO")

    def procesar_envio_masivo(self,):

        try:
            if self.actual == 0:
                # Crear tabla temporal si no existe
                self.db_manager.execute_query("""
                IF NOT EXISTS (SELECT * FROM sys.tables WHERE name = 'pal_temp_envio')
                CREATE TABLE pal_temp_envio (
                    numero_cliente NVARCHAR(50),
                    codigo_producto NVARCHAR(15),  -- Nuevo campo
                    descripcion NVARCHAR(255),
                    timestamp DATETIME DEFAULT GETDATE()  -- Nuevo campo
                    )
                """)
        
            # Insertar datos en la tabla temporal antes del envío
                for numero, codigos in self.clientes_lista:
                    for codigo in codigos:
                        cantidad_result = self.db_manager.fetch_data(
                            "SELECT n_cantidad FROM MA_DEPOPROD WHERE c_codarticulo = ? AND c_coddeposito = '0301'",
                            (codigo,)
                        )
                        if cantidad_result and cantidad_result[0][0] > 0:
                            desc_result = self.db_manager.fetch_data(
                                "SELECT C_DESCRI FROM MA_PRODUCTOS WHERE C_CODIGO = ?",
                                (codigo,)
                            )
                            if desc_result:
                                self.db_manager.execute_query(
                                    "INSERT INTO pal_temp_envio (numero_cliente, descripcion) VALUES (?, ?)",
                                    (numero, desc_result[0][0])
                                )
        finally:
            if self.actual >= self.total or self.enviando == False:
                self.toggle_buttons('normal')
                self.enviando = False
                if hasattr(self, 'progress') and self.progress and self.progress.winfo_exists():
                    self.progress.destroy()
                if hasattr(self, 'lbl_progreso') and self.lbl_progreso and self.lbl_progreso.winfo_exists():
                    self.lbl_progreso.destroy()
                self.log("Proceso de envío completado", "SUCCESS")
                try:
                    if hasattr(self, 'audit_db') and self.current_user:
                        self.audit_db.log_action(
                            accion='MSG_MASS_COMPLETE', usuario_id=self.current_user['id'], modulo='MENSAJES')
                except Exception:
                    pass

        numero = self.clientes_lista[self.actual][0]
    
        # Obtener todos los productos del cliente desde la tabla temporal
        productos_result = self.db_manager.fetch_data(
            "SELECT descripcion FROM pal_temp_envio WHERE numero_cliente = ?", (numero,)
        )
        productos = [row[0] for row in productos_result]
    
        # Registrar en logs todas las descripciones obtenidas
        self.audit_log.log_event("ENVIO_MASIVO_LOG", os.getlogin(), f"Cliente: {numero}, Productos: {productos}")
    
        if productos:
            self.enviar_mensaje_whatsapp(numero, productos)  # Enviar como un solo mensaje
            mensaje_estado = f"Enviando {self.actual + 1}/{self.total} | Cliente: {numero}"
        else:
            mensaje_estado = f"Omitido {self.actual + 1}/{self.total} | Sin stock: {numero}"
    
        # Eliminar los productos del cliente de la tabla temporal después de enviar el mensaje
        self.db_manager.execute_query("DELETE FROM pal_temp_envio WHERE numero_cliente = ?", (numero,))
    
        # Actualizar progreso
        self.actual += 1
        self.progress["value"] = self.actual
        self.lbl_progreso.config(text=mensaje_estado)
        self.root.after(7000, self.procesar_envio_masivo)

    def enviar_mensaje_whatsapp(self, numero_cliente: str, productos: list = None, tipo_envio: str = None) -> bool:
        """Envía mensaje por WhatsApp usando plantilla con lista de productos"""
        try:
            # Determinar plantilla según el tipo de envío
            if tipo_envio == "ENTREGA":
                template_name = "recordatorio_entrega"
                texto_plantilla = "📦 Tu entrega está programada para mañana. ¡Estaremos listos!"
            elif tipo_envio == "DISPONIBILIDAD":
                template_name = "sede"
                texto_plantilla = "🏪 ¡Tu producto ya está disponible en nuestra sede! Visítanos."
            else:
                # Lógica para notificaciones de stock
                template_name = "alerta_stock"
                
                # Validar productos para stock
                if productos is None:
                    codigo = self.obtener_codigo_producto_cliente(numero_cliente)
                    if not codigo:
                        self.log(f"Cliente {numero_cliente} sin producto asociado", "ERROR")
                        return False
                    
                    descripcion = self.obtener_descripcion_producto(codigo)
                    productos = [descripcion] if descripcion else []
    
                # Procesar lista de productos solo para stock
                MAX_ITEMS = 10
                MAX_LENGTH = 45
                SEPARADOR = " • "
                
                items_procesados = []
                for idx, producto in enumerate(productos[:MAX_ITEMS], 1):
                    producto_limpio = re.sub(r'[\n\t]', ' ', str(producto))
                    if len(producto_limpio) > MAX_LENGTH:
                        producto_limpio = producto_limpio[:MAX_LENGTH-3] + "..."
                    items_procesados.append(f"{idx}. {producto_limpio}")
                
                texto_plantilla = SEPARADOR.join(items_procesados)
                if len(productos) > MAX_ITEMS:
                    texto_plantilla += f"{SEPARADOR}... (+{len(productos) - MAX_ITEMS} productos más)"
    
            raw = numero_cliente or ""
            digits = re.sub(r'\D', '', raw)  # solo dígitos

            # Prefijos nacionales válidos
            prefijos = ('0412', '0424', '0414', '0416', '0426')

            if digits.startswith('0'):
                # Caso local: debe ser 11 dígitos y prefijo en la lista
                if not (len(digits) == 11 and digits.startswith(prefijos)):
                    self.log(f"Número local inválido, omitiendo: {raw}", "ERROR")
                    return False
                # Convertir 0XXXXXXXXXX → 58XXXXXXXXXX
                numero_formateado = '58' + digits[1:]
            else:
                # Caso internacional: dejar tal cual
                numero_formateado = digits

            # Validar token
            whatsapp_token = self.cred_manager.get_whatsapp_token()
            if not whatsapp_token:
                self.update_status('api_error', message="Token no configurado")
                self.audit_log.log_event(
                    "WHATSAPP_API_ERROR", 
                    os.getlogin(), 
                    "FAILED",
                    ErrorCode.INVALID_API_TOKEN
                )
                messagebox.showerror("Error", str(ErrorCode.INVALID_API_TOKEN))
                self.show_settings()
                return False
    
    
            # Configurar payload
            payload = {
                "messaging_product": "whatsapp",
                "to": numero_formateado,
                "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": "es"},
                    "components": [{
                        "type": "body",
                        "parameters": [{"type": "text", "text": texto_plantilla}]
                    }]
                }
            }
    
            # Enviar solicitud
            response = requests.post(
                "https://graph.facebook.com/v21.0/490677417472051/messages",
                headers={"Authorization": f"Bearer {whatsapp_token}", "Content-Type": "application/json"},
                json=payload
            )
    
            # Manejar respuesta
            if response.status_code == 200:
                self.log(f"Mensaje enviado a {numero_formateado}", "SUCCESS")
                self.audit_log.log_event(
                    "ENVIO_EXITOSO",
                    os.getlogin(),
                    f"Cliente: {numero_formateado} | Tipo: {tipo_envio or 'Stock'}"
                )
                return True
            
            # Manejar errores de API
            error_data = response.json().get('error', {})
            error_msg = error_data.get('message', 'Error desconocido')
            self.log(f"Error API: {error_msg}", "ERROR")
            messagebox.showerror("Error API", f"{error_msg}")
            return False
    
        except Exception as e:
            self.log(f"Error crítico: {str(e)}", "ERROR")
            self.audit_log.log_event(
                "ERROR_ENVIO",
                os.getlogin(),
                "CRITICAL",
                error_code=ErrorCode.WHATSAPP_API_FAILURE
            )
            return False
        
    def _save_modules_config(self):
        # Tomar valores de los BooleanVar y guardar en ini
        new_cfg = {k: v.get() for k, v in self.mod_vars.items()}
        save_modules_config(new_cfg)
        messagebox.showinfo(
            "Módulos Actualizados",
            "Los cambios se guardaron correctamente.\nReinicie la aplicación para que tengan efecto."
        )

    def actualizar_descripcion(self, texto: str):
        self.descripcion.config(state='normal')
        self.descripcion.delete(0, tk.END)
        self.descripcion.insert(0, texto)
        self.descripcion.config(state='readonly')
        self.descripcion.update_idletasks()

    # Helpers de Registros
    def obtener_descripcion_producto(self, codigo: str) -> str | None:
        try:
            rows = self.db_manager.fetch_data(
                "SELECT COALESCE(cu_descripcion_corta, 'SIN DESCRIPCIÓN') FROM dbo.MA_PRODUCTOS WHERE C_CODIGO = ?",
                (str(codigo),)
            )
            if rows and rows[0] and rows[0][0]:
                return str(rows[0][0])
            return None
        except Exception:
            return None

    def validar_stock_producto(self, codigo: str) -> bool:
        try:
            rows = self.db_manager.fetch_data(
                "SELECT ISNULL(n_cantidad,0) FROM dbo.MA_DEPOPROD WHERE c_codarticulo = ? AND c_coddeposito = '0301'",
                (str(codigo),)
            )
            cantidad = int(rows[0][0]) if rows and rows[0] else 0
            return cantidad > 0
        except Exception:
            return False

    # ====================
    # MBRP (Baja Rotación)
    # ====================
    def cargar_mbrp_base(self):
        try:
            # Validar rango
            fecha_inicio = self.mbrp_fecha_inicio_entry.get_date()
            fecha_fin = self.mbrp_fecha_fin_entry.get_date()
            sede = (self.mbrp_sede_var.get() or '').split(' - ')[0]
            
            # Log básico siempre visible al iniciar carga MBRP
            self.log(f"[MBRP] Iniciando carga: {fecha_inicio} a {fecha_fin}, Sede: {sede}", "INFO")

            # Evitar cargas simultáneas
            if getattr(self, 'mbrp_loader_thread', None) is not None and self.mbrp_loader_thread.is_alive():
                self.log("MBRP: Carga en curso; ignorando clic", "WARNING")
                return

            # Reset datos y cachés
            self.cached_ventas_mbrp = []
            self.mbrp_fecha_inicio = fecha_inicio
            self.mbrp_fecha_fin = fecha_fin
            # Mapear ICH a consulta global
            if sede == '00':
                sede = '%'
            self.mbrp_sede_codigo = sede
            
            # Invalidar caché de últimas ventas al cargar nuevos datos
            if hasattr(self, '_mbrp_ultimas_ventas_cache'):
                self._mbrp_ultimas_ventas_cache = {}
                self._mbrp_ultimas_ventas_time = 0
                # Detalle solo cuando debug MBRP está activo
                self.mbrp_debug_log("Caché de últimas ventas invalidado")

            # Estado UI
            try:
                self.api_status.config(text="MBRP: Iniciando carga...", foreground="#004C97")
                self.global_progress.pack(side=tk.RIGHT, padx=10)
                self.global_progress.config(mode="indeterminate")
                self.global_progress.start(5)
            except Exception:
                pass

            # Fila de carga
            if hasattr(self, 'mbrp_tree'):
                self.mbrp_tree.delete(*self.mbrp_tree.get_children())
                self.mbrp_tree.insert("", tk.END, values=(
                    "...", "Cargando productos de baja rotación...", "...", "...", "...", "...", "...", "..."
                ), tags=("loading",))

            # Lanzar hilo
            from threading import Thread
            self.mbrp_loader_thread = Thread(target=self._background_load_ventas_mbrp, daemon=True, name="mbrp_loader")
            self.mbrp_loader_thread.start()
        except Exception as e:
            self.log(f"Error iniciando MBRP: {e}", "ERROR")

    def _background_load_ventas_mbrp(self):
        try:
            if not all([self.mbrp_fecha_inicio, self.mbrp_fecha_fin, self.mbrp_sede_codigo]):
                self.log("MBRP: Faltan parámetros", "ERROR")
                return

            # Carga por chunks (reutiliza infra de TRA)
            start = 1
            # Adaptive chunk size controller
            from pal.core.chunks import AdaptiveChunkController
            controller = AdaptiveChunkController(initial=500, min_size=100, max_size=2000, target_latency=2.0)
            chunk_size = controller.size
            seen = set()
            acumulados = []
            
            # INFO: inicio de carga adaptativa siempre visible; detalles de chunks van por debug
            self.log(f"[MBRP] Iniciando carga adaptativa (chunk inicial: {chunk_size})", "INFO")

            chunk_count = 0
            # Obtener lista de almacenes tratables (Nuevo Motor)
            sedes_list = self.config_manager.get_tratables_by_sede(self.mbrp_sede_codigo)

            while True:
                chunk_count += 1
                
                chunk_t0 = time.perf_counter()
                rows = self.db_manager.obtener_ventas_por_producto_chunk(
                    fecha_inicio=self.mbrp_fecha_inicio,
                    fecha_fin=self.mbrp_fecha_fin,
                    sede_codigo=sedes_list,
                    start_row=start,
                    fetch_size=chunk_size,
                    exclude_depts=getattr(self, 'excluded_depts', [])
                )
                chunk_time = time.perf_counter() - chunk_t0
                if not rows:
                    self.log(f"[MBRP] Carga finalizada en chunk {chunk_count}", "INFO")
                    break
                    
                productos_nuevos = 0
                productos_duplicados = 0
                for r in rows:
                    codigo = str(r[0])
                    if codigo in seen:
                        productos_duplicados += 1
                        continue
                    seen.add(codigo)
                    acumulados.append(r)
                    productos_nuevos += 1
                
                # Solo loggear cada 5 chunks, al inicio, o al final
                should_log = chunk_count <= 2 or chunk_count % 5 == 0 or len(rows) < chunk_size
                if should_log:
                    self.mbrp_debug_log(
                        f"Chunk {chunk_count}: {len(rows)} filas | Nuevos: {productos_nuevos} | "
                        f"Total: {len(acumulados)} | Latencia: {chunk_time:.2f}s",
                        level="INFO",
                        throttle_key="chunk_progress",
                        throttle_seconds=3.0
                    )
                
                start += chunk_size
                time.sleep(0.05)

            self.log(f"[MBRP] Carga completada: {len(acumulados)} productos únicos", "SUCCESS")
            
            # Clasificar y filtrar por rotación usando lógica específica para MBRP
            from pal.services.mbrp import clasificar_rotacion_mbrp
            from pal.services.tra import clasificar_rotacion_tra
            try:
                # INFO general siempre visible
                self.log("[MBRP] Clasificando y filtrando productos de baja rotación...", "INFO")
                # Detalles solo en modo debug
                self.mbrp_debug_log(f"Datos acumulados antes de filtrar: {len(acumulados)} productos")
                
                if acumulados:
                    primera = acumulados[0]
                    self.mbrp_debug_log(f"Primera fila acumulada: {primera} (len={len(primera)})")
                
                # PASO 1: Clasificar con la lógica RI/TRA para aislar solo BAJA/SIN MOVIMIENTO/SIN CLASIFICAR
                try:
                    self.mbrp_debug_log("[MBRP] Clasificando rotación base (RI/TRA) para aislar BAJA/SIN MOVIMIENTO/SIN CLASIFICAR...")
                    base_clasificados = clasificar_rotacion_tra(acumulados)
                except Exception as e:
                    self.log(f"[MBRP] Error clasificando con RI/TRA: {e} - usando datos sin clasificar", "WARNING")
                    base_clasificados = list(acumulados)
                
                # Importante: incluir también "SIN CLASIFICAR" para contemplar casos de cero ventas
                productos_baja_base = [
                    r for r in base_clasificados
                    if len(r) > 6 and str(r[6]).upper() in {"BAJA", "SIN MOVIMIENTO", "SIN CLASIFICAR"}
                ]
                self.log(
                    f"[MBRP] Productos BAJA/SIN_MOVIMIENTO/SIN_CLASIFICAR (RI): {len(productos_baja_base)}/{len(base_clasificados)}",
                    "INFO"
                )
                
                if not productos_baja_base:
                    self.log("[MBRP] ADVERTENCIA: No hay productos BAJA/SIN_MOVIMIENTO/SIN_CLASIFICAR tras clasificación RI", "WARNING")
                    productos_baja_base = list(base_clasificados)  # fallback para no quedar vacíos
                
                # PASO 2: Dentro del conjunto de baja rotación, solo recalcular IM (sin excluir por umbral)
                #         Todos los productos BAJA/SIN_MOVIMIENTO/SIN_CLASIFICAR participan en el cálculo de IM%.
                productos_para_mbrp = list(productos_baja_base)
                self.log(
                    f"[MBRP] Productos incluidos para MBRP (BAJA/SIN_MOVIMIENTO): {len(productos_para_mbrp)}/{len(productos_baja_base)}",
                    "INFO",
                )
                
                # PASO 3: Clasificar MBRP (calcula IM internamente y etiqueta rotación específica MBRP)
                self.mbrp_debug_log(f"Iniciando clasificación de {len(productos_para_mbrp)} productos...")
                productos_clasificados = clasificar_rotacion_mbrp(productos_para_mbrp)
                self.log(f"[MBRP] Clasificados: {len(productos_clasificados)} productos", "INFO")
                
                if not productos_clasificados:
                    self.log(f"[MBRP] ADVERTENCIA: Clasificación retornó lista vacía", "WARNING")
                
                self.cached_ventas_mbrp = productos_clasificados
                self.log(f"[MBRP] Cache MBRP actualizado: {len(self.cached_ventas_mbrp)} productos", "INFO")
            except Exception as e:
                self.log(f"Error en clasificación MBRP: {e}", "ERROR")
                import traceback
                self.log(f"Traceback: {traceback.format_exc()}", "DEBUG")
                # En caso de error, usar datos sin clasificar
                self.log(f"[MBRP] Usando datos sin clasificar como fallback: {len(acumulados)} productos", "WARNING")
                self.cached_ventas_mbrp = list(acumulados)

            # Actualizar UI
            def _finish():
                try:
                    self.mbrp_debug_log("_finish() iniciado")
                    self.mbrp_debug_log(f"_finish() - cached_ventas_mbrp tiene {len(self.cached_ventas_mbrp)} productos")
                    
                    # Reconstruir vistas efectivas para MBRP
                    try:
                        self._rebuild_effective_views()
                    except Exception:
                        pass
                    self.aplicar_filtro_mbrp()
                    self.mbrp_debug_log("_finish() - aplicar_filtro_mbrp completado")
                    
                    try:
                        self.api_status.config(text="MBRP: Completo", foreground="green")
                        self.mbrp_debug_log("_finish() - api_status actualizado")
                    except Exception as e:
                        self.log(f"[MBRP] Error actualizando api_status: {e}", "WARNING")
                    
                    try:
                        self.global_progress.stop()
                        self.global_progress.pack_forget()
                        self.mbrp_debug_log("_finish() - progress bar detenido")
                    except Exception as e:
                        self.log(f"[MBRP] Error deteniendo progress bar: {e}", "WARNING")
                    
                    self.log("[MBRP] _finish() completado correctamente", "SUCCESS")
                except Exception as e:
                    self.log(f"[MBRP] ERROR en _finish: {e}", "ERROR")
                    import traceback
                    tb = traceback.format_exc()
                    self.mbrp_debug_log(f"Traceback en _finish: {tb}")
                    # Intentar detener progress bar de todas formas
                    try:
                        self.global_progress.stop()
                        self.global_progress.pack_forget()
                    except Exception:
                        pass
            self.root.after(0, _finish)
        except Exception as e:
            self.log(f"MBRP error en carga: {e}", "ERROR")
            try:
                self.root.after(0, lambda: (self.global_progress.stop(), self.global_progress.pack_forget()))
            except Exception:
                pass

    def aplicar_filtro_mbrp(self):
        try:
            if not hasattr(self, 'cached_ventas_mbrp') or not self.cached_ventas_mbrp:
                self.mbrp_debug_log(
                    "No hay datos MBRP en cache",
                    level="WARNING",
                    throttle_key="no_data",
                    throttle_seconds=5.0
                )
                return
            
            self.mbrp_debug_log(f"aplicar_filtro_mbrp() iniciado con {len(self.cached_ventas_mbrp)} productos en cache")
            
            # Inicializar diccionarios si no existen
            if not hasattr(self, 'mbrp_dept_dict'):
                self.log("[MBRP] Inicializando mbrp_dept_dict vacío", "WARNING")
                self.mbrp_dept_dict = {}
            if not hasattr(self, 'mbrp_group_dict'):
                self.mbrp_group_dict = {}
            if not hasattr(self, 'mbrp_sub_dict'):
                self.mbrp_sub_dict = {}
            
        # Exclusión por departamento (global) para MBRP
            excluded_set = getattr(self, '_excluded_depts_set', set())
            if excluded_set:
                datos_filtrados = [r for r in self.cached_ventas_mbrp if len(r) > 2 and str(r[2]) not in excluded_set]
            else:
                datos_filtrados = list(self.cached_ventas_mbrp)

            # Filtro por proveedor (si está seleccionado en MBRP)
            proveedor_cod = getattr(self, 'mbrp_proveedor_codigo', None)
            if proveedor_cod and datos_filtrados:
                codigos_prov = self._get_codigos_por_proveedor_cached(proveedor_cod)
                if codigos_prov:
                    datos_filtrados = [r for r in datos_filtrados if str(r[0]) in codigos_prov]
                else:
                    datos_filtrados = []

            # Filtro jerárquico (similar a TRA)
            dept_cod = self.mbrp_dept_dict.get(self.mbrp_dept_var.get()) if hasattr(self, 'mbrp_dept_var') else None
            group_cod = None
            sub_cod = None
            if dept_cod and hasattr(self, 'mbrp_group_var'):
                group_desc = self.mbrp_group_var.get()
                group_cod = self.mbrp_group_dict.get(dept_cod, {}).get(group_desc)
                if group_cod and hasattr(self, 'mbrp_sub_var'):
                    sub_desc = self.mbrp_sub_var.get()
                    # Usar string como key (formato: "dept|group")
                    key = f"{dept_cod}|{group_cod}"
                    sub_cod = self.mbrp_sub_dict.get(key, {}).get(sub_desc)

            texto = self.mbrp_search_var.get() if hasattr(self, 'mbrp_search_var') else ''
            
            self.mbrp_debug_log(
                f"Aplicando filtros - Dept: {dept_cod}, Group: {group_cod}, Sub: {sub_cod}, Texto: '{texto}'"
            )

            from pal.services.tra import filter_ventas_tra, paginate_tra
            datos_filtrados = filter_ventas_tra(
                ventas=datos_filtrados,
                dept_code=dept_cod,
                group_code=group_cod,
                sub_code=sub_cod,
                search_text=texto,
                filter_rotacion='TODAS',
                favoritos=self._get_favoritos_local(),
            )
            
            self.mbrp_debug_log(
                f"Después de filter_ventas_tra: {len(datos_filtrados)} productos filtrados"
            )

            if not hasattr(self, 'mbrp_current_page') or self.mbrp_current_page < 1:
                self.mbrp_current_page = 1
            if not hasattr(self, 'mbrp_page_size'):
                self.mbrp_page_size = 500
            datos_pagina, total_paginas, self.mbrp_current_page = paginate_tra(
                datos_filtrados, self.mbrp_current_page, self.mbrp_page_size
            )
            
            self.mbrp_debug_log(
                f"Página {self.mbrp_current_page}/{total_paginas} ({len(datos_pagina)} filas en página)"
            )
            
            self.mbrp_debug_log(f"Llamando a mostrar_mbrp_paginado con {len(datos_pagina)} productos")
            self.mostrar_mbrp_paginado(datos_pagina)
            self.mbrp_debug_log("mostrar_mbrp_paginado completado")
            
            self.actualizar_controles_paginacion_mbrp(total_paginas)
        except Exception as e:
            self.log(f"[MBRP] ERROR en aplicar_filtro_mbrp: {e}", "ERROR")
            import traceback
            self.mbrp_debug_log(f"Traceback en aplicar_filtro_mbrp: {traceback.format_exc()}")

    def mostrar_mbrp_paginado(self, datos):
        if not hasattr(self, 'mbrp_tree'):
            self.log("[MBRP] ERROR: mbrp_tree no existe", "ERROR")
            return
        try:
            self.mbrp_tree.delete(*self.mbrp_tree.get_children())
        except Exception as e:
            self.log(f"[MBRP] Error limpiando mbrp_tree: {e}", "ERROR")
            return
        
        if not datos:
            self.mbrp_debug_log("datos vacío en mostrar_mbrp_paginado")
            return
            
        # Importar servicios MBRP
        from pal.services.mbrp import calcular_indice_movilidad, obtener_ultimas_ventas_bulk, calcular_dias_sin_venta
        
        # Cache de stock rápido
        codigos = [r[0] for r in datos]
        # Alinear sede de stock con RI/TRA; si el filtro es ICH/global, se suman todas las sedes
        sede = None
        try:
            sede = getattr(self, 'mbrp_sede_codigo', None)
            if not sede and hasattr(self, 'sede_var'):
                sede = (self.sede_var.get() or '').split(' - ')[0]
        except Exception:
            sede = None
        # Por defecto usar ICH (global) para evitar caer a Cabudare cuando no hay sede explícita
        stock_map = self.obtener_stock_actual_bulk(codigos, sede)
        
        # Calcular Índices de Movilidad para todos los productos
        indices_movilidad = calcular_indice_movilidad(self.cached_ventas_mbrp or datos)
        
        # Cache de últimas ventas con TTL de 60 segundos
        # Se invalida automáticamente al presionar "Cargar"
        if not hasattr(self, '_mbrp_ultimas_ventas_cache') or not hasattr(self, '_mbrp_ultimas_ventas_time'):
            self._mbrp_ultimas_ventas_cache = {}
            self._mbrp_ultimas_ventas_time = 0
        
        current_time = time.time()
        cache_ttl = 60  # 60 segundos (reducido de 5 minutos para mayor frescura de datos)
        
        # Verificar si los códigos de la página actual están en caché
        codigos_faltantes = [c for c in codigos if c not in self._mbrp_ultimas_ventas_cache]
        cache_expirado = (current_time - self._mbrp_ultimas_ventas_time) > cache_ttl
        
        # Consultar si: 1) caché expirado, 2) caché vacío, o 3) hay códigos faltantes
        if cache_expirado or not self._mbrp_ultimas_ventas_cache or codigos_faltantes:
            # Consultar todos los códigos de la página actual
            ultimas_ventas_nuevas = obtener_ultimas_ventas_bulk(self.db_manager, codigos, sede)
            
            # Actualizar caché (merge con datos existentes si no está expirado)
            if cache_expirado or not self._mbrp_ultimas_ventas_cache:
                self._mbrp_ultimas_ventas_cache = ultimas_ventas_nuevas
                self.mbrp_debug_log(f"Últimas ventas consultadas: {len(ultimas_ventas_nuevas)} productos (caché renovado)")
            else:
                # Merge incremental
                self._mbrp_ultimas_ventas_cache.update(ultimas_ventas_nuevas)
                self.mbrp_debug_log(f"Últimas ventas actualizadas: {len(ultimas_ventas_nuevas)} productos nuevos")
            
            self._mbrp_ultimas_ventas_time = current_time
            ultimas_ventas = self._mbrp_ultimas_ventas_cache
        else:
            ultimas_ventas = self._mbrp_ultimas_ventas_cache
            tiempo_restante = int(cache_ttl - (current_time - self._mbrp_ultimas_ventas_time))
            self.mbrp_debug_log(f"Usando caché de últimas ventas ({tiempo_restante}s restantes)")
        
        for idx, fila in enumerate(datos):
            try:
                # Preservar formato original del código (mantener ceros a la izquierda)
                if isinstance(fila[0], str):
                    codigo = fila[0]
                else:
                    # Si es numérico, convertir a string sin perder ceros a la izquierda
                    codigo_str = str(fila[0])
                    # Si el código original tenía menos de 4 dígitos, rellenar con ceros
                    if len(codigo_str) < 4:
                        codigo = codigo_str.zfill(4)
                    else:
                        codigo = codigo_str
                desc = fila[1]
                neto = fila[5]
                rotacion = fila[6] if len(fila) > 6 else 'BAJA'
                precio = fila[7] if len(fila) > 7 else 0  # Precio unitario desde la consulta (después de clasificación)
                stock_actual = int(stock_map.get(codigo, 0) or 0)
                
                # Índice de Movilidad
                im_porcentaje = indices_movilidad.get(codigo, 0.0)
                
                # Calcular días desde última venta
                fecha_ultima_venta = ultimas_ventas.get(codigo)
                dias_sin_venta = calcular_dias_sin_venta(fecha_ultima_venta)
                
                # Formatear última venta para mostrar
                if dias_sin_venta == -1:
                    ultima_venta_texto = "Nunca"
                elif dias_sin_venta == 0:
                    ultima_venta_texto = "Hoy"
                elif dias_sin_venta == 1:
                    ultima_venta_texto = "1 día"
                else:
                    ultima_venta_texto = f"{dias_sin_venta} días"
                
                # Calcular Días de Stock (DS) usando el mismo concepto que TRA
                # DS = Stock Actual / Promedio Diario de Ventas en el período MBRP
                try:
                    dias_stock = self.calcular_dias_restantes(
                        stock_actual=int(stock_actual),
                        neto_ventas=float(neto or 0),
                        fecha_inicio=self.mbrp_fecha_inicio,
                        fecha_fin=self.mbrp_fecha_fin,
                    )
                except Exception:
                    dias_stock = 0
                
                # Determinar tags de color por rotación e Índice de Movilidad con filas alternadas
                # Reemplazar espacios y guiones para que coincidan con los tags configurados en mbrp.py
                tag_base_rotacion = str(rotacion).lower().replace(" ", "_")
                
                # Tag por Índice de Movilidad (prioridad sobre rotación)
                if im_porcentaje < 5.0:
                    tag_base = "im_critico"
                elif im_porcentaje <= 10.0:
                    tag_base = "im_muy_bajo" 
                elif im_porcentaje <= 20.0:
                    tag_base = "im_bajo"
                else:
                    tag_base = tag_base_rotacion  # Usar tag de rotación normal
                
                # Aplicar estilo alternado
                if idx % 2 == 0:
                    tag_final = tag_base  # Filas pares: colores claros
                else:
                    tag_final = f"{tag_base}_alt"  # Filas impares: colores oscuros
                
                # Formatear neto según el modo de display (con IVA incluido)
                neto_valor = float(neto or 0)
                mostrar_dolares = getattr(self, 'mbrp_mostrar_dolares_var', tk.BooleanVar()).get()
                
                if mostrar_dolares:
                    # Convertir a dólares usando precio con IVA desde la base de datos
                    precio_unitario = self._obtener_precio_producto(codigo)  # Ya incluye IVA
                    ventas_dolares = neto_valor * precio_unitario
                    neto_formateado = f"${ventas_dolares:,.2f}"
                else:
                    # Mostrar como unidades (entero)
                    neto_formateado = int(neto_valor)
                
                self.mbrp_tree.insert(
                    "", tk.END,
                    values=(codigo, desc, rotacion, neto_formateado, stock_actual, dias_stock, f"{im_porcentaje}%", ultima_venta_texto),
                    tags=(tag_final,)
                )

                # Si estamos en modo ICH, actualizar info de detalle para el primer elemento
                if idx == 0 and hasattr(self, '_update_mbrp_ich_info') and (
                    getattr(self, 'mbrp_sede_codigo', None) in (None, '%', '00', 'ICH', 'ALL')
                ):
                    try:
                        # Seleccionar visualmente la primera fila solo si no hay selección previa
                        if not self.mbrp_tree.selection():
                            first_item = self.mbrp_tree.get_children()[0]
                            self.mbrp_tree.selection_set(first_item)
                            self.mbrp_tree.focus(first_item)
                        self._update_mbrp_ich_info()
                    except Exception:
                        pass
            except Exception as e:
                self.log(f"Error procesando fila MBRP {fila}: {str(e)}", "ERROR")
                continue

    def actualizar_controles_paginacion_mbrp(self, total_paginas):
        if hasattr(self, 'mbrp_pagina_label'):
            self.mbrp_pagina_label.config(text=f"Página {self.mbrp_current_page}/{total_paginas}")
        if hasattr(self, 'mbrp_btn_prev'):
            self.mbrp_btn_prev['state'] = 'normal' if self.mbrp_current_page > 1 else 'disabled'
        if hasattr(self, 'mbrp_btn_next'):
            self.mbrp_btn_next['state'] = 'normal' if self.mbrp_current_page < total_paginas else 'disabled'

    def cambiar_pagina_mbrp(self, delta):
        self.mbrp_current_page += delta
        self.aplicar_filtro_mbrp()
    
    def actualizar_ultimas_ventas_mbrp(self):
        """
        Fuerza la actualización del caché de últimas ventas sin recargar todos los datos.
        Útil cuando se sabe que hubo ventas recientes.
        """
        if not hasattr(self, 'cached_ventas_mbrp') or not self.cached_ventas_mbrp:
            self.log("No hay datos MBRP cargados. Use 'Cargar' primero.", "WARNING")
            return
        
        try:
            # Invalidar caché
            if hasattr(self, '_mbrp_ultimas_ventas_cache'):
                self._mbrp_ultimas_ventas_cache = {}
                self._mbrp_ultimas_ventas_time = 0
            
            self.log("[MBRP] Actualizando últimas ventas...", "INFO")
            
            # Refrescar la vista actual (forzará nueva consulta)
            self.aplicar_filtro_mbrp()
            
            self.log("[MBRP] Últimas ventas actualizadas correctamente", "SUCCESS")
            
        except Exception as e:
            self.log(f"Error actualizando últimas ventas MBRP: {e}", "ERROR")

    def _map_deposito_to_sede(self, cod: str) -> str:
        """Mapea un código de depósito a una sede legible (Cabudare, Barinas, Guanare, Otra)."""
        try:
            c = (cod or "").strip()
            if c.startswith('03'):
                return 'Cabudare'
            if c.startswith('01'):
                return 'Barinas'
            if c.startswith('04'):
                return 'Guanare'
            return 'Otra'
        except Exception:
            return 'Otra'

    def _get_mbrp_last_sale_global(self, codigo: str):
        """Obtiene depósito y fecha de la última venta global (todas las sedes) para un producto."""
        try:
            sql = """
                SELECT TOP 1 c_Deposito, f_fecha
                FROM TR_INVENTARIO WITH (NOLOCK)
                WHERE c_Codarticulo = ?
                  AND c_Concepto = 'VEN'
                  AND n_Cantidad > 0
                ORDER BY f_fecha DESC
            """
            rows = self.db_manager.fetch_data(sql, (str(codigo),)) if hasattr(self, 'db_manager') and self.db_manager else None
            if rows:
                dep, fecha = rows[0]
                return (str(dep).strip() if dep is not None else None, fecha)
            return (None, None)
        except Exception as e:
            self.log(f"[MBRP] Error obteniendo última venta global para {codigo}: {e}", "ERROR")
            return (None, None)

    def _get_mbrp_stock_por_sede(self, codigo: str):
        """Obtiene stock global y distribución por sede para un producto (solo modo ICH)."""
        try:
            if not hasattr(self, 'db_manager') or not self.db_manager:
                return 0, {}
            sql = """
                SELECT c_coddeposito, SUM(n_cantidad)
                FROM MA_DEPOPROD WITH (NOLOCK)
                WHERE c_codarticulo = ?
                GROUP BY c_coddeposito
            """
            rows = self.db_manager.fetch_data(sql, (str(codigo),)) or []
            sede_totals = {}
            total = 0
            for dep, qty in rows:
                try:
                    dep_str = str(dep).strip()
                    sede = self._map_deposito_to_sede(dep_str)
                    q = int(qty or 0)
                except Exception:
                    continue
                total += q
                sede_totals[sede] = sede_totals.get(sede, 0) + q
            return total, sede_totals
        except Exception as e:
            self.log(f"[MBRP] Error obteniendo stock por sede para {codigo}: {e}", "ERROR")
            return 0, {}

    def _update_mbrp_ich_info(self):
        """Actualiza la etiqueta informativa de MBRP cuando el filtro de sede es ICH (global).

        Muestra:
          - Dónde fue la última venta (sede)
          - Stock global y cómo se reparte entre sedes
        """
        try:
            if not hasattr(self, 'mbrp_ich_info_var'):
                return

            # Solo aplica cuando MBRP está en modo ICH/global
            sede_codigo = getattr(self, 'mbrp_sede_codigo', None)
            if sede_codigo not in (None, '%', '00', 'ICH', 'ALL'):
                self.mbrp_ich_info_var.set("")
                return

            tree = getattr(self, 'mbrp_tree', None)
            if tree is None:
                self.mbrp_ich_info_var.set("")
                return

            sel = tree.selection()
            if not sel:
                self.mbrp_ich_info_var.set("")
                return

            values = tree.item(sel[0], 'values') or []
            if not values:
                self.mbrp_ich_info_var.set("")
                return

            codigo = str(values[0])

            # Última venta global (todas las sedes)
            dep_ult, fecha_ult = self._get_mbrp_last_sale_global(codigo)
            if dep_ult:
                sede_ult = self._map_deposito_to_sede(dep_ult)
                ultima_txt = f"Última venta en {sede_ult} ({dep_ult})"
            else:
                ultima_txt = "Sin ventas registradas (global)"

            # Stock por sede
            total_stock, sede_totals = self._get_mbrp_stock_por_sede(codigo)
            if not sede_totals:
                stock_txt = "Sin stock en ninguna sede"
            else:
                partes = []
                # Ordenar por nombre de sede para consistencia
                for sede_nombre in sorted(sede_totals.keys()):
                    partes.append(f"{sede_nombre}: {sede_totals[sede_nombre]}")
                stock_txt = "Stock por sede: " + "  •  ".join(partes)

            self.mbrp_ich_info_var.set(
                f"* {ultima_txt}  |  Stock global: {total_stock}  |  {stock_txt}"
            )
        except Exception as e:
            # No romper la UI por errores de detalle
            self.log(f"[MBRP] Error actualizando info ICH: {e}", "ERROR")

    def generar_reporte_mbrp(self):
        """Genera el "Reporte 0" para MBRP: productos que no se han vendido en X días.

        Criterio:
          - Se parte del universo actual MBRP (cargado en cached_ventas_mbrp)
          - Se consulta TR_INVENTARIO (obtener_ultimas_ventas_bulk) por sede para
             identificar productos sin ventas recientes (según criterio de días)
          - Por defecto muestra productos NUNCA vendidos, pero se puede configurar
             un umbral de días (ej: 30 días sin venta)
        """
        if not hasattr(self, 'cached_ventas_mbrp') or not self.cached_ventas_mbrp:
            messagebox.showwarning("Sin datos", "No hay datos MBRP disponibles. Cargue datos primero.")
            return
            
        try:
            from pal.services.mbrp import obtener_ultimas_ventas_bulk
            from pal.services.tra import filter_ventas_tra
            
            sede = self.mbrp_sede_codigo or '0301'

            # 1) Partir de todos los datos MBRP en caché
            datos_filtrados = list(self.cached_ventas_mbrp)

            # 2) Exclusión global por departamento (igual que en aplicar_filtro_mbrp)
            excluded_set = getattr(self, '_excluded_depts_set', set())
            if excluded_set:
                datos_filtrados = [r for r in datos_filtrados if len(r) > 2 and str(r[2]) not in excluded_set]

            # 3) Filtro por proveedor (si está seleccionado en MBRP)
            proveedor_cod = getattr(self, 'mbrp_proveedor_codigo', None)
            if proveedor_cod and datos_filtrados:
                codigos_prov = self._get_codigos_por_proveedor_cached(proveedor_cod)
                if codigos_prov:
                    codigos_prov = set(str(c) for c in codigos_prov)
                    datos_filtrados = [r for r in datos_filtrados if str(r[0]) in codigos_prov]
                else:
                    datos_filtrados = []

            # 4) Filtros jerárquicos y de búsqueda (mismos que la grilla MBRP)
            dept_cod = self.mbrp_dept_dict.get(self.mbrp_dept_var.get()) if hasattr(self, 'mbrp_dept_var') else None
            group_cod = None
            sub_cod = None
            if dept_cod and hasattr(self, 'mbrp_group_var'):
                group_desc = self.mbrp_group_var.get()
                group_cod = self.mbrp_group_dict.get(dept_cod, {}).get(group_desc)
                if group_cod and hasattr(self, 'mbrp_sub_var'):
                    sub_desc = self.mbrp_sub_var.get()
                    key = f"{dept_cod}|{group_cod}"
                    sub_cod = self.mbrp_sub_dict.get(key, {}).get(sub_desc)

            texto = self.mbrp_search_var.get() if hasattr(self, 'mbrp_search_var') else ''

            datos_filtrados = filter_ventas_tra(
                ventas=datos_filtrados,
                dept_code=dept_cod,
                group_code=group_cod,
                sub_code=sub_cod,
                search_text=texto,
                filter_rotacion='TODAS',
                favoritos=self._get_favoritos_local(),
            )

            if not datos_filtrados:
                messagebox.showinfo(
                    "Reporte 0 MBRP",
                    "No hay productos MBRP para el Reporte 0 con los filtros actuales.",
                )
                return

            # 5) Construir lista de códigos únicos del conjunto YA filtrado
            codigos = []
            cols_dept = set()
            cols_group = set()
            cols_sub = set()
            
            for item in datos_filtrados:
                try:
                    if item and len(item) > 0 and item[0] is not None:
                        codigos.append(str(item[0]))
                        if len(item) > 2 and item[2]: cols_dept.add(str(item[2]))
                        if len(item) > 3 and item[3]: cols_group.add(str(item[3]))
                        if len(item) > 4 and item[4]: cols_sub.add(str(item[4]))
                except Exception:
                    continue
            codigos_unicos = sorted(set(codigos))

            if not codigos_unicos:
                messagebox.showinfo(
                    "Reporte 0 MBRP",
                    "No hay códigos de producto en el dataset filtrado actual.",
                )
                return

            # Pre-cargar descripciones de Dept/Grupo/Subgrupo para el reporte
            mapa_dept = {}
            mapa_grupo = {}
            mapa_sub = {}
            
            try:
                # Cargar Departamentos
                if cols_dept:
                    placeholders = ','.join(['?' for _ in cols_dept])
                    rows = self.db_manager.fetch_data(f"SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS WHERE C_CODIGO IN ({placeholders})", list(cols_dept))
                    if rows: mapa_dept = {r[0].strip(): r[1].strip() for r in rows if r[0]}
                
                # Cargar Grupos
                if cols_group:
                    placeholders = ','.join(['?' for _ in cols_group])
                    rows = self.db_manager.fetch_data(f"SELECT C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WHERE C_CODIGO IN ({placeholders})", list(cols_group))
                    if rows: mapa_grupo = {r[0].strip(): r[1].strip() for r in rows if r[0]}

                # Cargar Subgrupos
                if cols_sub:
                    placeholders = ','.join(['?' for _ in cols_sub])
                    rows = self.db_manager.fetch_data(f"SELECT C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WHERE C_CODIGO IN ({placeholders})", list(cols_sub))
                    if rows: mapa_sub = {r[0].strip(): r[1].strip() for r in rows if r[0]}
                    
            except Exception as e:
                self.log(f"Error cargando descripciones para reporte: {e}", "WARNING")

            # 6) Consultar últimas ventas y STOCK ACTUAL
            from datetime import datetime
            start_time = datetime.now()
            ultimas_ventas = obtener_ultimas_ventas_bulk(self.db_manager, codigos_unicos, sede)
            
            # Consultar Stock Actual usando el método existente en la app
            stock_actual_map = {}
            try:
                if hasattr(self, 'obtener_stock_actual_bulk'):
                    stock_actual_map = self.obtener_stock_actual_bulk(codigos_unicos, sede)
                else:
                    self.log("[MBRP] obtener_stock_actual_bulk no encontrado en self", "ERROR")
            except Exception as e:
                self.log(f"[MBRP] Error obteniendo stock bulk: {e}", "ERROR")

            query_time = (datetime.now() - start_time).total_seconds()
            self.log(f"Reporte 0: Consultas (ventas+stock) completadas en {query_time:.2f}s", "INFO")

            # 7) Calcular días del período seleccionado en el módulo MBRP
            dias_criterio = -1  # Default: nunca vendidos
            if hasattr(self, 'mbrp_fecha_inicio') and hasattr(self, 'mbrp_fecha_fin') and self.mbrp_fecha_inicio and self.mbrp_fecha_fin:
                try:
                    fecha_inicio = self.mbrp_fecha_inicio if hasattr(self.mbrp_fecha_inicio, 'date') else datetime.strptime(str(self.mbrp_fecha_inicio), '%Y-%m-%d').date()
                    fecha_fin = self.mbrp_fecha_fin if hasattr(self.mbrp_fecha_fin, 'date') else datetime.strptime(str(self.mbrp_fecha_fin), '%Y-%m-%d').date()
                    dias_criterio = (fecha_fin - fecha_inicio).days
                    if dias_criterio < 0:
                        dias_criterio = 30  # Default si hay error
                except Exception:
                    dias_criterio = 30  # Default período de 30 días
            else:
                dias_criterio = 30  # Default período de 30 días

            # 8) Filtrar productos según criterio de días
            from pal.services.mbrp import calcular_dias_sin_venta
            productos_filtrados = []
            
            for item in datos_filtrados:
                if not item or len(item) < 1:
                    continue
                try:
                    codigo = str(item[0])
                except Exception:
                    continue
                
                # Obtener fecha de última venta
                fecha_ultima = ultimas_ventas.get(codigo)
                
                if dias_criterio == -1:
                    # Modo original: productos NUNCA vendidos
                    if codigo in ultimas_ventas:
                        continue  # tiene al menos una venta registrada
                    dias_sin_venta = -1
                else:
                    # Modo nuevo: productos sin venta en X días
                    if fecha_ultima is None:
                        # Nunca vendido, cumple cualquier criterio de días
                        dias_sin_venta = 9999
                    else:
                        dias_sin_venta = calcular_dias_sin_venta(fecha_ultima)
                        if dias_sin_venta < dias_criterio:
                            continue  # Tiene venta más reciente que el criterio
                
                # Estructura MBRP: (codigo, desc, dept, group, sub, neto, rotacion, ...)
                desc = str(item[1]) if len(item) > 1 and item[1] is not None else codigo
                
                # Resolver descripciones usando los mapas
                c_dept = str(item[2]).strip() if len(item) > 2 and item[2] else ''
                c_grupo = str(item[3]).strip() if len(item) > 3 and item[3] else ''
                c_sub = str(item[4]).strip() if len(item) > 4 and item[4] else ''
                
                dept_desc = mapa_dept.get(c_dept, c_dept)
                grupo_desc = mapa_grupo.get(c_grupo, c_grupo)
                sub_desc = mapa_sub.get(c_sub, c_sub)
                
                rotacion = item[6] if len(item) > 6 else 'SIN_MOVIMIENTO'
                current_stock = stock_actual_map.get(codigo, 0.0)
                
                productos_filtrados.append({
                    'codigo': codigo,
                    'descripcion': desc,
                    'dept': dept_desc,
                    'grupo': grupo_desc,
                    'sub': sub_desc,
                    'rotacion': rotacion,
                    'stock': current_stock,
                    'dias_sin_venta': dias_sin_venta,
                    'fecha_ultima': fecha_ultima,
                })

            if not productos_filtrados:
                criterio_texto = "nunca vendidos" if dias_criterio == -1 else f"sin venta en los últimos {dias_criterio} días"
                messagebox.showinfo(
                    "Reporte 0 MBRP",
                    f"No se encontraron productos {criterio_texto} para los filtros actuales.",
                )
                return

            # Generar archivo Excel en lugar de ventana de texto
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from datetime import datetime
                import os
                from tkinter import filedialog
                
                # Crear nombre de archivo por defecto
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                criterio_nombre = "NUNCA_VENDIDOS" if dias_criterio == -1 else f"SIN_VENTA_{dias_criterio}_DIAS"
                default_name = f"reporte_0_mbrp_{criterio_nombre}_{timestamp}.xlsx"

                # Permitir al usuario escoger dónde guardar el reporte 0
                try:
                    filename = filedialog.asksaveasfilename(
                        parent=self.root if hasattr(self, 'root') else None,
                        title="Guardar Reporte 0 MBRP como...",
                        defaultextension=".xlsx",
                        initialfile=default_name,
                        filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
                    )
                except Exception as e:
                    # Si falla el diálogo, caer al nombre por defecto en el directorio actual
                    self.log(f"[MBRP] Error mostrando diálogo de guardar Reporte 0: {e} - usando ruta por defecto", "WARNING")
                    filename = default_name

                if not filename:
                    # Usuario canceló el guardado
                    self.log("[MBRP] Usuario canceló el guardado del Reporte 0 MBRP", "INFO")
                    return
                
                # Crear workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Reporte 0 MBRP"
                
                # Título del reporte
                criterio_texto = "NUNCA VENDIDOS" if dias_criterio == -1 else f"SIN VENTA EN ÚLTIMOS {dias_criterio} DÍAS"
                ws['A1'] = f"REPORTE 0 MBRP - PRODUCTOS {criterio_texto}"
                ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                
                # Información del reporte
                ws['A2'] = f"Período MBRP: {self.mbrp_fecha_inicio} - {self.mbrp_fecha_fin}"
                ws['A3'] = f"Sede: {sede}"
                ws['A4'] = f"Fecha generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ws['A5'] = f"Total productos filtrados en MBRP: {len(datos_filtrados)}"
                ws['A6'] = f"Productos {criterio_texto.lower()}: {len(productos_filtrados)}"
                
                # Encabezados actualizados
                headers = ['Código', 'Descripción', 'Departamento', 'Grupo', 'Subgrupo', 'Rotación', 'Stock Actual', 'Días sin venta', 'Última venta']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=8, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for i, prod in enumerate(productos_filtrados, 9):
                    # Formatear fecha
                    fecha_str = "NUNCA"
                    if prod['fecha_ultima']:
                        fecha_str = prod['fecha_ultima'].strftime('%d-%m-%Y')
                        
                    ws.cell(row=i, column=1, value=prod['codigo'])
                    ws.cell(row=i, column=2, value=prod['descripcion'])
                    ws.cell(row=i, column=3, value=prod['dept'])
                    ws.cell(row=i, column=4, value=prod['grupo'])
                    ws.cell(row=i, column=5, value=prod['sub'])
                    ws.cell(row=i, column=6, value=prod['rotacion']).alignment = Alignment(horizontal='center')
                    ws.cell(row=i, column=7, value=prod['stock']).alignment = Alignment(horizontal='center')
                    ws.cell(row=i, column=8, value=prod['dias_sin_venta']).alignment = Alignment(horizontal='center')
                    ws.cell(row=i, column=9, value=fecha_str).alignment = Alignment(horizontal='center')

                # Ajustar columnas
                column_widths = [15, 60, 25, 25, 25, 12, 12, 15, 12]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
                # Guardar archivo
                wb.save(filename)
                
                messagebox.showinfo(
                    "Reporte 0 MBRP",
                    f"Reporte 0 MBRP generado exitosamente:\n\n"
                    f"• Archivo: {filename}\n"
                    f"• Productos: {len(productos_filtrados)}\n"
                    f"• Criterio: {criterio_texto.lower()}\n"
                )
                
                self.log(f"Reporte 0 MBRP Excel generado: {filename} - {len(productos_filtrados)} productos", "SUCCESS")
                
            except ImportError:
                messagebox.showerror(
                    "Error",
                    "Para generar el reporte Excel necesita instalar openpyxl:\n\n"
                    "pip install openpyxl\n\n"
                    "Instale la librería y vuelva a intentar."
                )
                return
            except Exception as e:
                self.log(f"Error generando reporte Excel: {str(e)}", "ERROR")
                messagebox.showerror("Error", f"Error generando reporte Excel: {str(e)}")
                return
            

            
        except Exception as e:
            self.log(f"Error generando reporte MBRP: {str(e)}", "ERROR")
            messagebox.showerror("Error", f"Error generando reporte: {str(e)}")

    def _create_admin_exclusions_tab(self, parent):
        """Crea la pestaña para gestionar exclusiones globales de departamentos."""
        prefs_excl = ttk.LabelFrame(parent, text="Excluir departamentos (no considerar en reportes)", padding=10)
        prefs_excl.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        row1 = ttk.Frame(prefs_excl)
        row1.pack(fill=tk.X, pady=(0,6))
        ttk.Label(row1, text="Departamento:").pack(side=tk.LEFT)
        dep_ex_var = tk.StringVar()
        dep_names = sorted(list(getattr(self, 'tra_dept_dict', {}).keys())) if hasattr(self, 'tra_dept_dict') and self.tra_dept_dict else []
        cb_dep_ex = ttk.Combobox(row1, textvariable=dep_ex_var, state='readonly', values=dep_names, width=32)
        cb_dep_ex.pack(side=tk.LEFT, padx=6)

        excl_listbox = tk.Listbox(prefs_excl, height=8)

        def _refresh_excl_list():
            # Reconstruir listado a partir de códigos
            excl_codes = list(set(str(x) for x in (getattr(self, 'excluded_depts', []) or [])))
            # Convertir a nombres usando tra_dept_dict
            names = []
            for code in sorted(excl_codes):
                desc = None
                for d, c in (getattr(self, 'tra_dept_dict', {}) or {}).items():
                    if str(c) == str(code):
                        desc = d; break
                names.append(desc or code)
            excl_listbox.delete(0, tk.END)
            for n in names:
                excl_listbox.insert(tk.END, n)

        def _prefs_exclude_add():
            sel_desc = (dep_ex_var.get() or '').strip()
            if not sel_desc:
                return
            code = (getattr(self, 'tra_dept_dict', {}) or {}).get(sel_desc)
            if not code:
                return
            if not hasattr(self, 'excluded_depts') or self.excluded_depts is None:
                self.excluded_depts = set()
            self.excluded_depts = set(str(x) for x in (self.excluded_depts or set()))
            self.excluded_depts.add(str(code))
            try:
                self._save_global_settings()
            except Exception:
                pass
            # Actualizar set y vistas efectivas
            self._update_excluded_set()
            self._rebuild_effective_views()
            _refresh_excl_list()
            try:
                self.aplicar_filtro_stock()
            except Exception:
                pass
            try:
                self.aplicar_filtro_tra()
            except Exception:
                pass

        def _prefs_exclude_remove():
            sel = excl_listbox.curselection()
            if not sel:
                return
            name = excl_listbox.get(sel[0])
            # map name back to code
            code = None
            for d, c in (getattr(self, 'tra_dept_dict', {}) or {}).items():
                if d == name:
                    code = c; break
            if not code:
                return
            codes = set(str(x) for x in (getattr(self, 'excluded_depts', []) or []))
            if str(code) in codes:
                codes.discard(str(code))
                self.excluded_depts = codes
                try:
                    self._save_global_settings()
                except Exception:
                    pass
                # Actualizar set y vistas efectivas
                self._update_excluded_set()
                self._rebuild_effective_views()
                _refresh_excl_list()
                try:
                    self.aplicar_filtro_stock()
                except Exception:
                    pass
                try:
                    self.aplicar_filtro_tra()
                except Exception:
                    pass

        def _prefs_exclude_clear():
            self.excluded_depts = set()
            try:
                self._save_global_settings()
            except Exception:
                pass
            # Actualizar set y vistas efectivas
            self._update_excluded_set()
            self._rebuild_effective_views()
            _refresh_excl_list()
            try:
                self.aplicar_filtro_stock()
            except Exception:
                pass
            try:
                self.aplicar_filtro_tra()
            except Exception:
                pass

        ttk.Button(row1, text="➖ Excluir", command=_prefs_exclude_add).pack(side=tk.LEFT, padx=6)

        # Listado actual
        list_frame = ttk.Frame(prefs_excl)
        list_frame.pack(fill=tk.BOTH, expand=True)
        excl_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0,6))
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=excl_listbox.yview)
        excl_listbox.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.LEFT, fill=tk.Y)

        # Botonera
        buttons = ttk.Frame(prefs_excl)
        buttons.pack(fill=tk.X, pady=(6,0))
        ttk.Button(buttons, text="🗑️ Quitar seleccionado", command=_prefs_exclude_remove).pack(side=tk.LEFT)
        ttk.Button(buttons, text="♻️ Limpiar todos", command=_prefs_exclude_clear).pack(side=tk.LEFT, padx=6)

        _refresh_excl_list()


    def mostrar_context_menu_tra(self, event):
        """Muestra menú contextual en la tabla TRA"""
        # Verificar permiso
        can_view = False
        try:
            if hasattr(self, 'permissions') and self.current_user:
                can_view = self.permissions.tiene_permiso(self.current_user['id'], 'TRA', 'ver_proveedores')
            if self.current_user and self.current_user.get('username','').lower() == 'admin':
                can_view = True
        except Exception:
            can_view = False
        
        if not can_view:
            return

        item = self.tra_tree.identify_row(event.y)
        if item:
            self.tra_tree.selection_set(item)
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="🔍 Ver proveedores", command=lambda: self.ver_proveedores(self.tra_tree, "tra"))
            menu.post(event.x_root, event.y_root)

    def mostrar_context_menu_mbrp(self, event):
        """Muestra menú contextual en la tabla MBRP"""
        # Verificar permiso
        can_view = False
        try:
            if hasattr(self, 'permissions') and self.current_user:
                can_view = self.permissions.tiene_permiso(self.current_user['id'], 'MBRP', 'ver_proveedores')
            if self.current_user and self.current_user.get('username','').lower() == 'admin':
                can_view = True
        except Exception:
            can_view = False
        
        if not can_view:
            return

        item = self.mbrp_tree.identify_row(event.y)
        if item:
            self.mbrp_tree.selection_set(item)
            menu = tk.Menu(self.root, tearoff=0)
            menu.add_command(label="🔍 Ver proveedores", command=lambda: self.ver_proveedores(self.mbrp_tree, "mbrp"))
            menu.post(event.x_root, event.y_root)

    def ver_proveedores(self, tree, module_type):
        """Carga y muestra proveedores para el producto seleccionado en modo jerárquico"""
        selected = tree.selection()
        if not selected:
            return
        
        parent_item = selected[0]
        # Si ya tiene hijos, colapsar/expandir o no hacer nada si ya están cargados
        if tree.get_children(parent_item):
            # Si ya tiene hijos, simplemente alternar expansión
            if tree.item(parent_item, "open"):
                tree.item(parent_item, open=False)
            else:
                tree.item(parent_item, open=True)
            return

        values = tree.item(parent_item, "values")
        if not values:
            return
            
        cod_producto = values[0] # El código siempre es la primera columna
        
        try:
            # Mostrar indicador de carga
            temp_id = tree.insert(parent_item, tk.END, values=("Cargando proveedores...", "", "", "", "", "", "", "", ""))
            tree.item(parent_item, open=True)
            self.root.update_idletasks()

            # Consultar base de datos
            proveedores = self.db_manager.obtener_proveedores_detalle_por_producto(cod_producto)
            
            # Eliminar indicador de carga
            tree.delete(temp_id)

            if not proveedores:
                tree.insert(parent_item, tk.END, values=("Sin proveedores registrados", "", "", "", "", "", "", "", ""))
                return

            # Insertar proveedores como hijos
            # El formato debe ajustarse a las columnas de cada treeview para que se vea bien
            # Columnas TRA: "Código", "Descripción", "Rotación", "Ventas", "Representación %", "Stock Actual", "Stock Ideal", "Días Restantes", "Estado Stock"
            # Columnas MBRP: "Código", "Descripción", "Rotación", "Ventas", "Stock Actual", "Días de Stock", "IM %", "Última Venta"
            
            for prov in proveedores:
                p_cod = prov[0]
                p_nom = prov[1]
                p_num = prov[2]
                p_fec = prov[3].strftime('%Y-%m-%d') if prov[3] else "N/A"
                p_cos = prov[4] if len(prov) > 4 else 0.0
                
                # Formatear valores para que quepan en las columnas existentes
                if module_type == "tra":
                    child_values = ("", f"PROV: {p_nom} ({p_cod})", f"COSTO: {p_cos:,.2f}", f"N° CP: {p_num}", f"FECHA: {p_fec}", "", "", "", "")
                else: # mbrp
                    child_values = ("", f"PROV: {p_nom} ({p_cod})", f"COSTO: {p_cos:,.2f}", f"N° CP: {p_num}", "", f"FECHA: {p_fec}", "", "")
                
                tree.insert(parent_item, tk.END, values=child_values, tags=('prov_child',))
            
            # Estilo para los hijos proveedores
            tree.tag_configure('prov_child', background='#F0F0F0', foreground='#555555')
            
        except Exception as e:
            self.log(f"Error al cargar proveedores: {e}", "ERROR")
            try: tree.delete(temp_id)
            except: pass
            tree.insert(parent_item, tk.END, values=("Error al cargar datos", "", "", "", "", "", "", "", ""))

if __name__ == "__main__":
    root = tk.Tk() 
    root.withdraw()
    app = DatabaseApp(root)
    root.mainloop() 


