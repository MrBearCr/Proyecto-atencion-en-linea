"""
Módulo de servicios de exportación para la aplicación PAL
"""
import csv
import os
from typing import List, Dict, Any, Optional
from collections.abc import Callable
from pal.core.log import get_logger
from pal.services.tra import _get_tra_neto

logger = get_logger("EXPORTS")

# Función auxiliar para limpiar texto para Excel
def clean_for_excel(text):
    """Limpia texto para que sea compatible con Excel"""
    if text is None:
        return ""
    # Convertir a string y limpiar caracteres de control
    text = str(text)
    # Remover caracteres de control (0x00-0x1F excepto 0x09, 0x0A, 0x0D)
    cleaned = ''.join(char for char in text if ord(char) >= 32 or char in '\t\n\r')
    # Truncar si es muy largo (Excel tiene límite de 32767 caracteres por celda)
    return cleaned[:32760] if len(cleaned) > 32760 else cleaned

# Intentar importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl no está disponible. Instale con: pip install openpyxl")


def export_stock_csv(filename: str, datos_exportar: List, seleccionadas: List[str], 
                    location_groups: Dict[str, List[str]], db_manager, 
                    progress_cb: Optional[Callable[[int, int], None]] = None) -> int:
    """
    Exporta datos de stock a un archivo CSV con existencias por ubicación
    
    Args:
        filename: Nombre del archivo CSV a crear
        datos_exportar: Lista de datos de stock a exportar
        seleccionadas: Lista de ubicaciones seleccionadas para incluir
        location_groups: Diccionario con grupos de ubicaciones
        db_manager: Instancia del gestor de base de datos
        progress_cb: Callback opcional para reportar progreso
        
    Returns:
        int: Número total de registros exportados
    """
    from datetime import datetime
    try:
        total_registros = len(datos_exportar)
        logger.info(f"Iniciando exportación de {total_registros} registros a {filename}")
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # ENCABEZADO DESCRIPTIVO DEL REPORTE
            writer.writerow([f'REPORTE DE QUIEBRES DE STOCK - GENERADO EL {datetime.now().strftime("%d/%m/%Y a las %H:%M:%S")}'])
            writer.writerow([''])
            writer.writerow([f'Total de productos en alerta: {total_registros}'])
            #writer.writerow([f'Ubicaciones incluidas: {", ".join(seleccionadas)}'])
            writer.writerow([f'Depósitos consultados: {sum(len(location_groups[u]) for u in seleccionadas)}'])
            writer.writerow([''])
            writer.writerow(['NIVELES DE ALERTA:'])
            writer.writerow(['- CRÍTICA: Stock menor a 8 unidades'])
            writer.writerow(['- MEDIA: Stock entre 8 y 14 unidades'])
            writer.writerow(['- LEVE: Stock entre 15 y 20 unidades'])
            writer.writerow([''])
            writer.writerow(['DETALLE DE PRODUCTOS:'])
            writer.writerow([''])
            
            # Preparar headers de datos
            headers = ['Código Producto', 'Descripción del Producto', 'Stock Depósito Principal (0301)', 'Nivel de Alerta']
            
            # Agregar columnas por ubicación seleccionada
            for ubicacion in seleccionadas:
                headers.append(f'Existencias en {ubicacion}')
            
            writer.writerow(headers)
            # Fila separadora
            writer.writerow(['-' * 15 for _ in headers])
            
            for i, (codigo, desc, stock, nivel) in enumerate(datos_exportar):
                try:
                    # Fila base con datos principales (formato más legible)
                    nivel_texto = {
                        'critica': '⚠️ CRÍTICA',
                        'media': '🟡 MEDIA', 
                        'leve': '🟢 LEVE'
                    }.get(nivel.lower(), nivel.upper())
                    
                    fila = [
                        codigo,
                        desc,
                        f'{stock} unidades',
                        nivel_texto
                    ]
                    
                    # Agregar stock por ubicación
                    for ubicacion in seleccionadas:
                        deps = location_groups.get(ubicacion, [])
                        try:
                            # Importar función para obtener existencias
                            from pal.services.stock import get_existencias_por_ubicacion
                            existencias = get_existencias_por_ubicacion(db_manager, codigo, deps)
                            fila.append(f'{existencias} unidades' if existencias > 0 else 'Sin stock')
                        except Exception as e:
                            logger.warning(f"Error consultando stock en {ubicacion} para {codigo}: {e}")
                            fila.append('Error consulta')
                    
                    writer.writerow(fila)
                    
                    # Callback de progreso
                    if progress_cb:
                        progress_cb(i + 1, total_registros)
                        
                except Exception as e:
                    logger.error(f"Error procesando registro {i}: {codigo} - {e}")
                    continue
            
            # PIE DEL REPORTE
            writer.writerow([''])
            writer.writerow(['===== FIN DEL REPORTE ====='])
            writer.writerow([f'Archivo generado por: Sistema PAL (Proyecto Atención en Línea)'])
            writer.writerow([f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
        
        logger.info(f"Exportación completada: {total_registros} registros en {filename}")
        return total_registros
        
    except Exception as e:
        logger.error(f"Error en exportación CSV: {e}")
        raise


def export_tra_csv(filename: str, datos_tra: List, progress_cb: Optional[Callable[[int, int], None]] = None) -> int:
    """Exporta datos de RI (Rotación de Inventario) a un archivo CSV.
    
    Args:
        filename: Nombre del archivo CSV a crear
        datos_tra: Lista de datos TRA a exportar
        progress_cb: Callback opcional para reportar progreso
        
    Returns:
        int: Número total de registros exportados
    """
    from datetime import datetime
    try:
        total_registros = len(datos_tra)
        logger.info(f"Iniciando exportación RI CSV de {total_registros} registros a {filename}")
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # ENCABEZADO DESCRIPTIVO DEL REPORTE RI (Rotación de Inventario)
            writer.writerow([f'REPORTE DE ROTACIÓN DE INVENTARIO (RI) - {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
            writer.writerow([''])
            writer.writerow([f'Total de productos analizados: {total_registros}'])
            writer.writerow([''])
            writer.writerow(['CLASIFICACIÓN DE ROTACIÓN:'])
            writer.writerow(['- ALTA: Productos con alta rotación de ventas'])
            writer.writerow(['- MEDIA: Productos con rotación moderada'])
            writer.writerow(['- BAJA: Productos con baja rotación'])
            writer.writerow(['- SIN MOVIMIENTO: Productos sin ventas en el período'])
            writer.writerow([''])
            writer.writerow(['DETALLE DE PRODUCTOS:'])
            writer.writerow([''])
            
            headers = ['Código Producto', 'Descripción del Producto', 'Departamento', 'Grupo', 'Subgrupo', 'Ventas Netas', 'Clasificación de Rotación']
            writer.writerow(headers)
            writer.writerow(['-' * 20 for _ in headers])
            
            for i, fila in enumerate(datos_tra):
                try:
                    # Manejar diferentes longitudes de fila
                    if len(fila) >= 7:
                        codigo, desc, dept, grupo, sub, neto, rotacion = fila[:7]
                    elif len(fila) >= 6:
                        codigo, desc, dept, grupo, sub, neto = fila[:6]
                        rotacion = "SIN CLASIFICAR"
                    else:
                        logger.warning(f"Fila con formato incorrecto: {fila}")
                        continue
                    
                    # Formatear datos para mejor legibilidad
                    rotacion_texto = {
                        'alta': '🟢 ALTA ROTACIÓN',
                        'media': '🟡 ROTACIÓN MEDIA',
                        'baja': '🔴 BAJA ROTACIÓN',
                        'sin_movimiento': '⚫ SIN MOVIMIENTO'
                    }.get(str(rotacion).lower(), str(rotacion).upper())
                    
                    writer.writerow([
                        codigo, 
                        desc, 
                        dept or 'Sin clasificar', 
                        grupo or 'Sin clasificar', 
                        sub or 'Sin clasificar', 
                        f'${round(float(neto or 0), 2):,.2f}', 
                        rotacion_texto
                    ])
                    
                    # Callback de progreso
                    if progress_cb:
                        progress_cb(i + 1, total_registros)
                        
                except Exception as e:
                    logger.error(f"Error procesando fila TRA {i}: {fila} - {e}")
                    continue
            
            # PIE DEL REPORTE RI
            writer.writerow([''])
            writer.writerow(['===== FIN DEL REPORTE RI (Rotación de Inventario) ====='])
            writer.writerow([f'Archivo generado por: Sistema PAL (Proyecto Atención en Línea)'])
            writer.writerow([f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
        
        logger.info(f"Exportación RI CSV completada: {total_registros} registros en {filename}")
        return total_registros
        
    except Exception as e:
        logger.error(f"Error en exportación RI CSV: {e}")
        raise


def export_mbrp_csv(filename: str, datos_mbrp: List, progress_cb: Optional[Callable[[int, int], None]] = None) -> int:
    """
    Exporta datos MBRP a un archivo CSV
    
    Args:
        filename: Nombre del archivo CSV a crear
        datos_mbrp: Lista de datos MBRP a exportar
        progress_cb: Callback opcional para reportar progreso
        
    Returns:
        int: Número total de registros exportados
    """
    from datetime import datetime
    try:
        total_registros = len(datos_mbrp)
        logger.info(f"Iniciando exportación MBRP de {total_registros} registros a {filename}")
        
        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # ENCABEZADO DESCRIPTIVO DEL REPORTE MBRP
            writer.writerow([f'REPORTE DE MERCANCÍA DE BAJA ROTACIÓN DE PRODUCTOS (MBRP) - {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
            writer.writerow([''])
            writer.writerow([f'Total de productos analizados: {total_registros}'])
            writer.writerow([''])
            writer.writerow(['DESCRIPCIÓN DEL REPORTE:'])
            writer.writerow(['- Identifica productos con baja movilidad en inventario'])
            writer.writerow(['- Analiza relación entre ventas, compras y rentabilidad'])
            writer.writerow(['- Ayuda a optimizar la gestión de inventario'])
            writer.writerow([''])
            writer.writerow(['DETALLE DE PRODUCTOS:'])
            writer.writerow([''])
            
            headers = ['Código Producto', 'Descripción del Producto', 'Monto Vendido', 'Monto Comprado', 'Diferencia (Ganancia/Pérdida)', 'Margen de Ganancia %']
            writer.writerow(headers)
            writer.writerow(['-' * 25 for _ in headers])
            
            for i, fila in enumerate(datos_mbrp):
                try:
                    if len(fila) >= 6:
                        codigo, desc, vendido, comprado, diferencia, margen = fila[:6]
                        
                        # Formatear para mejor legibilidad
                        vendido_fmt = f'${round(float(vendido or 0), 2):,.2f}'
                        comprado_fmt = f'${round(float(comprado or 0), 2):,.2f}'
                        diferencia_val = round(float(diferencia or 0), 2)
                        diferencia_fmt = f'${diferencia_val:,.2f}' if diferencia_val >= 0 else f'-${abs(diferencia_val):,.2f}'
                        margen_val = round(float(margen or 0), 2)
                        margen_fmt = f'{margen_val:,.2f}%' + (' 🟢' if margen_val > 20 else ' 🟡' if margen_val > 5 else ' 🔴')
                        
                        writer.writerow([
                            codigo, 
                            desc, 
                            vendido_fmt,
                            comprado_fmt,
                            diferencia_fmt,
                            margen_fmt
                        ])
                    else:
                        logger.warning(f"Fila MBRP con formato incorrecto: {fila}")
                        continue
                    
                    # Callback de progreso
                    if progress_cb:
                        progress_cb(i + 1, total_registros)
                        
                except Exception as e:
                    logger.error(f"Error procesando fila MBRP: {fila} - {e}")
                    continue
            
            # PIE DEL REPORTE MBRP
            writer.writerow([''])
            writer.writerow(['INTERPRETACIÓN DE INDICADORES:'])
            writer.writerow(['🟢 Margen > 20%: Excelente rentabilidad'])
            writer.writerow(['🟡 Margen 5-20%: Rentabilidad aceptable'])
            writer.writerow(['🔴 Margen < 5%: Requiere atención'])
            writer.writerow([''])
            writer.writerow(['===== FIN DEL REPORTE MBRP ====='])
            writer.writerow([f'Archivo generado por: Sistema PAL (Proyecto Atención en Línea)'])
            writer.writerow([f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'])
        
        logger.info(f"Exportación MBRP completada: {total_registros} registros en {filename}")
        return total_registros
        
    except Exception as e:
        logger.error(f"Error en exportación MBRP CSV: {e}")
        raise


# ============== FUNCIONES DE EXPORTACIÓN EXCEL ==============

def export_stock_excel(filename: str, datos_exportar: List, seleccionadas: List[str], 
                      location_groups: Dict[str, List[str]], db_manager, 
                      progress_cb: Optional[Callable[[int, int], None]] = None,
                      current_localidad: Optional[str] = None,
                      permissions_manager=None,
                      current_user_id: int = None) -> int:
    """
    Exporta datos de stock a un archivo Excel con formato profesional,
    filtros, formato condicional y múltiples hojas de análisis.
    
    Args:
        filename: Nombre del archivo Excel a crear
        datos_exportar: Lista de datos de stock a exportar
        seleccionadas: Lista de ubicaciones seleccionadas
        location_groups: Diccionario con grupos de ubicaciones
        db_manager: Instancia del gestor de base de datos
        progress_cb: Callback opcional para reportar progreso
        
    Returns:
        int: Número total de registros exportados
    """
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl no está disponible. Use export_stock_csv en su lugar.")
        raise ImportError("openpyxl es requerido para exportación Excel")
        
    from datetime import datetime
    import time
    
    try:
        tiempo_inicio = time.time()
        total_registros = len(datos_exportar)
        logger.info(f"[EXPORT TIMER] Iniciando exportación Excel de {total_registros} registros a {filename}")
        
        # Crear workbook
        tiempo_pre_wb = time.time()
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        ws_main = wb.create_sheet("Alertas de Stock")
        logger.info(f"[EXPORT TIMER] Workbook preparado en {time.time() - tiempo_pre_wb:.2f}s")
        
        # Encabezado del reporte
        ws_main['A1'] = f'REPORTE DE ALERTAS DE STOCK'
        ws_main['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M:%S")}'
        ws_main['A4'] = f'Total de productos: {total_registros}'
        ws_main['A5'] = f'Ubicaciones: {", ".join(seleccionadas)}'
        
        # Formato del encabezado
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_main['A1'].font = header_font
        ws_main['A1'].fill = header_fill
        ws_main.merge_cells('A1:G1')
        
        # Obtener nombres de depósitos desde la BD antes de crear headers
        depositos_info = {}
        # Mapeos de jerarquía: Código -> Descripción
        dept_desc_map = {}
        group_desc_map = {}
        sub_desc_map = {}
        # Mapeo de producto -> jerarquía (código -> codes)
        product_hierarchy_map = {}
        
        # Mapeo de precios, costos y proveedores (just-in-time)
        precio_map_stock = {}
        costo_map_stock = {}
        impuestos_map_stock = {}
        proveedor_map_stock = {}

        mostrar_costo_utilidad = False
        mostrar_proveedores = False
        if permissions_manager and current_user_id:
            try:
                mostrar_costo_utilidad = permissions_manager.tiene_permiso(current_user_id, 'STOCK', 'ver_costo_utilidad')
                mostrar_proveedores = permissions_manager.tiene_permiso(current_user_id, 'STOCK', 'ver_proveedores')
            except Exception as e:
                logger.warning(f"Error verificando permisos en STOCK: {e}")
        
        try:
            if db_manager:
                # 1. Obtener depósitos
                try:
                    depositos_result = db_manager.obtener_depositos()
                    depositos_info = {codigo: descripcion for codigo, descripcion in depositos_result}
                except Exception: pass
                
                # 2. Obtener nombres de departamentos, grupos y subgrupos
                try:
                    depts = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS WITH (NOLOCK)")
                    dept_desc_map = {str(c).strip(): str(d).strip() for c, d in depts if c}
                    
                    groups = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WITH (NOLOCK)")
                    group_desc_map = {str(c).strip(): str(d).strip() for c, d in groups if c}
                    
                    subs = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WITH (NOLOCK)")
                    sub_desc_map = {str(c).strip(): str(d).strip() for c, d in subs if c}
                except Exception as e:
                    logger.warning(f"Error cargando nombres de jerarquía: {e}")

                # 3. Obtener códigos de jerarquía para los productos a exportar (en batch)
                codigos_export = []
                for item in datos_exportar:
                    if isinstance(item, dict): codigos_export.append(str(item.get('codigo', '')).strip())
                    else: codigos_export.append(str(item[0]).strip())
                
                if codigos_export:
                    # SQL Server parameter limit (approx 2100)
                    BATCH_SIZE = 1800
                    for i in range(0, len(codigos_export), BATCH_SIZE):
                        batch = codigos_export[i:i + BATCH_SIZE]
                        placeholders = ','.join('?' * len(batch))
                        
                        # 3.1 Metadata de jerarquía
                        query_h = f"""
                            SELECT RTRIM(LTRIM(C_CODIGO)), 
                                   COALESCE(C_DEPARTAMENTO, ''), 
                                   COALESCE(C_GRUPO, ''), 
                                   COALESCE(C_SUBGRUPO, ''), 
                                   COALESCE(c_marca, '')
                            FROM MA_PRODUCTOS WITH (NOLOCK)
                            WHERE C_CODIGO IN ({placeholders})
                        """
                        rows_h = db_manager.fetch_data(query_h, batch)
                        for r in rows_h:
                            product_hierarchy_map[r[0]] = {
                                'dept': r[1],
                                'group': r[2],
                                'sub': r[3],
                                'marca': r[4]
                            }

                        # 3.2 Precios, costos e impuestos (JIT)
                        query_eco = f"""
                            SELECT C_CODIGO, COALESCE(n_precio1, 0), COALESCE(n_costoact, 0), COALESCE(n_impuesto1, 0)
                            FROM MA_PRODUCTOS WITH (NOLOCK)
                            WHERE C_CODIGO IN ({placeholders})
                        """
                        rows_eco = db_manager.fetch_data(query_eco, batch) or []
                        for cod, p, c, imp in rows_eco:
                            c_str = str(cod).strip()
                            precio_map_stock[c_str] = float(p or 0)
                            costo_map_stock[c_str] = float(c or 0)
                            impuestos_map_stock[c_str] = float(imp or 0)

                        # 3.3 Último Proveedor (JIT)
                        if mostrar_proveedores:
                            try:
                                proveedor_map_stock = db_manager.obtener_ultimas_compras_bulk(batch)
                                logger.info(f"[EXPORT STOCK] Cargados {len(proveedor_map_stock)} proveedores para el batch")
                            except Exception as e:
                                logger.warning(f"Error cargando proveedores bulk en STOCK: {e}")
                
                logger.info(f"[EXPORT TIMER] Metadatos, Precios y Proveedores cargados")
        except Exception as e:
            logger.error(f"Error crítico cargando metadatos para exportación: {e}")
        
        # Definición de grupos de depósitos para alineación con TRA (RI)
        GRUPOS_DEPOSITOS = {
            'Cabudare': ['0301', '0302'],
            'Barinas':  ['0101', '0102', '0108'],
            'Guanare':  ['0401', '0402'],
            'CDT':      ['0106'],
            'Transito': ['0104', '0110', '0112']
        }
        
        # Columnas modernas alineadas con RI (TRA)
        headers = [
            'Código', 'Descripción', 'Departamento', 'Grupo', 'Subgrupo', 'Marca', 
            'Sede de Quiebre', 'Unid. Perdidas', 'Días Quiebre', 'Últ. Liquidación', 'Últ. Venta',
            'Stock Cabudare', 'Stock Barinas', 'Stock Guanare', 'CDT', 'Transito SEDES', 'Stock Total'
        ]

        if mostrar_proveedores:
            headers.append('Últ. Proveedor')
        
        # Columnas de análisis económico (según permisos)
        if mostrar_costo_utilidad:
            headers.extend(['Precio + IVA', 'Costo', 'Utilidad %'])
        
        start_row = 8
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos de productos - OPTIMIZADO: Obtener todas las existencias en una sola consulta
        tiempo_pre_datos = time.time()
        
        # Obtener todos los códigos de productos
        codigos_productos = []
        for item in datos_exportar:
            if isinstance(item, dict):
                codigos_productos.append(item.get('codigo', ''))
            else:
                codigos_productos.append(item[0])
        
        # Obtener existencias en BATCH con chunking para evitar límites de parámetros SQL
        logger.info(f"[EXPORT TIMER] Obteniendo existencias batch para {len(codigos_productos)} productos x {len(seleccionadas)} depósitos...")
        tiempo_pre_batch = time.time()
        
        try:
            # Inicializar mapa de existencias
            existencias_map = {codigo: {} for codigo in codigos_productos}
            for deposito in seleccionadas:
                for codigo in codigos_productos:
                    existencias_map[codigo][deposito] = 0
            
            # SQL Server tiene un límite de ~2100 parámetros
            # Con N depósitos, podemos usar (2000 - N) / 1 productos por chunk para seguridad
            max_params_per_chunk = 2000 - len(seleccionadas)
            chunk_size = max(100, max_params_per_chunk)  # Al menos 100, pero respetar límite
            
            total_chunks = (len(codigos_productos) + chunk_size - 1) // chunk_size
            logger.info(f"[EXPORT TIMER] Procesando en {total_chunks} chunks de máximo {chunk_size} productos")
            
            # Procesar en chunks con progreso
            result_batch = []
            # Reservar 10% del progreso total para la carga de chunks
            chunk_progress_weight = 0.10
            data_progress_weight = 0.90
            
            for chunk_idx in range(0, len(codigos_productos), chunk_size):
                chunk_productos = codigos_productos[chunk_idx:chunk_idx + chunk_size]
                chunk_num = (chunk_idx // chunk_size) + 1
                
                # Construir consulta para este chunk
                codigo_placeholders = ','.join('?' * len(chunk_productos))
                deposito_placeholders = ','.join('?' * len(seleccionadas))
                
                sql_chunk = (
                    f"SELECT c_codarticulo, c_coddeposito, ISNULL(SUM(n_cantidad), 0) as total "
                    f"FROM MA_DEPOPROD WITH (NOLOCK) "
                    f"WHERE c_codarticulo IN ({codigo_placeholders}) "
                    f"AND c_coddeposito IN ({deposito_placeholders}) "
                    f"GROUP BY c_codarticulo, c_coddeposito "
                    f"HAVING ISNULL(SUM(n_cantidad), 0) != 0"
                )
                
                params_chunk = chunk_productos + seleccionadas
                logger.info(f"[EXPORT TIMER] Ejecutando chunk {chunk_num}/{total_chunks} ({len(chunk_productos)} productos, {len(params_chunk)} params)")
                
                chunk_result = db_manager.fetch_data(sql_chunk, params_chunk)
                if chunk_result:
                    result_batch.extend(chunk_result)
                
                # Reportar progreso de carga de chunks (0-10% del total)
                if progress_cb:
                    chunk_progress = int((chunk_num / total_chunks) * chunk_progress_weight * total_registros)
                    progress_cb(chunk_progress, total_registros)
            
            # Poblar mapa con resultados
            if result_batch:
                for codigo, deposito, cantidad in result_batch:
                    if codigo in existencias_map:
                        existencias_map[codigo][deposito] = int(cantidad or 0)
            
            tiempo_post_batch = time.time()
            logger.info(f"[EXPORT TIMER] Existencias batch obtenidas en {tiempo_post_batch - tiempo_pre_batch:.2f}s ({total_chunks} consultas chunked, {len(result_batch)} resultados)")
            
        except Exception as e:
            logger.error(f"Error en consulta batch: {e}")
            existencias_map = {codigo: {dep: 0 for dep in seleccionadas} for codigo in codigos_productos}
        
        # Ahora procesar los datos usando el mapa precargado
        data_start_row = start_row + 1

        # Reordenar datos por unidades perdidas (descendente)
        def _local_stock(codigo):
            dep_qty = existencias_map.get(codigo, {})
            prefijo = '03' if current_localidad == 'Cabudare' else '01' if current_localidad == 'Barinas' else '04'
            deps_local = [d for d in seleccionadas if str(d).startswith(prefijo)]
            return sum(int(dep_qty.get(d, 0)) for d in deps_local)

        datos_exportar = sorted(
            datos_exportar,
            key=lambda r: (
                float(r.get('unidades_perdidas') if isinstance(r, dict) else (r[3] if len(r) > 3 else 0)),
                -_local_stock(r.get('codigo') if isinstance(r, dict) else r[0])
            ),
            reverse=True
        )

        # Precalcular depósitos seleccionados por sede (dinámico desde location_groups)
        if isinstance(location_groups, dict) and location_groups:
            selected_by_sede = {sede: [d for d in (location_groups.get(sede, []) or []) if d in seleccionadas] for sede in location_groups.keys()}
        else:
            selected_by_sede = {
                'Cabudare': [d for d in seleccionadas if str(d).startswith('03')],
                'Barinas':  [d for d in seleccionadas if str(d).startswith('01')],
                'Guanare':  [d for d in seleccionadas if str(d).startswith('04')],
            }
        
        for i, item in enumerate(datos_exportar):
            try:
                row = data_start_row + i
                # Manejar tanto diccionarios (nuevos) como tuplas (antiguos)
                if isinstance(item, dict):
                    codigo = item.get('codigo', '')
                    desc = item.get('descripcion', '')
                    sede_q = item.get('sede_detectada', '')
                    unid_p = item.get('unidades_perdidas', 0)
                    dias_q = item.get('dias_quiebre', 0)
                    u_comp = item.get('ultima_compra', '')
                    u_vent = item.get('ultima_venta', '')
                else:
                    # Estructura q: (codigo, descripcion, sede, unidades_perdidas, dias_quiebre, ultima_compra, ultima_venta)
                    codigo = item[0]
                    desc = item[1]
                    sede_q = item[2] if len(item) > 2 else ''
                    unid_p = item[3] if len(item) > 3 else 0
                    dias_q = item[4] if len(item) > 4 else 0
                    u_comp = item[5] if len(item) > 5 else ''
                    u_vent = item[6] if len(item) > 6 else ''

                # Resolver jerarquía desde el mapa interno
                h_info = product_hierarchy_map.get(str(codigo).strip(), {})
                dept_code = h_info.get('dept', '')
                group_code = h_info.get('group', '')
                sub_code = h_info.get('sub', '')
                brand = h_info.get('marca', '')

                # Resolver nombres desde mapeos
                dept = dept_desc_map.get(str(dept_code).strip(), dept_code)
                group = group_desc_map.get(str(group_code).strip(), group_code)
                sub = sub_desc_map.get(str(sub_code).strip(), sub_code)

                # Escribir 11 columnas fijas (Quiebre de Stock)
                from pal.services.exports import clean_for_excel
                ws_main.cell(row=row, column=1, value=clean_for_excel(codigo))
                ws_main.cell(row=row, column=2, value=clean_for_excel(desc))
                ws_main.cell(row=row, column=3, value=clean_for_excel(dept))
                ws_main.cell(row=row, column=4, value=clean_for_excel(group))
                ws_main.cell(row=row, column=5, value=clean_for_excel(sub))
                ws_main.cell(row=row, column=6, value=clean_for_excel(brand))
                # Sede de quiebre (col 7)
                ws_main.cell(row=row, column=7, value=clean_for_excel(sede_q))
                # Unid. Perdidas (col 8)
                ws_main.cell(row=row, column=8, value=float(unid_p))
                # Días Quiebre (col 9)
                ws_main.cell(row=row, column=9, value=int(dias_q))
                # Últs (col 10, 11)
                ws_main.cell(row=row, column=10, value=str(u_comp))
                ws_main.cell(row=row, column=11, value=str(u_vent))

                # Sumar por grupos alineados con TRA
                codigo_s = str(codigo).strip()
                dep_qty = existencias_map.get(codigo_s, {})
                s_cab = sum(int(dep_qty.get(d, 0)) for d in GRUPOS_DEPOSITOS['Cabudare'])
                s_bar = sum(int(dep_qty.get(d, 0)) for d in GRUPOS_DEPOSITOS['Barinas'])
                s_gua = sum(int(dep_qty.get(d, 0)) for d in GRUPOS_DEPOSITOS['Guanare'])
                s_cdt = sum(int(dep_qty.get(d, 0)) for d in GRUPOS_DEPOSITOS['CDT'])
                s_tra = sum(int(dep_qty.get(d, 0)) for d in GRUPOS_DEPOSITOS['Transito'])
                s_tot = s_cab + s_bar + s_gua + s_cdt + s_tra
                
                ws_main.cell(row=row, column=12, value=s_cab)
                ws_main.cell(row=row, column=13, value=s_bar)
                ws_main.cell(row=row, column=14, value=s_gua)
                ws_main.cell(row=row, column=15, value=s_cdt)
                ws_main.cell(row=row, column=16, value=s_tra)
                ws_main.cell(row=row, column=17, value=s_tot)

                current_dyn_col = 18
                if mostrar_proveedores:
                    ws_main.cell(row=row, column=current_dyn_col, value=proveedor_map_stock.get(codigo_s, 'SIN COMPRAS'))
                    current_dyn_col += 1

                # Columnas económicas
                if mostrar_costo_utilidad:
                    precio_base = precio_map_stock.get(codigo_s, 0.0)
                    costo_val = costo_map_stock.get(codigo_s, 0.0)
                    iva_pct = impuestos_map_stock.get(codigo_s, 0.0)
                    
                    precio_iva = precio_base * (1 + (iva_pct / 100))
                    utilidad_pct = ((precio_base - costo_val) / precio_base * 100) if precio_base > 0 else 0
                    
                    c_p = ws_main.cell(row=row, column=current_dyn_col, value=precio_iva)
                    c_p.number_format = '#,##0.00'
                    current_dyn_col += 1
                    
                    c_c = ws_main.cell(row=row, column=current_dyn_col, value=costo_val)
                    c_c.number_format = '#,##0.00'
                    current_dyn_col += 1
                    
                    c_u = ws_main.cell(row=row, column=current_dyn_col, value=utilidad_pct)
                    c_u.number_format = '0.00 "%"'
                    current_dyn_col += 1

                # Reportar progreso
                if progress_cb:
                    base_p = int(0.10 * total_registros)
                    curr_p = int(((i + 1) / total_registros) * 0.90 * total_registros)
                    progress_cb(base_p + curr_p, total_registros)
                    
            except Exception as e:
                logger.error(f"Error procesando registro {i}: {codigo} - {e}")
                continue
        
        tiempo_post_datos = time.time()
        logger.info(f"[EXPORT TIMER] Procesamiento de datos: {tiempo_post_datos - tiempo_pre_datos:.2f}s (1 consulta batch)")
        
        # Crear tabla con filtros
        end_col_letter = get_column_letter(len(headers))
        table_range = f"A{start_row}:{end_col_letter}{data_start_row + total_registros - 1}"
        from openpyxl.worksheet.table import Table, TableStyleInfo
        table = Table(displayName="TablaStock", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws_main.add_table(table)
        
        # Ajustar anchos de columna
        ws_main.column_dimensions['A'].width = 12  # Código
        ws_main.column_dimensions['B'].width = 40  # Descripción
        ws_main.column_dimensions['C'].width = 20  # Departamento
        ws_main.column_dimensions['D'].width = 20  # Grupo
        ws_main.column_dimensions['E'].width = 20  # Subgrupo
        ws_main.column_dimensions['F'].width = 20  # Marca
        ws_main.column_dimensions['G'].width = 18  # Sede de Quiebre
        ws_main.column_dimensions['H'].width = 15  # Unid. Perdidas
        ws_main.column_dimensions['I'].width = 12  # Días Quiebre
        ws_main.column_dimensions['J'].width = 18  # Últ. Liquidación
        ws_main.column_dimensions['K'].width = 18  # Últ. Venta
        
        # Columnas de Stock y económicas
        for col_idx in range(12, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws_main.column_dimensions[col_letter].width = 15
        
        # === FIN DE REPORTE ===
        # Guardar archivo
        tiempo_pre_save = time.time()
        wb.save(filename)
        tiempo_post_save = time.time()
        logger.info(f"[EXPORT TIMER] Guardado de archivo: {tiempo_post_save - tiempo_pre_save:.2f}s")
        
        tiempo_total = tiempo_post_save - tiempo_inicio
        logger.info(f"[EXPORT TIMER] TOTAL exportación Excel: {tiempo_total:.2f}s para {total_registros} registros")
        logger.info(f"Exportación Excel completada: {total_registros} registros en {filename}")
        return total_registros
        
    except Exception as e:
        logger.error(f"Error en exportación Excel: {e}")
        raise


def export_tra_excel(filename: str, datos_tra: List, db_manager=None, progress_cb: Optional[Callable[[int, int], None]] = None, 
                     permissions_manager=None, current_user_id: int = None,
                     provider_label: Optional[str] = None,
                     sede_codigo: Optional[str] = None,
                     fecha_inicio=None, fecha_fin=None) -> int:
    """Exporta datos de RI (Rotación de Inventario) a un archivo Excel con formato profesional y múltiples hojas de análisis.
    
    Args:
        filename: Nombre del archivo Excel a crear
        datos_tra: Lista de datos TRA a exportar
        db_manager: Gestor de base de datos
        progress_cb: Callback opcional para reportar progreso
        permissions_manager: Gestor de permisos (opcional, para verificar ver_costo_utilidad)
        current_user_id: ID del usuario actual (opcional, para verificar permisos)
        fecha_inicio: Fecha inicio del periodo (datetime/date)
        fecha_fin: Fecha fin del periodo (datetime/date)
        
    Returns:
        int: Número total de registros exportados
    """
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl no está disponible. Use export_tra_csv en su lugar.")
        raise ImportError("openpyxl es requerido para exportación Excel")
        
    from datetime import datetime
    try:
        total_registros = len(datos_tra)
        logger.info(f"Iniciando exportación RI Excel de {total_registros} registros a {filename}")

        # Calcular días del periodo para "Estado Stock"
        dias_periodo = 1
        if fecha_inicio and fecha_fin:
            try:
                # Asegurar compatibilidad si son strings o dates
                fi = fecha_inicio if hasattr(fecha_inicio, 'year') else datetime.now() # Fallback simple
                ff = fecha_fin if hasattr(fecha_fin, 'year') else datetime.now()
                # Si son date, convertir a datetime o restar directo
                diff = (ff - fi).days
                dias_periodo = diff if diff > 0 else 1
            except Exception as e:
                logger.warning(f"Error calculando días de periodo: {e}")
                dias_periodo = 1

        # Determinar si el filtro de sede es global (modo ICH)
        ich_mode = sede_codigo in (None, '%', '00', 'ICH', 'ALL')
        
        # Calcular total de ventas una sola vez usando helper robusto
        try:
            total_ventas = sum(_get_tra_neto(f) for f in datos_tra if f)
        except Exception as e:
            logger.warning(f"[EXPORT TRA] Error calculando total_ventas, se usará 0: {e}")
            total_ventas = 0.0
        
        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto

        mostrar_costo_utilidad = False
        mostrar_proveedores = False
        if permissions_manager and current_user_id:
            try:
                mostrar_costo_utilidad = permissions_manager.tiene_permiso(current_user_id, 'TRA', 'ver_costo_utilidad')
                logger.info(f"Permiso ver_costo_utilidad para TRA: {mostrar_costo_utilidad}")
                mostrar_proveedores = permissions_manager.tiene_permiso(current_user_id, 'TRA', 'ver_proveedores')
                logger.info(f"Permiso ver_proveedores para TRA: {mostrar_proveedores}")
            except Exception as e:
                logger.warning(f"Error verificando permisos TRA: {e}")
        
        # Crear diccionarios de mapeo código -> descripción para jerarquía
        dept_desc_map = {}
        group_desc_map = {}
        sub_desc_map = {}
        # Mapa de impuestos por producto (código -> n_impuesto1)
        impuestos_map = {}
        
        if db_manager and db_manager.ensure_connection():
            try:
                # Obtener departamentos
                depts = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL")
                dept_desc_map = {codigo: desc for codigo, desc in depts if codigo and desc}
                
                # Obtener grupos
                groups = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL")
                group_desc_map = {codigo: desc for codigo, desc in groups if codigo and desc}
                
                # Obtener subgrupos
                subs = db_manager.fetch_data("SELECT C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL")
                sub_desc_map = {codigo: desc for codigo, desc in subs if codigo and desc}
                
                # Obtener marcas
                marcas = db_manager.fetch_data("SELECT RTRIM(LTRIM(C_CODIGO)), COALESCE(c_marca, '') FROM MA_PRODUCTOS WITH (NOLOCK)")
                marca_map = {str(codigo).strip(): marca for codigo, marca in marcas if codigo}
                
                logger.info(f"Mapeos cargados - Departamentos: {len(dept_desc_map)}, Grupos: {len(group_desc_map)}, Subgrupos: {len(sub_desc_map)}, Marcas: {len(marca_map)}")
            except Exception as e:
                logger.warning(f"No se pudieron cargar las descripciones de jerarquía/marcas: {e}")
            
            # Cargar impuestos (IVA) por producto desde MA_PRODUCTOS.n_impuesto1
            codigos_productos = sorted({
                str(f[0]).strip() for f in datos_tra
                if f and len(f) > 0 and f[0] is not None
            })
            
            # Mapa de precios y costos (just-in-time para evitar persistencia de datos sensibles)
            precio_map_tra = {}
            costo_map_tra = {}
            
            try:
                if codigos_productos:
                    # Evitar límite de 2100 parámetros de SQL Server
                    BATCH_SIZE = 1800
                    for i in range(0, len(codigos_productos), BATCH_SIZE):
                        batch = codigos_productos[i:i + BATCH_SIZE]
                        placeholders = ','.join('?' * len(batch))
                        
                        # Cargar impuestos
                        query_imp = f"""
                            SELECT C_CODIGO, COALESCE(n_impuesto1, 0) AS impuesto1
                            FROM MA_PRODUCTOS WITH (NOLOCK)
                            WHERE C_CODIGO IN ({placeholders})
                        """
                        rows_imp = db_manager.fetch_data(query_imp, batch) or []
                        for cod, imp in rows_imp:
                            try:
                                impuestos_map[str(cod).strip()] = float(imp or 0)
                            except Exception:
                                continue
                                
                        # Cargar precios y costos just-in-time si el usuario tiene permiso
                        if mostrar_costo_utilidad:
                            query_eco = f"""
                                SELECT C_CODIGO, COALESCE(n_precio1, 0), COALESCE(n_costoact, 0)
                                FROM MA_PRODUCTOS WITH (NOLOCK)
                                WHERE C_CODIGO IN ({placeholders})
                            """
                            rows_eco = db_manager.fetch_data(query_eco, batch) or []
                            for cod, p, c in rows_eco:
                                try:
                                    c_str = str(cod).strip()
                                    precio_map_tra[c_str] = float(p or 0)
                                    costo_map_tra[c_str] = float(c or 0)
                                except Exception:
                                    continue
                    
                    logger.info(f"[EXPORT TRA] Datos económicos cargados - Impuestos: {len(impuestos_map)}, Precios/Costos: {len(precio_map_tra)}")
            except Exception as e:
                logger.warning(f"No se pudieron cargar los datos económicos para TRA: {e}")
        
        # === HOJA 1: DATOS PRINCIPALES RI ===
        ws_main = wb.create_sheet("Datos RI")
        
        # Encabezado del reporte
        ws_main['A1'] = f'REPORTE DE ROTACIÓN DE INVENTARIO (RI)'
        ws_main['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M:%S")}'
        if provider_label:
            ws_main['A3'] = f'Proveedor: {provider_label}'
        ws_main['A4'] = f'Total de productos analizados: {total_registros}'
        
        # Formato del encabezado
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_main['A1'].font = header_font
        ws_main['A1'].fill = header_fill
        
        
        # Helper para obtener stock actual por producto
        def _get_stock_actual_bulk_tra(codigos: List[str], deposito: Optional[str]) -> Dict[str, int]:
            """Obtiene el stock actual por código para TRA (optimizado)."""
            resultados: Dict[str, int] = {}
            if not db_manager or not db_manager.ensure_connection() or not codigos:
                return resultados
            try:
                MAX_IN = 2000
                global_query = deposito in (None, '%', '00', 'ICH', 'ALL')
                
                for cod in codigos:
                    resultados[str(cod)] = 0
                
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    if global_query:
                        sql = (
                            f"SELECT c_codarticulo, SUM(n_cantidad) "
                            f"FROM MA_DEPOPROD WITH (NOLOCK) "
                            f"WHERE c_codarticulo IN ({placeholders}) "
                            f"GROUP BY c_codarticulo"
                        )
                        params = chunk
                    else:
                        sql = (
                            f"SELECT c_codarticulo, SUM(n_cantidad) "
                            f"FROM MA_DEPOPROD WITH (NOLOCK) "
                            f"WHERE c_coddeposito = ? AND c_codarticulo IN ({placeholders}) "
                            f"GROUP BY c_codarticulo"
                        )
                        params = [deposito] + chunk
                    rows = db_manager.fetch_data(sql, params)
                    for cod, sum_qty in (rows or []):
                        try:
                            resultados[str(cod)] = int(sum_qty or 0)
                        except Exception:
                            pass
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT TRA] Error obteniendo stock actual: {e}")
                return resultados
        
        def _map_deposito_to_sede_tra(dep: str) -> str:
            """Mapea un código de depósito a una sede legible."""
            try:
                c = (dep or "").strip()
                if c.startswith('03'):
                    return 'Cabudare'
                if c.startswith('01'):
                    return 'Barinas'
                if c.startswith('04'):
                    return 'Guanare'
                return 'Otra'
            except Exception:
                return 'Otra'
        
        def _get_stock_por_sede_tra(codigos: List[str]) -> Dict[str, Dict[str, int]]:
            """
            Obtiene distribución de stock por sede/grupo para cada código,
            agrupando múltiples depósitos según reglas específicas.
            """
            resultados: Dict[str, Dict[str, int]] = {}
            if not db_manager or not db_manager.ensure_connection() or not codigos:
                return resultados
            try:
                # Definir grupos de depósitos
                # Barinas: 0101, 0102, 0108
                # Cabudare: 0301, 0302
                # Guanare: 0401, 0402
                # CDT: 0106
                # Transito SEDES: 0104, 0110, 0112
                GRUPOS_DEPOSITOS = {
                    'Barinas':  ['0101', '0102', '0108'],
                    'Cabudare': ['0301', '0302'],
                    'Guanare':  ['0401', '0402'],
                    'CDT':      ['0106'],
                    'Transito': ['0104', '0110', '0112']
                }

                # Crear un mapa inverso: deposito -> nombre_grupo
                DEPOSITO_A_GRUPO = {}
                for grupo, lista_deps in GRUPOS_DEPOSITOS.items():
                    for d in lista_deps:
                        DEPOSITO_A_GRUPO[d] = grupo
                
                MAX_IN = 2000
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    
                    # Filtrar solo por depósitos que nos interesan
                    todos_deps = list(DEPOSITO_A_GRUPO.keys())
                    depos_placeholders = "'" + "','".join(todos_deps) + "'"
                    
                    sql = (
                        f"SELECT c_codarticulo, c_coddeposito, SUM(n_cantidad) "
                        f"FROM MA_DEPOPROD WITH (NOLOCK) "
                        f"WHERE c_codarticulo IN ({placeholders}) "
                        f"AND c_coddeposito IN ({depos_placeholders}) "
                        f"GROUP BY c_codarticulo, c_coddeposito"
                    )
                    rows = db_manager.fetch_data(sql, chunk) or []
                    for cod, dep, qty in rows:
                        try:
                            cod_str = str(cod).strip()
                            dep_str = str(dep).strip()
                            
                            # Identificar grupo
                            grupo_nombre = DEPOSITO_A_GRUPO.get(dep_str)
                            if not grupo_nombre:
                                continue
                            
                            q = int(qty or 0)
                        except Exception:
                            continue
                        
                        if cod_str not in resultados:
                            resultados[cod_str] = {}
                        
                        # Sumar al acumulador del grupo
                        resultados[cod_str][grupo_nombre] = resultados[cod_str].get(grupo_nombre, 0) + q
                
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT TRA] Error obteniendo stock por sede: {e}")
                return resultados
        
        # Obtener códigos únicos para consultar stock
        codigos_unicos_tra = sorted({
            str(f[0]).strip() for f in datos_tra
            if f and len(f) > 0 and f[0] is not None
        })
        
        # Cargar stock según modo ICH o sede específica
        stock_map_tra: Dict[str, int] = {}
        stock_por_sede_map_tra: Dict[str, Dict[str, int]] = {}
        if codigos_unicos_tra and db_manager:
            sede_stock = sede_codigo or '0301'
            stock_map_tra = _get_stock_actual_bulk_tra(codigos_unicos_tra, sede_stock)
            if ich_mode:
                stock_por_sede_map_tra = _get_stock_por_sede_tra(codigos_unicos_tra)
            
            # Cargar últimos proveedores si se tiene permiso
            proveedores_map_tra = {}
            if mostrar_proveedores:
                try:
                    proveedores_map_tra = db_manager.obtener_ultimas_compras_bulk(codigos_unicos_tra)
                    logger.info(f"[EXPORT TRA] Cargados {len(proveedores_map_tra)} proveedores")
                except Exception as e:
                    logger.warning(f"Error cargando proveedores para export: {e}")
        
        # Headers de la tabla (fila 7) - Orden: Ventas Netas, Stock/Stocks, Estado Stock, Rotación
        headers = ['Código', 'Descripción', 'Marca', 'Departamento', 'Grupo', 'Subgrupo', 'Ventas Netas']
        
        # Agregar columnas de stock según modo ICH
        if ich_mode:
                headers.extend(['Stock Cabudare', 'Stock Barinas', 'Stock Guanare', 'CDT', 'Transito SEDES', 'Stock Total'])
        else:
            headers.append('Stock')
        
        # NUEVA COLUMNA: Estado Stock
        headers.append('Estado Stock')
        
        # Agregar Rotación y Representación %
        headers.extend(['Rotación', 'Representación %'])
        
        if mostrar_costo_utilidad:
            # Mostrar precio con IVA incluido (sin exponer la columna % IVA)
            headers.extend(['Precio + IVA', 'Costo', 'Utilidad %'])
        
        if mostrar_proveedores:
            headers.append('Último Proveedor')
        
        # Ajustar merge del título según cantidad total de columnas
        from openpyxl.utils import get_column_letter as _col_letter
        last_col_letter = _col_letter(len(headers))
        merge_range = f'A1:{last_col_letter}1'
        ws_main.merge_cells(merge_range)
        
        start_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos de productos
        data_start_row = start_row + 1
        logger.info(f"[EXPORT DEBUG] Iniciando procesamiento de {len(datos_tra)} filas")
        
        for i, fila in enumerate(datos_tra):
            try:
                row = data_start_row + i
                # Extraer código base primero para resolución JIT
                codigo = fila[0] if fila and len(fila) > 0 else None
                if codigo is None:
                    continue
                codigo_str = str(codigo).strip()
                
                # Manejar diferentes longitudes de fila según estructura real:
                if len(fila) >= 10:
                    codigo, desc, dept, grupo, sub, neto = fila[:6]
                    rotacion = fila[6]
                    precio = fila[7]
                    costo = fila[9]
                elif len(fila) == 9:
                    codigo, desc, dept, grupo, sub, neto, rotacion, precio, costo = fila[:9]
                elif len(fila) == 8:
                    codigo, desc, dept, grupo, sub, neto, precio, costo = fila[:8]
                    rotacion = "SIN CLASIFICAR"
                elif len(fila) == 7:
                    codigo, desc, dept, grupo, sub, neto, rotacion = fila[:7]
                    precio = costo = 0
                elif len(fila) >= 6:
                    codigo, desc, dept, grupo, sub, neto = fila[:6]
                    rotacion = "SIN CLASIFICAR"
                    precio = costo = 0
                else:
                    continue
                
                # Resolucion just-in-time de precios/costos si vienen en 0 (Nodo Maestro)
                if mostrar_costo_utilidad:
                    if precio == 0:
                        precio = precio_map_tra.get(codigo_str, 0)
                    if costo == 0:
                        costo = costo_map_tra.get(codigo_str, 0)
                
                # Convertir códigos a descripciones legibles
                dept_desc = dept_desc_map.get(dept, dept) if dept else 'Sin clasificar'
                grupo_desc = group_desc_map.get(grupo, grupo) if grupo else 'Sin clasificar'
                sub_desc = sub_desc_map.get(sub, sub) if sub else 'Sin clasificar'
                marca_val = marca_map.get(str(codigo).strip(), '')
                
                ws_main.cell(row=row, column=1, value=clean_for_excel(codigo))
                ws_main.cell(row=row, column=2, value=clean_for_excel(desc))
                ws_main.cell(row=row, column=3, value=clean_for_excel(marca_val)) # Marca
                ws_main.cell(row=row, column=4, value=clean_for_excel(dept_desc))
                ws_main.cell(row=row, column=5, value=clean_for_excel(grupo_desc))
                ws_main.cell(row=row, column=6, value=clean_for_excel(sub_desc))
                
                # Usar helper robusto para obtener el neto
                neto_valor = _get_tra_neto(fila)
                neto_formateado = int(neto_valor) if neto_valor == int(neto_valor) else round(neto_valor, 2)
                
                # Columna 7: Ventas Netas
                ws_main.cell(row=row, column=7, value=neto_formateado)
                
                # Obtener código para consultar stock
                try:
                    codigo_str = str(codigo).strip() if codigo is not None else ''
                except Exception:
                    codigo_str = ''
                
                current_col = 8 # Starts after Ventas Netas (Col 7)
                stock_para_calculo = 0
                
                # Columnas de stock según modo ICH
                if ich_mode:
                    dist = stock_por_sede_map_tra.get(codigo_str, {})
                    stock_cabudare = dist.get('Cabudare', 0)
                    stock_barinas = dist.get('Barinas', 0)
                    stock_guanare = dist.get('Guanare', 0)
                    stock_cdt = dist.get('CDT', 0)
                    stock_transito = dist.get('Transito', 0)
                    stock_total = int(stock_map_tra.get(codigo_str, 0) or 0)
                    
                    ws_main.cell(row=row, column=current_col, value=stock_cabudare)
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=stock_barinas)
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=stock_guanare)
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=stock_cdt)
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=stock_transito)
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=stock_total)
                    current_col += 1
                    stock_para_calculo = stock_total
                else:
                    stock_actual = int(stock_map_tra.get(codigo_str, 0) or 0)
                    ws_main.cell(row=row, column=current_col, value=stock_actual)
                    current_col += 1
                    stock_para_calculo = stock_actual
                
                # CALCULO DE ESTADO STOCK
                estado_stock_texto = "N/A"
                if dias_periodo > 0:
                    promedio_diario = neto_valor / dias_periodo
                    if promedio_diario > 0:
                        dias_restantes = stock_para_calculo / promedio_diario
                        if dias_restantes < 25:
                            estado_stock_texto = "Posible quiebre"
                        elif 25 <= dias_restantes <= 59:
                            estado_stock_texto = "Alerta Compra"
                        elif 60 <= dias_restantes <= 90:
                            estado_stock_texto = "Optimo"
                        elif 91 <= dias_restantes <= 119:
                            estado_stock_texto = "Critico"
                        else:
                            estado_stock_texto = "Sobre Stock"
                    else:
                        if stock_para_calculo <= 0:
                            estado_stock_texto = "Sin Stock/Ventas"
                        else:
                             estado_stock_texto = "Sin Ventas" # Stock > 0 pero ventas 0 = infinito
                
                # Escribir columna Estado Stock
                ws_main.cell(row=row, column=current_col, value=estado_stock_texto)
                current_col += 1

                # Columna siguiente: Rotación
                ws_main.cell(row=row, column=current_col, value=clean_for_excel(str(rotacion).upper()))
                current_col += 1
                
                # Columna siguiente: Representación %
                porcentaje = (neto_valor / total_ventas * 100) if total_ventas > 0 else 0
                ws_main.cell(row=row, column=current_col, value=round(porcentaje, 2))
                current_col += 1
                
                # Agregar costo, precio (con IVA) y utilidad si el usuario tiene permiso
                if mostrar_costo_utilidad:
                    try:
                        precio_base = float(precio or 0)
                    except (ValueError, TypeError):
                        precio_base = 0.0
                    try:
                        costo_val = float(costo or 0)
                    except (ValueError, TypeError):
                        costo_val = 0.0
                    
                    iva_pct = float(impuestos_map.get(codigo_str, 0.0)) if impuestos_map else 0.0
                    # Precio con IVA incluido (si n_impuesto1 es porcentaje, por ejemplo 16 para 16%)
                    precio_con_iva = precio_base * (1.0 + (iva_pct / 100.0)) if precio_base > 0 else 0.0
                    
                    # Precio + IVA
                    cell_precio = ws_main.cell(row=row, column=current_col, value=precio_con_iva)
                    cell_precio.number_format = '0.00'
                    current_col += 1
                    # Costo
                    ws_main.cell(row=row, column=current_col, value=round(costo_val, 2))
                    current_col += 1
                    
                    # Calcular utilidad porcentual
                    if precio_base > 0:
                        utilidad_raw = (costo_val / precio_base) * 100 - 100
                        utilidad_pct = abs(utilidad_raw)
                        ws_main.cell(row=row, column=current_col, value=round(utilidad_pct, 2))
                        current_col += 1
                    else:
                        ws_main.cell(row=row, column=current_col, value=0)
                        current_col += 1
                
                # Columna Último Proveedor
                if mostrar_proveedores:
                    prov_nombre = proveedores_map_tra.get(codigo_str, 'Sin proveedor')
                    ws_main.cell(row=row, column=current_col, value=clean_for_excel(prov_nombre))
                    current_col += 1
                
                if progress_cb:
                    progress_cb(i + 1, total_registros)
                    
            except Exception as e:
                logger.error(f"Error procesando fila TRA {i}: {fila} - {e}")
                continue
        
        # Crear tabla con filtros
        end_col_letter = get_column_letter(len(headers))
        table_range = f"A{start_row}:{end_col_letter}{data_start_row + total_registros - 1}"
        table = Table(displayName="TablaTRA", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        ws_main.add_table(table)
        
        # Formato condicional para rotación (columna después de estado stock)
        # Sin ICH: Código(1), Descripción(2), Marca(3), Dept(4), Grupo(5), Subgrupo(6), Ventas Netas(7), Stock(8), Estado Stock(9), Rotación(10)
        # Con ICH: Código(1), Descripción(2), Marca(3), Dept(4), Grupo(5), Subgrupo(6), Ventas Netas(7), Stock Cabudare(8), ..., Stock Total(13), Estado Stock(14), Rotación(15)
        rotacion_col = 10 if not ich_mode else 15
        rotacion_col_letter = get_column_letter(rotacion_col)
        rotacion_range = f"{rotacion_col_letter}{data_start_row}:{rotacion_col_letter}{data_start_row + total_registros - 1}"
        
        # Alta = Verde
        ws_main.conditional_formatting.add(rotacion_range, 
            CellIsRule(operator='equal', formula=['"ALTA"'], 
                      fill=PatternFill(start_color="4CAF50", end_color="4CAF50")))
        
        # Media = Amarillo  
        ws_main.conditional_formatting.add(rotacion_range,
            CellIsRule(operator='equal', formula=['"MEDIA"'],
                      fill=PatternFill(start_color="FF9800", end_color="FF9800")))
        
        # Baja = Rojo
        ws_main.conditional_formatting.add(rotacion_range,
            CellIsRule(operator='equal', formula=['"BAJA"'],
                      fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B")))
                      
        # Formato condicional para Estado Stock (columna anterior a Rotación)
        estado_col = rotacion_col - 1
        estado_col_letter = get_column_letter(estado_col)
        estado_range = f"{estado_col_letter}{data_start_row}:{estado_col_letter}{data_start_row + total_registros - 1}"
        
        # Posible quiebre (<25) = Rojo
        ws_main.conditional_formatting.add(estado_range,
            CellIsRule(operator='equal', formula=['"Posible quiebre"'],
                      fill=PatternFill(start_color="FF6B6B", end_color="FF6B6B")))
        
        # Alerta Compra (25-59) = Amarillo
        ws_main.conditional_formatting.add(estado_range,
            CellIsRule(operator='equal', formula=['"Alerta Compra"'],
                      fill=PatternFill(start_color="FFD93D", end_color="FFD93D")))
                      
        # Optimo (60-90) = Verde
        ws_main.conditional_formatting.add(estado_range,
            CellIsRule(operator='equal', formula=['"Optimo"'],
                      fill=PatternFill(start_color="6BCF7F", end_color="6BCF7F")))
        
        # Ajustar anchos de columna
        ws_main.column_dimensions['A'].width = 12  # Código
        ws_main.column_dimensions['B'].width = 40  # Descripción
        ws_main.column_dimensions['C'].width = 15  # Marca
        ws_main.column_dimensions['D'].width = 20  # Departamento
        ws_main.column_dimensions['E'].width = 20  # Grupo
        ws_main.column_dimensions['F'].width = 20  # Subgrupo
        ws_main.column_dimensions['G'].width = 15  # Ventas Netas
        
        # Determinar letra de columna Último Proveedor y ajustar su ancho a 350 píxeles
        # En openpyxl, el ancho no es en píxeles, sino en caracteres aproximados.
        # 350 píxeles / 7 píxeles por carácter promedio ≈ 50 unidades de ancho.
        # Ajustar ancho Último Proveedor a 350 píxeles
        try:
            if mostrar_proveedores:
                prov_col_idx = len(headers)
                prov_col_letter = get_column_letter(prov_col_idx)
                ws_main.column_dimensions[prov_col_letter].width = 50 # Aprox 350px
        except Exception:
            pass

        if ich_mode:
            ws_main.column_dimensions['H'].width = 15  # Stock Cabudare
            ws_main.column_dimensions['I'].width = 15  # Stock Barinas
            ws_main.column_dimensions['J'].width = 15  # Stock Guanare
            ws_main.column_dimensions['K'].width = 15  # Stock Total
            ws_main.column_dimensions['L'].width = 12  # Rotación
            ws_main.column_dimensions['M'].width = 15  # Representación %
            
            if mostrar_costo_utilidad:
                ws_main.column_dimensions['N'].width = 15  # Precio + IVA
                ws_main.column_dimensions['O'].width = 15  # Costo
                ws_main.column_dimensions['P'].width = 15  # Utilidad %
        else:
            ws_main.column_dimensions['H'].width = 15  # Stock
            ws_main.column_dimensions['I'].width = 12  # Rotación
            ws_main.column_dimensions['J'].width = 15  # Representación %
            
            if mostrar_costo_utilidad:
                ws_main.column_dimensions['K'].width = 15  # Precio + IVA
                ws_main.column_dimensions['L'].width = 15  # Costo
                ws_main.column_dimensions['M'].width = 15  # Utilidad %
        
        # === HOJA 2: RESUMEN POR ROTACIÓN ===
        ws_summary = wb.create_sheet("Resumen por Rotación")
        
        # Contar por rotación
        rotacion_counts = {'ALTA': 0, 'MEDIA': 0, 'BAJA': 0, 'SIN MOVIMIENTO': 0, 'SIN CLASIFICAR': 0}
        for fila in datos_tra:
            if len(fila) >= 7:
                rotacion = str(fila[6]).upper()
                if rotacion in rotacion_counts:
                    rotacion_counts[rotacion] += 1
            else:
                rotacion_counts['SIN CLASIFICAR'] += 1
        
        # Crear tabla resumen
        ws_summary['A1'] = 'RESUMEN POR ROTACIÓN TRA'
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary.merge_cells('A1:C1')
        
        ws_summary['A3'] = 'Rotación'
        ws_summary['B3'] = 'Cantidad'
        ws_summary['C3'] = 'Porcentaje'
        
        for i, cell in enumerate(['A3', 'B3', 'C3']):
            ws_summary[cell].font = Font(bold=True, color="FFFFFF")
            ws_summary[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4")
        
        row = 4
        for rotacion, count in rotacion_counts.items():
            porcentaje = (count / total_registros * 100) if total_registros > 0 else 0
            ws_summary.cell(row=row, column=1, value=rotacion)
            ws_summary.cell(row=row, column=2, value=count)
            ws_summary.cell(row=row, column=3, value=f"{porcentaje:.1f}%")
            row += 1
        
        # === HOJA 3: PRODUCTOS DE BAJA ROTACIÓN ===
        ws_low = wb.create_sheet("Baja Rotación")
        
        productos_baja = [fila for fila in datos_tra if len(fila) >= 7 and str(fila[6]).upper() == 'BAJA']
        
        ws_low['A1'] = f'PRODUCTOS DE BAJA ROTACIÓN ({len(productos_baja)} productos)'
        ws_low['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_low['A1'].fill = PatternFill(start_color="F44336", end_color="F44336")
        ws_low.merge_cells('A1:E1')
        
        # Headers
        headers_baja = ['Código', 'Descripción', 'Ventas Netas', 'Departamento', 'Acción Recomendada']
        for col, header in enumerate(headers_baja, 1):
            cell = ws_low.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800")
        
        # Datos de baja rotación
        for i, fila in enumerate(productos_baja, 4):
            # Convertir código de departamento a descripción
            dept_codigo = fila[2] if len(fila) > 2 else None
            dept_desc = dept_desc_map.get(dept_codigo, dept_codigo) if dept_codigo else 'Sin clasificar'
            
            ws_low.cell(row=i, column=1, value=clean_for_excel(fila[0]))
            ws_low.cell(row=i, column=2, value=clean_for_excel(fila[1]))
            # Aplicar el mismo formateo que en la interfaz
            neto_valor = float(fila[5] or 0)
            neto_formateado = int(neto_valor) if neto_valor == int(neto_valor) else round(neto_valor, 2)
            ws_low.cell(row=i, column=3, value=neto_formateado)
            ws_low.cell(row=i, column=4, value=clean_for_excel(dept_desc))
            ws_low.cell(row=i, column=5, value="REVISAR ESTRATEGIA")
        
        # Ajustar columnas
        ws_low.column_dimensions['A'].width = 12
        ws_low.column_dimensions['B'].width = 40
        ws_low.column_dimensions['C'].width = 15
        ws_low.column_dimensions['D'].width = 20
        ws_low.column_dimensions['E'].width = 20
        
        # === HOJA 4: RESUMEN JERÁRQUICO (Depto → Grupo → Sub) ===
        try:
            from openpyxl.chart import PieChart, BarChart, Reference
            chart_available = True
        except Exception as chart_err:
            logger.warning(f"No se pudo importar openpyxl.chart: {chart_err}")
            PieChart = None
            BarChart = None
            Reference = None
            chart_available = False
        
        ws_h = wb.create_sheet("Resumen Jerárquico")
        ws_h['A1'] = 'RESUMEN JERÁRQUICO DE REPRESENTACIÓN (RI)'
        ws_h['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws_h['A1'].fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_h.merge_cells('A1:H1')
        
        # Función interna para escribir una tabla de resumen y opcionalmente un gráfico
        def _write_summary(title, rows, start_row, chart_type='pie'):
            """Escribe tabla con columnas: Elemento, Neto, Porcentaje y agrega gráfico a la derecha.
            
            Args:
                title: Título de la sección
                rows: Lista de tuplas (nombre, valor)
                start_row: Fila donde comenzar
                chart_type: 'pie' o 'bar' para el tipo de gráfico
            """
            if not rows:
                logger.warning(f"No hay datos para la sección: {title}")
                return start_row
            
            logger.info(f"Generando sección '{title}' con {len(rows)} elementos (gráfico: {chart_type})")
            
            # Título de la sección
            ws_h.cell(row=start_row, column=1, value=title).font = Font(size=12, bold=True)
            ws_h.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            start_row += 1
            
            # Encabezados
            header_row = start_row
            ws_h.cell(row=header_row, column=1, value="Elemento").font = Font(bold=True, color="FFFFFF")
            ws_h.cell(row=header_row, column=2, value="Neto").font = Font(bold=True, color="FFFFFF")
            ws_h.cell(row=header_row, column=3, value="Porcentaje").font = Font(bold=True, color="FFFFFF")
            for col in (1,2,3):
                ws_h.cell(row=header_row, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws_h.cell(row=header_row, column=col).alignment = Alignment(horizontal="center")
            start_row += 1
            
            # Datos
            total = sum(v for _, v in rows) or 1.0
            first_data_row = start_row
            for name, neto in rows:
                ws_h.cell(row=start_row, column=1, value=clean_for_excel(name))
                ws_h.cell(row=start_row, column=2, value=float(neto))
                # Porcentaje como valor decimal (Excel lo formateará como %)
                pct_cell = ws_h.cell(row=start_row, column=3, value=(float(neto) / total))
                pct_cell.number_format = '0.0%'
                start_row += 1
            last_row = start_row - 1
            
            # Ajuste de anchos
            ws_h.column_dimensions['A'].width = max(ws_h.column_dimensions.get('A', type('obj', (), {'width': 12})).width or 12, 30)
            ws_h.column_dimensions['B'].width = 16
            ws_h.column_dimensions['C'].width = 14
            
            # Generar gráfico si hay suficientes datos
            num_categories = last_row - first_data_row + 1
            logger.info(f"Categorías disponibles para gráfico: {num_categories}")
            
            if chart_available and num_categories >= 1:  # Cambiar de 2 a 1 para permitir un solo elemento
                try:
                    # Decidir tipo de gráfico
                    if chart_type == 'bar' or num_categories > 8:
                        # Usar gráfico de barras si hay muchas categorías
                        chart = BarChart()
                        chart.type = "col"  # Barras verticales
                        chart.title = title
                        chart.y_axis.title = "Ventas Netas"
                        chart.x_axis.title = "Categorías"
                        
                        # Referencias para datos (sin incluir encabezado)
                        labels = Reference(ws_h, min_col=1, min_row=first_data_row, max_row=last_row)
                        data = Reference(ws_h, min_col=2, min_row=header_row, max_row=last_row)  # Incluir encabezado "Neto"
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        chart.height = 10  # Altura en cm
                        chart.width = 18   # Ancho en cm
                        
                    else:
                        # Usar gráfico de pastel para pocas categorías
                        chart = PieChart()
                        chart.title = title
                        
                        # Referencias para datos
                        labels = Reference(ws_h, min_col=1, min_row=first_data_row, max_row=last_row)
                        data = Reference(ws_h, min_col=2, min_row=header_row, max_row=last_row)  # Incluir encabezado "Neto"
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(labels)
                        chart.height = 10
                        chart.width = 15
                    
                    # Colocar el gráfico a la derecha de la tabla (columna E)
                    anchor_col = 5
                    anchor_row = header_row - 1  # Comenzar una fila arriba del encabezado
                    ws_h.add_chart(chart, f"{get_column_letter(anchor_col)}{anchor_row}")
                    logger.info(f"Gráfico '{chart_type}' añadido exitosamente en {get_column_letter(anchor_col)}{anchor_row}")
                    
                except Exception as chart_err:
                    logger.error(f"Error al crear gráfico para '{title}': {chart_err}")
                    import traceback
                    traceback.print_exc()
            else:
                if not chart_available:
                    logger.warning(f"Gráficos no disponibles (openpyxl.chart no importado)")
                else:
                    logger.warning(f"Insuficientes categorías ({num_categories}) para generar gráfico")
            
            # Espacio para próxima sección (más espacio si hay gráfico)
            spacing = 15 if chart_available and num_categories >= 1 else 3
            return last_row + spacing
        
        # Construir jerarquía interactiva drill-down con agrupación de filas
        try:
            # Mapeos para nombres (fallback a código si no hay descripción)
            dept_name = lambda code: (dept_desc_map.get(code) if code in dept_desc_map else str(code)) if code else 'Sin clasificar'
            group_name = lambda code: (group_desc_map.get(code) if code in group_desc_map else str(code)) if code else 'Sin clasificar'
            sub_name = lambda code: (sub_desc_map.get(code) if code in sub_desc_map else str(code)) if code else 'Sin clasificar'
            
            # Totales por depto/grupo/sub
            dept_tot = {}
            group_tot = {}
            sub_tot = {}
            for r in datos_tra:
                if len(r) < 6:
                    continue
                dept = r[2] if len(r) > 2 else None
                group = r[3] if len(r) > 3 else None
                sub = r[4] if len(r) > 4 else None
                try:
                    neto = float(r[5] or 0)
                except Exception:
                    neto = 0.0
                if dept:
                    dept_tot[str(dept)] = dept_tot.get(str(dept), 0.0) + neto
                if dept and group:
                    key_g = (str(dept), str(group))
                    group_tot[key_g] = group_tot.get(key_g, 0.0) + neto
                if dept and group and sub:
                    key_s = (str(dept), str(group), str(sub))
                    sub_tot[key_s] = sub_tot.get(key_s, 0.0) + neto
            
            # Construir estructura jerárquica completa
            total_general = sum(dept_tot.values()) or 1.0
            
            # Título y encabezados
            row_ptr = 3
            ws_h.cell(row=row_ptr, column=1, value="JERARQUÍA INTERACTIVA DE VENTAS").font = Font(size=12, bold=True)
            ws_h.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
            row_ptr += 2
            
            # Instrucciones para el usuario
            ws_h.cell(row=row_ptr, column=1, value="➡️ Haga clic en los botones [+]/[-] a la izquierda para expandir/colapsar departamentos y grupos").font = Font(size=10, italic=True, color="0066CC")
            ws_h.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
            row_ptr += 2
            
            # Encabezados de columnas
            header_row = row_ptr
            headers_cols = ['Categoría', 'Ventas Netas', 'Porcentaje', '% Acumulado']
            for col_idx, header in enumerate(headers_cols, 1):
                cell = ws_h.cell(row=header_row, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row_ptr += 1
            
            # Construir jerarquía drill-down
            dept_rows_sorted = sorted(dept_tot.items(), key=lambda x: x[1], reverse=True)
            acum_pct = 0.0
            
            for dept_code, dept_neto in dept_rows_sorted:
                dept_pct = (dept_neto / total_general) * 100
                acum_pct += dept_pct
                
                # Fila de DEPARTAMENTO (nivel 0 - no indentada)
                dept_row = row_ptr
                ws_h.cell(row=dept_row, column=1, value=f"📦 {dept_name(dept_code)}").font = Font(bold=True, size=11)
                ws_h.cell(row=dept_row, column=2, value=float(dept_neto))
                ws_h.cell(row=dept_row, column=3, value=dept_pct/100).number_format = '0.0%'
                ws_h.cell(row=dept_row, column=4, value=acum_pct/100).number_format = '0.0%'
                
                # Aplicar formato de departamento
                for col in range(1, 5):
                    ws_h.cell(row=dept_row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    ws_h.cell(row=dept_row, column=col).border = Border(
                        top=Side(style='thin'), bottom=Side(style='thin'),
                        left=Side(style='thin'), right=Side(style='thin')
                    )
                row_ptr += 1
                
                # Obtener grupos de este departamento
                dept_groups = sorted(
                    [(g, v) for (d, g), v in group_tot.items() if d == dept_code],
                    key=lambda x: x[1], reverse=True
                )
                
                if dept_groups:
                    group_start_row = row_ptr
                    
                    for group_code, group_neto in dept_groups:
                        group_pct = (group_neto / dept_neto) * 100 if dept_neto > 0 else 0
                        
                        # Fila de GRUPO (nivel 1 - indentada)
                        group_row = row_ptr
                        ws_h.cell(row=group_row, column=1, value=f"  📋 {group_name(group_code)}").font = Font(size=10)
                        ws_h.cell(row=group_row, column=2, value=float(group_neto))
                        ws_h.cell(row=group_row, column=3, value=group_pct/100).number_format = '0.0%'
                        ws_h.cell(row=group_row, column=4, value="-")
                        
                        # Aplicar formato de grupo
                        for col in range(1, 5):
                            ws_h.cell(row=group_row, column=col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        row_ptr += 1
                        
                        # Obtener subgrupos de este grupo
                        group_subs = sorted(
                            [(s, v) for (d, g, s), v in sub_tot.items() if d == dept_code and g == group_code],
                            key=lambda x: x[1], reverse=True
                        )
                        
                        if group_subs:
                            sub_start_row = row_ptr
                            
                            for sub_code, sub_neto in group_subs:
                                sub_pct = (sub_neto / group_neto) * 100 if group_neto > 0 else 0
                                
                                # Fila de SUBGRUPO (nivel 2 - más indentada)
                                sub_row = row_ptr
                                ws_h.cell(row=sub_row, column=1, value=f"    📄 {sub_name(sub_code)}").font = Font(size=9, color="666666")
                                ws_h.cell(row=sub_row, column=2, value=float(sub_neto))
                                ws_h.cell(row=sub_row, column=3, value=sub_pct/100).number_format = '0.0%'
                                ws_h.cell(row=sub_row, column=4, value="-")
                                row_ptr += 1
                            
                            # Agrupar subgrupos bajo el grupo (nivel 2)
                            if row_ptr > sub_start_row:
                                ws_h.row_dimensions.group(sub_start_row, row_ptr - 1, outline_level=2, hidden=True)
                    
                    # Agrupar grupos bajo el departamento (nivel 1)
                    if row_ptr > group_start_row:
                        ws_h.row_dimensions.group(group_start_row, row_ptr - 1, outline_level=1, hidden=True)
            
            # Mensaje de ayuda al final
            row_ptr += 2
            help_row = row_ptr
            ws_h.cell(row=help_row, column=1, value="ℹ️ NOTA: Si no ve los botones [+]/[-], vaya a Vista → Mostrar → Active 'Esquema'").font = Font(size=9, italic=True, color="999999")
            ws_h.merge_cells(start_row=help_row, start_column=1, end_row=help_row, end_column=4)
            
            # Ajustar anchos de columnas
            ws_h.column_dimensions['A'].width = 45
            ws_h.column_dimensions['B'].width = 16
            ws_h.column_dimensions['C'].width = 14
            ws_h.column_dimensions['D'].width = 16
            
            # Agregar gráfico de departamentos a la derecha
            if chart_available and len(dept_rows_sorted) >= 1:
                try:
                    if len(dept_rows_sorted) > 8:
                        chart = BarChart()
                        chart.type = "col"
                        chart.title = "Distribución de Ventas por Departamento"
                        chart.y_axis.title = "Ventas Netas"
                    else:
                        chart = PieChart()
                        chart.title = "Distribución de Ventas por Departamento"
                    
                    # Referenciar solo filas de departamento (sin incluir grupos/subs)
                    # Calcular las filas que corresponden a departamentos
                    dept_row_indices = []
                    current_row = header_row + 1
                    for _ in dept_rows_sorted:
                        dept_row_indices.append(current_row)
                        # Contar cuántas filas tiene este departamento (grupos + subgrupos)
                        dept_code_temp = dept_rows_sorted[len(dept_row_indices)-1][0]
                        num_groups = len([1 for (d, g), v in group_tot.items() if d == dept_code_temp])
                        num_subs = len([1 for (d, g, s), v in sub_tot.items() if d == dept_code_temp])
                        current_row += 1 + num_groups + num_subs
                    
                    # Crear referencias no contiguas para departamentos
                    # Por simplicidad, usamos solo los primeros N departamentos
                    max_chart_items = min(10, len(dept_rows_sorted))
                    chart_start = header_row + 1
                    chart_data_rows = []
                    temp_row = chart_start
                    for idx in range(max_chart_items):
                        chart_data_rows.append(temp_row)
                        dept_code_temp = dept_rows_sorted[idx][0]
                        num_groups = len([1 for (d, g), v in group_tot.items() if d == dept_code_temp])
                        num_subs = len([1 for (d, g, s), v in sub_tot.items() if d == dept_code_temp])
                        temp_row += 1 + num_groups + num_subs
                    
                    # Para el gráfico, simplemente usar un rango contiguo desde header hasta primera agrupación
                    # Excel no maneja bien referencias no contiguas en gráficos, así que mostramos resumen simple
                    labels = Reference(ws_h, min_col=1, min_row=header_row+1, max_row=header_row+len(dept_rows_sorted))
                    data = Reference(ws_h, min_col=2, min_row=header_row, max_row=header_row+len(dept_rows_sorted))
                    
                    # Nota: El gráfico solo mostrará departamentos si no están colapsados
                    # Para simplificar, usamos una tabla resumida en columnas F-H
                    summary_start = 3
                    ws_h.cell(row=summary_start, column=6, value="RESUMEN PARA GRÁFICO").font = Font(bold=True, size=10)
                    ws_h.cell(row=summary_start+1, column=6, value="Departamento").font = Font(bold=True)
                    ws_h.cell(row=summary_start+1, column=7, value="Ventas").font = Font(bold=True)
                    
                    for idx, (dept_code, dept_neto) in enumerate(dept_rows_sorted, summary_start+2):
                        ws_h.cell(row=idx, column=6, value=dept_name(dept_code))
                        ws_h.cell(row=idx, column=7, value=float(dept_neto))
                    
                    # Referencias para gráfico desde tabla resumida
                    labels_summary = Reference(ws_h, min_col=6, min_row=summary_start+2, max_row=summary_start+1+len(dept_rows_sorted))
                    data_summary = Reference(ws_h, min_col=7, min_row=summary_start+1, max_row=summary_start+1+len(dept_rows_sorted))
                    
                    if isinstance(chart, PieChart):
                        chart.add_data(data_summary, titles_from_data=True)
                        chart.set_categories(labels_summary)
                        chart.height = 12
                        chart.width = 18
                    else:
                        chart.add_data(data_summary, titles_from_data=True)
                        chart.set_categories(labels_summary)
                        chart.height = 12
                        chart.width = 18
                    
                    ws_h.add_chart(chart, "I3")
                    logger.info(f"Gráfico de departamentos añadido con éxito")
                    
                except Exception as chart_err:
                    logger.error(f"Error al crear gráfico de departamentos: {chart_err}")
                    import traceback
                    traceback.print_exc()
                    
        except Exception as e:
            logger.error(f"Error construyendo jerarquía interactiva: {e}")
            import traceback
            traceback.print_exc()
        
        # Guardar archivo
        wb.save(filename)
        logger.info(f"Exportación RI Excel completada: {total_registros} registros en {filename}")
        return total_registros
        
    except Exception as e:
        logger.error(f"Error en exportación TRA Excel: {e}")
        raise


def export_mbrp_excel(filename: str, datos_mbrp: List, db_manager=None, progress_cb: Optional[Callable[[int, int], None]] = None,
                      permissions_manager=None, current_user_id: int = None,
                      provider_label: Optional[str] = None,
                      sede_codigo: Optional[str] = None,
                      fecha_inicio=None,
                      fecha_fin=None) -> int:
    """
    Exporta datos MBRP a un archivo Excel con formato profesional y análisis de rentabilidad.
    
    Args:
        filename: Nombre del archivo Excel a crear
        datos_mbrp: Lista de datos MBRP a exportar
        db_manager: Gestor de base de datos
        progress_cb: Callback opcional para reportar progreso
        permissions_manager: Gestor de permisos (opcional, para verificar ver_costo_utilidad)
        current_user_id: ID del usuario actual (opcional, para verificar permisos)
        
    Returns:
        int: Número total de registros exportados
    """
    if not EXCEL_AVAILABLE:
        logger.error("openpyxl no está disponible. Use export_mbrp_csv en su lugar.")
        raise ImportError("openpyxl es requerido para exportación Excel")
        
    from datetime import datetime
    try:
        total_registros_entrada = len(datos_mbrp)
        export_start_time = datetime.now()
        logger.info(f"[EXPORT MBRP] INICIANDO EXPORTACIÓN - {total_registros_entrada} registros a {filename}")
        logger.info(f"[EXPORT MBRP] Inicio: {export_start_time.strftime('%H:%M:%S.%f')[:-3]}")
        
        # Determinar si el filtro de sede es global (modo ICH)
        ich_mode = sede_codigo in (None, '%', '00', 'ICH', 'ALL')

        # Crear workbook
        wb = Workbook()
        wb.remove(wactive := wb.active)  # Remover hoja por defecto
        
        # === HOUGHT 1: DATOS PRINCIPALES MBRP ===
        ws_main = wb.create_sheet("Datos MBRP")
        
        # Encabezado del reporte
        ws_main['A1'] = f'REPORTE MERCANCÍA DE BAJA ROTACIÓN (MBRP)'
        ws_main['A2'] = f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M:%S")}'
        if provider_label:
            ws_main['A3'] = f'Proveedor: {provider_label}'
        ws_main['A4'] = f'Total de productos analizados: {total_registros_entrada}'
        
        mostrar_costo_utilidad = False
        mostrar_proveedores = False
        if permissions_manager and current_user_id:
            try:
                mostrar_costo_utilidad = permissions_manager.tiene_permiso(current_user_id, 'MBRP', 'ver_costo_utilidad')
                logger.info(f"Permiso ver_costo_utilidad para MBRP: {mostrar_costo_utilidad}")
                mostrar_proveedores = permissions_manager.tiene_permiso(current_user_id, 'MBRP', 'ver_proveedores')
                logger.info(f"Permiso ver_proveedores para MBRP: {mostrar_proveedores}")
            except Exception as e:
                logger.warning(f"Error verificando permisos MBRP: {e}")
        
        # Formato del encabezado
        header_font = Font(size=14, bold=True, color="FFFFFF")
        header_fill = new_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        ws_main['A1'].font = header_font
        ws_main['A1'].fill = header_fill

        # Headers de la tabla (fila 7) alineados con el módulo MBRP
        # Incluimos jerarquía para permitir filtros por Departamento/Grupo/Subgrupo
        headers = [
            'Código', 'Descripción', 'Marca', 'Departamento', 'Grupo', 'Subgrupo',
            'Rotación', 'Ventas', 'Stock Actual'
        ]
        
        # Columnas de stock por sede (solo modo ICH)
        if ich_mode:
            headers.extend(['Stock Cabudare', 'Stock Barinas', 'Stock Guanare'])
        
        # Columnas adicionales de análisis
        headers.extend(['Días de Stock', 'IM %', 'Última Venta (DIAS)'])
        
        # Columnas adicionales de análisis económico
        if mostrar_costo_utilidad:
            headers.extend(['Precio + IVA', 'Costo', 'Utilidad %'])
        
        if mostrar_proveedores:
            headers.append('Último Proveedor')
        
        # Columna final de detalle ICH (solo última venta)
        if ich_mode:
            headers.append('Detalle ICH (última venta)')

        # Ajustar merge del título según cantidad total de columnas
        from openpyxl.utils import get_column_letter as _col_letter
        last_col_letter = _col_letter(len(headers))
        ws_main.merge_cells(f"A1:{last_col_letter}1")

        start_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

# Índice de la columna de detalle ICH (si aplica)
        # Con ICH y costo/utilidad: columna R (18)
        # Con ICH sin costo/utilidad: columna O (15) 
        # Sin ICH: no hay detalle ICH
        if ich_mode:
            detalle_ich_col = 19 if mostrar_costo_utilidad else 16
        else:
            detalle_ich_col = None

        # Datos de productos
        data_start_row = start_row + 1

        # Helper seguro para convertir a float sin lanzar excepciones por valores no numéricos
        def _safe_float(value, default: float = 0.0) -> float:
            try:
                return float(value or 0)
            except (ValueError, TypeError):
                return default

        # Helper para calcular días de stock (misma lógica que en la app)
        def _calcular_dias_restantes(stock_actual: int, neto_ventas: float) -> int:
            """Calcula días de stock restantes usando el promedio diario de ventas
            basado en fecha_inicio y fecha_fin del período MBRP."""
            try:
                if not fecha_inicio or not fecha_fin:
                    return 0
                dias_periodo = (fecha_fin - fecha_inicio).days or 1
                promedio_diario = neto_ventas / dias_periodo if dias_periodo else 0
                if promedio_diario <= 0:
                    return 0
                return max(0, int(stock_actual // promedio_diario))
            except Exception:
                return 0

        # Helper para stock actual por producto (similar a obtener_stock_actual_bulk del app)
        def _get_stock_actual_bulk_mbrp(codigos: List[str], deposito: Optional[str]) -> Dict[str, int]:
            """Obtiene el stock actual por código para MBRP (optimizado).

            Si deposito es global ('%', '00', 'ICH', 'ALL' o None), suma en todas las sedes;
            de lo contrario filtra por c_coddeposito.
            """
            resultados: Dict[str, int] = {}
            if not db_manager or not db_manager.ensure_connection() or not codigos:
                return resultados
            try:
                MAX_IN = 2000  # Optimizado: aumentado de 900 a 2000
                global_query = deposito in (None, '%', '00', 'ICH', 'ALL')
                
                # Pre-inicializar resultados para evitar setdefault en el loop
                for cod in codigos:
                    resultados[str(cod)] = 0
                
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    if global_query:
                        sql = (
                            f"SELECT c_codarticulo, SUM(n_cantidad) "
                            f"FROM MA_DEPOPROD WITH (NOLOCK) "
                            f"WHERE c_codarticulo IN ({placeholders}) "
                            f"GROUP BY c_codarticulo"
                        )
                        params = chunk
                    else:
                        sql = (
                            f"SELECT c_codarticulo, SUM(n_cantidad) "
                            f"FROM MA_DEPOPROD WITH (NOLOCK) "
                            f"WHERE c_coddeposito = ? AND c_codarticulo IN ({placeholders}) "
                            f"GROUP BY c_codarticulo"
                        )
                        params = [deposito] + chunk
                    rows = db_manager.fetch_data(sql, params)
                    for cod, sum_qty in (rows or []):
                        try:
                            resultados[str(cod)] = int(sum_qty or 0)
                        except Exception:
                            pass  # Ya está inicializado en 0
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT MBRP] Error obteniendo stock actual: {e}")
                return resultados

        def _map_deposito_to_sede(dep: str) -> str:
            """Mapea un código de depósito a una sede legible."""
            try:
                c = (dep or "").strip()
                if c.startswith('03'):
                    return 'Cabudare'
                if c.startswith('01'):
                    return 'Barinas'
                if c.startswith('04'):
                    return 'Guanare'
                return 'Otra'
            except Exception:
                return 'Otra'

        def _get_stock_por_sede_mbrp(codigos: List[str]) -> Dict[str, Dict[str, int]]:
            """Obtiene distribución de stock por sede para cada código (solo modo ICH, optimizado)."""
            resultados: Dict[str, Dict[str, int]] = {}
            if not db_manager or not db_manager.ensure_connection() or not codigos:
                return resultados
            try:
                MAX_IN = 2000  # Optimizado: aumentado de 900 a 2000
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    sql = (
                        f"SELECT c_codarticulo, c_coddeposito, SUM(n_cantidad) "
                        f"FROM MA_DEPOPROD WITH (NOLOCK) "
                        f"WHERE c_codarticulo IN ({placeholders}) "
                        f"GROUP BY c_codarticulo, c_coddeposito"
                    )
                    rows = db_manager.fetch_data(sql, chunk) or []
                    for cod, dep, qty in rows:
                        try:
                            cod_str = str(cod).strip()
                            dep_str = str(dep).strip()
                            sede_nombre = _map_deposito_to_sede(dep_str)
                            q = int(qty or 0)
                        except Exception:
                            continue
                        if cod_str not in resultados:
                            resultados[cod_str] = {}
                        resultados[cod_str][sede_nombre] = resultados[cod_str].get(sede_nombre, 0) + q
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT MBRP] Error obteniendo stock por sede: {e}")
                return resultados

        def _get_last_sale_deposito_bulk(codigos: List[str]) -> Dict[str, str]:
            """Obtiene el depósito de la última venta global (todas las sedes) para cada código (optimizado)."""
            resultados: Dict[str, str] = {}
            if not db_manager or not db_manager.ensure_connection() or not codigos:
                return resultados
            try:
                MAX_IN = 2000  # Optimizado: aumentado de 900 a 2000
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    query = f"""
                        WITH ult AS (
                            SELECT c_Codarticulo, c_Deposito, f_fecha,
                                   ROW_NUMBER() OVER (PARTITION BY c_Codarticulo ORDER BY f_fecha DESC) AS rn
                            FROM TR_INVENTARIO WITH (NOLOCK)
                            WHERE c_Codarticulo IN ({placeholders})
                              AND c_Concepto = 'VEN'
                              AND n_Cantidad > 0
                        )
                        SELECT c_Codarticulo, c_Deposito
                        FROM ult
                        WHERE rn = 1
                    """
                    rows = db_manager.fetch_data(query, chunk) or []
                    for cod, dep in rows:
                        try:
                            cod_str = str(cod).strip()
                            dep_str = str(dep).strip() if dep is not None else ''
                            resultados[cod_str] = dep_str
                        except Exception:
                            continue
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT MBRP] Error obteniendo último depósito de venta: {e}")
                return resultados
            try:
                MAX_IN = 900
                for i in range(0, len(codigos), MAX_IN):
                    chunk = codigos[i:i + MAX_IN]
                    placeholders = ','.join(['?'] * len(chunk))
                    query = f"""
                        WITH ult AS (
                            SELECT c_Codarticulo, c_Deposito, f_fecha,
                                   ROW_NUMBER() OVER (PARTITION BY c_Codarticulo ORDER BY f_fecha DESC) AS rn
                            FROM TR_INVENTARIO WITH (NOLOCK)
                            WHERE c_Codarticulo IN ({placeholders})
                              AND c_Concepto = 'VEN'
                              AND n_Cantidad > 0
                        )
                        SELECT c_Codarticulo, c_Deposito
                        FROM ult
                        WHERE rn = 1
                    """
                    rows = db_manager.fetch_data(query, chunk) or []
                    for cod, dep in rows:
                        try:
                            cod_str = str(cod).strip()
                            dep_str = str(dep).strip() if dep is not None else ''
                        except Exception:
                            continue
                        if cod_str and dep_str and cod_str not in resultados:
                            resultados[cod_str] = dep_str
                return resultados
            except Exception as e:
                logger.error(f"[EXPORT MBRP] Error obteniendo última venta por depósito: {e}")
                return resultados

        # Mapas de descripciones jerárquicas (Depto/Grupo/Subgrupo) para MBRP
        dept_desc_map: Dict[str, str] = {}
        group_desc_map: Dict[str, str] = {}
        sub_desc_map: Dict[str, str] = {}


        costo_map_mbrp: Dict[str, float] = {}

        # Cargar costos por producto
        try:
            codigos_mbrp = sorted({
                str(f[0]).strip() for f in datos_mbrp
                if f and len(f) > 0 and f[0] is not None
            })
            if codigos_mbrp:
                BATCH_SIZE = 1800
                for i_b in range(0, len(codigos_mbrp), BATCH_SIZE):
                    batch = codigos_mbrp[i_b:i_b + BATCH_SIZE]
                    placeholders = ','.join('?' * len(batch))
                    query_costo = f"""
                    SELECT C_CODIGO, COALESCE(n_costoact, 0) AS costo_actual
                    FROM MA_PRODUCTOS WITH (NOLOCK)
                    WHERE C_CODIGO IN ({placeholders})
                """
                rows_costo = db_manager.fetch_data(query_costo, batch) or []
                for cod, costo in rows_costo:
                    try:
                        costo_map_mbrp[str(cod).strip()] = float(costo or 0)
                    except Exception:
                        continue
            logger.info(f"[EXPORT MBRP] Costos cargados para {len(costo_map_mbrp)} productos")
        except Exception as e:
            logger.warning(f"No se pudieron cargar los costos para MBRP: {e}")
            costo_map_mbrp = {}

        # Mapa de impuestos (IVA) por producto para MBRP (se usa solo para calcular Precio + IVA)
        impuestos_map_mbrp: Dict[str, float] = {}
        if db_manager and db_manager.ensure_connection():
            # Cargar descripciones de jerarquía
            try:
                depts = db_manager.fetch_data(
                    "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_DEPARTAMENTOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                dept_desc_map = {str(cod).strip(): desc for cod, desc in (depts or []) if cod and desc}

                groups = db_manager.fetch_data(
                    "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_GRUPOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                group_desc_map = {str(cod).strip(): desc for cod, desc in (groups or []) if cod and desc}

                subs = db_manager.fetch_data(
                    "SELECT C_CODIGO, C_DESCRIPCIO FROM MA_SUBGRUPOS WHERE C_CODIGO IS NOT NULL AND C_DESCRIPCIO IS NOT NULL"
                )
                sub_desc_map = {str(cod).strip(): desc for cod, desc in (subs or []) if cod and desc}

                logger.info(
                    f"[EXPORT MBRP] Mapeos cargados - Departamentos: {len(dept_desc_map)}, "
                    f"Grupos: {len(group_desc_map)}, Subgrupos: {len(sub_desc_map)}"
                )
                
                # Obtener marcas
                marcas = db_manager.fetch_data("SELECT RTRIM(LTRIM(C_CODIGO)), COALESCE(c_marca, '') FROM MA_PRODUCTOS WITH (NOLOCK)")
                marca_map_mbrp = {str(codigo).strip(): marca for codigo, marca in marcas if codigo}
                logger.info(f"[EXPORT MBRP] Marcas cargadas: {len(marca_map_mbrp)}")
            except Exception as e:
                logger.warning(f"[EXPORT MBRP] No se pudieron cargar descripciones de jerarquía/marcas: {e}")

            # Cargar impuestos (IVA) por producto
            try:
                codigos_mbrp = sorted({
                    str(f[0]).strip() for f in datos_mbrp
                    if f and len(f) > 0 and f[0] is not None
                })
                if codigos_mbrp:
                    BATCH_SIZE = 1800
                    for i_b in range(0, len(codigos_mbrp), BATCH_SIZE):
                        batch = codigos_mbrp[i_b:i_b + BATCH_SIZE]
                        placeholders = ','.join('?' * len(batch))
                        query_imp = f"""
                            SELECT C_CODIGO, COALESCE(n_impuesto1, 0) AS impuesto1
                            FROM MA_PRODUCTOS WITH (NOLOCK)
                            WHERE C_CODIGO IN ({placeholders})
                        """
                        rows_imp = db_manager.fetch_data(query_imp, batch) or []
                        for cod, imp in rows_imp:
                            try:
                                impuestos_map_mbrp[str(cod).strip()] = float(imp or 0)
                            except Exception:
                                continue
                    logger.info(f"[EXPORT MBRP] Impuestos cargados para {len(impuestos_map_mbrp)} productos")
            except Exception as e:
                logger.warning(f"No se pudieron cargar los impuestos (IVA) para MBRP: {e}")

        # Precalcular métricas de movilidad, stock y últimas ventas para replicar el Treeview MBRP
        from pal.services.mbrp import calcular_indice_movilidad, obtener_ultimas_ventas_bulk, calcular_dias_sin_venta

        # Extraer códigos únicos una sola vez
        codigos_unicos = sorted({
            str(f[0]).strip() for f in datos_mbrp
            if f and len(f) > 0 and f[0] is not None
        })
        
        if not codigos_unicos:
            logger.warning("[EXPORT MBRP] No hay códigos válidos para procesar")
            return 0
            
        sede = sede_codigo or '0301'
        
        # Optimización: Consultas combinadas y batch sizes optimizados
        start_time = datetime.now()
        
        # 1) Calcular índices de movilidad (proceso en memoria, rápido)
        indices_movilidad = calcular_indice_movilidad(datos_mbrp) if datos_mbrp else {}
        
        # 2) Consultas paralelas de datos maestros (costo + precio en una sola consulta)
        costo_map_mbrp = {}
        precio_map_mbrp = {}
        if mostrar_costo_utilidad or True:  # Siempre cargamos precios para el reporte
            try:
                BATCH_SIZE = 2000  # Optimizado para SQL Server
                for i_b in range(0, len(codigos_unicos), BATCH_SIZE):
                    batch = codigos_unicos[i_b:i_b + BATCH_SIZE]
                    placeholders = ','.join('?' * len(batch))
                    
                    # Consulta combinada para costo y precio
                    query_combined = f"""
                        SELECT C_CODIGO, 
                               COALESCE(n_costoact, 0) AS costo_actual,
                               COALESCE(n_precio1, 0) AS precio_base
                        FROM MA_PRODUCTOS WITH (NOLOCK)
                        WHERE C_CODIGO IN ({placeholders})
                    """
                    rows_combined = db_manager.fetch_data(query_combined, batch) or []
                    for cod, costo, precio in rows_combined:
                        try:
                            cod_str = str(cod).strip()
                            if mostrar_costo_utilidad:
                                costo_map_mbrp[cod_str] = float(costo or 0)
                            precio_map_mbrp[cod_str] = float(precio or 0)
                        except Exception:
                            continue
                            
                logger.info(f"[EXPORT MBRP] Datos maestros cargados para {len(costo_map_mbrp if mostrar_costo_utilidad else precio_map_mbrp)} productos")
            except Exception as e:
                logger.warning(f"No se pudieron cargar los datos maestros para MBRP: {e}")
                costo_map_mbrp = {}
                precio_map_mbrp = {}

        # 3) Consultas de inventario y ventas (optimizadas)
        stock_map: Dict[str, int] = {}
        ultimas_ventas: Dict[str, Any] = {}
        stock_por_sede_map: Dict[str, Dict[str, int]] = {}
        last_dep_map: Dict[str, str] = {}
        
        if db_manager and db_manager.ensure_connection():
            # Stock actual (batch size optimizado)
            stock_map = _get_stock_actual_bulk_mbrp(codigos_unicos, sede)
            
            # Últimas ventas
            ultimas_ventas = obtener_ultimas_ventas_bulk(db_manager, codigos_unicos, sede)
            
            # Datos adicionales solo para modo ICH
            if ich_mode:
                stock_por_sede_map = _get_stock_por_sede_mbrp(codigos_unicos)
                last_dep_map = _get_last_sale_deposito_bulk(codigos_unicos)
            
            # Cargar últimos proveedores si se tiene permiso
            proveedores_map_mbrp = {}
            if mostrar_proveedores:
                try:
                    proveedores_map_mbrp = db_manager.obtener_ultimas_compras_bulk(codigos_unicos)
                    logger.info(f"[EXPORT MBRP] Cargados {len(proveedores_map_mbrp)} proveedores")
                except Exception as e:
                    logger.warning(f"Error cargando proveedores para export MBRP: {e}")
        
        load_time = (datetime.now() - start_time).total_seconds()
        logger.info(f"[EXPORT MBRP] Datos cargados en {load_time:.2f}s - Stock: {len(stock_map)}, Ventas: {len(ultimas_ventas)}")

        # Optimización: Preparar datos en lote para reducir overhead de Excel
        start_export_time = datetime.now()
        filas_exportadas = 0
        
        # Pre-calcular todas las filas de datos en memoria primero
        filas_datos = []
        for i, fila in enumerate(datos_mbrp):
            try:
                # Normalizar a lista para poder inspeccionar la estructura
                fila_list = list(fila) if not isinstance(fila, list) else fila

                if not fila_list or len(fila_list) < 6:
                    continue

                codigo = fila_list[0]
                codigo_str = str(codigo).strip()
                desc = fila_list[1]

                # Jerarquía (códigos) para filtros: depto, grupo, subgrupo
                dept = fila_list[2] if len(fila_list) > 2 else None
                grupo = fila_list[3] if len(fila_list) > 3 else None
                sub = fila_list[4] if len(fila_list) > 4 else None

                # Convertir a descripciones legibles si están disponibles
                dept_desc = dept_desc_map.get(str(dept).strip(), dept) if dept else 'Sin clasificar'
                grupo_desc = group_desc_map.get(str(grupo).strip(), grupo) if grupo else 'Sin clasificar'
                sub_desc = sub_desc_map.get(str(sub).strip(), sub) if sub else 'Sin clasificar'
                marca_val = marca_map_mbrp.get(codigo_str, '')

                # Neto de ventas siempre en posición 5 en la estructura TRA/MBRP
                neto_valor = _safe_float(fila_list[5])

                # Rotación clasificada añadida por clasificar_rotacion_mbrp en posición 6
                rotacion = fila_list[6] if len(fila_list) > 6 else 'SIN_CLASIFICAR'

                try:
                    codigo_str = str(codigo).strip() if codigo is not None else ''
                except Exception:
                    codigo_str = ''

                stock_actual = int(stock_map.get(codigo_str, 0) or 0)

                # Regla solicitada: excluir de la exportación productos con stock actual = 0
                if stock_actual <= 0:
                    if progress_cb:
                        # Seguimos reportando progreso sobre el total de entradas,
                        # aunque este registro no se exporte.
                        progress_cb(i + 1, total_registros_entrada)
                    continue

                im_valor = float(indices_movilidad.get(codigo_str, 0.0))

                # Días desde última venta (valor numérico para filtrar en Excel)
                fecha_ultima = ultimas_ventas.get(codigo_str)
                dias_sin_venta = calcular_dias_sin_venta(fecha_ultima) if fecha_ultima is not None else -1
                # Para productos sin ventas, usar un valor grande (9999) para que aparezcan al final al ordenar ascendente
                dias_ultima_venta_excel = dias_sin_venta if dias_sin_venta is not None and dias_sin_venta >= 0 else 9999

                # Días de Stock basados en promedio diario del período
                dias_stock = _calcular_dias_restantes(stock_actual, neto_valor)

                # Almacenar fila de datos para escritura en lote
                fila_datos = {
                    'codigo': codigo,
                    'desc': desc,
                    'marca': marca_val, # Añadido
                    'dept_desc': dept_desc,
                    'grupo_desc': grupo_desc,
                    'sub_desc': sub_desc,
                    'rotacion': rotacion,
                    'neto_valor': neto_valor,
                    'stock_actual': stock_actual,
                    'dias_stock': dias_stock,
                    'im_valor': im_valor,
                    'dias_ultima_venta_excel': dias_ultima_venta_excel,
                    'codigo_str': codigo_str,
                    'fecha_ultima': fecha_ultima
                }
                
                # Agregar datos de costo/precio si aplica
                if mostrar_costo_utilidad:
                    fila_datos['precio_base'] = _safe_float(precio_map_mbrp.get(codigo_str, 0.0))
                    fila_datos['costo_val'] = _safe_float(costo_map_mbrp.get(codigo_str, 0.0))
                    fila_datos['iva_pct'] = float(impuestos_map_mbrp.get(codigo_str, 0.0)) if impuestos_map_mbrp else 0.0
                
                # Agregar datos ICH si aplica
                if ich_mode:
                    fila_datos['dep_ult'] = last_dep_map.get(codigo_str)
                    fila_datos['dist'] = stock_por_sede_map.get(codigo_str, {})
                
                filas_datos.append(fila_datos)
                filas_exportadas += 1
                
                # Reportar progreso
                if progress_cb and i % 100 == 0:
                    progress_cb(i + 1, total_registros_entrada)
                    
            except Exception as e:
                logger.warning(f"[EXPORT MBRP] Error procesando fila: {e}")
                continue
        
        # Reportar progreso final
        if progress_cb:
            progress_cb(total_registros_entrada, total_registros_entrada)
            
        export_time = (datetime.now() - start_export_time).total_seconds()
        logger.info(f"[EXPORT MBRP] Exportación completada en {export_time:.2f}s - {filas_exportadas} filas")
        
        # Escritura optimizada en lote al Excel
        logger.info(f"[EXPORT MBRP] Escribiendo {len(filas_datos)} filas en Excel...")
        for i, fila_datos in enumerate(filas_datos):
            try:
                row = data_start_row + i
                
                # Escribir columnas principales
                ws_main.cell(row=row, column=1, value=clean_for_excel(fila_datos['codigo']))
                ws_main.cell(row=row, column=2, value=clean_for_excel(fila_datos['desc']))
                ws_main.cell(row=row, column=3, value=clean_for_excel(fila_datos.get('marca', ''))) # Marca añadida
                ws_main.cell(row=row, column=4, value=clean_for_excel(fila_datos['dept_desc']))
                ws_main.cell(row=row, column=5, value=clean_for_excel(fila_datos['grupo_desc']))
                ws_main.cell(row=row, column=6, value=clean_for_excel(fila_datos['sub_desc']))
                ws_main.cell(row=row, column=7, value=clean_for_excel(str(fila_datos['rotacion'])))
                ws_main.cell(row=row, column=8, value=fila_datos['neto_valor'])
                ws_main.cell(row=row, column=9, value=fila_datos['stock_actual'])
                
                current_col = 9
                
                # Columnas de stock por sede (solo modo ICH)
                if ich_mode:
                    dist = fila_datos.get('dist', {})
                    stock_cabudare = dist.get('Cabudare', 0)
                    stock_barinas = dist.get('Barinas', 0)
                    stock_guanare = dist.get('Guanare', 0)
                    
                    ws_main.cell(row=row, column=10, value=stock_cabudare)
                    ws_main.cell(row=row, column=11, value=stock_barinas)
                    ws_main.cell(row=row, column=12, value=stock_guanare)
                    current_col = 12
                else:
                    # Si no es modo ICH, dejamos espacios para las columnas de sede
                    current_col = 9
                
                # Columnas adicionales de análisis
                current_col += 1
                ws_main.cell(row=row, column=current_col, value=fila_datos['dias_stock'])
                current_col += 1
                ws_main.cell(row=row, column=current_col, value=round(fila_datos['im_valor'], 2))
                current_col += 1
                ws_main.cell(row=row, column=current_col, value=fila_datos['dias_ultima_venta_excel'])
                current_col += 1

                # Columnas de análisis económico
                if mostrar_costo_utilidad:
                    precio_base = fila_datos.get('precio_base', 0)
                    costo_val = fila_datos.get('costo_val', 0)
                    iva_pct = fila_datos.get('iva_pct', 0)
                    precio_con_iva = precio_base * (1.0 + (iva_pct / 100.0)) if precio_base > 0 else 0.0

                    ws_main.cell(row=row, column=current_col, value=round(precio_con_iva, 2))
                    current_col += 1
                    ws_main.cell(row=row, column=current_col, value=round(costo_val, 2))
                    current_col += 1

                    if precio_base > 0:
                        utilidad_raw = (costo_val / precio_base) * 100 - 100
                        utilidad_pct = abs(utilidad_raw)
                    else:
                        utilidad_pct = 0.0
                    ws_main.cell(row=row, column=current_col, value=round(utilidad_pct, 2))
                    current_col += 1
                
                # Columna Último Proveedor
                if mostrar_proveedores:
                    prov_nombre = proveedores_map_mbrp.get(fila_datos['codigo_str'], 'Sin proveedor')
                    ws_main.cell(row=row, column=current_col, value=clean_for_excel(prov_nombre))
                    current_col += 1

                # Columna final de detalle ICH (solo modo ICH) - Solo última venta
                if ich_mode:
                    dep_ult = fila_datos.get('dep_ult')
                    if dep_ult:
                        sede_ult = _map_deposito_to_sede(dep_ult)
                        # Obtener fecha de la última venta para más contexto
                        fecha_ultima_str = ""
                        try:
                            if fila_datos.get('fecha_ultima'):
                                fecha_ultima_str = f" el {fila_datos['fecha_ultima'].strftime('%d/%m/%Y')}"
                        except Exception:
                            pass
                        detalle_ich = f"Última venta: {sede_ult}{fecha_ultima_str} (dep: {dep_ult})"
                    else:
                        detalle_ich = "Sin ventas registradas"
                    
                    ws_main.cell(row=row, column=current_col, value=clean_for_excel(detalle_ich))



            except Exception as e:
                logger.error(f"Error procesando fila MBRP: {fila_datos} - {e}")
                continue

        # Crear tabla con filtros sólo si hubo filas exportadas
        end_col_letter = get_column_letter(len(headers))
        if filas_exportadas > 0:
            last_data_row = data_start_row + filas_exportadas - 1
            table_range = f"A{start_row}:{end_col_letter}{last_data_row}"
            table = Table(displayName="TablaMBRP", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws_main.add_table(table)
        # Formato condicional para Índice de Movilidad (IM %)
        if filas_exportadas > 0:
            # Determinar columna de IM % dinámicamente
            try:
                im_col_idx = headers.index('IM %') + 1
                im_col_letter = get_column_letter(im_col_idx)
            except Exception:
                im_col_letter = 'N' if ich_mode else 'K' # Fallback
                
            im_range = f"{im_col_letter}{data_start_row}:{im_col_letter}{data_start_row + filas_exportadas - 1}"
            
            # IM < 5% = Rojo (crítico)
            ws_main.conditional_formatting.add(im_range,
                CellIsRule(operator='lessThan', formula=[5],
                          fill=PatternFill(start_color="F44336", end_color="F44336")))
            
            # 5% <= IM <= 10% = Naranja (muy bajo)
            ws_main.conditional_formatting.add(im_range,
                CellIsRule(operator='between', formula=[5, 10],
                          fill=PatternFill(start_color="FF9800", end_color="FF9800")))
            
            # 10% < IM <= 20% = Amarillo (bajo)
            ws_main.conditional_formatting.add(im_range,
                CellIsRule(operator='between', formula=[10, 20],
                          fill=PatternFill(start_color="FFEB3B", end_color="FFEB3B")))
            
            # Formato condicional especial para modo ICH
            if ich_mode and filas_exportadas > 0:
                # Columna de stock actual (siempre columna I ahora que Marca está en C)
                stock_range = f"I{data_start_row}:I{data_start_row + filas_exportadas - 1}"
                ws_main.conditional_formatting.add(stock_range,
                    CellIsRule(operator='equal', formula=[0],
                              fill=PatternFill(start_color="FFE5E5", end_color="FFE5E5"),
                              font=Font(color="CC0000")))
                
                # Columna de días sin venta (siempre columna N)
                dias_col_letter = 'N'  # Última Venta (DIAS)
                dias_range = f"{dias_col_letter}{data_start_row}:{dias_col_letter}{data_start_row + filas_exportadas - 1}"
                ws_main.conditional_formatting.add(dias_range,
                    CellIsRule(operator='greaterThan', formula=[180],
                              fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2"),
                              font=Font(color="B71C1C")))
                
                # Resaltar stock bajo en Cabudare (columna I)
                cabudare_range = f"I{data_start_row}:I{data_start_row + filas_exportadas - 1}"
                ws_main.conditional_formatting.add(cabudare_range,
                    CellIsRule(operator='equal', formula=[0],
                              fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2"),
                              font=Font(color="B71C1C")))
        
# Ajustar anchos de columna - Actualizado para mapeo correcto
        ws_main.column_dimensions['A'].width = 12  # Código
        ws_main.column_dimensions['B'].width = 40  # Descripción
        ws_main.column_dimensions['C'].width = 14  # Departamento
        ws_main.column_dimensions['D'].width = 14  # Grupo
        ws_main.column_dimensions['E'].width = 14  # Subgrupo
        ws_main.column_dimensions['F'].width = 12  # Rotación
        ws_main.column_dimensions['G'].width = 15  # Ventas
        ws_main.column_dimensions['H'].width = 15  # Stock Actual
        
        if ich_mode:
            # Columnas de stock por sede
            ws_main.column_dimensions['I'].width = 15  # Stock Cabudare
            ws_main.column_dimensions['J'].width = 15  # Stock Barinas
            ws_main.column_dimensions['K'].width = 15  # Stock Guanare
            # Columnas de análisis (después de stock por sede)
            ws_main.column_dimensions['L'].width = 15  # Días de Stock
            ws_main.column_dimensions['M'].width = 12  # IM %
            ws_main.column_dimensions['N'].width = 18  # Última Venta (DIAS)
        else:
            # Sin ICH: columnas de análisis después de Stock Actual
            ws_main.column_dimensions['I'].width = 15  # Días de Stock
            ws_main.column_dimensions['J'].width = 12  # IM %
            ws_main.column_dimensions['K'].width = 18  # Última Venta (DIAS)
        
        # Columnas de costo/utilidad (siempre al final)
        if mostrar_costo_utilidad:
            if ich_mode:
                # Con ICH: O, P, Q
                ws_main.column_dimensions['O'].width = 15  # Precio + IVA
                ws_main.column_dimensions['P'].width = 15  # Costo
                ws_main.column_dimensions['Q'].width = 15  # Utilidad %
                # Detalle ICH es la última columna R
                ws_main.column_dimensions['R'].width = 60  # Detalle ICH (más ancho para multilinea)
            else:
                # Sin ICH: L, M, N
                ws_main.column_dimensions['L'].width = 15  # Precio + IVA
                ws_main.column_dimensions['M'].width = 15  # Costo
                ws_main.column_dimensions['N'].width = 15  # Utilidad %
        
        # === HOJA 2: RESUMEN POR MOVILIDAD ===
        ws_summary = wb.create_sheet("Resumen por Movilidad")
        
        # Clasificar productos por rangos de Índice de Movilidad
        im_values = list(indices_movilidad.values()) if indices_movilidad else []
        critico = sum(1 for v in im_values if v < 5.0)
        muy_bajo = sum(1 for v in im_values if 5.0 <= v <= 10.0)
        bajo = sum(1 for v in im_values if 10.0 < v <= 20.0)
        medio_alto = sum(1 for v in im_values if 20.0 < v <= 50.0)
        alto = sum(1 for v in im_values if v > 50.0)
        
        # Crear tabla resumen
        ws_summary['A1'] = 'RESUMEN POR ÍNDICE DE MOVILIDAD (IM%) MBRP'
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary.merge_cells('A1:C1')
        
        ws_summary['A3'] = 'Categoría'
        ws_summary['B3'] = 'Cantidad'
        ws_summary['C3'] = 'Porcentaje'
        
        for i, cell in enumerate(['A3', 'B3', 'C3']):
            ws_summary[cell].font = Font(bold=True, color="FFFFFF")
            ws_summary[cell].fill = PatternFill(start_color="4472C4", end_color="4472C4")
        
        categorias = [
            ('IM < 5% (CRÍTICO)', critico),
            ('5% ≤ IM ≤ 10% (MUY BAJO)', muy_bajo),
            ('10% < IM ≤ 20% (BAJO)', bajo),
            ('20% < IM ≤ 50% (MEDIO)', medio_alto),
            ('IM > 50% (ACEPTABLE)', alto),
        ]
        
        for i, (categoria, count) in enumerate(categorias, 4):
            porcentaje = (count / total_registros_entrada * 100) if total_registros_entrada > 0 else 0
            ws_summary.cell(row=i, column=1, value=categoria)
            ws_summary.cell(row=i, column=2, value=count)
            ws_summary.cell(row=i, column=3, value=f"{porcentaje:.1f}%")
        
        # === HOJA 3: PRODUCTOS CRÍTICOS ===
        ws_critical = wb.create_sheet("Productos Críticos")
        
        # Productos críticos: IM muy bajo y muchos días sin venta
        productos_criticos = []
        for fila in datos_mbrp:
            fila_list = list(fila) if not isinstance(fila, list) else fila
            if not fila_list or len(fila_list) < 2:
                continue
            codigo = str(fila_list[0])
            desc = fila_list[1]
            im_val = float(indices_movilidad.get(codigo, 0.0))
            fecha_ult = ultimas_ventas.get(codigo)
            dias_sin_venta = calcular_dias_sin_venta(fecha_ult) if fecha_ult is not None else -1
            if im_val <= 5.0 and dias_sin_venta > 90:
                productos_criticos.append((codigo, desc, im_val, dias_sin_venta))
        
        ws_critical['A1'] = f'PRODUCTOS CRÍTICOS - IM ≤ 5% y > 90 días sin venta ({len(productos_criticos)} productos)'
        ws_critical['A1'].font = Font(size=12, bold=True, color="FFFFFF")
        ws_critical['A1'].fill = PatternFill(start_color="F44336", end_color="F44336")
        ws_critical.merge_cells('A1:E1')
        
        # Headers
        headers_criticos = ['Código', 'Descripción', 'IM %', 'Días sin venta', 'Acción Requerida']
        for col, header in enumerate(headers_criticos, 1):
            cell = ws_critical.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF9800", end_color="FF9800")
        
        # Datos críticos
        for i, (codigo, desc, im_val, dias_sin_venta) in enumerate(productos_criticos, 4):
            ws_critical.cell(row=i, column=1, value=clean_for_excel(codigo))
            ws_critical.cell(row=i, column=2, value=clean_for_excel(desc))
            ws_critical.cell(row=i, column=3, value=round(im_val, 2))
            ws_critical.cell(row=i, column=4, value=dias_sin_venta)
            
            # Acción sugerida según IM y días sin venta
            if dias_sin_venta > 180:
                accion = "DESCONTINUAR / LIQUIDAR"
            elif dias_sin_venta > 90:
                accion = "REVISAR ESTRATEGIA"
            else:
                accion = "MONITOREAR"
            
            ws_critical.cell(row=i, column=5, value=accion)
        
        # Ajustar columnas
        ws_critical.column_dimensions['A'].width = 12
        ws_critical.column_dimensions['B'].width = 40
        ws_critical.column_dimensions['C'].width = 12
        ws_critical.column_dimensions['D'].width = 15
        ws_critical.column_dimensions['E'].width = 20
        
        # Guardar archivo
        wb.save(filename)
        
        # Calcular tiempo total de exportación
        export_end_time = datetime.now()
        total_time = (export_end_time - export_start_time).total_seconds()
        
        logger.info(f"Exportación MBRP Excel completada: {filas_exportadas} registros exportados (de {total_registros_entrada}) en {filename}")
        logger.info(f"[EXPORT MBRP] FIN: {export_end_time.strftime('%H:%M:%S.%f')[:-3]}")
        logger.info(f"[EXPORT MBRP] TIEMPO TOTAL: {total_time:.2f} segundos ({total_time/60:.1f} minutos)")
        
        return filas_exportadas
        
    except Exception as e:
        logger.error(f"Error en exportación MBRP Excel: {e}")
        raise
